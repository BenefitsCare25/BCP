"""Flexible-Benefits scheme configuration.

A broker uploads a heterogeneous Flex document (PDF/image/email); AI extracts the
four parameter groups (family-status tiers, limits, employee-type eligibility,
benefit statements) into one normalized ``scheme`` bag. The broker reviews/edits
it, then confirms. Confirm validates and flips status — it does NOT materialize
catalog rows or run matching (deferred to a later phase).

- POST   /policy-years/{id}/flex-scheme/extract   — upload doc → AI extract → upsert draft
- GET    /policy-years/{id}/flex-scheme            — fetch the scheme (404 if none)
- PUT    /policy-years/{id}/flex-scheme            — save edited scheme bag
- POST   /policy-years/{id}/flex-scheme/confirm    — validate + confirm
- DELETE /policy-years/{id}/flex-scheme            — discard the draft

Tenant scoping rides on `load_policy_year` (the scheme is keyed by the already
tenant-checked policy year), matching the product-setup / recommendations routers.
"""
from __future__ import annotations

import logging
import re
from dataclasses import asdict
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Request,
    Response,
    UploadFile,
    status,
)
from openpyxl import Workbook
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from starlette.concurrency import run_in_threadpool

from app.core.audit import write_audit
from app.core.auth import CurrentUser, get_current_user
from app.core.deps import (
    assert_policy_year_editable,
    load_policy_year,
    require_client_id,
)
from app.core.rate_limit import limiter
from app.core.uploads import FLEX_SUFFIXES, saved_upload
from app.db.session import get_db
from app.models import FlexScheme, PolicyYear
from app.models.flex_scheme import FlexSchemeOrigin, FlexSchemeStatus
from app.services.ai_breaker import CircuitOpenError
from app.services.ai_extractor import (
    FLEX_FAMILY_STATUS_CODES,
    AINotConfiguredError,
    AIParseError,
)
from app.services.ai_gateway import AIBudgetExceededError, extract_flex_scheme
from app.services.flex_assignment import FlexAssignmentSummary, assign_and_audit
from app.services.flex_intake import FlexIntakeError, normalize_flex_document
from app.services.flex_membership import (
    COVERAGE_PREVIEW_CAP,
    ResolvedEmployee,
    _is_catch_all,
    cap_bucket,
    compute_flex_coverage,
    compute_flex_membership,
    resolve_roster,
    roster_vocabulary,
)
from app.services.flex_pricing_resolver import _is_age
from app.services.flex_proration import proration_errors
from app.services.flex_reconcile import seed_tier_match_sets

logger = logging.getLogger(__name__)

router = APIRouter(tags=["flex-scheme"])

_CURRENCY_RE = re.compile(r"^[A-Z]{3}$")
# Top-level sections of the scheme bag, and the kind each must be. The PUT
# write-boundary guard rejects malformed sections so a non-list `tiers` can't
# persist and later white-screen the reader.
_SECTION_KINDS: dict[str, tuple[type, ...]] = {
    "meta": (dict,),
    "tiers": (list,),
    "eligibility": (dict, type(None)),
    "dependant_def": (dict, type(None)),
}
_SCHEME_SECTIONS = tuple(_SECTION_KINDS)


def _section_shape_errors(body: dict[str, Any]) -> list[str]:
    """Validate the kind of each provided top-level section (write boundary)."""
    errs: list[str] = []
    for key, kinds in _SECTION_KINDS.items():
        if key in body and not isinstance(body[key], kinds):
            want = " or ".join("null" if k is type(None) else k.__name__ for k in kinds)
            errs.append(f"'{key}' must be {want}.")
    return errs


def _reopen_as_draft(row: FlexScheme) -> None:
    """Reset a confirmed scheme back to an editable draft (idempotent)."""
    if row.status == FlexSchemeStatus.confirmed:
        row.status = FlexSchemeStatus.draft
        row.confirmed_at = None
        row.confirmed_by = None


def _tier_sig(tier: dict[str, Any]) -> str:
    """Identity of a tier for dedupe across files: country + eligibility + name."""
    et = tier.get("employee_type") if isinstance(tier.get("employee_type"), dict) else {}
    return "|".join(
        s.strip().lower()
        for s in (
            str(tier.get("country") or ""),
            str(et.get("raw") or ""),
            str(tier.get("name") or ""),
        )
    )


def _merge_schemes(existing: dict[str, Any], new: list[dict[str, Any]]) -> dict[str, Any]:
    """Merge newly-extracted schemes into the existing one.

    A complete Flex program often spans several documents (e.g. the JG8-17 tier,
    the JG18+ tier, and the per-country tables). Tiers accumulate — deduped by
    country+eligibility+name, with a re-extracted tier replacing its prior
    version — so uploading more files builds out one scheme rather than replacing
    it. Scalar meta/eligibility/dependant fields keep the existing value, else
    take the first non-empty value the new documents provide.
    """
    merged_meta = (
        dict(existing.get("meta")) if isinstance(existing.get("meta"), dict) else {}
    )
    eligibility = existing.get("eligibility")
    dependant = existing.get("dependant_def")

    tiers_by_sig: dict[str, dict[str, Any]] = {}
    for t in existing.get("tiers") or []:
        if isinstance(t, dict):
            tiers_by_sig[_tier_sig(t)] = t

    for sch in new:
        meta = sch.get("meta") if isinstance(sch.get("meta"), dict) else {}
        for k, v in meta.items():
            if v not in (None, "", []) and not merged_meta.get(k):
                merged_meta[k] = v
        if eligibility is None:
            eligibility = sch.get("eligibility")
        if dependant is None:
            dependant = sch.get("dependant_def")
        for t in sch.get("tiers") or []:
            if isinstance(t, dict):
                tiers_by_sig[_tier_sig(t)] = t

    return {
        "meta": merged_meta,
        "tiers": list(tiers_by_sig.values()),
        "eligibility": eligibility,
        "dependant_def": dependant,
    }


class FlexSchemeOut(BaseModel):
    id: str
    policy_year_id: str
    status: str
    origin: str
    scheme: dict[str, Any]
    source_ref: str | None
    confidence: float | None
    confirmed_at: datetime | None

    @classmethod
    def from_model(cls, row: FlexScheme) -> FlexSchemeOut:
        return cls(
            id=row.id,
            policy_year_id=row.policy_year_id,
            status=row.status,
            origin=row.origin,
            scheme=row.scheme or {},
            source_ref=row.source_ref,
            confidence=row.confidence,
            confirmed_at=row.confirmed_at,
        )


class FlexSchemeIn(BaseModel):
    scheme: dict[str, Any]


class FlexTierHeadcountOut(BaseModel):
    name: str
    country: str | None
    currency: str | None
    eligible: int
    by_family_status: dict[str, int]
    wallet_by_family_status: dict[str, float | None]


class FlexAssignmentOut(BaseModel):
    employee_id: str
    family_status: str | None
    source: str
    spouse_count: int
    child_count: int
    tier_name: str | None
    currency: str | None
    wallet_amount: float | None


class FlexMembershipOut(BaseModel):
    """Family-status headcounts + per-employee flex assignment for a policy year.

    ``family_status_counts`` is keyed by S/M/M1C/M2C/M3C (plus ``unknown``);
    ``source_counts`` by how each was resolved (dependants / roster / none).
    """

    employees_total: int
    family_status_counts: dict[str, int]
    source_counts: dict[str, int]
    tiers: list[FlexTierHeadcountOut]
    assignments: list[FlexAssignmentOut]
    scheme_status: str | None
    # Active employees who matched no tier, and the designations that didn't match.
    ineligible_count: int = 0
    ineligible_designations: dict[str, int] = {}
    # Active employees whose grade/designation satisfies more than one reconciled
    # tier (assigned to the first; overlap surfaced so the broker can tighten).
    ambiguous_count: int = 0
    ambiguous_examples: list[dict[str, Any]] = []


class CoverageRowOut(BaseModel):
    staff_id: str | None
    name: str | None
    label: str | None
    detail: str


class CoverageBucketOut(BaseModel):
    key: str
    label: str
    kind: str          # "employee" | "dependant"
    count: int
    rows: list[CoverageRowOut]
    truncated: bool


class FlexCoverageOut(BaseModel):
    """"Is anyone left out?" reconciliation of the flex-eligible roster.

    Every active employee lands in exactly one of {ok, no family status, not in
    any tier}; ``multiple_tiers`` is an assigned-but-review warning. Every active
    dependant lands in {classified, unclassified, orphaned, inactive link}. Each
    exception bucket carries a capped preview of WHO — the full list is in the
    ``/coverage/export`` .xlsx."""

    employees_total: int
    employees_ok: int
    dependants_total: int
    dependants_ok: int
    has_tiers: bool
    scheme_status: str | None
    buckets: list[CoverageBucketOut]
    preview_cap: int


class VocabValueOut(BaseModel):
    value: str
    count: int
    claimed: bool


class RosterVocabOut(BaseModel):
    """Distinct employee-type + job-grade values on the active roster — the
    vocabulary the flex-tier match-set pickers select from."""

    employees_total: int
    designations: list[VocabValueOut]
    grades: list[VocabValueOut]


class FlexConfirmIn(BaseModel):
    # Confirming with unmatched employees (and no catch-all tier) warns first;
    # resend with acknowledge=true to proceed anyway (they get no wallet).
    acknowledge: bool = False


class FlexAssignOut(BaseModel):
    """Outcome of persisting Flex wallets across a policy year's roster."""

    employees_total: int
    employees_assigned: int
    employees_with_status: int
    by_tier: dict[str, int]
    duration_ms: int

    @classmethod
    def from_summary(cls, s: FlexAssignmentSummary) -> FlexAssignOut:
        return cls(
            employees_total=s.employees_total,
            employees_assigned=s.employees_assigned,
            employees_with_status=s.employees_with_status,
            by_tier=s.by_tier,
            duration_ms=s.duration_ms,
        )




def _get_scheme(db: Session, policy_year_id: str) -> FlexScheme | None:
    return db.execute(
        select(FlexScheme).where(FlexScheme.policy_year_id == policy_year_id)
    ).scalar_one_or_none()


def _validate_tier(
    tier: Any, index: int, has_system_cap: bool
) -> list[str]:
    """Validate one eligibility tier; returns its errors (empty == valid)."""
    label = (tier.get("name") if isinstance(tier, dict) else None) or f"Tier {index + 1}"
    if not isinstance(tier, dict):
        return [f"{label}: malformed tier."]

    errors: list[str] = []

    # Effective currency always resolves (tier → scheme default → platform SGD),
    # so it's never "required"; only an explicitly-entered value is format-checked.
    tier_currency = str(tier.get("currency") or "").strip().upper()
    if tier_currency and not _CURRENCY_RE.match(tier_currency):
        errors.append(f"{label}: currency '{tier_currency}' is not a 3-letter ISO code.")

    emp = tier.get("employee_type") if isinstance(tier.get("employee_type"), dict) else {}
    md, mg = emp.get("match_designations"), emp.get("match_grades")
    has_desig = isinstance(md, list) and any(str(x).strip() for x in md)
    has_grade_set = isinstance(mg, list) and any(str(x).strip() for x in mg)
    lo, hi = emp.get("job_grade_min"), emp.get("job_grade_max")
    has_band = isinstance(lo, int) or isinstance(hi, int)
    has_raw = bool(str(emp.get("raw") or "").strip())
    if not (has_desig or has_grade_set or has_band or has_raw):
        errors.append(
            f"{label}: pick at least one job title or job grade for eligibility."
        )
    if isinstance(lo, int) and isinstance(hi, int) and lo > hi:
        errors.append(f"{label}: job-grade min ({lo}) is greater than max ({hi}).")

    tier_cap = tier.get("system_cap")
    has_tier_cap = isinstance(tier_cap, (int, float))
    if has_tier_cap and tier_cap < 0:
        errors.append(f"{label}: flat annual cap must be ≥ 0.")

    limits = tier.get("limits") if isinstance(tier.get("limits"), list) else []
    if not limits and not has_tier_cap and not has_system_cap:
        errors.append(
            f"{label}: needs at least one family-status limit row, or a flat annual cap."
        )
    seen: set[str] = set()
    for row in limits:
        if not isinstance(row, dict):
            errors.append(f"{label}: malformed limit row.")
            continue
        fs = str(row.get("family_status") or "")
        if fs not in FLEX_FAMILY_STATUS_CODES:
            errors.append(
                f"{label}: family status '{fs}' is not one of "
                f"{', '.join(FLEX_FAMILY_STATUS_CODES)}."
            )
        elif fs in seen:
            errors.append(f"{label}: duplicate family status '{fs}'.")
        else:
            seen.add(fs)
        amount = row.get("amount")
        if not isinstance(amount, (int, float)) or amount < 0:
            errors.append(f"{label}: limit amount for '{fs}' must be a number ≥ 0.")

    raw_cats = tier.get("benefit_categories")
    cats = raw_cats if isinstance(raw_cats, list) else []
    if not cats:
        errors.append(f"{label}: needs at least one benefit category.")
    for cat in cats:
        if not isinstance(cat, dict):
            errors.append(f"{label}: malformed benefit category.")
            continue
        if not str(cat.get("name") or "").strip():
            errors.append(f"{label}: a benefit category is missing a name.")
        if not isinstance(cat.get("claimable"), bool):
            errors.append(
                f"{label}: category '{cat.get('name') or '(unnamed)'}' "
                "must set claimable true/false."
            )
        sub = cat.get("sub_limit")
        if sub is not None and (not isinstance(sub, (int, float)) or sub < 0):
            errors.append(f"{label}: a category sub-limit must be a number ≥ 0.")

    return errors


def validate_scheme(scheme: dict[str, Any]) -> list[str]:
    """Return a list of human-readable validation errors (empty == valid)."""
    errors: list[str] = []
    meta = scheme.get("meta") if isinstance(scheme.get("meta"), dict) else {}
    tiers = scheme.get("tiers") if isinstance(scheme.get("tiers"), list) else []

    if not tiers:
        errors.append("Scheme must have at least one eligibility tier.")

    # Currency is per-tier (a scheme can span countries); meta.currency is an
    # optional default. Validate the default here if present; resolve each tier's
    # effective currency in the tier loop.
    default_currency = str(meta.get("currency") or "").strip().upper()
    if default_currency and not _CURRENCY_RE.match(default_currency):
        errors.append(f"Default currency '{default_currency}' is not a 3-letter ISO code.")

    # Effective period is optional (blank inherits the policy year window), but an
    # entered date must be a real ISO date and the window must not be inverted.
    bounds: dict[str, date | None] = {}
    for field in ("effective_start", "effective_end"):
        raw = meta.get(field)
        bounds[field] = None
        if raw in (None, ""):
            continue
        try:
            bounds[field] = date.fromisoformat(str(raw).strip())
        except ValueError:
            errors.append(
                f"Effective {'start' if field.endswith('start') else 'end'} date "
                f"'{raw}' must be an ISO date (YYYY-MM-DD)."
            )
    if (
        bounds["effective_start"] is not None
        and bounds["effective_end"] is not None
        and bounds["effective_start"] > bounds["effective_end"]
    ):
        errors.append("Effective start date must be on or before the effective end date.")

    # GST config is optional; when present, gst_included must be a boolean and
    # gst_rate a percentage in [0, 100].
    if "gst_included" in meta and not isinstance(meta.get("gst_included"), (bool, type(None))):
        errors.append("GST included must be true or false.")
    gst_rate = meta.get("gst_rate")
    if gst_rate is not None and (
        isinstance(gst_rate, bool)
        or not isinstance(gst_rate, (int, float))
        or not 0 <= gst_rate <= 100
    ):
        errors.append("GST rate must be a percentage between 0 and 100.")

    has_system_cap = isinstance(meta.get("system_cap"), (int, float))

    for i, tier in enumerate(tiers):
        errors.extend(_validate_tier(tier, i, has_system_cap))

    # Scheme-level dependant age-limit default (age next-birthday per role). This
    # feeds the real eligibility engine via get_pricing's __dep_age__ stamp; a
    # product's Flex-pricing entry can override it.
    dep_limits = meta.get("dependant_age_limits")
    if isinstance(dep_limits, dict):
        for role in ("spouse", "child"):
            win = dep_limits.get(role)
            if not isinstance(win, dict):
                continue
            lo, hi = win.get("min"), win.get("max")
            for field, val in (("min", lo), ("max", hi)):
                if val is not None and not _is_age(val):
                    errors.append(
                        f"Dependant {role} age {field} must be a non-negative whole number."
                    )
            if _is_age(lo) and _is_age(hi) and lo > hi:
                errors.append(f"Dependant {role} age min must be ≤ max.")

    return errors


@router.post("/policy-years/{policy_year_id}/flex-scheme/extract", response_model=FlexSchemeOut)
@limiter.limit("20/minute")
async def extract_flex(
    request: Request,
    policy_year_id: str,
    files: Annotated[
        list[UploadFile],
        File(description="One or more Flex benefit documents (PDF/PNG/JPG/MSG)"),
    ],
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexSchemeOut:
    """Upload one or more Flex documents, AI-extract each, and merge into the draft.

    A complete program usually spans several documents (different eligibility
    bands, per-country tables); their tiers accumulate into one scheme.
    Locked once the policy year is activated (configuration is snapshotted).
    """
    assert_policy_year_editable(policy_year)
    client_id = require_client_id(user)
    if not files:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_CONTENT, "No files uploaded.")

    new_schemes: list[dict[str, Any]] = []
    confidences: list[float] = []
    filenames: list[str] = []
    for file in files:
        suffix = Path(file.filename or "").suffix.lower()
        # Offload CPU-bound normalization (PDF raster / image decode) off the loop.
        async with saved_upload(file, FLEX_SUFFIXES) as tmp_path:
            try:
                text, images = await run_in_threadpool(
                    normalize_flex_document, tmp_path, suffix
                )
            except FlexIntakeError as exc:
                raise HTTPException(
                    status.HTTP_422_UNPROCESSABLE_CONTENT, f"{file.filename}: {exc}"
                ) from exc

        try:
            # The AI call is a blocking multi-second round-trip — offload it too.
            result = await run_in_threadpool(
                lambda text=text, images=images: extract_flex_scheme(
                    db,
                    client_id=client_id,
                    policy_year_id=policy_year_id,
                    text=text,
                    images=images,
                )
            )
        except AINotConfiguredError as exc:
            raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, str(exc)) from exc
        except AIBudgetExceededError as exc:
            raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc)) from exc
        except CircuitOpenError as exc:
            raise HTTPException(
                status.HTTP_503_SERVICE_UNAVAILABLE,
                "AI provider temporarily unavailable (circuit open). Try again shortly.",
            ) from exc
        except AIParseError as exc:
            logger.exception("Flex extraction parse failure")
            raise HTTPException(
                status.HTTP_502_BAD_GATEWAY,
                f"{file.filename}: AI returned an unparseable Flex scheme.",
            ) from exc

        new_schemes.append(result.scheme)
        conf = result.metadata.get("confidence")
        if isinstance(conf, (int, float)):
            confidences.append(float(conf))
        if file.filename:
            filenames.append(file.filename)

    row = _get_scheme(db, policy_year_id)
    # Snapshot the prior scheme so a re-extract is recoverable from the audit log.
    before: dict[str, Any] | None = None
    existing_scheme: dict[str, Any] = {}
    if row is None:
        row = FlexScheme(policy_year_id=policy_year_id)
        db.add(row)
    else:
        before = {
            "status": row.status,
            "source_ref": row.source_ref,
            "scheme": row.scheme or {},
        }
        existing_scheme = row.scheme or {}
        _reopen_as_draft(row)

    merged = _merge_schemes(existing_scheme, new_schemes)
    # Reconcile the extracted eligibility against the actual roster: pre-fill each
    # new tier's roster-anchored match sets and flag terms that map to nothing.
    seed_tier_match_sets(
        merged, roster_vocabulary(db, policy_year_id, policy_year.client_id)
    )
    row.scheme = merged
    row.origin = FlexSchemeOrigin.upload
    row.source_ref = (", ".join(filenames))[:512] or row.source_ref
    # Surface the least-confident extraction so the UI flags what to review.
    row.confidence = min(confidences) if confidences else row.confidence
    flag_modified(row, "scheme")

    write_audit(
        db, user, "flex_scheme.extract", "flex_scheme", policy_year_id,
        before=before,
        after={"files": filenames, "tiers": len(merged.get("tiers") or [])},
    )
    db.commit()
    db.refresh(row)
    return FlexSchemeOut.from_model(row)


@router.get("/policy-years/{policy_year_id}/flex-scheme", response_model=FlexSchemeOut)
def get_flex(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    db: Session = Depends(get_db),
) -> FlexSchemeOut:
    row = _get_scheme(db, policy_year_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No flex scheme for this policy year")
    return FlexSchemeOut.from_model(row)


@router.get(
    "/policy-years/{policy_year_id}/flex-scheme/membership",
    response_model=FlexMembershipOut,
)
def flex_membership(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexMembershipOut:
    """Resolve every employee's family status (from the dependant + employee
    listings) and flex-tier assignment, and aggregate the headcounts.

    Read-only and scheme-tolerant: the family-status counts don't need a scheme;
    tier headcounts are empty until one exists. Tenant scoping rides on
    ``load_policy_year``.
    """
    result = compute_flex_membership(db, policy_year_id, policy_year.client_id)
    # The dataclass fields mirror FlexMembershipOut 1:1 (nested tiers/assignments
    # included); Pydantic coerces the nested dicts into their models.
    return FlexMembershipOut(**asdict(result))


@router.get(
    "/policy-years/{policy_year_id}/flex-scheme/coverage",
    response_model=FlexCoverageOut,
)
def flex_coverage(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexCoverageOut:
    """Coverage validation: which active employees / dependants are left out of
    the flex sizing, and who they are (capped preview — full list via export).

    Read-only, scheme-tolerant (family-status checks run without a scheme; tier
    checks appear once one exists). Tenant scoping rides on ``load_policy_year``.
    """
    roster = resolve_roster(db, policy_year_id, policy_year.client_id)
    cov = compute_flex_coverage(roster)
    return FlexCoverageOut(
        employees_total=cov.employees_total,
        employees_ok=cov.employees_ok,
        dependants_total=cov.dependants_total,
        dependants_ok=cov.dependants_ok,
        has_tiers=cov.has_tiers,
        scheme_status=cov.scheme_status,
        buckets=[asdict(cap_bucket(b)) for b in cov.buckets],
        preview_cap=COVERAGE_PREVIEW_CAP,
    )


def _sanitize_sheet_title(label: str) -> str:
    """Excel sheet titles: ≤31 chars, none of []:*?/\\."""
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", label).strip()
    return (cleaned or "Sheet")[:31]


def _employee_issues(r: ResolvedEmployee, has_tiers: bool) -> str:
    issues: list[str] = []
    if r.family_status is None:
        issues.append("no family status")
    if has_tiers and r.tier_idx is None:
        issues.append("not in any tier")
    if len(r.overlap_tiers) > 1:
        issues.append("multiple tiers")
    return "; ".join(issues)


def _build_coverage_workbook(
    cov, resolved: list[ResolvedEmployee], label: str
) -> bytes:
    """Full (uncapped) coverage report: summary + one sheet per non-empty
    exception bucket + the complete resolved roster."""
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Flex coverage validation", label])
    summary.append(["Scheme status", cov.scheme_status or "—"])
    summary.append([])
    summary.append(["Employees", ""])
    summary.append(["Total active", cov.employees_total])
    summary.append(["Fully resolved & assigned", cov.employees_ok])
    summary.append(["Need attention", cov.employees_total - cov.employees_ok])
    summary.append([])
    summary.append(["Dependants", ""])
    summary.append(["Total active", cov.dependants_total])
    summary.append(["Classified (spouse/child)", cov.dependants_ok])
    summary.append(["Need attention", cov.dependants_total - cov.dependants_ok])
    summary.append([])
    summary.append(["Exception", "Count"])
    for b in cov.buckets:
        summary.append([b.label, b.count])

    for b in cov.buckets:
        if b.count == 0:
            continue
        ws = wb.create_sheet(_sanitize_sheet_title(b.label))
        if b.kind == "dependant":
            ws.append(["Employee Staff ID", "Employee Name", "Dependant", "Reason"])
        else:
            ws.append(["Staff ID", "Name", "Designation", "Reason"])
        for row in b.rows:
            ws.append([row.staff_id or "", row.name or "", row.label or "", row.detail])

    roster_ws = wb.create_sheet("Full roster")
    roster_ws.append([
        "Staff ID", "Name", "Designation", "Grade", "Nationality",
        "Family status", "Source", "Spouse", "Children",
        "Tier", "Wallet", "Currency", "Issues",
    ])
    for r in resolved:
        roster_ws.append([
            r.staff_id, r.name or "", r.designation or "", r.grade or "",
            r.nationality or "", r.family_status or "", r.source,
            r.spouse_count, r.child_count, r.tier_name or "",
            r.wallet_amount if r.wallet_amount is not None else "",
            r.currency or "", _employee_issues(r, cov.has_tiers),
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/policy-years/{policy_year_id}/flex-scheme/coverage/export")
def flex_coverage_export(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    """Download the full flex coverage report as an .xlsx (all exceptions + the
    complete resolved roster). Audited — the workbook carries employee PII."""
    roster = resolve_roster(db, policy_year_id, policy_year.client_id)
    cov = compute_flex_coverage(roster)
    label = f"{policy_year.start_date:%Y-%m-%d} to {policy_year.end_date:%Y-%m-%d}"
    content = _build_coverage_workbook(cov, roster.resolved, label)

    write_audit(
        db, user, "flex_coverage.export", "flex_scheme", policy_year_id,
        after={
            "employees_total": cov.employees_total,
            "employees_attention": cov.employees_total - cov.employees_ok,
            "dependants_attention": cov.dependants_total - cov.dependants_ok,
        },
    )
    db.commit()

    return Response(
        content=content,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="flex-coverage-{policy_year.start_date:%Y}.xlsx"'
            )
        },
    )


@router.get(
    "/policy-years/{policy_year_id}/flex-scheme/roster-vocab",
    response_model=RosterVocabOut,
)
def flex_roster_vocab(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> RosterVocabOut:
    """Distinct employee-type (designation) + job-grade values on the active
    roster, each flagged whether some tier already selects it. Powers the tier
    match-set dropdowns and their coverage hints. Tenant scoping rides on
    ``load_policy_year``."""
    vocab = roster_vocabulary(db, policy_year_id, policy_year.client_id)
    return RosterVocabOut(**asdict(vocab))


@router.post(
    "/policy-years/{policy_year_id}/flex-scheme/suggest-matches",
    response_model=FlexSchemeOut,
)
def flex_suggest_matches(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexSchemeOut:
    """Re-seed each unreconciled tier's match sets from the current roster.

    Seeding also runs automatically on extract, but the roster is often uploaded
    AFTER the flex document — this lets the broker pull suggestions in once it
    exists. Tiers already carrying match sets (reconciled by the broker) are left
    untouched. Locked once the policy year is activated."""
    assert_policy_year_editable(policy_year)
    row = _get_scheme(db, policy_year_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No flex scheme for this policy year")

    scheme = dict(row.scheme or {})
    seed_tier_match_sets(
        scheme, roster_vocabulary(db, policy_year_id, policy_year.client_id)
    )
    row.scheme = scheme
    _reopen_as_draft(row)  # a match-set change reopens a confirmed scheme as draft
    flag_modified(row, "scheme")
    write_audit(db, user, "flex_scheme.suggest_matches", "flex_scheme", policy_year_id)
    db.commit()
    db.refresh(row)
    return FlexSchemeOut.from_model(row)


@router.put("/policy-years/{policy_year_id}/flex-scheme", response_model=FlexSchemeOut)
def save_flex(
    policy_year_id: str,
    body: FlexSchemeIn,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexSchemeOut:
    """Save edited scheme answers. Shallow-merges top-level sections so a partial
    body cannot wipe a section the client didn't send."""
    assert_policy_year_editable(policy_year)
    # Fail fast at the write boundary: a malformed section (e.g. a non-list
    # `tiers`) must not persist and break later reads.
    shape_errors = _section_shape_errors(body.scheme)
    # Pro-ration decides money, so a save that carries an unrecognised basis must
    # be refused rather than silently read as "no pro-ration" — a broker would
    # believe it was on. Checked HERE and not in `validate_scheme`, so the error
    # attaches to the save that caused it: an inherited value from an older
    # extraction must never block a confirm of a scheme nobody edited
    # (test_legacy_eligibility_keys_are_tolerated).
    if "eligibility" in body.scheme:
        shape_errors = [*shape_errors, *proration_errors(body.scheme)]
    if shape_errors:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail={"message": "Malformed flex scheme.", "errors": shape_errors},
        )

    row = _get_scheme(db, policy_year_id)
    if row is None:
        row = FlexScheme(policy_year_id=policy_year_id, origin=FlexSchemeOrigin.manual)
        db.add(row)

    merged = dict(row.scheme or {})
    for key in _SCHEME_SECTIONS:
        if key in body.scheme:
            merged[key] = body.scheme[key]
    row.scheme = merged
    _reopen_as_draft(row)  # an edit reopens a confirmed scheme as a draft
    flag_modified(row, "scheme")

    write_audit(db, user, "flex_scheme.save", "flex_scheme", policy_year_id)
    db.commit()
    db.refresh(row)
    return FlexSchemeOut.from_model(row)


@router.post("/policy-years/{policy_year_id}/flex-scheme/confirm", response_model=FlexSchemeOut)
def confirm_flex(
    policy_year_id: str,
    body: FlexConfirmIn | None = None,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexSchemeOut:
    assert_policy_year_editable(policy_year)
    row = _get_scheme(db, policy_year_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No flex scheme to confirm")

    # Idempotent: an already-confirmed scheme is returned untouched. Any edit
    # reopens it as a draft (``_reopen_as_draft``), so a genuine re-confirm always
    # starts from draft — this only short-circuits a duplicate confirm (double
    # click / retry), preventing a redundant full-roster re-assignment.
    if row.status == FlexSchemeStatus.confirmed:
        return FlexSchemeOut.from_model(row)

    errors = validate_scheme(row.scheme or {})
    if errors:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail={"message": "Flex scheme is incomplete.", "errors": errors},
        )

    # Coverage guard: warn (not block) if active employees would match no tier and
    # there's no catch-all to absorb them — they'd silently get no wallet. The
    # broker resends with acknowledge=true to proceed. Skipped once acknowledged.
    if not (body and body.acknowledge):
        tiers = [t for t in ((row.scheme or {}).get("tiers") or []) if isinstance(t, dict)]
        has_catch_all = any(_is_catch_all(t) for t in tiers)
        if not has_catch_all:
            membership = compute_flex_membership(db, policy_year_id, policy_year.client_id)
            if membership.ineligible_count > 0:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    detail={
                        "code": "unmatched_employees",
                        "message": (
                            f"{membership.ineligible_count} active employee(s) match "
                            "no eligibility tier and would receive no flex wallet. "
                            "Add or widen a tier, add a catch-all, or confirm anyway."
                        ),
                        "ineligible_count": membership.ineligible_count,
                        "ineligible_designations": membership.ineligible_designations,
                    },
                )

    row.status = FlexSchemeStatus.confirmed
    row.confirmed_at = datetime.now(tz=UTC)
    row.confirmed_by = user.user_id
    write_audit(db, user, "flex_scheme.confirm", "flex_scheme", policy_year_id)
    # Confirming a scheme immediately materializes each eligible employee's wallet
    # onto the roster, so the benefit statement reflects Flex without a separate
    # step (mirrors product-setup confirm re-running insured matching).
    assign_and_audit(
        db, user, policy_year_id, policy_year.client_id, trigger="confirm"
    )
    db.commit()
    db.refresh(row)
    return FlexSchemeOut.from_model(row)


@router.post(
    "/policy-years/{policy_year_id}/flex-scheme/assign", response_model=FlexAssignOut
)
@limiter.limit("10/minute")
def assign_flex(
    request: Request,
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> FlexAssignOut:
    """(Re-)assign Flex wallets across the roster from the confirmed scheme.

    Resolves every employee's family status + eligibility tier and persists the
    resulting wallet onto the employee rows. Requires a confirmed scheme — wallets
    are only meaningful once the configuration is locked. Use after the roster or
    dependant listing changes to refresh the snapshot.
    """
    row = _get_scheme(db, policy_year_id)
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No flex scheme for this policy year")
    if row.status != FlexSchemeStatus.confirmed:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            "Confirm the flex scheme before assigning wallets.",
        )

    summary = assign_and_audit(db, user, policy_year_id, policy_year.client_id)
    db.commit()
    return FlexAssignOut.from_summary(summary)


@router.delete("/policy-years/{policy_year_id}/flex-scheme", status_code=status.HTTP_204_NO_CONTENT)
def discard_flex(
    policy_year_id: str,
    policy_year: PolicyYear = Depends(load_policy_year),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> None:
    assert_policy_year_editable(policy_year)
    row = _get_scheme(db, policy_year_id)
    if row is not None:
        db.delete(row)
        write_audit(db, user, "flex_scheme.discard", "flex_scheme", policy_year_id)
        db.commit()

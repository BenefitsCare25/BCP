"""Composite reports — one workbook, several NAMED sheets.

A broker's filing problem was never that a report was missing. It was that one
submission arrived as five files called ``Sheet1``: the roster went out as an
employee listing, a dependant listing, two coverage exports and a portal-access
export, and reconciling them meant opening five windows and remembering which
five. The zip "report sets" that preceded this module existed only because a
workbook had not been allowed more than one sheet — a zip of five workbooks and
a workbook of five sheets carry the same bytes, but only one of them can be
read, filtered and cross-referenced in place, and only one keeps the sheet
names attached to the file after it is emailed on.

**Every sheet here is GRAFTED from the existing single-sheet builder, never
reimplemented.** ``graft`` copies a built worksheet's values into a named sheet
of the composite. That is the whole mechanism, and it is deliberate: the
composite is then provably the same rows as the standalone file, so the two can
never disagree about a member's cover — which is the single thing a broker uses
them together to check. Re-deriving a sheet's rows in a second place is the
defect this codebase keeps finding (two flex panels that could disagree, one
figure computed twice); a copy cannot drift.

Grafting rather than refactoring every builder to write into a supplied
worksheet is the same decision made the cheap way. The builders' output is
plain rows plus a bold first row plus column widths — ``_bold_header`` and
``_autosize`` are the entirety of their styling — so a value copy followed by
re-applying both is faithful, and no builder signature changes.

Cell values are copied VERBATIM. `safe_cell`'s formula-injection guard was
applied when the source row was written, so re-running `append_safe` over an
already-escaped value would double the apostrophe and corrupt it.
"""
from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models import PolicyYear
from app.services.insurer_reports import autosize, bold_header

# Excel's own limit, and the characters it forbids in a sheet name. A title that
# breaks either raises deep inside openpyxl at save time — i.e. after the whole
# report has been built — so it is clamped at the point the name is chosen.
_MAX_TITLE = 31
_ILLEGAL_TITLE_CHARS = set(r"[]:*?/\\")


def safe_title(title: str) -> str:
    cleaned = "".join(" " if c in _ILLEGAL_TITLE_CHARS else c for c in title)
    return cleaned.strip()[:_MAX_TITLE] or "Sheet"


def graft(target: Workbook, title: str, source: Workbook) -> Worksheet:
    """Copy ``source``'s active worksheet into ``target`` under ``title``."""
    src = source.active
    ws = target.create_sheet(safe_title(title))
    for row in src.iter_rows(values_only=True):
        # Verbatim — already formula-guarded by the source builder.
        ws.append(row)
    bold_header(ws)
    autosize(ws)
    # Header stays visible while a 650-row roster scrolls. Worth doing here and
    # not in the single-file builders: a composite is the one a broker actually
    # reads on screen rather than re-imports.
    ws.freeze_panes = "A2"
    return ws


@dataclass(frozen=True)
class BuildContext:
    """Everything a sheet might be scoped by, in one argument.

    A single context rather than a positional per parameter: most sheets ignore
    most of it (only the insurer listings read ``insurer``, only the activity
    sheets read the dates), and a uniform signature is what keeps the composer
    from needing a branch per sheet. Adding a future scope is then a field here
    and nothing else.
    """

    masked: bool = True
    insurer: str | None = None
    start: date | None = None
    end: date | None = None
    # Which slice of the roster the member listings cover ("all" | "active").
    # Carried here rather than defaulted inside the sheet because it is the
    # incumbent's own report parameter and the two populations differ by more
    # than a hundred rows on a real roster — dropping it in the composite would
    # silently narrow a report a broker reconciles against theirs.
    employee_status: str = "all"


@dataclass(frozen=True)
class SheetSpec:
    """One sheet of a composite.

    ``build`` takes (db, py, ctx) and returns a single-sheet Workbook — the
    SAME builder the standalone download uses, never a reimplementation.
    """

    title: str
    build: Callable[..., Workbook]
    # What the sheet holds, in one line. SERVED to the Reports page so a broker
    # knows what is inside a workbook before downloading it — the reason this
    # is data and not a frontend constant is that a sheet added here must not
    # need a matching edit in TypeScript to be described.
    description: str = ""


@dataclass(frozen=True)
class WorkbookSpec:
    key: str
    label: str
    description: str
    sheets: list[SheetSpec] = field(default_factory=list)
    # Produced once PER INSURER. The endpoint 400s without one rather than
    # emitting a workbook whose coverage columns are all blank, which reads as
    # "there is nothing to submit".
    requires_insurer: bool = False
    # NRIC/FIN masking applies. False where no sheet carries an identification
    # number, so the Reports page can omit a toggle that would govern nothing.
    supports_masking: bool = True
    # Bounded by a date range instead of the benefit year (the activity logs).
    supports_date_range: bool = False
    # Honours the `all | active` roster slice. Declared so the page offers the
    # control on exactly the workbook it changes, rather than above a tab where
    # it would read as scoping reports it does nothing to.
    supports_employee_status: bool = False


# ── Sheet builders ───────────────────────────────────────────────────────────
#
# Imported inside each function, exactly as the zip bundles did: the report
# modules pull in the whole pricing/underwriting/statement stack between them,
# and importing that at module scope would put it on the API's startup path for
# a feature most requests never touch.


def _employee_listing(db, py, ctx: BuildContext):
    from app.services.insurer_listings import build_employee_listing

    return build_employee_listing(db, py, ctx.insurer, masked=ctx.masked)


def _dependant_listing(db, py, ctx: BuildContext):
    from app.services.insurer_listings import build_dependant_listing

    return build_dependant_listing(db, py, ctx.insurer, masked=ctx.masked)


def _benefit_selection(db, py, ctx: BuildContext):
    from app.services.insurer_reports import build_benefit_selection_workbook

    return build_benefit_selection_workbook(db, py, masked=ctx.masked)


def _built_in_employees(db, py, ctx: BuildContext):
    from app.services.built_in_listings import build_built_in_employee_listing

    return build_built_in_employee_listing(
        db, py, masked=ctx.masked, employee_status=ctx.employee_status
    )


def _built_in_dependants(db, py, ctx: BuildContext):
    from app.services.built_in_listings import build_built_in_dependant_listing

    return build_built_in_dependant_listing(
        db, py, masked=ctx.masked, employee_status=ctx.employee_status
    )


def _employee_coverage(db, py, ctx: BuildContext):
    from app.services.roster_reports import build_employee_report_workbook

    return build_employee_report_workbook(db, py.id)


def _dependant_coverage(db, py, ctx: BuildContext):
    from app.services.roster_reports import build_dependant_report_workbook

    return build_dependant_report_workbook(db, py.id)


def _portal_access(db, py, ctx: BuildContext):
    from app.services.portal_access_report import build_portal_access_workbook

    return build_portal_access_workbook(db, py)


def _all_claims(db, py, ctx: BuildContext):
    from app.services.claims_reports import build_insurance_claims_workbook

    return build_insurance_claims_workbook(db, py, scope="all", masked=ctx.masked)


def _inpatient_claims(db, py, ctx: BuildContext):
    from app.services.claims_reports import build_insurance_claims_workbook

    return build_insurance_claims_workbook(
        db, py, scope="inpatient", masked=ctx.masked
    )


def _outpatient_claims(db, py, ctx: BuildContext):
    from app.services.claims_reports import build_insurance_claims_workbook

    return build_insurance_claims_workbook(
        db, py, scope="outpatient", masked=ctx.masked
    )


def _employee_claims(db, py, ctx: BuildContext):
    from app.services.claims_reports import build_employee_claims_workbook

    return build_employee_claims_workbook(db, py, masked=ctx.masked)


def _adjudication(db, py, ctx: BuildContext):
    from app.services.claims_register import build_claims_register_workbook

    return build_claims_register_workbook(db, py)


def _wallet_summary(db, py, ctx: BuildContext):
    from app.services.flex_ledger import build_utilisation_summary_workbook

    return build_utilisation_summary_workbook(db, py, masked=ctx.masked)


def _wallet_ledger(db, py, ctx: BuildContext):
    from app.services.flex_ledger import build_utilisation_workbook

    return build_utilisation_workbook(db, py, masked=ctx.masked)


def _leaver_summary(db, py, ctx: BuildContext):
    from app.services.leaver_reports import build_leaver_summary_workbook

    return build_leaver_summary_workbook(db, py, masked=ctx.masked)


def _leaver_details(db, py, ctx: BuildContext):
    from app.services.leaver_reports import build_leaver_details_workbook

    return build_leaver_details_workbook(db, py, masked=ctx.masked)


def _underwriting(db, py, ctx: BuildContext):
    from app.services.underwriting import adopt_orphan_cases
    from app.services.underwriting_report import build_underwriting_report

    # The same lazy adoption the queue GET performs. Rows written before the
    # insurer-grouped model carry no review, so without this their workflow
    # status exports blank — and a broker can reach this download without ever
    # having opened the queue that would otherwise have adopted them. Flush
    # only; the endpoint owns the commit, like every other service here.
    adopt_orphan_cases(db, py.id)
    return build_underwriting_report(db, py, masked=ctx.masked)


def _portal_activity(db, py, ctx: BuildContext):
    from app.services.activity_reports import build_portal_activity_workbook

    return build_portal_activity_workbook(db, py, ctx.start, ctx.end)


def _company_activity(db, py, ctx: BuildContext):
    from app.services.activity_reports import build_company_activity_workbook

    return build_company_activity_workbook(db, py, ctx.start, ctx.end)


# ── The composites ───────────────────────────────────────────────────────────

WORKBOOKS: dict[str, WorkbookSpec] = {
    "insurer-submission": WorkbookSpec(
        key="insurer-submission",
        label="Insurer Submission",
        description=(
            "Everything one insurer receives, as one workbook: the employee and "
            "dependant listings and the benefit-selection record."
        ),
        requires_insurer=True,
        sheets=[
            SheetSpec(
                "Employees",
                _employee_listing,
                "Covered employees with member IDs and sums insured.",
            ),
            SheetSpec(
                "Dependants",
                _dependant_listing,
                "Covered spouses and children, with relationship and cover.",
            ),
            SheetSpec(
                "Benefit Selection",
                _benefit_selection,
                "What each member elected, plus buy/sell leave.",
            ),
        ],
    ),
    "member-register": WorkbookSpec(
        key="member-register",
        label="Member Register",
        description=(
            "The full roster across every insurer — who is on file and what "
            "each product covers them for."
        ),
        supports_employee_status=True,
        sheets=[
            SheetSpec(
                "Employees",
                _built_in_employees,
                "Everyone on file, with each product's default plan.",
            ),
            SheetSpec(
                "Dependants",
                _built_in_dependants,
                "Every dependant on file, including those nobody covers yet.",
            ),
            # These two builders take no masking parameter — they always mask.
            # Said in the description because the workbook DOES offer an
            # unmasked pull for the two sheets above it, and a control that
            # silently governs half a file is the mislabelled-scope problem the
            # whole page was rearranged to fix. (It fails closed, so this is a
            # wording fix, not a leak.)
            SheetSpec(
                "Employee Coverage",
                _employee_coverage,
                "Matched products per employee. Always NRIC-masked.",
            ),
            SheetSpec(
                "Dependant Coverage",
                _dependant_coverage,
                "Covered dependants grouped by their employee. Always masked.",
            ),
        ],
    ),
    "claims-register": WorkbookSpec(
        key="claims-register",
        label="Claims Register",
        description=(
            "The year's claims with their full servicing history — the whole "
            "book, then split by setting, then per member."
        ),
        sheets=[
            SheetSpec(
                "All Claims",
                _all_claims,
                "Every insured claim: reference, SLA dates, insurer and payment.",
            ),
            SheetSpec(
                "Inpatient",
                _inpatient_claims,
                "Hospitalisation and day surgery, with sector and admission.",
            ),
            SheetSpec(
                "Outpatient",
                _outpatient_claims,
                "GP, specialist and dental, with the referral position.",
            ),
            SheetSpec(
                "By Employee",
                _employee_claims,
                "Insured AND flex together — a member's whole year on one page.",
            ),
            # NOT redundant with All Claims, and its own row on the Reports page
            # was deleted in the consolidation — which left the endpoint behind
            # it reachable from nothing, the failure mode
            # `docs/ORPHANED_UI_RECOVERY.md` exists for. It is the only sheet
            # carrying the claim id and the INVOICE NUMBER, which is the key a
            # broker reconciles a disputed line against.
            SheetSpec(
                "Adjudication",
                _adjudication,
                "Flat register keyed by claim id and invoice number.",
            ),
        ],
    ),
    "flex-wallet": WorkbookSpec(
        key="flex-wallet",
        label="Flex Wallet",
        description=(
            "Each member's wallet position and the dated movements behind it."
        ),
        sheets=[
            SheetSpec(
                "Summary",
                _wallet_summary,
                "One row per member: allocated, spent, in flight, left.",
            ),
            SheetSpec(
                "Ledger",
                _wallet_ledger,
                "Every dated movement: allocation, price tags, leave, claims.",
            ),
        ],
    ),
    "leavers": WorkbookSpec(
        key="leavers",
        label="Leavers",
        description=(
            "Everyone who left in the period: their final wallet position, and "
            "the claims still in flight when cover ended."
        ),
        sheets=[
            SheetSpec(
                "Summary",
                _leaver_summary,
                "Cover window and closing wallet position per leaver.",
            ),
            SheetSpec(
                "Claims",
                _leaver_details,
                "Their claims, including anything unsettled at cover end.",
            ),
        ],
    ),
    "underwriting": WorkbookSpec(
        key="underwriting",
        label="Underwriting Register",
        description=(
            "One row per life and product above the Non-Evidence Limit, with "
            "the insurer case status, decision and sums insured. Every insurer."
        ),
        sheets=[
            SheetSpec(
                "Underwriting",
                _underwriting,
                "Cases, decisions and the guaranteed / accepted sums insured.",
            ),
        ],
    ),
    "activity-access": WorkbookSpec(
        key="activity-access",
        label="Activity & Access",
        description=(
            "Who signed in, who changed what, and who can reach the portal at "
            "all — the three questions a security review opens together."
        ),
        # Bounded by a date range, not the benefit year: `auth_events` is
        # high-volume and append-only across every surface.
        supports_date_range=True,
        # No sheet here carries an identification number, so a masking toggle
        # would govern nothing. Saying so lets the page omit the control rather
        # than offer one that silently does nothing.
        supports_masking=False,
        sheets=[
            SheetSpec(
                "Portal Sign-ins",
                _portal_activity,
                "Sign-ins across every surface, including failures and lockouts.",
            ),
            SheetSpec(
                "Company Changes",
                _company_activity,
                "Configuration and administration changes, and who made them.",
            ),
            SheetSpec(
                "Portal Access",
                _portal_access,
                "Who is provisioned, unsent, or has never signed in.",
            ),
        ],
    ),
}


def spec_for(key: str) -> WorkbookSpec | None:
    return WORKBOOKS.get((key or "").strip().lower())


def build_workbook(
    db: Session,
    py: PolicyYear,
    spec: WorkbookSpec,
    ctx: BuildContext | None = None,
) -> Workbook:
    """Render every sheet of the composite into one workbook, in order.

    One source workbook is held at a time — a sheet is grafted and its source
    dropped before the next is built — which is what keeps a 650-member
    roster's five sheets inside a normal request's memory.
    """
    context = ctx or BuildContext()
    wb = Workbook()
    # openpyxl seeds a blank default sheet; the first graft supplies the real
    # one, so remove it rather than leaving an empty "Sheet" in the file.
    wb.remove(wb.active)
    for sheet in spec.sheets:
        graft(wb, sheet.title, sheet.build(db, py, context))
    return wb


def _slug(value: str) -> str:
    cleaned = "".join(c if c.isalnum() else "-" for c in (value or "").lower())
    return "-".join(p for p in cleaned.split("-") if p)


def workbook_filename(
    spec: WorkbookSpec,
    py: PolicyYear,
    insurer: str | None = None,
    today: date | None = None,
) -> str:
    """Benefit year AND export date in the name, so two years' copies of the
    same report never collide in a downloads folder."""
    parts = [spec.key]
    if spec.requires_insurer and insurer:
        parts.append(_slug(insurer))
    parts += [str(py.year), f"{(today or date.today()):%Y%m%d}"]
    return "-".join(parts) + ".xlsx"

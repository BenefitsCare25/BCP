"""Report version retention + movement diffs.

Persists a generated Reports Center document as an immutable (versioned) or
supersede-in-place (latest) record — bytes in the storage backend, metadata +
membership manifest in ``report_versions``. See ``report_registry.py`` for the
per-report classification and ``models/report_version.py`` for the row.
"""
from __future__ import annotations

import hashlib
import json
import zipfile
from collections.abc import Callable, Sequence
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.auth import CurrentUser
from app.core.pagination import MAX_LIMIT
from app.core.storage import (
    MAX_REPORT_BYTES,
    REPORT_SUFFIXES,
    document_path,
    get_storage,
)
from app.db.base import new_uuid
from app.models import (
    Category,
    Dependant,
    Employee,
    Enrollment,
    EnrollmentElection,
    EnrollmentWindow,
    LeaveElection,
    Plan,
    PolicyYear,
    User,
)
from app.models.report_version import (
    MODE_LATEST,
    ReportVersion,
)
from app.services.insurer_listings import membership_manifest
from app.services.insurer_reports import append_safe, autosize
from app.services.report_registry import (
    REGISTRY,
    ReportSpec,
    build_report_bytes,
    mime_for,
    scope_key_for,
    spec_for,
)

# Tables whose changes make a report stale, beyond the shared roster + config
# set (Employee/Dependant/Category/Plan). Keyed by report_type.
_EXTRA_STALENESS_MODELS = {
    "benefit_selection": (
        Enrollment,
        EnrollmentElection,
        LeaveElection,
        EnrollmentWindow,
    ),
}


# Retries when a concurrent submission takes the version number first. Two is
# ample: each attempt re-reads the series max, so a third collision means real
# contention, not the read-then-write window.
_VERSION_NO_ATTEMPTS = 3


class ReportTooLargeError(Exception):
    """The generated report exceeds MAX_REPORT_BYTES."""


def _manifest_hash(manifest: dict[str, Any] | None) -> str:
    members = sorted((manifest or {}).get("members", []), key=lambda m: m["key"])
    payload = json.dumps(members, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


# NOTE: there is deliberately no manifest-based dedup key any more. It hashed
# only member identity, so it could not see the underwriting/salary-derived
# columns the listing renders and wrongly reported "unchanged".
#
# Masking is not in the key either — masked and unmasked bytes differ, so
# `_content_signature` separates them on its own. What it does NOT do is stop
# them displacing each other, which is why the no-op guard compares against the
# newest version sharing this masking choice (`latest_comparable`) rather than
# the newest version outright.


# The only OOXML package part openpyxl / python-docx stamp with a write
# timestamp (dcterms:created/modified); everything else serialises
# deterministically for identical data. Excluding it makes the package content
# a stable fingerprint.
_VOLATILE_OOXML = frozenset({"docProps/core.xml"})


def _ooxml_signature(blob_bytes: bytes) -> bytes:
    """Stable digest of one XLSX/DOCX package, excluding its timestamps."""
    with zipfile.ZipFile(BytesIO(blob_bytes)) as package:
        digest = hashlib.sha256()
        for name in sorted(package.namelist()):
            if name in _VOLATILE_OOXML:
                continue
            digest.update(name.encode("utf-8"))
            digest.update(b"\x00")
            digest.update(package.read(name))
        return digest.digest()


def _content_signature(spec: ReportSpec, blob_bytes: bytes) -> str | None:
    """A stable, data-only fingerprint of a non-listing artifact, used to skip a
    no-op "save". Hashes the package's data parts directly — deterministic for
    identical data once the volatile ``docProps/core.xml`` timestamp is excluded
    — which is stable AND avoids re-parsing the freshly built document (openpyxl
    ``load_workbook``) on the save/download hot path. Masking is captured
    naturally (the masked NRIC lives in the cells). Returns None when no stable
    signature can be computed."""
    if spec.fmt not in ("xlsx", "docx", "zip"):
        return None
    try:
        if spec.fmt != "zip":
            return _ooxml_signature(blob_bytes).hex()
        with zipfile.ZipFile(BytesIO(blob_bytes)) as archive:
            digest = hashlib.sha256()
            for name in sorted(archive.namelist()):
                digest.update(name.encode("utf-8"))
                digest.update(b"\x00")
                member = archive.read(name)
                digest.update(
                    _ooxml_signature(member)
                    if name.casefold().endswith(".xlsx")
                    else member
                )
            return digest.hexdigest()
    except (KeyError, zipfile.BadZipFile):
        return None


def _summary(
    spec: ReportSpec,
    manifest: dict[str, Any] | None,
    params: dict[str, Any],
) -> dict[str, Any]:
    out: dict[str, Any] = {"masked": bool(params.get("masked", True))}
    if manifest is not None:
        members = manifest.get("members", [])
        out["member_count"] = len(members)
        out["employee_count"] = sum(1 for m in members if m["role"] == "employee")
        out["dependant_count"] = sum(1 for m in members if m["role"] == "dependant")
        out["manifest_hash"] = _manifest_hash(manifest)
    return out


def _default_filename(
    report_type: str, scope_key: str | None, version_no: int, fmt: str
) -> str:
    stem = report_type.replace("_", "-")
    if scope_key:
        safe = "".join(c if c.isalnum() else "-" for c in scope_key).strip("-")
        stem = f"{stem}-{safe}"
    return f"{stem}-v{version_no}.{fmt}"


def latest_version(
    db: Session, py: PolicyYear, report_type: str, scope_key: str | None
) -> ReportVersion | None:
    return db.execute(
        select(ReportVersion)
        .where(
            ReportVersion.client_id == py.client_id,
            ReportVersion.policy_year_id == py.id,
            ReportVersion.report_type == report_type,
            ReportVersion.scope_key.is_(scope_key)
            if scope_key is None
            else ReportVersion.scope_key == scope_key,
        )
        .order_by(ReportVersion.version_no.desc())
        .limit(1)
    ).scalar_one_or_none()


def list_versions(
    db: Session, py: PolicyYear, report_type: str, scope_key: str | None
) -> list[ReportVersion]:
    stmt = (
        select(ReportVersion)
        .where(
            ReportVersion.client_id == py.client_id,
            ReportVersion.policy_year_id == py.id,
            ReportVersion.report_type == report_type,
        )
        .order_by(ReportVersion.version_no.desc())
        .limit(MAX_LIMIT)
    )
    # A None scope_key means "the unscoped series", exactly as `latest_version`
    # and `previous_version` read it — NOT "every scope". Treating it as no
    # filter interleaved every insurer's series into one list, so the history
    # showed repeated version_no values that looked like duplicates.
    stmt = stmt.where(
        ReportVersion.scope_key.is_(None)
        if scope_key is None
        else ReportVersion.scope_key == scope_key
    )
    return list(db.execute(stmt).scalars().all())


def latest_comparable(
    db: Session,
    py: PolicyYear,
    report_type: str,
    scope_key: str | None,
    params: dict[str, Any],
) -> ReportVersion | None:
    """The newest version in the series produced with the SAME masking choice.

    The no-op guard compares against this, not against the newest version
    outright, and the difference only started to matter once retention moved
    onto the download. Masked and unmasked bytes always differ, so a broker
    flipping the NRIC toggle twice — with no data change at all — minted v1, v2
    and v3. Comparing like with like means the second masked pull recognises v1
    and stores nothing.

    The version NUMBER still comes from the series as a whole: one chronological
    record of what was sent, in the order it was sent.
    """
    wanted = bool(params.get("masked", True))
    for rv in list_versions(db, py, report_type, scope_key):
        if bool((rv.summary or {}).get("masked", True)) == wanted:
            return rv
    return None


def prune_series(
    db: Session, py: PolicyYear, report_type: str, scope_key: str | None, keep: int
) -> list[str]:
    """Drop all but the newest ``keep`` versions of a series.

    Returns the storage paths of the deleted blobs — the CALLER deletes them
    only after its commit, for the same reason `create_version` defers a
    superseded blob: a file removed before the transaction lands is orphaned
    from a row that still references it if the transaction rolls back.

    Retention was unbounded before the download became the retention event, when
    a version cost somebody a deliberate click. It now costs a data change, so a
    year of weekly roster edits on one insurer would keep every intermediate
    copy of a workbook carrying the whole roster's NRICs.
    """
    versions = list_versions(db, py, report_type, scope_key)
    doomed = versions[keep:]
    paths = [rv.storage_path for rv in doomed]
    for rv in doomed:
        db.delete(rv)
    return paths


def previous_version(db: Session, rv: ReportVersion) -> ReportVersion | None:
    """The version immediately below ``rv`` in its series (default movement
    baseline = "since the last submission")."""
    return db.execute(
        select(ReportVersion)
        .where(
            ReportVersion.client_id == rv.client_id,
            ReportVersion.policy_year_id == rv.policy_year_id,
            ReportVersion.report_type == rv.report_type,
            ReportVersion.scope_key.is_(rv.scope_key)
            if rv.scope_key is None
            else ReportVersion.scope_key == rv.scope_key,
            ReportVersion.version_no < rv.version_no,
        )
        .order_by(ReportVersion.version_no.desc())
        .limit(1)
    ).scalar_one_or_none()


def create_version(
    db: Session,
    user: CurrentUser,
    py: PolicyYear,
    report_type: str,
    params: dict[str, Any],
    label: str | None = None,
    blob_bytes: bytes | None = None,
) -> tuple[ReportVersion, bool, str | None]:
    """Generate the report, retain the bytes, and record a version row. Returns
    ``(version, created, superseded_blob_path)`` — ``created`` is False when the
    content is identical to the latest version (no-op guard), in which case the
    existing version is returned unchanged. ``superseded_blob_path`` is the
    storage path of a latest-mode predecessor whose ROW was deleted in this
    transaction; the CALLER must physically delete that blob only AFTER it
    commits (deleting earlier would orphan a still-referenced file on rollback).

    Versioned reports append a new ``version_no``; latest-mode reports supersede
    the prior row + blob. Caller owns the audit write + commit.
    """
    spec = spec_for(report_type)
    # Enforce the retained-report format allowlist (defense in depth: a report
    # spec must produce a known document type, never an arbitrary blob).
    if f".{spec.fmt}" not in REPORT_SUFFIXES:
        raise ValueError(
            f"Report format {spec.fmt!r} is not an allowed retained-report type "
            f"({sorted(REPORT_SUFFIXES)})."
        )
    scope_key = scope_key_for(spec, params)
    prior = latest_version(db, py, report_type, scope_key)
    # Numbering follows the whole series; the no-op comparison follows the same
    # masking choice — see `latest_comparable`.
    comparable = latest_comparable(db, py, report_type, scope_key, params)
    prior_hash = (comparable.summary or {}).get("content_hash") if comparable else None

    manifest = (
        membership_manifest(db, py, params["insurer"]) if spec.has_movement else None
    )

    # No-op guard: if nothing changed since the last saved version, don't create
    # a duplicate (stops "save" spam from piling up identical rows).
    #
    # This ALWAYS builds the bytes and fingerprints them, including for the
    # insurer listings. The membership manifest used to short-circuit here
    # without building the workbook, but it only carries member identity
    # (staff_id / name / member id / status / plan+grouping) — while
    # `build_employee_listing` renders ~33 further columns plus per-product
    # `Eligible / Pending U/W / Last Accepted Sum Insured` derived from
    # underwriting cases and free-cover limits. So accepting a UW case or
    # correcting a salary left the manifest identical, the save answered
    # "unchanged", and the retained record of what was submitted silently
    # diverged from what the report now produces.
    # ``blob_bytes`` is supplied when the caller has ALREADY built this artifact
    # — the download path, which retains exactly the bytes it streamed. Building
    # a second copy there would double the cost of every download of a report
    # that sweeps the whole roster, and would make the retained file a
    # re-derivation of the one that was sent rather than the one that was sent.
    if blob_bytes is None:
        blob_bytes = build_report_bytes(db, py, report_type, params)
    if len(blob_bytes) > MAX_REPORT_BYTES:
        raise ReportTooLargeError(
            f"Report is {len(blob_bytes)} bytes (max {MAX_REPORT_BYTES})."
        )

    content_hash = _content_signature(spec, blob_bytes)
    if (
        content_hash is not None
        and content_hash == prior_hash
        and comparable is not None
    ):
        # `comparable`, not `prior` — the identical bytes belong to THAT version,
        # and it is the one the download is being recorded against.
        return comparable, False, None

    next_no = (prior.version_no + 1) if prior else 1
    version_id = new_uuid()
    # Path segment must be filesystem-safe (no ':' on Windows, no separators);
    # the raw scope_key is preserved on the row, this is only the blob folder.
    scope_seg = (
        "".join(c if c.isalnum() else "-" for c in scope_key).strip("-")
        if scope_key
        else "_"
    )
    path = document_path(
        user.broker_firm_id,
        py.client_id,
        "report_version",
        f"{report_type}-{scope_seg}",
        version_id,
        f".{spec.fmt}",
    )
    saved = get_storage().save(BytesIO(blob_bytes), path)

    # Latest-mode keeps a single retained copy: drop the prior ROW now (same
    # transaction, so a failed commit restores it), but hand its blob path back
    # to the caller to delete only AFTER commit. Deleting the blob here would
    # physically remove a still-referenced file if the caller's commit fails.
    superseded_path: str | None = None
    if spec.mode == MODE_LATEST and prior is not None:
        superseded_path = prior.storage_path
        db.delete(prior)
        db.flush()

    summary = _summary(spec, manifest, params)
    if content_hash is not None:
        summary["content_hash"] = content_hash

    # `next_no` was read before this insert, so a concurrent submission can
    # already hold it. `ix_report_versions_series` is UNIQUE precisely so that
    # surfaces as an IntegrityError instead of leaving two rows sharing a
    # number — which `previous_version`'s strict `<` then steps over, dropping a
    # whole submission out of "what changed since last time".
    #
    # The insert runs inside a SAVEPOINT so only IT rolls back: a plain
    # `db.rollback()` here would discard whatever the CALLER had already done in
    # this transaction. The blob is saved once and reused across attempts, so a
    # retry costs a row insert and not another upload.
    for _ in range(_VERSION_NO_ATTEMPTS):
        rv = ReportVersion(
            id=version_id,
            client_id=py.client_id,
            policy_year_id=py.id,
            report_type=report_type,
            scope_key=scope_key,
            version_no=next_no,
            mode=spec.mode,
            label=label,
            params=dict(params),
            summary=summary,
            manifest=manifest,
            file_name=_default_filename(report_type, scope_key, next_no, spec.fmt),
            mime_type=mime_for(spec.fmt),
            size_bytes=saved.size_bytes,
            sha256=saved.sha256,
            storage_path=saved.path,
            generated_by_user_id=user.user_id,
        )
        try:
            with db.begin_nested():
                db.add(rv)
                db.flush()
            return rv, True, superseded_path
        except IntegrityError:
            db.expunge(rv)
            if _series_holds(db, py, report_type, scope_key, next_no) is None:
                # Not a version_no clash — some other constraint failed, and
                # swallowing it into a retry would loop on a real fault.
                raise
            taken = latest_version(db, py, report_type, scope_key)
            next_no = (taken.version_no + 1) if taken else 1
    raise RuntimeError(
        f"Could not allocate a version number for {report_type!r} after "
        f"{_VERSION_NO_ATTEMPTS} attempts."
    )


def _series_holds(
    db: Session,
    py: PolicyYear,
    report_type: str,
    scope_key: str | None,
    version_no: int,
) -> ReportVersion | None:
    """The row occupying ``version_no``, if the number really was taken.

    Re-raising when it was not is what stops an unrelated constraint failure
    being swallowed into a retry loop.
    """
    return db.execute(
        select(ReportVersion).where(
            ReportVersion.client_id == py.client_id,
            ReportVersion.policy_year_id == py.id,
            ReportVersion.report_type == report_type,
            ReportVersion.scope_key.is_(scope_key)
            if scope_key is None
            else ReportVersion.scope_key == scope_key,
            ReportVersion.version_no == version_no,
        )
    ).scalars().first()


def load_version_blob(rv: ReportVersion) -> bytes:
    return get_storage().read(rv.storage_path)


def _max_data_change(
    db: Session,
    py: PolicyYear,
    extra_models: tuple[Any, ...] = (),
) -> datetime | None:
    """Newest ``updated_at`` across the roster + config rows that feed the
    reports (plus any report-specific ``extra_models``). ADC inserts (new
    ``created_at`` == ``updated_at``) and soft-terminations (status/
    terminated_effective bump ``updated_at``) both move this, so it is a cheap
    gate for "did anything change since the version".
    """
    times: list[datetime] = []
    for model in (Employee, Dependant, Category, Plan, *extra_models):
        t = db.execute(
            select(func.max(model.updated_at)).where(model.policy_year_id == py.id)
        ).scalar()
        if t is not None:
            times.append(t)
    return max(times) if times else None


def is_stale(db: Session, py: PolicyYear, rv: ReportVersion) -> bool:
    """True when live data changed after the retained version, so a new one is due.

    Deliberately errs toward "stale". It used to narrow the answer for the
    insurer listings by re-hashing the membership manifest — but the manifest
    covers only member identity, not the underwriting- and salary-derived
    columns the listing renders, so an accepted UW case reported "up to date"
    for a document that had genuinely changed. A wrong "up to date" on a record
    of what was submitted to an insurer is the dangerous direction; a wrong
    "update available" merely offers a save, and `create_version` fingerprints
    the real bytes and returns `unchanged` if nothing moved.
    """
    extra = _EXTRA_STALENESS_MODELS.get(rv.report_type, ())
    latest_change = _max_data_change(db, py, extra)
    return latest_change is not None and latest_change > rv.created_at


def actor_names(db: Session, versions: list[ReportVersion]) -> dict[str, str]:
    """Display name per `generated_by_user_id`, for the history list.

    The id has always been stored and served and was rendered nowhere, so "who
    sent this to the insurer" — the first question asked of a submission record
    — was answerable only by looking up a UUID by hand.
    """
    ids = {rv.generated_by_user_id for rv in versions if rv.generated_by_user_id}
    if not ids:
        return {}
    rows = db.execute(
        select(User.id, User.display_name, User.email).where(User.id.in_(ids))
    ).all()
    return {uid: (name or email or uid) for uid, name, email in rows}


def report_status(
    db: Session, py: PolicyYear, report_type: str, scope_key: str | None
) -> dict[str, Any]:
    """Drive the UI: the latest retained version (if any) + whether it is stale."""
    spec = spec_for(report_type)
    rv = latest_version(db, py, report_type, scope_key)
    if rv is None:
        return {"latest": None, "is_stale": False, "has_movement": spec.has_movement}
    return {
        # WITH the actor's name: "last sent by whom" is half of what the status
        # line says, and resolving it in the list endpoint alone would leave the
        # one place it is always read showing a bare UUID.
        "latest": version_out(rv, actor_names(db, [rv])),
        "is_stale": is_stale(db, py, rv),
        "has_movement": spec.has_movement,
    }


def version_out(
    rv: ReportVersion, names: dict[str, str] | None = None
) -> dict[str, Any]:
    return {
        "id": rv.id,
        "report_type": rv.report_type,
        "report_label": REGISTRY[rv.report_type].label
        if rv.report_type in REGISTRY
        else rv.report_type,
        "scope_key": rv.scope_key,
        "version_no": rv.version_no,
        "mode": rv.mode,
        "label": rv.label,
        "file_name": rv.file_name,
        "size_bytes": rv.size_bytes,
        "summary": rv.summary,
        "generated_by_user_id": rv.generated_by_user_id,
        "generated_by": (names or {}).get(rv.generated_by_user_id or "") or None,
        "created_at": rv.created_at.isoformat() if rv.created_at else None,
    }


# ── Movement report (insurer listings) ──────────────────────────────────────

_TERMINATED = "terminated"
_CHANGE_FIELDS = ("member_id", "relationship")


def _newly_terminated(old: dict[str, Any], new: dict[str, Any]) -> bool:
    """A member that stayed on the listing but just went terminated — a mid-year
    leaver reads as a DELETION on the movement report, not a field change (they
    remain on the full listing for the insurer to off-bill)."""
    return new.get("status") == _TERMINATED and old.get("status") != _TERMINATED


def _member_diffs(old: dict[str, Any], new: dict[str, Any]) -> list[str]:
    diffs = []
    for f in _CHANGE_FIELDS:
        if old.get(f) != new.get(f):
            diffs.append(f"{f}: {old.get(f) or '—'} → {new.get(f) or '—'}")
    if old.get("coverage") != new.get("coverage"):
        diffs.append("coverage changed")
    return diffs


def _baseline_and_target(
    db: Session, rv: ReportVersion, since: ReportVersion | str | None
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    """Resolve (old_members, new_members, old_label, new_label) for a movement.

    ``since`` is a prior ReportVersion → diff(since → rv); the string ``"live"``
    → diff(rv → current roster) ("changes since this version"); ``None`` → diff
    from an empty baseline → rv (initial submission, everything is an addition).
    """
    insurer = (rv.params or {}).get("insurer") or ""
    target = (rv.manifest or {}).get("members", [])
    if since == "live":
        py = db.get(PolicyYear, rv.policy_year_id)
        live = membership_manifest(db, py, insurer).get("members", []) if py else []
        return target, live, f"v{rv.version_no}", "current roster"
    if since is None:
        return [], target, "initial", f"v{rv.version_no}"
    if isinstance(since, str):
        raise ValueError(f"Unsupported movement baseline {since!r}.")
    prior = (since.manifest or {}).get("members", [])
    return prior, target, f"v{since.version_no}", f"v{rv.version_no}"


def compute_movement(
    db: Session, rv: ReportVersion, since: ReportVersion | str | None
) -> Workbook:
    """Movement (adds/deletions/changes) of ``rv`` versus a baseline (see
    ``_baseline_and_target`` for the ``since`` semantics)."""
    insurer = (rv.params or {}).get("insurer") or ""
    old_members, new_members, old_label, new_label = _baseline_and_target(db, rv, since)

    old_by = {m["key"]: m for m in old_members}
    new_by = {m["key"]: m for m in new_members}
    common = old_by.keys() & new_by.keys()
    newly_terminated = {k for k in common if _newly_terminated(old_by[k], new_by[k])}

    additions = [new_by[k] for k in new_by.keys() - old_by.keys()]
    # Gone from the listing, plus mid-year leavers (still listed but now
    # terminated) — both are removals from the insurer's active cover.
    deletions = [old_by[k] for k in old_by.keys() - new_by.keys()]
    deletions += [new_by[k] for k in newly_terminated]
    changes = [
        (new_by[k], _member_diffs(old_by[k], new_by[k]))
        for k in common - newly_terminated
        if _member_diffs(old_by[k], new_by[k])
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Movement"
    append_safe(ws, [f"Movement report — {insurer or 'insurer'} ({old_label} → {new_label})"])
    append_safe(ws, [])

    def _section(
        title: str,
        rows: list[Any],
        cols: Sequence[object],
        to_row: Callable[[Any], Sequence[object]],
    ) -> None:
        append_safe(ws, [f"{title} ({len(rows)})"])
        append_safe(ws, cols)
        # Bold this section's column-header row (the one just written), not row 1.
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for r in rows:
            append_safe(ws, to_row(r))
        append_safe(ws, [])

    base_cols = ["Role", "Staff ID", "Name", "Member ID", "Relationship", "Status"]
    _section(
        "ADDITIONS", additions,
        [*base_cols, "Effective"],
        lambda m: [
            m["role"], m.get("staff_id") or "", m.get("name") or "",
            m.get("member_id") or "", m.get("relationship") or "",
            m.get("status") or "", "",
        ],
    )
    _section(
        "DELETIONS", deletions,
        [*base_cols, "Deletion Date"],
        lambda m: [
            m["role"], m.get("staff_id") or "", m.get("name") or "",
            m.get("member_id") or "", m.get("relationship") or "",
            m.get("status") or "", m.get("terminated_effective") or "",
        ],
    )
    _section(
        "CHANGES", changes,
        [*base_cols[:4], "Changed"],
        lambda pair: [
            pair[0]["role"], pair[0].get("staff_id") or "", pair[0].get("name") or "",
            pair[0].get("member_id") or "", "; ".join(pair[1]),
        ],
    )
    autosize(ws)
    return wb


def movement_summary(
    db: Session, rv: ReportVersion, since: ReportVersion | str | None
) -> dict[str, Any]:
    """Counts only (for the staleness banner) — same diff as compute_movement."""
    old_members, new_members, _ol, _nl = _baseline_and_target(db, rv, since)
    old_by = {m["key"]: m for m in old_members}
    new_by = {m["key"]: m for m in new_members}
    common = old_by.keys() & new_by.keys()
    newly_terminated = {k for k in common if _newly_terminated(old_by[k], new_by[k])}
    changed = sum(
        1 for k in common - newly_terminated if _member_diffs(old_by[k], new_by[k])
    )
    return {
        "added": len(new_by.keys() - old_by.keys()),
        "removed": len(old_by.keys() - new_by.keys()) + len(newly_terminated),
        "changed": changed,
    }

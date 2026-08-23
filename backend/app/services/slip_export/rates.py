"""The Rate section, the annual-premium total, and the per-product terms rows.

Which table a category lands in is decided by the ``rate_basis`` it actually
stores, never by its product code: a tiered rate map gets the tier grid, a
per-head rate with a separate dependant rate gets the collapsed per-plan table,
and everything else gets the flat table — whose amount/rate column headings are
themselves derived from the stored basis, so a travel policy priced as one flat
annual premium is not headed "Rate per S$1000 sum insured".

Each writer returns the sum of the premiums it actually PRINTED, so the total
below the tables is by construction the sum of what is on the page.
"""
from __future__ import annotations

import json
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.models import Category, Plan, ProductTerm
from app.services.slip_export.basis import (
    TIER_ORDER,
    insured_text,
    plan_assignments,
)
from app.services.slip_export.context import SlipContext
from app.services.slip_export.styles import (
    CENTER_WRAP,
    COUNT,
    HEADER,
    MIDDLE_WRAP,
    MONEY,
    SECTION,
    border_row,
    spacer_row,
    style_row,
)

# Rate bases whose stored premium is one figure for the WHOLE policy rather than
# a per-member or per-S$1000 rate.
_POLICY_LEVEL = frozenset({"annual_flat"})


def gst_suffix(term: ProductTerm | None) -> str:
    if term is not None and term.gst_included is True:
        return "(sbj to GST)"
    if term is not None and term.gst_included is False:
        return "(GST exempt)"
    return "(GST-exclusive)"


def _tier_codes(categories: list[Category]) -> list[str]:
    codes: set[str] = set()
    for c in categories:
        tiers = plan_assignments(c).get("rate_tiers")
        if isinstance(tiers, dict):
            codes.update(tiers)
    return sorted(codes, key=lambda c: (TIER_ORDER.get(c, 99), c))


def _write_tiered_rates(
    ws: Worksheet,
    categories: list[Category],
    codes: list[str],
    blank: bool,
    insured_default: str,
) -> float | None:
    """Render the tiered rate table; return the sum of the premiums SHOWN."""
    # Row 1: tier codes above each Rate/Premium pair; row 2: the pair headers.
    head1 = ["Rate :", "", ""]
    head2 = ["", "Insured", "Category", "Plan"]
    for code in codes:
        head1 += [code, ""]
        head2 += ["Rate", "Premium"]
    last_col = 4 + 2 * len(codes)
    ws.append([*head1, ""])
    style_row(ws, font=SECTION)
    border_row(ws, 4, last_col)
    ws.append(head2)
    style_row(ws, font=HEADER)
    border_row(ws, 2, last_col)

    running = 0.0
    found = False
    prev_insured: str | None = None
    prev_name = ""
    for c in categories:
        pa = plan_assignments(c)
        tiers = pa.get("rate_tiers")
        if not isinstance(tiers, dict) or not tiers:
            continue
        insured = insured_text(pa) or insured_default
        row: list[Any] = [
            "",
            insured if insured != prev_insured else "",
            c.display_name if c.display_name != prev_name else "",
            str(pa.get("plan_code") or ""),
        ]
        for code in codes:
            cell = tiers.get(code) or {}
            premium = cell.get("premium") if isinstance(cell, dict) else None
            if blank or not premium:
                # A stored zero premium is the parser's "no committed premium"
                # (voluntary tiers) — the slips leave those cells empty.
                rate = "" if blank else (cell.get("rate") if isinstance(cell, dict) else "")
                row += [rate, ""]
            else:
                row += [cell.get("rate"), premium]
                running += float(premium)
                found = True
        ws.append(row)
        r = style_row(ws, wrap_cols=(2, 3))
        border_row(ws, 2, last_col)
        for col in (2, 3):
            ws.cell(row=r, column=col).alignment = MIDDLE_WRAP
        for col in range(4, last_col + 1):
            ws.cell(row=r, column=col).alignment = CENTER_WRAP
        for i in range(len(codes)):
            ws.cell(row=r, column=6 + 2 * i).number_format = MONEY
        prev_insured = insured
        prev_name = c.display_name
    return running if found else None


def _derived_premium(pa: dict[str, Any], sum_insured: float | None) -> float | None:
    """SI-rated premium the slip leaves implicit: sum insured / 1,000 x rate.

    Only when the category is rated per S$1,000 SI and no stored premium exists
    (GPA-style sheets state the rate but not per-row premiums). The sum insured
    is the RESOLVED one, so a premium derived here is priced off the same
    headcount the Basis-of-Cover table published.
    """
    if pa.get("annual_premium") is not None or pa.get("rate_basis") != "per_1000_si":
        return None
    rate = pa.get("premium_rate")
    if rate is None or sum_insured is None or sum_insured == 0:
        return None
    return round(float(sum_insured) / 1000.0 * float(rate), 2)


def _flat_headers(categories: list[Category]) -> tuple[str, str, str]:
    """(amount header, rate header, amount kind) for the flat rate table.

    Read from the categories' own ``rate_basis`` / stored amounts so each
    product's table is headed by what it is actually rated on.
    """
    pas = [plan_assignments(c) for c in categories]
    if any(pa.get("estimated_annual_earnings") is not None for pa in pas):
        return "* Estimated annual earnings", "Rate", "earnings"
    bases = {pa.get("rate_basis") for pa in pas if pa.get("rate_basis")}
    if bases and bases <= _POLICY_LEVEL:
        # One flat annual premium for the whole policy — there is no amount it
        # is rated on and no per-S$1000 rate to quote.
        return "", "", "none"
    if bases == {"per_member"}:
        return "* No. of members", "Rate per member", "members"
    return "Sum Insured ( SI )", "Rate per S$1000 sum insured", "sum_insured"


def _write_flat_rates(
    ws: Worksheet,
    categories: list[Category],
    ctx: SlipContext,
    with_label: bool,
    blank: bool,
    insured_default: str,
) -> float | None:
    """Render the flat rate table; return the sum of the premiums SHOWN.

    Two premium kinds are treated differently, matching the slips:
    * Block-level (``annual_premium`` / ``premium_note``): the parser copies one
      figure onto every category in a block, so it prints once — blank
      consecutive repeats for the same insured, and count it once.
    * Derived per-S$1,000-SI: a genuine per-row figure — never blanked (two
      cohorts with equal SI still each carry it) and always summed.
    """
    amount_header, rate_header, amount_kind = _flat_headers(categories)
    if with_label:
        ws.append(["Rate :"])
        style_row(ws, font=SECTION)
    show_amount = amount_kind != "none"
    header = ["", "Insured", "Category"]
    if show_amount:
        header += [amount_header, rate_header]
    header.append("Annual Premium")
    ws.append(header)
    style_row(ws, font=HEADER)
    last_col = len(header)
    border_row(ws, 2, last_col)
    premium_col = last_col
    amount_col = 4 if show_amount else None

    prev_insured: str | None = None
    prev_display: Any = object()
    running = 0.0
    found = False
    amount_total = 0.0
    amount_found = False
    printed = 0
    for c in categories:
        pa = plan_assignments(c)
        figures = ctx.figures_for(c)
        insured = insured_text(pa) or insured_default
        if amount_kind == "earnings":
            amount: Any = pa.get("estimated_annual_earnings")
        elif amount_kind == "members":
            amount = figures.members
        elif amount_kind == "sum_insured":
            amount = figures.sum_insured
        else:
            amount = None
        # `premium` is what the cell shows; `numeric` is what it contributes to
        # the total (None when blanked or non-numeric, e.g. an annotated note).
        numeric: float | None = None
        if blank:
            rate: Any = ""
            premium: Any = ""
        else:
            rate = pa.get("premium_rate")
            derived = _derived_premium(pa, figures.sum_insured)
            if derived is not None:
                premium = derived
                numeric = derived
            else:
                # An annotated premium ("S$3,169.80 subject to minimum…") carries
                # its full text; the plain number is the fallback.
                display = pa.get("premium_note") or pa.get("annual_premium")
                stored = pa.get("annual_premium")
                if (display is not None and display == prev_display
                        and insured == prev_insured):
                    premium = ""  # block copy already printed above
                else:
                    premium = display if display is not None else ""
                    if display is not None:
                        prev_display = display
                    numeric = float(stored) if stored is not None else None
        row: list[Any] = ["", insured if insured != prev_insured else "", c.display_name]
        if show_amount:
            row += ["" if amount is None else amount, rate]
        row.append(premium)
        ws.append(row)
        r = style_row(ws, wrap_cols=(2, 3, premium_col))
        border_row(ws, 2, last_col)
        for col in (2, 3):
            ws.cell(row=r, column=col).alignment = MIDDLE_WRAP
        for col in range(4, last_col + 1):
            ws.cell(row=r, column=col).alignment = CENTER_WRAP
        if amount_col is not None and isinstance(amount, (int, float)):
            ws.cell(row=r, column=amount_col).number_format = COUNT
            amount_total += float(amount)
            amount_found = True
        if not isinstance(premium, str):
            ws.cell(row=r, column=premium_col).number_format = MONEY
        if numeric is not None:
            running += numeric
            found = True
        prev_insured = insured
        printed += 1

    # The insurer rates off the AGGREGATE cover, so state it rather than making
    # them add up the rows (the reference slips quote exactly this one figure).
    # The premium cell stays blank here — the "Annual Premium" line directly
    # below is the premium total, and printing it twice invites the two to
    # disagree after an edit.
    if printed > 1 and amount_found and amount_col is not None:
        ws.append(["", "", "Total", amount_total, "", ""][:last_col])
        r = style_row(ws, font=HEADER)
        border_row(ws, 2, last_col)
        ws.cell(row=r, column=amount_col).number_format = COUNT
    return running if found else None


def _write_per_member_rates(
    ws: Worksheet,
    categories: list[Category],
    with_label: bool,
    blank: bool,
    insured_default: str,
) -> float | None:
    """Per-member products: one rate per head, with a separate dependant rate.

    The slips price these per PLAN, not per category — the same rate row is
    replicated across every cohort sharing the plan — so collapse to one block
    per plan: "1 - Employees" / "1 - Dependents" (combined when the two rates
    match, mirroring the reference slips).
    """
    if with_label:
        ws.append(["Rate :"])
        style_row(ws, font=SECTION)
    ws.append(["", "Insured", "Plan", "Rate", "Premium"])
    style_row(ws, font=HEADER)
    border_row(ws, 2, 5)
    seen_plans: set[str] = set()
    prev_insured: str | None = None
    running = 0.0
    found = False
    for c in categories:
        pa = plan_assignments(c)
        plan = str(pa.get("plan_code") or "")
        if plan in seen_plans:
            continue
        seen_plans.add(plan)
        insured = insured_text(pa) or insured_default
        emp_rate = pa.get("premium_rate")
        dep_rate = pa.get("dependant_rate")
        stored = pa.get("annual_premium")
        premium: Any = "" if blank else (
            pa.get("premium_note") or pa.get("annual_premium") or ""
        )
        show_insured = insured if insured != prev_insured else ""
        combined = dep_rate is not None and dep_rate == emp_rate
        label = f"{plan} - Employees / Dependents" if combined else f"{plan} - Employees"
        ws.append(["", show_insured, label, "" if blank else emp_rate, premium])
        row = style_row(ws, wrap_cols=(2, 5))
        border_row(ws, 2, 5)
        ws.cell(row=row, column=2).alignment = MIDDLE_WRAP
        for col in range(3, 6):
            ws.cell(row=row, column=col).alignment = CENTER_WRAP
        if not isinstance(premium, str):
            ws.cell(row=row, column=5).number_format = MONEY
        if not blank and stored is not None:
            running += float(stored)
            found = True
        if dep_rate is not None and not combined:
            ws.append(["", "", f"{plan} - Dependents", "" if blank else dep_rate, ""])
            row = style_row(ws)
            for col in range(2, 6):
                ws.cell(row=row, column=col).alignment = CENTER_WRAP
            border_row(ws, 2, 5)
        prev_insured = insured
    return running if found else None


def _write_voluntary_rates(
    ws: Worksheet, categories: list[Category], blank: bool
) -> None:
    # Many voluntary tiers usually share ONE published rate table (the slips
    # print it once) — dedupe identical band lists; a per-category heading only
    # appears when genuinely distinct tables exist.
    blocks: list[tuple[str, list[Any]]] = []
    seen: set[str] = set()
    for c in categories:
        bands = plan_assignments(c).get("voluntary_rates")
        if not isinstance(bands, list) or not bands:
            continue
        key = json.dumps(bands, sort_keys=True, default=str)
        if key in seen:
            continue
        seen.add(key)
        blocks.append((c.display_name, bands))
    if not blocks:
        return
    spacer_row(ws)
    ws.append(["", "Voluntary Rates"])
    style_row(ws, font=SECTION)
    for name, bands in blocks:
        if len(blocks) > 1:
            ws.append(["", name])
            style_row(ws, font=HEADER)
        ws.append(["", "Based on Age Last Birthday", "Rate per 1,000 Sum assured (S$)"])
        style_row(ws, font=HEADER)
        border_row(ws, 2, 3)
        for band in bands:
            if not isinstance(band, dict):
                continue
            ws.append(["", str(band.get("label") or ""), "" if blank else band.get("rate")])
            row = style_row(ws)
            for col in range(2, 4):
                ws.cell(row=row, column=col).alignment = CENTER_WRAP
            border_row(ws, 2, 3)


def _has_rate_signal(cat: Category) -> bool:
    pa = plan_assignments(cat)
    return any(
        pa.get(k) is not None
        for k in ("premium_rate", "annual_premium", "premium_note")
    )


def _is_priceable(cat: Category, ctx: SlipContext) -> bool:
    """True when this category gives the insurer something to price.

    A premium already on file, or an AMOUNT the rate applies to. A rate on its
    own is not enough: a voluntary option states the per-member cover a member
    could elect, and until someone elects it there is no sum insured to rate —
    so it can only ever print an empty row. The slips leave those out, and
    twenty blank rows (a GPA sheet's option ladder) bury the rows that do need
    a quote.
    """
    pa = plan_assignments(cat)
    if pa.get("annual_premium") is not None or pa.get("premium_note") is not None:
        return True
    figures = ctx.figures_for(cat)
    if figures.sum_insured is not None:
        return True
    if pa.get("estimated_annual_earnings") is not None:
        return True
    # A per-head rate prices off the headcount rather than a cover amount.
    return pa.get("rate_basis") == "per_member" and bool(figures.members)


def write_rate_section(
    ws: Worksheet,
    categories: list[Category],
    term: ProductTerm | None,
    ctx: SlipContext,
    insured_default: str,
) -> None:
    blank = ctx.blank_rates
    tiered = [
        c for c in categories
        if isinstance(plan_assignments(c).get("rate_tiers"), dict)
        and plan_assignments(c).get("rate_tiers")
    ]
    # Per-member products publish a head rate + separate dependant rate, priced
    # per plan — they get the collapsed per-plan table.
    per_member = [
        c for c in categories
        if c not in tiered
        and plan_assignments(c).get("dependant_rate") is not None
        and plan_assignments(c).get("sum_insured") is None
    ]
    rest = [c for c in categories if c not in tiered and c not in per_member]
    flat = [c for c in rest if _is_priceable(c, ctx)]
    if not flat and not tiered and not per_member:
        # Nothing has an amount to rate against, but the product still carries
        # rates — show those rows rather than emitting a headless Rate section.
        flat = [c for c in rest if _has_rate_signal(c)]
    has_voluntary = any(
        isinstance(plan_assignments(c).get("voluntary_rates"), list)
        and plan_assignments(c)["voluntary_rates"]
        for c in categories
    )
    if not tiered and not per_member and not flat and not has_voluntary:
        return
    subtotals: list[float] = []
    spacer_row(ws)
    if tiered:
        t = _write_tiered_rates(ws, tiered, _tier_codes(tiered), blank, insured_default)
        if t is not None:
            subtotals.append(t)
    if per_member:
        if tiered:
            spacer_row(ws)
        t = _write_per_member_rates(
            ws, per_member, with_label=not tiered, blank=blank,
            insured_default=insured_default,
        )
        if t is not None:
            subtotals.append(t)
    if flat:
        if tiered or per_member:
            spacer_row(ws)
        t = _write_flat_rates(
            ws, flat, ctx, with_label=not (tiered or per_member), blank=blank,
            insured_default=insured_default,
        )
        if t is not None:
            subtotals.append(t)
    # A product placed purely on voluntary age-banded rates still needs the
    # section label above its table.
    if has_voluntary and not tiered and not per_member and not flat:
        ws.append(["Rate :"])
        style_row(ws, font=SECTION)
    _write_voluntary_rates(ws, categories, blank)

    spacer_row(ws)
    label = f"Annual Premium {gst_suffix(term)} :"
    total = None if blank or not subtotals else sum(subtotals)
    ws.append([label, "", total if total is not None else ""])
    r = style_row(ws, font=HEADER)
    ws.cell(row=r, column=3).alignment = CENTER_WRAP
    if total is not None:
        ws.cell(row=r, column=3).number_format = MONEY


def product_premium_total(
    categories: list[Category], ctx: SlipContext
) -> float | None:
    """The product's annual premium, computed the same way the sheet prints it.

    Used by the Overview so the summary can never disagree with the per-product
    sheet. Mirrors the writers' rules: tiered premiums sum per tier cell,
    block-level flat premiums count once per (insured, figure), and a
    per-S$1,000 premium is derived from the resolved sum insured.
    """
    if ctx.blank_rates:
        return None
    total = 0.0
    found = False
    prev_display: Any = object()
    prev_insured: str | None = None
    seen_plans: set[str] = set()
    for c in categories:
        pa = plan_assignments(c)
        tiers = pa.get("rate_tiers")
        if isinstance(tiers, dict) and tiers:
            for cell in tiers.values():
                premium = cell.get("premium") if isinstance(cell, dict) else None
                if premium:
                    total += float(premium)
                    found = True
            continue
        if pa.get("dependant_rate") is not None and pa.get("sum_insured") is None:
            plan = str(pa.get("plan_code") or "")
            if plan in seen_plans:
                continue
            seen_plans.add(plan)
            stored = pa.get("annual_premium")
            if stored is not None:
                total += float(stored)
                found = True
            continue
        if not _is_priceable(c, ctx):
            continue
        derived = _derived_premium(pa, ctx.figures_for(c).sum_insured)
        if derived is not None:
            total += derived
            found = True
            continue
        insured = insured_text(pa) or ""
        display = pa.get("premium_note") or pa.get("annual_premium")
        if display is not None and display == prev_display and insured == prev_insured:
            prev_insured = insured
            continue
        if display is not None:
            prev_display = display
        stored = pa.get("annual_premium")
        if stored is not None:
            total += float(stored)
            found = True
        prev_insured = insured
    return total if found else None


def _nel_text(term: ProductTerm | None) -> str:
    """The product's Non-Evidence Limit, as configured.

    Both gates are optional and independent: a sum-insured ceiling above which
    the insurer underwrites, and an age at/above which it underwrites regardless.
    """
    if term is None:
        return ""
    parts: list[str] = []
    if term.free_cover_limit is not None:
        parts.append(f"S${term.free_cover_limit:,.0f}")
    if term.nel_age_limit is not None:
        parts.append(f"underwriting from age {term.nel_age_limit} (ANB)")
    return "; ".join(parts)


def _shared_policy_limit(plans: list[Plan]) -> str:
    """The per-insured maximum, when every plan that states one agrees.

    Plans that differ leave it blank rather than publishing one plan's ceiling
    as though it applied to all.
    """
    limits = {
        str(p.annual_policy_limit).strip()
        for p in plans
        if str(p.annual_policy_limit or "").strip()
    }
    return next(iter(limits)) if len(limits) == 1 else ""


def term_rows(term: ProductTerm | None, plans: list[Plan]) -> list[tuple[str, str]]:
    """The slip's terms ladder, filled from configuration where it exists.

    The remaining rows are commercial terms the platform has no field for; they
    stay as labelled blanks for the broker to complete in Excel, exactly as the
    reference slips leave their unknown cells.
    """
    return [
        ("Non Evidence Limit :", _nel_text(term)),
        ("Maximum Limit Per Insured Person :", _shared_policy_limit(plans)),
        ("Experience Refund Formula / Maximum Loss Ratio :", ""),
        ("Product Rated Together :", ""),
        ("Policyholder(s) Rated Together :", ""),
    ]

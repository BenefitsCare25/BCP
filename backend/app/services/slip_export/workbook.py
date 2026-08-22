"""Sheet assembly: the Overview index plus one sheet per configured product."""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models import Category, Plan, PolicyYear, Product, ProductTerm
from app.services.matching_engine import insured_names
from app.services.product_terms import envelope_for
from app.services.slip_export.basis import (
    insured_text,
    plan_assignments,
    write_basis_of_cover,
)
from app.services.slip_export.context import Mode, SlipContext, load_context
from app.services.slip_export.header import (
    captured_answers,
    coverage_window,
    fmt_window,
    write_header_block,
)
from app.services.slip_export.rates import (
    product_premium_total,
    term_rows,
    write_rate_section,
)
from app.services.slip_export.sob import shared_cover, write_plan_details, write_sob
from app.services.slip_export.styles import (
    COUNT,
    HEADER,
    MONEY,
    TITLE,
    border_row,
    finalize_sheet,
    set_compact_product_widths,
    set_widths,
    style_row,
)

_SHEET_BAD_CHARS = set("[]:*?/\\")


def _sheet_title(base: str, taken: set[str]) -> str:
    """Excel-legal (≤31 chars, no []:*?/\\) and unique within the workbook."""
    clean = "".join(c for c in base if c not in _SHEET_BAD_CHARS).strip() or "Product"
    title = clean[:31]
    n = 2
    while title.lower() in taken:
        suffix = f" ({n})"
        title = clean[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(title.lower())
    return title


def _distinct_insured(
    categories: list[Category], product: Product | None = None
) -> list[str]:
    """Every distinct entity covered, in first-seen order.

    Mirrors the matching gate's precedence (`_build_product_indices`): the
    product's own `entities` when set, otherwise whatever the categories name.
    Without the product arm this line goes out BLANK for any product configured
    through the setup header — the per-category `insured` has no editor, so a
    manually built product never populates it.
    """
    product_names = insured_names(
        (product.product_metadata or {}).get("entities") if product else None
    )
    if product_names:
        return product_names
    seen: list[str] = []
    for c in categories:
        for name in insured_names(plan_assignments(c).get("insured")):
            if name not in seen:
                seen.append(name)
    return seen


def _product_members(categories: list[Category], ctx: SlipContext) -> int | None:
    """Lives covered by this product, counted once per member.

    A member matches exactly ONE category per product, so summing the roster's
    per-category counts is the product's headcount. Stated slip figures must not
    be added on top of them: an unmatched category's stated lives are, in
    practice, members the roster already counted under a sibling category (a
    grade the rule missed, a catch-all that absorbed them), so mixing the two
    over-states the product. Roster figures win outright when any exist; a
    product the roster never matched falls back to the stated figures alone.
    """
    live = [f.members for c in categories if (f := ctx.figures_for(c)).from_roster]
    if live:
        return sum(m for m in live if m)
    stated = [f.members for c in categories if (f := ctx.figures_for(c)).members]
    return sum(stated) if stated else None


def _write_product_sheet(
    ws: Worksheet,
    ctx: SlipContext,
    product: Product | None,
    categories: list[Category],
    plans: list[Plan],
    term: ProductTerm | None,
) -> None:
    py = ctx.policy_year
    ws.append([product.display_name if product else "Unassigned categories"])
    style_row(ws, font=TITLE)
    ws.append([])

    insured = ", ".join(_distinct_insured(categories, product))
    answers = ctx.answers_for(product)
    write_header_block(
        ws, py, product, term, answers, insured, quotation=ctx.blank_rates
    )
    ws.append([])

    if categories:
        # Per-block Insured cells fall back to the product-level entities for
        # the same reason the header line does; when configuration names none,
        # the slip's own captured wording stands in.
        default_insured = insured or captured_answers(answers).get("insured", "")
        write_basis_of_cover(ws, categories, ctx, default_insured)
        write_rate_section(ws, categories, term, ctx, default_insured)

    if product is None:
        return
    # The terms ladder belongs to the PRODUCT, not its plans — a product still
    # awaiting its schedule has a non-evidence limit and needs the remaining
    # rows as labelled blanks, so it is never gated on plans existing.
    ws.append([])
    for label, value in term_rows(term, plans):
        ws.append([label, "", value])
        row = style_row(ws)
        ws.cell(row=row, column=1).font = HEADER
    if plans:
        cover = shared_cover(plans)
        write_plan_details(ws, plans, cover)
        write_sob(
            ws,
            plans,
            cover,
            answers=answers,
            quotation=ctx.blank_rates,
        )
    set_compact_product_widths(ws)


def _write_overview(wb: Workbook, ctx: SlipContext, db_envelope: tuple[Any, ...]) -> None:
    py = ctx.policy_year
    doc_name = "Quotation Slip" if ctx.blank_rates else "Placement Slip"
    overview = wb.active
    overview.title = "Overview"
    set_widths(overview, {1: 26, 2: 40, 3: 22, 4: 28, 5: 12, 6: 8, 7: 16, 8: 18})
    overview.append([f"{doc_name} — Configured Products"])
    overview.cell(row=1, column=1).font = TITLE
    # The legal policyholder wherever a product captured it — the client record
    # holds an internal short name, which must not go out to an insurer.
    policyholder = next(
        (
            name
            for answers in ctx.answers_by_code.values()
            if (name := captured_answers(answers).get("policyholder"))
        ),
        py.client.name if py.client else "",
    )
    overview.append(["Policyholder", policyholder])
    overview.append(["Policy Year", str(py.year)])
    # Company-level period = earliest product start → latest product end (the
    # shared envelope), not the bare policy-year span — products can renew
    # off-cycle via ProductTerm coverage overrides.
    overview.append(["Period of Insurance", fmt_window(*db_envelope)])
    overview.append(["Note", "GST treatment is shown on each product sheet."])
    for r in range(2, 6):
        overview.cell(row=r, column=1).font = HEADER
    overview.append([])
    overview.append([
        "Code", "Product", "Insurer", "Coverage Period", "Categories", "Plans",
        "Members", "Annual Premium",
    ])
    for cell in overview[overview.max_row]:
        cell.font = HEADER
    border_row(overview, 1, 8)

    premium_total = 0.0
    premium_found = False
    for product in ctx.products:
        categories = ctx.cats_by_product.get(product.id, [])
        members = _product_members(categories, ctx)
        premium = product_premium_total(categories, ctx)
        overview.append([
            product.code,
            product.display_name,
            "" if ctx.blank_rates else ctx.insurer_for(product),
            coverage_window(py, ctx.terms.get(product.id)),
            len(categories),
            len(ctx.plans_by_product.get(product.id, [])),
            members if members is not None else "",
            premium if premium is not None else "",
        ])
        row = overview.max_row
        border_row(overview, 1, 8)
        if members is not None:
            overview.cell(row=row, column=7).number_format = COUNT
        if premium is not None:
            overview.cell(row=row, column=8).number_format = MONEY
            premium_total += premium
            premium_found = True

    # Members are NOT summed across products — the same person is covered by
    # several — so only the premium (a genuinely additive figure) is totalled,
    # and only when there is one to show (a quotation has none by design).
    if premium_found:
        overview.append(["Total", "", "", "", "", "", "", premium_total])
        row = overview.max_row
        style_row(overview, font=HEADER)
        border_row(overview, 1, 8)
        overview.cell(row=row, column=8).number_format = MONEY


def build(db: Session, py: PolicyYear, mode: Mode) -> Workbook:
    ctx = load_context(db, py, mode)
    envelope = envelope_for(db, py)

    wb = Workbook()
    _write_overview(wb, ctx, envelope)
    overview = wb["Overview"]
    doc_name = "Quotation Slip" if ctx.blank_rates else "Placement Slip"

    taken: set[str] = {"overview"}
    for product in ctx.products:
        ws = wb.create_sheet(_sheet_title(product.code, taken))
        _write_product_sheet(
            ws,
            ctx,
            product,
            ctx.cats_by_product.get(product.id, []),
            ctx.plans_by_product.get(product.id, []),
            ctx.terms.get(product.id),
        )
        finalize_sheet(
            ws,
                doc_name=doc_name,
        )

    unassigned = ctx.cats_by_product.get(None, [])
    if unassigned:
        ws = wb.create_sheet(_sheet_title("Unassigned", taken))
        _write_product_sheet(ws, ctx, None, unassigned, [], None)
        finalize_sheet(
            ws,
                doc_name=doc_name,
        )

    overview.freeze_panes = "A8"
    overview.auto_filter.ref = f"A7:H{overview.max_row}"
    finalize_sheet(overview, doc_name=doc_name)

    return wb


__all__ = ["build", "insured_text"]

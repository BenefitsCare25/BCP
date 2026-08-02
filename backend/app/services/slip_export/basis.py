"""The Basis-of-Cover table: who is covered, how many, and for how much.

Every column is derived from what the product's own categories actually carry —
there is no per-product column list here. A product that states a plan code gets
a Plan column; one that states earnings gets an earnings column; one whose
member counts are split by tier gets a count block sub-divided by exactly the
tiers present, in the registry's canonical order. Nothing is emitted for data
the product doesn't have, and nothing a product does have is dropped.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.models import Category
from app.services import product_registry
from app.services.matching_engine import insured_names
from app.services.slip_export.context import SlipContext
from app.services.slip_export.styles import (
    CENTER,
    COUNT,
    HEADER,
    NOTE,
    SECTION,
    border_row,
    numeric_or_text,
    style_row,
)

# Canonical tier order, from the product registry so a new tier scheme there is
# printed here without any change to this module.
TIER_ORDER: dict[str, int] = {
    key: i for i, key in enumerate(product_registry.tier_order())
}

_DISCLAIMER_TAIL = (
    "Actual figures for billing must be provided within 30 days from policy "
    "inception."
)

# Column kinds. "pa" reads the category's stored plan_assignments; the rest read
# the resolved figures (roster-first) so every count on the sheet agrees.
_PA = "pa"
_MEMBERS = "members"
_TIER = "tier"
_DEPENDANTS = "dependants"
_SUM_INSURED = "sum_insured"


@dataclass(frozen=True)
class _Col:
    header: str
    sub: str = ""
    kind: str = _PA
    key: str = ""
    numeric: bool = False


def plan_assignments(cat: Category) -> dict[str, Any]:
    return cat.plan_assignments if isinstance(cat.plan_assignments, dict) else {}


def insured_text(pa: dict[str, Any]) -> str:
    """The category's insured entities as they must appear ON THE SLIP.

    The export reproduces the LEGAL spelling verbatim — this document goes to
    an insurer — so it renders the stored names, never a normalized or
    alias-resolved form. Storage is a token list; a legacy comma-joined string
    round-trips unchanged.
    """
    return ", ".join(insured_names(pa.get("insured")))


def participation_text(c: Category) -> str:
    detail = c.participation_detail if isinstance(c.participation_detail, dict) else {}
    raw = str(detail.get("raw") or "").strip()
    text = " ".join(raw.split()) if raw else (c.participation_model or "")
    pa = plan_assignments(c)
    # The reference slips fold the scope into Participation ("Compulsory -
    # SG Office") and mark dependant-only rows ("Voluntary - Dependents").
    scope = str(pa.get("location_scope") or "").strip()
    if scope and scope.lower() not in text.lower():
        text = f"{text} - {scope}" if text else scope
    if not text and pa.get("member_scope") == "dependant":
        text = "Voluntary - Dependents"
    return text


def _tier_keys(categories: list[Category], ctx: SlipContext) -> list[str]:
    """Every tier the product's own counts are split by, canonically ordered."""
    keys: set[str] = set()
    for c in categories:
        keys.update(ctx.figures_for(c).tier_counts)
    return sorted(keys, key=lambda k: (TIER_ORDER.get(k, 99), k))


def _columns(categories: list[Category], ctx: SlipContext) -> list[_Col]:
    pas = [plan_assignments(c) for c in categories]
    figures = [ctx.figures_for(c) for c in categories]
    cols: list[_Col] = []
    if any(pa.get("plan_code") not in (None, "") for pa in pas):
        cols.append(_Col("Plan", kind=_PA, key="plan_code"))

    tiers = _tier_keys(categories, ctx)
    if tiers:
        # Split block: one column per tier present, then the total — the shape
        # the medical slips print ("* Number" spanning EO/ES/EC/EF).
        cols += [
            _Col("* No. of members", sub=key, kind=_TIER, key=key, numeric=True)
            for key in tiers
        ]
        cols.append(_Col("* No. of members", sub="Total", kind=_MEMBERS, numeric=True))
    else:
        cols.append(_Col("* No. of employees", kind=_MEMBERS, numeric=True))

    if any(f.dependants for f in figures):
        cols.append(_Col("* No. of dependants", kind=_DEPENDANTS, numeric=True))
    if any(pa.get("basis") for pa in pas):
        cols.append(_Col("Basis", kind=_PA, key="basis", numeric=True))
    if any(f.sum_insured is not None for f in figures):
        cols.append(_Col("* Sum Insured (S$)", kind=_SUM_INSURED, numeric=True))
    if any(pa.get("estimated_annual_earnings") is not None for pa in pas):
        cols.append(
            _Col(
                "* Estimated annual earnings",
                kind=_PA,
                key="estimated_annual_earnings",
                numeric=True,
            )
        )
    return cols


def _value(col: _Col, cat: Category, ctx: SlipContext) -> Any:
    figures = ctx.figures_for(cat)
    if col.kind == _MEMBERS:
        return figures.members if figures.members is not None else ""
    if col.kind == _TIER:
        return figures.tier_counts.get(col.key, "")
    if col.kind == _DEPENDANTS:
        return figures.dependants if figures.dependants else ""
    if col.kind == _SUM_INSURED:
        return figures.sum_insured if figures.sum_insured is not None else ""
    value = plan_assignments(cat).get(col.key)
    if value is None:
        return ""
    if col.key == "plan_code":
        return str(value)
    if col.key == "basis":
        # A pure amount renders as a number ("2000000.0" → 2,000,000);
        # salary-multiple expressions stay text.
        return numeric_or_text(value)
    return value


def _disclaimer(categories: list[Category], ctx: SlipContext) -> str:
    """The footnote under the table — states where its figures came from.

    A slip that silently mixes live and stated headcounts is worse than one that
    admits it, so the fallbacks are counted on the face of the document.

    Deliberately carries NO generated-on date. ``report_versions`` fingerprints
    the workbook's data parts to skip re-saving an unchanged report; a date here
    would change that hash every calendar day and defeat the guard, rewriting
    the retained blob daily. The version row already records when it was made.
    """
    stale = ctx.stale_categories(categories)
    live = sum(1 for c in categories if ctx.figures_for(c).from_roster)
    if live and not stale:
        lead = "* Member counts are from the current roster."
    elif live and stale:
        lead = (
            "* Member counts are from the current roster, except "
            f"{stale} categor{'y' if stale == 1 else 'ies'} that matched no "
            "member and show the figures stated on the placement slip."
        )
    elif stale:
        lead = (
            "* Member counts are the figures stated on the placement slip — no "
            "roster member matched these categories."
        )
    else:
        lead = "* Figures above are for illustration only."
    # A live headcount beside a sum insured that could NOT be rebuilt from it
    # (salary-relative basis, member missing a salary) describes two different
    # populations in one row. Say so rather than letting the reader assume the
    # whole row moved together.
    unrebuilt = sum(1 for c in categories if ctx.figures_for(c).cover_is_stale)
    if unrebuilt:
        lead += (
            f" Sum insured for {unrebuilt} categor"
            f"{'y' if unrebuilt == 1 else 'ies'} could not be recomputed from "
            "the roster and remains as stated on the placement slip."
        )
    return f"{lead} {_DISCLAIMER_TAIL}"


def write_basis_of_cover(
    ws: Worksheet,
    categories: list[Category],
    ctx: SlipContext,
    insured_default: str,
) -> None:
    ws.append(["Basis of Cover :"])
    style_row(ws, font=SECTION)
    cols = _columns(categories, ctx)
    last_col = 4 + len(cols)
    # A split count block needs a second header row for the tier codes.
    needs_sub = any(col.sub for col in cols)

    def _header() -> None:
        ws.append(["", "Insured", "Category", "Participation"] + [c.header for c in cols])
        row = style_row(ws, font=HEADER)
        border_row(ws, 2, last_col)
        if not needs_sub:
            return
        # Blank the repeated block label so the span reads once, as on the slips.
        seen: set[str] = set()
        for i, col in enumerate(cols):
            if not col.sub:
                continue
            cell = ws.cell(row=row, column=5 + i)
            if col.header in seen:
                cell.value = ""
            seen.add(col.header)
        ws.append(["", "", "", ""] + [c.sub for c in cols])
        sub_row = style_row(ws, font=HEADER)
        for i, col in enumerate(cols):
            if col.sub:
                ws.cell(row=sub_row, column=5 + i).alignment = CENTER
        border_row(ws, 2, last_col)

    prev_key: tuple[str, str] | None = None
    prev_name = ""
    for c in categories:
        pa = plan_assignments(c)
        insured = insured_text(pa) or insured_default
        participation = participation_text(c)
        key = (insured, participation)
        new_block = key != prev_key
        if new_block:
            if prev_key is not None:
                ws.append([])
            _header()
            prev_name = ""
        values = [_value(col, c, ctx) for col in cols]
        # Continuation rows (same category, next plan code) leave the category
        # blank, mirroring the slips' carried-down cells.
        name = c.display_name if c.display_name != prev_name else ""
        ws.append([
            "",
            insured if new_block else "",
            name,
            participation if new_block else "",
            *values,
        ])
        row = style_row(ws, wrap_cols=(2, 3, 4))
        border_row(ws, 2, last_col)
        for i, (col, value) in enumerate(zip(cols, values, strict=True)):
            if col.numeric and isinstance(value, (int, float)):
                ws.cell(row=row, column=5 + i).number_format = COUNT
        prev_key = key
        prev_name = c.display_name
    ws.append(["", _disclaimer(categories, ctx)])
    style_row(ws, font=NOTE)

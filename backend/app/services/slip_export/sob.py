"""Schedule of Benefits + Plan Details.

Plans fold to columns through the same model the SOB editor uses, so what the
insurer receives is what the broker sees: plans whose value vectors match
collapse into one column, and a column is headed by every plan code it covers.
"""
from __future__ import annotations

import re
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Plan
from app.services.slip_export.styles import (
    HEADER,
    SECTION,
    WRAP,
    border_row,
    style_row,
)
from app.services.sob_columns import sob_from_plan_items

_MIRRORED_INTO_LIMITS = frozenset(
    {"maximum_days", "qualification_period", "co_insurance", "surgical_schedule"}
)
_PROPERTY_LABELS = {
    "per_visit": "Per visit",
    "co_payment": "Co-payment",
    "per_policy_year": "Per policy year",
    "per_visit_restructured": "Per visit — Restructured Hospital",
    "per_visit_private": "Per visit — Private Hospital",
    "co_payment_restructured": "Co-payment — Restructured Hospital",
    "co_payment_private": "Co-payment — Private Hospital",
    "per_disability": "Per disability",
}


def _natural_code_key(value: str) -> tuple[tuple[int, int | str], ...]:
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", str(value or ""))
        if part
    )


def shared_cover(plans: list[Plan]) -> str | None:
    """The product-level "Cover :" sentence, when the plans agree on one.

    Slip-parsed GHS-style products stamp the SAME cover sentence on every plan
    ("Cover: Reimbursement of eligible inpatient expenses…") — that belongs on
    the Cover line above the SOB, not repeated per plan. A single plan whose
    description is a basis expression ("24x basic monthly salary") stays in
    Plan Details."""
    descs = {
        (p.cover_description or "").strip()
        for p in plans
        if (p.cover_description or "").strip()
    }
    if len(descs) != 1:
        return None
    text = next(iter(descs))
    n = sum(1 for p in plans if (p.cover_description or "").strip())
    if n > 1 or text.lower().startswith("cover"):
        return text
    return None


def _sob_value(entry: dict[str, Any], col_id: str, first: bool) -> Any:
    if first:
        return entry.get("base_value")
    overrides = entry.get("overrides") or {}
    return overrides.get(col_id, entry.get("base_value"))


def _sob_col_label(col: dict[str, Any]) -> str:
    """Insurer-facing label for exactly the available plans in this column."""
    codes = [str(c) for c in (col.get("plan_codes") or []) if str(c)]
    label = str(col.get("label") or "").strip()
    # Preserve the platform's source header (for example
    # "PLAN 1/U01/U04/U06") instead of replacing it with a synthetic label.
    if label and label.casefold() != "all plans" and not re.search(r" \+\d+$", label):
        return label
    if len(codes) == 1:
        return label or f"Plan {codes[0]}"
    if label.casefold() == "all plans":
        return f"All configured plans ({_compact_codes(codes)})"
    return "Plan " + "/".join(codes)


def _compact_codes(codes: list[str]) -> str:
    """Compact naturally ordered numeric plan codes into readable ranges."""
    numeric = sorted({int(code) for code in codes if code.isdigit()})
    other = sorted(
        {code for code in codes if not code.isdigit()},
        key=_natural_code_key,
    )
    parts: list[str] = []
    start = previous = None
    for number in [*numeric, None]:
        if start is None:
            start = previous = number
            continue
        if number is not None and previous is not None and number == previous + 1:
            previous = number
            continue
        parts.append(str(start) if start == previous else f"{start}-{previous}")
        start = previous = number
    parts.extend(other)
    return ", ".join(parts)


def _apply_setup_column_labels(
    sob: dict[str, Any], answers: dict[str, Any], available_codes: set[str]
) -> None:
    """Restore the labels the broker sees in Product Setup without using its
    potentially shorter draft row set as the exported schedule."""
    source = answers.get("sob") if isinstance(answers, dict) else None
    if not isinstance(source, dict):
        return
    labels: dict[frozenset[str], str] = {}
    for column in source.get("columns") or []:
        if not isinstance(column, dict):
            continue
        codes = frozenset(
            str(code) for code in (column.get("plan_codes") or [])
            if str(code) in available_codes
        )
        label = str(column.get("label") or "").strip()
        if codes and label:
            labels[codes] = label
    for column in sob.get("columns") or []:
        codes = frozenset(str(code) for code in column.get("plan_codes") or [])
        source_label = labels.get(codes)
        if source_label is not None:
            column["label"] = source_label


def _property_label(key: str) -> str:
    if key in _PROPERTY_LABELS:
        return _PROPERTY_LABELS[key]
    words = key.replace("_", " ").strip()
    return words[:1].upper() + words[1:]


def _display_properties(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, dict):
        return {}
    return {
        str(key): value
        for key, value in raw.items()
        if key not in _MIRRORED_INTO_LIMITS and str(value or "").strip()
    }


def write_sob(
    ws: Worksheet,
    plans: list[Plan],
    cover: str | None,
    answers: dict[str, Any] | None = None,
    quotation: bool = False,
) -> None:
    with_items = [
        {
            "code": p.code,
            "label": p.display_name,
            "benefit_items": (p.benefit_schedule or {}).get("items"),
        }
        for p in sorted(plans, key=lambda plan: _natural_code_key(plan.code))
        if isinstance(p.benefit_schedule, dict)
        and isinstance(p.benefit_schedule.get("items"), list)
        and p.benefit_schedule["items"]
    ]
    if not with_items:
        return
    # These are RESOLVED plan schedules, not a parsed slip grid: every cell was
    # already decided per plan, and none carries the parser's `na` flag. A blank
    # here is therefore an explicit blank — the broker cleared the cell, or the
    # slip said "NA" and it flattened to null on the way in — so it must NOT
    # inherit the base column's value into an insurer-facing document.
    sob = sob_from_plan_items(with_items, blank_inherits=False)
    _apply_setup_column_labels(
        sob, answers or {}, {str(plan.code) for plan in plans}
    )
    columns = sob.get("columns") or []
    items = sob.get("items") or []
    if not columns or not items:
        return

    # A compact single-plan setup can leave only enough room for the SOB title
    # and header at the foot of page one. Keep that header with its data. This
    # is driven by the rendered sheet shape, not by a product code or plan name.
    if len(with_items) == 1 and 20 <= ws.max_row <= 40:
        ws.row_breaks.append(Break(id=ws.max_row))

    ws.append([])
    # The stored sentence usually carries its own "Cover:" prefix — strip it,
    # the label cell already says it.
    cover_text = re.sub(r"(?i)^cover\s*:\s*", "", cover).strip() if cover else ""
    ws.append(["Cover :", cover_text])
    style_row(ws, font=HEADER, wrap_cols=(2,))
    ws.cell(row=ws.max_row, column=2).font = Font(bold=False)
    title = (
        "SCHEDULE OF BENEFITS / DEFINITIONS / INSURER RESPONSE"
        if quotation
        else "SCHEDULE OF BENEFITS / PLAN"
    )
    ws.append([title])
    style_row(ws, font=SECTION)
    response = ["Insurer Response"] if quotation else []
    ws.append(
        ["No.", "Benefit", "Details / Qualifiers"]
        + [_sob_col_label(col) for col in columns]
        + response
    )
    style_row(ws, font=HEADER)
    sob_last_col = 3 + len(columns) + len(response)
    ws.row_dimensions[ws.max_row].height = 28
    for column in range(1, sob_last_col + 1):
        ws.cell(row=ws.max_row, column=column).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    border_row(ws, 1, sob_last_col)
    value_cols = range(4, 4 + len(columns))

    def _value_row(number: str, name: str, entry: dict[str, Any]) -> None:
        values = [
            _sob_value(entry, col.get("id"), i == 0)
            for i, col in enumerate(columns)
        ]
        ws.append(
            [number, name, str(entry.get("note") or ""), *values]
            + ([""] if quotation else [])
        )
        row = ws.max_row
        ws.cell(row=row, column=2).alignment = WRAP
        ws.cell(row=row, column=3).alignment = WRAP
        for col in value_cols:
            ws.cell(row=row, column=col).alignment = WRAP
        border_row(ws, 1, sob_last_col)

    def _limit_rows(limits: Any, indent: str) -> None:
        for lim in limits or []:
            if not isinstance(lim, dict):
                continue
            label = str(lim.get("label") or "").strip()
            value = str(lim.get("value") or "").strip()
            if label or value:
                ws.append(
                    ["", f"{indent}• {label}", value]
                    + [""] * (len(columns) + len(response))
                )
                ws.cell(row=ws.max_row, column=2).alignment = WRAP
                ws.cell(row=ws.max_row, column=3).alignment = WRAP
                border_row(ws, 1, sob_last_col)

    def _property_rows(entry: dict[str, Any], indent: str) -> None:
        shared = _display_properties(entry.get("properties"))
        for key, value in shared.items():
            ws.append(
                ["", f"{indent}• {_property_label(key)}", value]
                + [""] * (len(columns) + len(response))
            )
            border_row(ws, 1, sob_last_col)

        per_column = entry.get("column_properties")
        if not isinstance(per_column, dict):
            return
        keys: list[str] = []
        for props in per_column.values():
            for key in _display_properties(props):
                if key not in keys:
                    keys.append(key)
        for key in keys:
            values = [
                _display_properties(per_column.get(col.get("id"))).get(key, "")
                for col in columns
            ]
            ws.append(
                ["", f"{indent}• {_property_label(key)}", "", *values]
                + ([""] if quotation else [])
            )
            row = ws.max_row
            for col in range(2, sob_last_col + 1):
                ws.cell(row=row, column=col).alignment = WRAP
            border_row(ws, 1, sob_last_col)

    for it in items:
        _value_row(str(it.get("number") or ""), str(it.get("name") or ""), it)
        _property_rows(it, "    ")
        # Non-copay `properties` are machine-derived duplicates of the limits
        # (maximum_days ↔ "Maximum no. of days") — never rendered.
        _limit_rows(it.get("limits"), "    ")
        for sub in it.get("sub_items") or []:
            _value_row(
                str(sub.get("number") or sub.get("key") or ""),
                f"    {sub.get('name') or ''!s}",
                sub,
            )
            _property_rows(sub, "        ")
            _limit_rows(sub.get("limits"), "        ")


def write_plan_details(
    ws: Worksheet, plans: list[Plan], cover: str | None
) -> None:
    """Cover-basis metadata (GCI-style plan tiers) — only when a plan carries
    something the Cover line doesn't already say. A cover sentence shared by
    every plan is hoisted onto the Cover line, so rows whose only content is
    that repeated sentence are dropped entirely."""
    def _own_cover(p: Plan) -> str:
        desc = (p.cover_description or "").strip()
        return "" if cover is not None and desc == cover else desc

    rows = [
        (p, own)
        for p in plans
        if (own := _own_cover(p)) or p.annual_policy_limit or p.report_label
    ]
    if not rows:
        return
    ws.append([])
    ws.append(["Plan Details"])
    style_row(ws, font=SECTION)
    ws.append(["Plan", "Cover Description", "Annual Policy Limit", "Report Label"])
    style_row(ws, font=HEADER)
    border_row(ws, 1, 4)
    for p, own in rows:
        ws.append([
            p.code, own,
            p.annual_policy_limit or "", p.report_label or "",
        ])
        style_row(ws, wrap_cols=(2,))
        border_row(ws, 1, 4)

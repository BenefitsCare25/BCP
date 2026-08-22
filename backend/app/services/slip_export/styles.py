"""Shared worksheet styling + row helpers for the slip export.

The reference slips are gridded tables, not floating text, so every table row
gets a full thin border and the header/label blocks share one vocabulary of
fonts. Keeping it here means the Basis, Rate and Schedule sections can't drift
apart visually.
"""
from __future__ import annotations

import math
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TITLE = Font(bold=True, size=14, color="FFFFFF")
SECTION = Font(bold=True, size=11, color="FFFFFF")
HEADER = Font(bold=True, color="6F0B1B")
NOTE = Font(italic=True, size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
MIDDLE_WRAP = Alignment(vertical="center", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

MONEY = "#,##0.00"
COUNT = "#,##0"

_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
TITLE_FILL = PatternFill("solid", fgColor="9F1239")
SECTION_FILL = PatternFill("solid", fgColor="C8102E")
HEADER_FILL = PatternFill("solid", fgColor="FCE7EC")


def set_widths(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def set_compact_product_widths(ws: Worksheet) -> None:
    """Size product sheets like the reference slips, based on table width.

    Column A is principally the SOB index and remains narrow. Header labels
    span A:B and their values span C:last-column during finalisation, so neither
    needs an oversized standalone column. Narrower sheets can spend more width
    on long benefit definitions; wide medical grids keep each plan compact.
    """
    last_col = max(ws.max_column, 3)
    if last_col <= 6:
        leading = {1: 9, 2: 32, 3: 44, 4: 20}
        trailing = 18
    elif last_col <= 8:
        leading = {1: 9, 2: 30, 3: 32, 4: 22}
        trailing = 17
    else:
        leading = {1: 8, 2: 28, 3: 24, 4: 20}
        trailing = 15
    set_widths(
        ws,
        {
            col: leading.get(col, trailing)
            for col in range(1, last_col + 1)
        },
    )


def style_row(
    ws: Worksheet, font: Font | None = None, wrap_cols: tuple[int, ...] = ()
) -> int:
    row = int(ws.max_row)
    if font is not None:
        for cell in ws[row]:
            cell.font = font
    for col in wrap_cols:
        ws.cell(row=row, column=col).alignment = WRAP
    return row


def spacer_row(ws: Worksheet, height: float = 6) -> int:
    """Append a real blank row and size that row, never the row above it.

    ``Worksheet.append([])`` advances openpyxl's internal cursor without
    creating a cell, so ``ws.max_row`` can still point at the preceding content
    row.  Using ``[None]`` registers the spacer before its compact height is
    assigned.
    """
    ws.append([None])
    row = int(ws.max_row)
    ws.row_dimensions[row].height = height
    return row


def border_row(ws: Worksheet, first: int, last: int, row: int | None = None) -> None:
    """Grid the table cells so the workbook reads like the reference slips."""
    r = row or ws.max_row
    for col in range(first, last + 1):
        ws.cell(row=r, column=col).border = BORDER


def label_value_rows(ws: Worksheet, rows: list[tuple[str, str] | None]) -> None:
    """Header block: label in col A, value in col C (reference layout).

    ``None`` emits a spacer row.
    """
    for entry in rows:
        if entry is None:
            spacer_row(ws)
            continue
        label, value = entry
        ws.append([label, "", value])
        row = ws.max_row
        ws.cell(row=row, column=1).font = HEADER
        ws.cell(row=row, column=1).alignment = MIDDLE_WRAP
        ws.cell(row=row, column=3).alignment = MIDDLE_WRAP


def numeric_or_text(v: Any) -> Any:
    """A pure amount becomes a number (so Excel formats it); anything else — a
    salary-multiple expression, a qualifier — stays verbatim text."""
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return v


def _cell_width(ws: Worksheet, row: int, col: int) -> float:
    """Effective width of a cell, including a merge that starts at it."""
    end_col = col
    for merged in ws.merged_cells.ranges:
        if merged.min_row == row == merged.max_row and merged.min_col == col:
            end_col = merged.max_col
            break
    return sum(
        float(ws.column_dimensions[get_column_letter(c)].width or 13)
        for c in range(col, end_col + 1)
    )


def _content_height(ws: Worksheet, row: int, populated: list[Any]) -> float:
    """Approximate Excel's wrapped auto-height without clipping definitions."""
    lines = 1
    for cell in populated:
        if not isinstance(cell.value, str):
            continue
        width = max(_cell_width(ws, row, cell.column), 1)
        cell_lines = sum(
            max(1, math.ceil(len(part) / max(width * 1.05, 1)))
            for part in cell.value.splitlines() or [""]
        )
        lines = max(lines, cell_lines)
    return min(300, max(18, lines * 15 + 3))


def finalize_sheet(ws: Worksheet, doc_name: str) -> None:
    """Apply consistent workbook-level print and visual formatting."""
    last_col = max(ws.max_column, 3)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"
    ws.print_title_rows = "1:1"
    ws.oddFooter.center.text = "Page &P"
    ws.oddFooter.right.text = doc_name
    ws.sheet_properties.tabColor = "C8102E"

    section_labels = {
        "Basis of Cover :",
        "Rate :",
        "Plan Details",
        "SCHEDULE OF BENEFITS / PLAN",
        "SCHEDULE OF BENEFITS / DEFINITIONS / INSURER RESPONSE",
    }
    for row in range(1, ws.max_row + 1):
        populated = [cell for cell in ws[row] if cell.value not in (None, "")]
        if not populated:
            continue
        for cell in populated:
            if isinstance(cell.value, str) and len(cell.value) > 32:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical or "center",
                    wrap_text=True,
                )

        if row == 1:
            if last_col > 1:
                ws.merge_cells(
                    start_row=1, start_column=1, end_row=1, end_column=last_col
                )
            cell = ws.cell(row=1, column=1)
            cell.font = TITLE
            cell.fill = TITLE_FILL
            cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 26
            continue

        first = ws.cell(row=row, column=1)
        if len(populated) == 1 and first.value in section_labels:
            if last_col > 1:
                ws.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=last_col,
                )
            first.font = SECTION
            first.fill = SECTION_FILL
            first.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 22
            continue

        # Header/term label rows use A and C. Merge the intentionally blank A:B
        # span even when the value is blank; quotations deliberately blank
        # Insurer, Policy No. and commercial terms, and those labels still need
        # the full label width and a correctly calculated row height.
        if (
            first.value not in (None, "")
            and first.font.bold
            and ws.cell(row=row, column=2).value in (None, "")
            and all(
                ws.cell(row=row, column=col).value in (None, "")
                for col in range(4, last_col + 1)
            )
            and not any(
                merged.min_row == row == merged.max_row
                and merged.min_col <= 2 <= merged.max_col
                for merged in ws.merged_cells.ranges
            )
        ):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            first.alignment = MIDDLE_WRAP
            if ws.cell(row=row, column=3).value not in (None, "") and last_col > 3:
                ws.merge_cells(
                    start_row=row,
                    start_column=3,
                    end_row=row,
                    end_column=last_col,
                )
            ws.cell(row=row, column=3).alignment = MIDDLE_WRAP

        if len(populated) >= 2 and all(cell.font.bold for cell in populated):
            first_col = min(cell.column for cell in populated)
            row_last_col = max(cell.column for cell in populated)
            for col in range(first_col, row_last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            ws.row_dimensions[row].height = max(
                ws.row_dimensions[row].height or 15, 20
            )

        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = _content_height(ws, row, populated)

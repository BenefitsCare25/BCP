"""Full member-listing upload template (company-profile setup).

Two-sheet .xlsx (Employees / Dependants) whose headers round-trip through
``roster_parser.EMPLOYEE_COLUMN_MAP`` / ``DEPENDANT_COLUMN_MAP`` — every column
written here re-imports as the intended attribute key. Existing roster rows are
pre-filled so the same file doubles as an update template (HR completes the
new columns without retyping the roster). One "<Insurer> Member ID" column is
appended per insurer configured on the year's products.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Dependant, Employee, PolicyYear
from app.models.dependant import DEPENDANT_STATUS_ACTIVE
from app.models.employee import EMPLOYEE_STATUS_ACTIVE
from app.services.insurer_listings import configured_insurers_for_year
from app.services.roster_parser import INSURER_MEMBER_ID_KEY

# (header, attribute key) pairs, in insurer-listing column order. ``staff_id``
# and ``employee_name`` are real Employee columns, marked with a leading "@".
_EMPLOYEE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Entity", "entity"),
    ("Staff ID", "@staff_id"),
    ("Employee Name", "@employee_name"),
    ("Identification No. (NRIC/FIN)", "id_no"),
    ("Date of Birth", "date_of_birth"),
    ("Gender", "gender"),
    ("Marital Status", "marital_status"),
    ("Employment Status", "employment_status"),
    ("Designation", "designation"),
    ("Country of Work", "country_of_work"),
    ("Foreigner Employment Pass", "pass"),
    ("Nationality", "nationality"),
    ("Date of Hire", "date_of_hire"),
    ("Confirmation Date", "confirmation_date"),
    ("Effective Date", "effective_date"),
    ("Last Day of Service", "last_day_of_service"),
    ("Category", "category"),
    ("Job Grade", "job_grade"),
    ("Division", "division"),
    ("Department", "department"),
    ("Cost Centre", "cost_centre"),
    ("Email", "email"),
    ("Mobile", "mobile"),
    ("Bank Code", "bank_code"),
    ("Branch Code", "branch_code"),
    ("Bank Account No.", "bank_account_no"),
    ("Has Insurance Cover Last Year", "prior_year_cover"),
    ("Eligible to Sell Leave", "leave_sell_eligible"),
    ("Monthly Salary", "salary"),
    ("Currency", "currency"),
    ("Remarks", "remarks"),
)

_DEPENDANT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Entity", "entity"),
    ("Staff ID", "employee_staff_id"),
    ("Employee Name", "employee_name"),
    ("Employee's Identification No. (NRIC/FIN)", "employee_id_no"),
    ("Dependant Name", "dependant_name"),
    ("Dependant's Identification No.", "dependant_id_no"),
    ("Relationship", "relationship"),
    ("Date of Marriage", "date_of_marriage"),
    ("Gender", "gender"),
    ("Date of Birth", "date_of_birth"),
    ("Effective Date", "effective_date"),
    ("Termination Date", "termination_date"),
    ("Remarks", "remarks"),
)

# Identifier columns forced to Text format so Excel keeps leading zeros
# ("081") and doesn't float-ify account numbers.
_TEXT_COLUMNS = {
    "id_no", "employee_id_no", "dependant_id_no", "mobile",
    "bank_code", "branch_code", "bank_account_no",
}


def _write_sheet(
    ws: Worksheet,
    columns: tuple[tuple[str, str], ...],
    insurers: list[str],
    rows: list[tuple[dict[str, Any], dict[str, Any]]],
) -> None:
    """``rows`` = (attribute_values, {"@field": value}) per person."""
    headers = [h for h, _ in columns] + [f"{ins} Member ID" for ins in insurers]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    text_cols = [
        i + 1
        for i, (_, key) in enumerate(columns)
        if key.lstrip("@") in _TEXT_COLUMNS
    ] + [len(columns) + i + 1 for i in range(len(insurers))]

    for attrs, fields in rows:
        member_ids = attrs.get(INSURER_MEMBER_ID_KEY) or {}
        values: list[Any] = []
        for _, key in columns:
            if key.startswith("@"):
                values.append(fields.get(key, ""))
            else:
                v = attrs.get(key)
                values.append("" if v is None else v)
        values.extend(member_ids.get(ins, "") for ins in insurers)
        ws.append(values)

    for col_idx in text_cols:
        for row_idx in range(1, max(ws.max_row, 2) + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "@"

    for col in ws.columns:
        width = max(
            (len(str(c.value)) for c in col if c.value is not None), default=10
        )
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 40)


def build_member_listing_template(db: Session, policy_year: PolicyYear) -> Workbook:
    insurers = configured_insurers_for_year(db, policy_year)

    employees = list(
        db.execute(
            select(Employee)
            .where(
                Employee.policy_year_id == policy_year.id,
                Employee.status == EMPLOYEE_STATUS_ACTIVE,
            )
            .order_by(Employee.staff_id)
        ).scalars().all()
    )
    dependants = list(
        db.execute(
            select(Dependant)
            .where(
                Dependant.policy_year_id == policy_year.id,
                Dependant.status == DEPENDANT_STATUS_ACTIVE,
            )
        ).scalars().all()
    )

    wb = Workbook()
    ws_emp = wb.active
    ws_emp.title = "Employees"
    _write_sheet(
        ws_emp,
        _EMPLOYEE_COLUMNS,
        insurers,
        [
            (
                e.attribute_values or {},
                {"@staff_id": e.staff_id, "@employee_name": e.employee_name or ""},
            )
            for e in employees
        ],
    )

    ws_dep = wb.create_sheet("Dependants")
    _write_sheet(
        ws_dep,
        _DEPENDANT_COLUMNS,
        insurers,
        [(d.attribute_values or {}, {}) for d in dependants],
    )
    return wb

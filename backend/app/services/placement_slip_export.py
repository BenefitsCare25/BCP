"""Placement/quotation-slip export of a policy year's configured products (.xlsx).

Reverses the intake direction: where the parser turns an insurer's placement
slip into Category/Plan rows, this workbook renders those rows back into a
slip-shaped document modelled on the broker's real slips (reference: the CDL
2026 placement + quotation workbooks) — one sheet per product with the header
label block, the Basis of Cover table, the Rate section, voluntary age-banded
rates, and the Schedule of Benefits.

Two modes share the renderer:

* ``placement`` — the placed state: insurer named, rates and premiums filled.
* ``quotation`` — the shopping document that accompanies the Fact-Find form:
  same structure and figures, but the insurer, every rate/premium cell and the
  annual-premium total are left BLANK for the quoting insurer to complete.

Fields the platform doesn't store (address, business, eligibility wording,
policy number, NEL, refund formula…) are emitted as labelled blank rows so the
broker completes them in Excel — exactly how the reference slips leave their
unknown cells. Config-only: no member PII, and premium figures are emitted as
stored (GST-exclusive — grossing is a display concern, never re-applied here).
"""
from __future__ import annotations

import json
import re
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.deps import tenant_or_global
from app.models import Category, Plan, PolicyYear, Product, ProductTerm
from app.services.matching_engine import insured_names
from app.services.product_terms import envelope_for
from app.services.sob_columns import sob_from_plan_items

Mode = Literal["placement", "quotation"]

_TITLE = Font(bold=True, size=13)
_SECTION = Font(bold=True, size=11)
_HEADER = Font(bold=True)
_NOTE = Font(italic=True, size=9)
_WRAP = Alignment(wrap_text=True, vertical="top")
_MONEY = "#,##0.00"
_COUNT = "#,##0"
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Canonical tier vocabulary (mirror of product_registry TIER_SCHEMES labels):
# composite employee tiers first, dependant-only tiers after. A slip-specific
# label persisted in plan_assignments.tier_labels overrides the canonical one.
_TIER_LABELS = {
    "EO": "Employee Only",
    "ES": "Employee + Spouse",
    "EC": "Employee + Children",
    "EF": "Employee + Family",
    "SO": "Spouse Only",
    "CO": "Child(ren) Only",
    "SC": "Spouse & Child(ren)",
    "FO": "Family Only",
}
_TIER_ORDER = {code: i for i, code in enumerate(_TIER_LABELS)}

_DISCLAIMER = (
    "* Figures above are for illustration only. Actual figures for billing "
    "must be provided within 30 days from policy inception."
)

# Slip terms the platform doesn't store — labelled blanks for the broker.
_TERM_LABELS = (
    "Non Evidence Limit :",
    "Maximum Limit Per Insured Person :",
    "Experience Refund Formula / Maximum Loss Ratio :",
    "Product Rated Together :",
    "Policyholder(s) Rated Together :",
)

_SHEET_BAD_CHARS = set("[]:*?/\\")


def _sheet_title(base: str, taken: set[str]) -> str:
    """Excel-legal (≤31 chars, no []:*?/\\) and unique within the workbook."""
    clean = "".join(c for c in base if c not in _SHEET_BAD_CHARS).strip() or "Product"
    title = clean[:31]
    n = 2
    while title.lower() in taken:
        suffix = f" ({n})"
        title = clean[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(title.lower())
    return title


def _pa(c: Category) -> dict[str, Any]:
    return c.plan_assignments if isinstance(c.plan_assignments, dict) else {}


def _numeric_or_text(v: Any) -> Any:
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return v


def _participation_text(c: Category) -> str:
    detail = c.participation_detail if isinstance(c.participation_detail, dict) else {}
    raw = str(detail.get("raw") or "").strip()
    text = " ".join(raw.split()) if raw else (c.participation_model or "")
    pa = _pa(c)
    # The reference slips fold the scope into Participation ("Compulsory -
    # SG Office") and mark dependant-only rows ("Voluntary - Dependents").
    scope = str(pa.get("location_scope") or "").strip()
    if scope and scope.lower() not in text.lower():
        text = f"{text} - {scope}" if text else scope
    if not text and pa.get("member_scope") == "dependant":
        text = "Voluntary - Dependents"
    return text


def _insured_text(pa: dict) -> str:
    """The category's insured entities as they must appear ON THE SLIP.

    The export reproduces the LEGAL spelling verbatim — this document goes to
    an insurer — so it renders the stored names, never a normalized or
    alias-resolved form. Storage is a token list; a legacy comma-joined string
    round-trips unchanged.
    """
    return ", ".join(insured_names(pa.get("insured")))


def _distinct_insured(categories: list[Category]) -> list[str]:
    """Every distinct entity named across the categories, in first-seen order."""
    seen: list[str] = []
    for c in categories:
        for name in insured_names(_pa(c).get("insured")):
            if name not in seen:
                seen.append(name)
    return seen


def _fmt_window(start: Any, end: Any) -> str:
    if not start or not end:
        return ""
    return f"{start:%d %b %Y} to {end:%d %b %Y}"


def _coverage_window(py: PolicyYear, term: ProductTerm | None) -> str:
    start = term.coverage_start if term and term.coverage_start else py.start_date
    end = term.coverage_end if term and term.coverage_end else py.end_date
    return _fmt_window(start, end)


def _gst_suffix(term: ProductTerm | None) -> str:
    if term is not None and term.gst_included is True:
        return "(sbj to GST)"
    if term is not None and term.gst_included is False:
        return "(GST exempt)"
    return "(GST-exclusive)"


def _set_widths(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _style_row(ws: Worksheet, font: Font | None = None, wrap_cols: tuple[int, ...] = ()) -> int:
    row = ws.max_row
    if font is not None:
        for cell in ws[row]:
            cell.font = font
    for col in wrap_cols:
        ws.cell(row=row, column=col).alignment = _WRAP
    return row


def _border_row(ws: Worksheet, first: int, last: int, row: int | None = None) -> None:
    """Grid the table cells — every table row gets a full thin border so the
    workbook reads like the reference slips, not floating text."""
    r = row or ws.max_row
    for col in range(first, last + 1):
        ws.cell(row=r, column=col).border = _BORDER


def _label_value_rows(ws: Worksheet, rows: list[tuple[str, str] | None]) -> None:
    """Header block: label in col A, value in col C (reference layout)."""
    for entry in rows:
        if entry is None:
            ws.append([])
            continue
        label, value = entry
        ws.append([label, "", value])
        row = ws.max_row
        ws.cell(row=row, column=1).font = _HEADER
        ws.cell(row=row, column=3).alignment = _WRAP


# ── Basis of Cover ───────────────────────────────────────────────────────────


def _basis_columns(categories: list[Category]) -> list[tuple[str, str]]:
    """(header, plan_assignments key) — only the columns this product uses."""
    pas = [_pa(c) for c in categories]
    cols: list[tuple[str, str]] = []
    if any(pa.get("plan_code") not in (None, "") for pa in pas):
        cols.append(("Plan", "plan_code"))
    cols.append(("* No. of employees", "num_employees"))
    if any(pa.get("basis") for pa in pas):
        cols.append(("Basis", "basis"))
    if any(pa.get("sum_insured") is not None for pa in pas):
        cols.append(("* Sum Insured (S$)", "sum_insured"))
    if any(pa.get("estimated_annual_earnings") is not None for pa in pas):
        cols.append(("* Estimated annual earnings", "estimated_annual_earnings"))
    return cols


def _write_basis_of_cover(ws: Worksheet, categories: list[Category]) -> None:
    ws.append(["Basis of Cover :"])
    _style_row(ws, font=_SECTION)
    cols = _basis_columns(categories)

    last_col = 4 + len(cols)

    def _header() -> None:
        ws.append(["", "Insured", "Category", "Participation"] + [h for h, _ in cols])
        _style_row(ws, font=_HEADER)
        _border_row(ws, 2, last_col)

    prev_key: tuple[str, str] | None = None
    prev_name = ""
    for c in categories:
        pa = _pa(c)
        insured = _insured_text(pa)
        participation = _participation_text(c)
        key = (insured, participation)
        new_block = key != prev_key
        if new_block:
            if prev_key is not None:
                ws.append([])
            _header()
            prev_name = ""
        values = []
        for _, k in cols:
            v = pa.get(k)
            if v is None:
                v = ""
            elif k == "plan_code":
                v = str(v)
            elif k == "basis":
                # Stored as string; a pure amount renders as a number ("2000000.0"
                # → 2,000,000), salary-multiple expressions stay text.
                v = _numeric_or_text(v)
            values.append(v)
        # Continuation rows (same category, next plan code) leave the category
        # blank, mirroring the slips' carried-down cells.
        name = c.display_name if c.display_name != prev_name else ""
        ws.append([
            "",
            insured if new_block else "",
            name,
            participation if new_block else "",
            *values,
        ])
        _style_row(ws, wrap_cols=(2, 3, 4))
        _border_row(ws, 2, last_col)
        for i, ((header, _), value) in enumerate(zip(cols, values, strict=True)):
            if ("Sum Insured" in header or "earnings" in header or header == "Basis") \
                    and isinstance(value, (int, float)):
                ws.cell(row=ws.max_row, column=5 + i).number_format = _COUNT
        prev_key = key
        prev_name = c.display_name
    ws.append(["", _DISCLAIMER])
    _style_row(ws, font=_NOTE)


# ── Rate section ─────────────────────────────────────────────────────────────


def _tier_codes(categories: list[Category]) -> list[str]:
    codes: set[str] = set()
    for c in categories:
        tiers = _pa(c).get("rate_tiers")
        if isinstance(tiers, dict):
            codes.update(tiers)
    return sorted(codes, key=lambda c: (_TIER_ORDER.get(c, 99), c))


def _write_tiered_rates(
    ws: Worksheet, categories: list[Category], codes: list[str], blank: bool
) -> float | None:
    """Render the tiered rate table; return the sum of the premiums actually
    SHOWN (None if none), so the caller's total always reconciles with the
    printed rows."""
    # Row 1: tier codes above each Rate/Premium pair; row 2: the pair headers.
    head1 = ["Rate :", "", ""]
    head2 = ["", "Insured", "Category", "Plan"]
    for code in codes:
        head1 += [code, ""]
        head2 += ["Rate", "Premium"]
    last_col = 4 + 2 * len(codes)
    ws.append([*head1, ""])
    _style_row(ws, font=_SECTION)
    _border_row(ws, 4, last_col)
    ws.append(head2)
    _style_row(ws, font=_HEADER)
    _border_row(ws, 2, last_col)

    running = 0.0
    found = False
    prev_insured: str | None = None
    prev_name = ""
    for c in categories:
        pa = _pa(c)
        tiers = pa.get("rate_tiers")
        if not isinstance(tiers, dict) or not tiers:
            continue
        insured = _insured_text(pa)
        row: list[Any] = [
            "",
            insured if insured != prev_insured else "",
            c.display_name if c.display_name != prev_name else "",
            str(pa.get("plan_code") or ""),
        ]
        for code in codes:
            cell = tiers.get(code) or {}
            premium = cell.get("premium") if isinstance(cell, dict) else None
            if blank or not premium:
                # A stored zero premium is the parser's "no committed premium"
                # (voluntary tiers) — the slips leave those cells empty.
                rate = "" if blank else (cell.get("rate") if isinstance(cell, dict) else "")
                row += [rate, ""]
            else:
                row += [cell.get("rate"), premium]
                running += float(premium)
                found = True
        ws.append(row)
        r = _style_row(ws, wrap_cols=(2, 3))
        _border_row(ws, 2, last_col)
        for i in range(len(codes)):
            ws.cell(row=r, column=6 + 2 * i).number_format = _MONEY
        prev_insured = insured
        prev_name = c.display_name
    return running if found else None


def _derived_premium(pa: dict[str, Any]) -> float | None:
    """SI-rated premium the slip leaves implicit: sum insured / 1,000 x rate.
    Only when the category is rated per S$1,000 SI and no stored premium exists
    (GPA-style sheets state the rate but not per-row premiums)."""
    if pa.get("annual_premium") is not None or pa.get("rate_basis") != "per_1000_si":
        return None
    rate, si = pa.get("premium_rate"), pa.get("sum_insured")
    if rate is None or si in (None, 0):
        return None
    return round(float(si) / 1000.0 * float(rate), 2)


def _write_flat_rates(
    ws: Worksheet,
    categories: list[Category],
    earnings_based: bool,
    with_label: bool,
    blank: bool,
) -> float | None:
    """Render the flat rate table; return the sum of the premiums actually
    SHOWN (None if none), so the caller's total reconciles with the rows.

    Two premium kinds are treated differently, matching the slips:
    * Block-level (``annual_premium`` / ``premium_note``): the parser copies one
      figure onto every category in a block, so it prints once — blank
      consecutive repeats for the same insured, and count it once.
    * Derived per-S$1,000-SI (``_derived_premium``): a genuine per-row figure —
      never blanked (two cohorts with equal SI still each carry it) and always
      summed.
    """
    amount_header = (
        "* Estimated annual earnings" if earnings_based else "Sum Insured ( SI )"
    )
    rate_header = "Rate" if earnings_based else "Rate per S$1000 sum insured"
    if with_label:
        ws.append(["Rate :"])
        _style_row(ws, font=_SECTION)
    ws.append(["", "Insured", "Category", amount_header, rate_header, "Annual Premium"])
    _style_row(ws, font=_HEADER)
    _border_row(ws, 2, 6)
    prev_insured: str | None = None
    prev_display: Any = object()
    running = 0.0
    found = False
    for c in categories:
        pa = _pa(c)
        insured = _insured_text(pa)
        amount = (
            pa.get("estimated_annual_earnings")
            if earnings_based
            else pa.get("sum_insured")
        )
        # `premium` is what the cell shows; `numeric` is what it contributes to
        # the total (None when blanked or non-numeric, e.g. an annotated note).
        numeric: float | None = None
        if blank:
            rate: Any = ""
            premium: Any = ""
        else:
            rate = pa.get("premium_rate")
            derived = _derived_premium(pa)
            if derived is not None:
                premium = derived
                numeric = derived
            else:
                # An annotated premium ("S$3,169.80 subject to minimum…") carries
                # its full text; the plain number is the fallback.
                display = pa.get("premium_note") or pa.get("annual_premium")
                stored = pa.get("annual_premium")
                if (display is not None and display == prev_display
                        and insured == prev_insured):
                    premium = ""  # block copy already printed above
                else:
                    premium = display if display is not None else ""
                    if display is not None:
                        prev_display = display
                    numeric = float(stored) if stored is not None else None
        ws.append([
            "",
            insured if insured != prev_insured else "",
            c.display_name,
            amount,
            rate,
            premium,
        ])
        r = _style_row(ws, wrap_cols=(2, 3, 6))
        _border_row(ws, 2, 6)
        ws.cell(row=r, column=4).number_format = _COUNT
        if not isinstance(premium, str):
            ws.cell(row=r, column=6).number_format = _MONEY
        if numeric is not None:
            running += numeric
            found = True
        prev_insured = insured
    return running if found else None


def _write_per_member_rates(
    ws: Worksheet, categories: list[Category], with_label: bool, blank: bool
) -> float | None:
    """Per-member products (GCGP/GCSP/GD): one rate per head, with a separate
    dependant rate. The slips price these per PLAN, not per category — the
    same rate row is replicated across every cohort sharing the plan — so
    collapse to one block per plan: "1 - Employees" / "1 - Dependents"
    (combined when the two rates match, mirroring the reference slips).

    Returns the sum of the premiums SHOWN (None if none) so the caller's total
    reconciles with the rows."""
    if with_label:
        ws.append(["Rate :"])
        _style_row(ws, font=_SECTION)
    ws.append(["", "Insured", "Plan", "Rate", "Premium"])
    _style_row(ws, font=_HEADER)
    _border_row(ws, 2, 5)
    seen_plans: set[str] = set()
    prev_insured: str | None = None
    running = 0.0
    found = False
    for c in categories:
        pa = _pa(c)
        plan = str(pa.get("plan_code") or "")
        if plan in seen_plans:
            continue
        seen_plans.add(plan)
        insured = _insured_text(pa)
        emp_rate = pa.get("premium_rate")
        dep_rate = pa.get("dependant_rate")
        stored = pa.get("annual_premium")
        premium: Any = "" if blank else (
            pa.get("premium_note") or pa.get("annual_premium") or ""
        )
        show_insured = insured if insured != prev_insured else ""
        combined = dep_rate is not None and dep_rate == emp_rate
        label = f"{plan} - Employees / Dependents" if combined else f"{plan} - Employees"
        ws.append(["", show_insured, label, "" if blank else emp_rate, premium])
        row = _style_row(ws, wrap_cols=(2, 5))
        _border_row(ws, 2, 5)
        if not isinstance(premium, str):
            ws.cell(row=row, column=5).number_format = _MONEY
        if not blank and stored is not None:
            running += float(stored)
            found = True
        if dep_rate is not None and not combined:
            ws.append(["", "", f"{plan} - Dependents", "" if blank else dep_rate, ""])
            _border_row(ws, 2, 5)
        prev_insured = insured
    return running if found else None


def _write_voluntary_rates(
    ws: Worksheet, categories: list[Category], blank: bool
) -> None:
    # Many voluntary tiers usually share ONE published rate table (the slips
    # print it once) — dedupe identical band lists; a per-category heading only
    # appears when genuinely distinct tables exist.
    blocks: list[tuple[str, list]] = []
    seen: set[str] = set()
    for c in categories:
        bands = _pa(c).get("voluntary_rates")
        if not isinstance(bands, list) or not bands:
            continue
        key = json.dumps(bands, sort_keys=True, default=str)
        if key in seen:
            continue
        seen.add(key)
        blocks.append((c.display_name, bands))
    if not blocks:
        return
    ws.append([])
    ws.append(["", "Voluntary Rates"])
    _style_row(ws, font=_SECTION)
    for name, bands in blocks:
        if len(blocks) > 1:
            ws.append(["", name])
            _style_row(ws, font=_HEADER)
        ws.append(["", "Based on Age Last Birthday", "Rate per 1,000 Sum assured (S$)"])
        _style_row(ws, font=_HEADER)
        _border_row(ws, 2, 3)
        for band in bands:
            if not isinstance(band, dict):
                continue
            ws.append(["", str(band.get("label") or ""), "" if blank else band.get("rate")])
            _border_row(ws, 2, 3)


def _write_rate_section(
    ws: Worksheet,
    categories: list[Category],
    term: ProductTerm | None,
    mode: Mode,
) -> None:
    blank = mode == "quotation"
    tiered = [
        c for c in categories
        if isinstance(_pa(c).get("rate_tiers"), dict) and _pa(c).get("rate_tiers")
    ]
    # Per-member products publish a head rate + separate dependant rate, priced
    # per plan (GCGP/GCSP/GD) — they get the collapsed per-plan table.
    per_member = [
        c for c in categories
        if c not in tiered
        and _pa(c).get("dependant_rate") is not None
        and _pa(c).get("sum_insured") is None
    ]
    flat = [
        c for c in categories
        if c not in tiered
        and c not in per_member
        and (
            blank
            or any(
                _pa(c).get(k) is not None
                for k in ("premium_rate", "annual_premium", "premium_note")
            )
        )
    ]
    has_voluntary = any(
        isinstance(_pa(c).get("voluntary_rates"), list) and _pa(c)["voluntary_rates"]
        for c in categories
    )
    if not tiered and not per_member and not flat and not has_voluntary:
        return
    # Each writer returns the sum of the premiums it actually printed (None if
    # none), so the "Annual Premium" total below is the sum of what's on the
    # page — the rows always reconcile with the total by construction.
    subtotals: list[float] = []
    ws.append([])
    if tiered:
        t = _write_tiered_rates(ws, tiered, _tier_codes(tiered), blank)
        if t is not None:
            subtotals.append(t)
    if per_member:
        if tiered:
            ws.append([])
        t = _write_per_member_rates(ws, per_member, with_label=not tiered, blank=blank)
        if t is not None:
            subtotals.append(t)
    if flat:
        if tiered or per_member:
            ws.append([])
        earnings_based = any(
            _pa(c).get("estimated_annual_earnings") is not None for c in flat
        )
        t = _write_flat_rates(
            ws, flat, earnings_based,
            with_label=not (tiered or per_member), blank=blank,
        )
        if t is not None:
            subtotals.append(t)
    # A product placed purely on voluntary age-banded rates still needs the
    # section label above its table.
    if has_voluntary and not tiered and not per_member and not flat:
        ws.append(["Rate :"])
        _style_row(ws, font=_SECTION)
    _write_voluntary_rates(ws, categories, blank)

    ws.append([])
    label = f"Annual Premium {_gst_suffix(term)} :"
    total = None if blank or not subtotals else sum(subtotals)
    ws.append([label, "", total if total is not None else ""])
    r = _style_row(ws, font=_HEADER)
    if total is not None:
        ws.cell(row=r, column=3).number_format = _MONEY


# ── Schedule of Benefits ─────────────────────────────────────────────────────


def _sob_value(entry: dict[str, Any], col_id: str, first: bool) -> Any:
    if first:
        return entry.get("base_value")
    overrides = entry.get("overrides") or {}
    return overrides.get(col_id, entry.get("base_value"))


def _shared_cover(plans: list[Plan]) -> str | None:
    """The product-level "Cover :" sentence, when the plans agree on one.

    Slip-parsed GHS-style products stamp the SAME cover sentence on every plan
    ("Cover: Reimbursement of eligible inpatient expenses…") — that belongs on
    the Cover line above the SOB, not repeated per plan. A single plan whose
    description is a basis expression ("24x basic monthly salary") stays in
    Plan Details."""
    descs = {
        (p.cover_description or "").strip()
        for p in plans
        if (p.cover_description or "").strip()
    }
    if len(descs) != 1:
        return None
    text = next(iter(descs))
    n = sum(1 for p in plans if (p.cover_description or "").strip())
    if n > 1 or text.lower().startswith("cover"):
        return text
    return None


def _sob_col_label(col: dict[str, Any]) -> str:
    """Reference-style column header: the base plan plus every plan sharing its
    schedule ("Plan 1/U01/U04/U06"), not the fold's "+N" shorthand."""
    codes = [str(c) for c in (col.get("plan_codes") or []) if str(c)]
    if len(codes) > 1:
        return "Plan " + "/".join(codes)
    return str(col.get("label") or (f"Plan {codes[0]}" if codes else ""))


def _write_sob(ws: Worksheet, plans: list[Plan], cover: str | None) -> None:
    """Schedule of Benefits, plans folded to columns (matching value vectors
    collapse — the same model the SOB editor uses)."""
    with_items = [
        {
            "code": p.code,
            "label": p.display_name,
            "benefit_items": (p.benefit_schedule or {}).get("items"),
        }
        for p in plans
        if isinstance(p.benefit_schedule, dict)
        and isinstance(p.benefit_schedule.get("items"), list)
        and p.benefit_schedule["items"]
    ]
    if not with_items:
        return
    sob = sob_from_plan_items(with_items)
    columns = sob.get("columns") or []
    items = sob.get("items") or []
    if not columns or not items:
        return

    ws.append([])
    # The stored sentence usually carries its own "Cover:" prefix — strip it,
    # the label cell already says it.
    cover_text = re.sub(r"(?i)^cover\s*:\s*", "", cover).strip() if cover else ""
    ws.append(["Cover :", cover_text])
    _style_row(ws, font=_HEADER, wrap_cols=(2,))
    ws.cell(row=ws.max_row, column=2).font = Font(bold=False)
    ws.append(["SCHEDULE OF BENEFITS / INSURER / PLAN"])
    _style_row(ws, font=_SECTION)
    ws.append(["No.", "Benefit"] + [_sob_col_label(col) for col in columns])
    _style_row(ws, font=_HEADER)
    sob_last_col = 2 + len(columns)
    _border_row(ws, 1, sob_last_col)
    value_cols = range(3, 3 + len(columns))

    def _value_row(number: str, name: str, entry: dict[str, Any]) -> None:
        values = [
            _sob_value(entry, col.get("id"), i == 0)
            for i, col in enumerate(columns)
        ]
        ws.append([number, name, *values])
        row = ws.max_row
        ws.cell(row=row, column=2).alignment = _WRAP
        for col in value_cols:
            ws.cell(row=row, column=col).alignment = _WRAP
        _border_row(ws, 1, sob_last_col)

    def _limit_rows(limits: Any, indent: str) -> None:
        for lim in limits or []:
            if not isinstance(lim, dict):
                continue
            label = str(lim.get("label") or "").strip()
            value = str(lim.get("value") or "").strip()
            text = f"{label}: {value}" if value else label
            if text:
                ws.append(["", f"{indent}· {text}"])
                _border_row(ws, 1, sob_last_col)

    for it in items:
        name = str(it.get("name") or "")
        if it.get("note"):
            name = f"{name} — {it['note']}"
        _value_row(str(it.get("number") or ""), name, it)
        # Copay axes vary per column (dash-group copay) — one row per property.
        col_props = it.get("column_properties")
        if isinstance(col_props, dict) and col_props:
            keys: list[str] = []
            for props in col_props.values():
                for k in props or {}:
                    if k not in keys:
                        keys.append(k)
            for key in keys:
                values = [
                    (col_props.get(col.get("id")) or {}).get(key, "")
                    for col in columns
                ]
                ws.append(["", f"    · {key}", *values])
                _border_row(ws, 1, sob_last_col)
        # Non-copay `properties` are machine-derived duplicates of the limits
        # (maximum_days ↔ "Maximum no. of days") — never rendered.
        _limit_rows(it.get("limits"), "    ")
        for sub in it.get("sub_items") or []:
            sub_name = str(sub.get("name") or "")
            if sub.get("note"):
                sub_name = f"{sub_name} — {sub['note']}"
            _value_row("", f"    {sub_name}", sub)
            _limit_rows(sub.get("limits"), "        ")


def _write_plan_details(
    ws: Worksheet, plans: list[Plan], shared_cover: str | None
) -> None:
    """Cover-basis metadata (GCI-style plan tiers) — only when a plan carries
    something the Cover line doesn't already say. A cover sentence shared by
    every plan is hoisted onto the Cover line, so rows whose only content is
    that repeated sentence are dropped entirely."""
    def _own_cover(p: Plan) -> str:
        desc = (p.cover_description or "").strip()
        return "" if shared_cover is not None and desc == shared_cover else desc

    # Resolve each plan's own-cover once, then keep the rows that carry anything
    # the Cover line doesn't already say.
    rows = [
        (p, own)
        for p in plans
        if (own := _own_cover(p)) or p.annual_policy_limit or p.report_label
    ]
    if not rows:
        return
    ws.append([])
    ws.append(["Plan Details"])
    _style_row(ws, font=_SECTION)
    ws.append(["Plan", "Cover Description", "Annual Policy Limit", "Report Label"])
    _style_row(ws, font=_HEADER)
    _border_row(ws, 1, 4)
    for p, own in rows:
        ws.append([
            p.code, own,
            p.annual_policy_limit or "", p.report_label or "",
        ])
        _style_row(ws, wrap_cols=(2,))
        _border_row(ws, 1, 4)


# ── Sheet assembly ───────────────────────────────────────────────────────────


def _write_product_sheet(
    ws: Worksheet,
    py: PolicyYear,
    product: Product | None,
    categories: list[Category],
    plans: list[Plan],
    term: ProductTerm | None,
    mode: Mode,
) -> None:
    _set_widths(ws, {
        1: 34, 2: 36, 3: 46, 4: 26, 5: 12, 6: 14, 7: 22, 8: 16, 9: 16, 10: 16,
    })
    title = (
        f"{product.display_name}" if product else "Unassigned categories"
    )
    ws.append([title])
    _style_row(ws, font=_TITLE)
    ws.append([])

    insured = ", ".join(_distinct_insured(categories))
    # A quotation goes out to prospective insurers — the incumbent's identity
    # and issued policy number stay blank alongside the rates.
    insurer = "" if mode == "quotation" or product is None else (product.insurer or "")
    policy_no = (
        "" if mode == "quotation" or term is None else (term.policy_number or "")
    )
    _label_value_rows(ws, [
        ("Group :", ""),
        ("Policyholder :", py.client.name if py.client else ""),
        ("Insured :", insured),
        ("Address :", ""),
        ("Business :", ""),
        ("Period of Insurance :", _coverage_window(py, term)),
        ("Insurer :", insurer),
        ("Pool :", ""),
        ("Policy No. :", policy_no),
        None,
        ("Eligibility :", ""),
        ("Eligibility Date :", ""),
        ("Last entry age :", ""),
        None,
        ("Type of Administration :", ""),
    ])
    ws.append([])

    if categories:
        _write_basis_of_cover(ws, categories)
        _write_rate_section(ws, categories, term, mode)

    if plans:
        ws.append([])
        for label in _TERM_LABELS:
            ws.append([label])
            _style_row(ws, font=_HEADER)
        shared_cover = _shared_cover(plans)
        _write_plan_details(ws, plans, shared_cover)
        _write_sob(ws, plans, shared_cover)


def _build(db: Session, py: PolicyYear, mode: Mode) -> Workbook:
    categories = list(
        db.execute(
            select(Category)
            .where(Category.policy_year_id == py.id)
            .order_by(Category.priority)
        ).scalars()
    )
    plans = list(
        db.execute(
            select(Plan).where(Plan.policy_year_id == py.id).order_by(Plan.code)
        ).scalars()
    )
    terms = {
        t.product_id: t
        for t in db.execute(
            select(ProductTerm).where(ProductTerm.policy_year_id == py.id)
        ).scalars()
    }

    product_ids = {c.product_id for c in categories if c.product_id}
    product_ids |= {p.product_id for p in plans}
    products = sorted(
        db.execute(
            select(Product).where(
                Product.id.in_(product_ids),
                # Product mixes global + per-client rows — scope to this tenant
                # (plus global catalog) even though the ids come from this year's
                # already-scoped categories/plans (defense in depth).
                tenant_or_global(Product.client_id, py.client_id),
            )
        ).scalars()
        if product_ids
        else [],
        key=lambda p: p.code,
    )
    cats_by_product: dict[str | None, list[Category]] = {}
    for c in categories:
        cats_by_product.setdefault(c.product_id, []).append(c)
    plans_by_product: dict[str, list[Plan]] = {}
    for p in plans:
        plans_by_product.setdefault(p.product_id, []).append(p)

    doc_name = "Quotation Slip" if mode == "quotation" else "Placement Slip"
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    _set_widths(overview, {1: 26, 2: 40, 3: 22, 4: 28, 5: 12, 6: 8})
    overview.append([f"{doc_name} — Configured Products"])
    overview.cell(row=1, column=1).font = _TITLE
    overview.append(["Client", py.client.name if py.client else ""])
    overview.append(["Policy Year", str(py.year)])
    # Company-level period = earliest product start → latest product end (the
    # shared envelope), not the bare policy-year span — products can renew
    # off-cycle via ProductTerm coverage overrides.
    overview.append(["Period of Insurance", _fmt_window(*envelope_for(db, py))])
    overview.append(["Status", py.status.value])
    overview.append([
        "Note", "All premium amounts are GST-exclusive, as extracted/configured.",
    ])
    for r in range(2, 7):
        overview.cell(row=r, column=1).font = _HEADER
    overview.append([])
    overview.append(
        ["Code", "Product", "Insurer", "Coverage Period", "Categories", "Plans"]
    )
    for cell in overview[overview.max_row]:
        cell.font = _HEADER
    _border_row(overview, 1, 6)
    for product in products:
        overview.append([
            product.code,
            product.display_name,
            "" if mode == "quotation" else (product.insurer or ""),
            _coverage_window(py, terms.get(product.id)),
            len(cats_by_product.get(product.id, [])),
            len(plans_by_product.get(product.id, [])),
        ])
        _border_row(overview, 1, 6)

    taken: set[str] = {overview.title.lower()}
    for product in products:
        ws = wb.create_sheet(_sheet_title(product.code, taken))
        _write_product_sheet(
            ws, py, product,
            cats_by_product.get(product.id, []),
            plans_by_product.get(product.id, []),
            terms.get(product.id),
            mode,
        )

    unassigned = cats_by_product.get(None, [])
    if unassigned:
        ws = wb.create_sheet(_sheet_title("Unassigned", taken))
        _write_product_sheet(ws, py, None, unassigned, [], None, mode)

    return wb


def build_placement_slip_workbook(db: Session, py: PolicyYear) -> Workbook:
    return _build(db, py, "placement")


def build_quotation_slip_workbook(db: Session, py: PolicyYear) -> Workbook:
    return _build(db, py, "quotation")

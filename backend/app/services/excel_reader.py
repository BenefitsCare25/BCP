"""Excel reader abstraction over openpyxl (.xlsx) and xlrd (.xls).

Both implementations return uniform `Sheet` and `Workbook` views so the parser
never branches on file format. This matters because placement slips arrive as
either format in the wild.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Protocol

Cell = str | int | float | bool | None

# Hard bound on columns read per sheet. Real slips use < 20 columns; some
# workbooks in the wild report thousands of phantom columns (stray formatting
# pushes openpyxl's max_column to ~16k), which would balloon every row with
# None cells. Generous enough for any legitimate slip, small enough to keep
# phantom-column sheets cheap.
MAX_SCAN_COLS = 256


@dataclass(frozen=True)
class Sheet:
    name: str
    rows: list[list[Cell]]


class Workbook(Protocol):
    @property
    def sheet_names(self) -> list[str]: ...

    def sheet(self, name: str) -> Sheet: ...

    def close(self) -> None: ...

    def __enter__(self) -> Workbook: ...

    def __exit__(self, *exc: object) -> None: ...


def open_workbook(path: Path | str) -> Workbook:
    """Open a workbook. Always use as a context manager so file handles release
    on Windows before the caller tries to delete the source file.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _XlrdWorkbook(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _OpenpyxlWorkbook(path)
    raise ValueError(f"Unsupported file extension: {suffix}")


def _coerce(value: object) -> Cell:
    if value is None:
        return None
    # Date/datetime cells (openpyxl yields datetime objects) → ISO date string,
    # so they don't stringify with a misleading "00:00:00" time tail downstream.
    # datetime is a subclass of date, so check it first.
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


class _XlrdWorkbook:
    def __init__(self, path: Path) -> None:
        import xlrd

        # on_demand=False (the default) reads the whole file into memory at
        # open time and releases the OS file handle immediately. Critical on
        # Windows where a held handle blocks the caller from deleting the
        # source temp file.
        self._wb = xlrd.open_workbook(str(path), formatting_info=False, on_demand=False)

    @property
    def sheet_names(self) -> list[str]:
        return list(self._wb.sheet_names())

    def sheet(self, name: str) -> Sheet:
        ws = self._wb.sheet_by_name(name)
        rows: list[list[Cell]] = []
        ncols = min(ws.ncols, MAX_SCAN_COLS)
        for r in range(ws.nrows):
            row: list[Cell] = []
            for c in range(ncols):
                cell = ws.cell(r, c)
                val = cell.value
                if val == "" or val is None:
                    row.append(None)
                else:
                    row.append(_coerce(val))
            rows.append(row)
        return Sheet(name=name, rows=rows)

    def close(self) -> None:
        # xlrd with on_demand=False already released the handle; nothing to do.
        pass

    def __enter__(self) -> _XlrdWorkbook:
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()


class _OpenpyxlWorkbook:
    def __init__(self, path: Path) -> None:
        from openpyxl import load_workbook

        # read_only=True keeps the underlying zip handle open. Caller must
        # close() / use as a context manager so Windows releases the handle
        # before the source file is deleted.
        self._wb = load_workbook(filename=str(path), read_only=True, data_only=True)

    @property
    def sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    def sheet(self, name: str) -> Sheet:
        ws = self._wb[name]
        # A sheet's size is DECLARED in its XML (`<dimension ref="A1:AK4807"/>`)
        # and `read_only=True` trusts that declaration rather than counting.
        # Several non-Excel writers — Go Excelize, which the incumbent
        # platform's exports come from — stamp a placeholder `ref="A1"` and
        # never update it, so a 4,807-row roster reads back as ONE cell.
        # Nothing raises: the header row is a single "Entity", no column maps
        # to Staff ID, every row is skipped, and the upload reports zero
        # records with no error to explain it.
        #
        # `reset_dimensions()` drops the declared bounds so openpyxl computes
        # them while streaming. Only taken when the declaration is degenerate,
        # so the normal path keeps its phantom-column cap (some workbooks in
        # the wild report ~16k columns) — and a genuinely 1x1 sheet costs one
        # cheap extra pass.
        if (ws.max_row or 0) <= 1 or (ws.max_column or 0) <= 1:
            ws.reset_dimensions()
            return Sheet(
                name=name,
                rows=[
                    [_coerce(v) for v in row[:MAX_SCAN_COLS]]
                    for row in ws.iter_rows(values_only=True)
                ],
            )
        rows: list[list[Cell]] = []
        max_col = min(ws.max_column or MAX_SCAN_COLS, MAX_SCAN_COLS)
        for row in ws.iter_rows(max_col=max_col, values_only=True):
            rows.append([_coerce(v) for v in row])
        return Sheet(name=name, rows=rows)

    def close(self) -> None:
        self._wb.close()

    def __enter__(self) -> _OpenpyxlWorkbook:
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

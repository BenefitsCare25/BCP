"""Built-in (non-insurer) member listings — the broker's own full roster.

The insurer listings in ``insurer_listings.py`` answer "what do we submit to
AIA". These answer "what is on file", across every insurer at once, and they are
what a broker hands to the client rather than to an underwriter.

**This is a projection of the insurer listing, not a second implementation.**
The demographic block, the coverage resolution and the underwriting amounts all
come from ``insurer_listings`` — ``product_blocks(db, py)`` with no insurer
filter is the whole of the difference. Rebuilding any of it here would let the
two sheets disagree about one member's cover, which is the single thing a
broker uses them together to check.

What genuinely differs is the product columns. An insurer wants the sum it is
being asked to carry (`Eligible` / `Pending U/W` / `Last Accepted`); a broker
wants the level the member sits at (`Default Plan ID` / `Default Group Option`),
which is what the incumbent platform's built-in listing prints.

**The leading block of each sheet is a COLUMN-FOR-COLUMN clone of the
incumbent's own built-in listing** (`REFERENCE_EMPLOYEE_HEADER`, columns 1-28;
`REFERENCE_DEPENDANT_HEADER`, columns 1-13) — same names, same order, including
its vocabulary where it differs from ours ("User ID" for the staff id on the
employee sheet but "Staff ID" on the dependant one; "Current Job Grade" for the
job grade). These two files are diffed side by side during the migration, so a
renamed or reordered column reads as a data difference. Anything we carry that
the incumbent does not follows AFTER that block (`EXTRA_*_HEADER`, then one
`{Insurer} Member ID` per insurer, then the product columns) rather than being
interleaved into it.

Each extra column is here because something ELSE in the platform reads the
field, so a broker chasing that behaviour can see its input:
`Designation` drives flex tier matching + leave rates, `Eligible to Sell Leave`
gates the sell-leave election, `Has Insurance Cover Last Year` is COMPUTED from
last year's roster (not a column anyone typed), `Employee Status` is the only
leaver/active signal on a sheet that defaults to every person on file, the
member ids are what the portal quotes on a claim, and `Dependant Status` is
what makes a pending self-add visible. Employment Status, Country of Work and
salary Currency were dropped WITH the incumbent's layout: nothing in the app
reads them, and they still ship on the insurer listing.

The upload template (`member_listing_template.py`) carries the SAME leading
block, so a file exported from the incumbent uploads here untouched and HR
learns one vocabulary. Both write `Deletion Date` and both mean the roster's
`termination_date`.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.clock import today as business_today
from app.models import Dependant, Employee, PolicyYear
from app.models.employee import EMPLOYEE_STATUS_ACTIVE
from app.services.insurer_listings import (
    _ROLE_CODES,
    _dependant_amount,
    _employee_coverage,
    _flag,
    _ident,
    _money,
    _prior_cover_flag,
    _prior_year_people,
    configured_insurers_for_year,
    member_id_for_insurer,
    product_blocks,
)
from app.services.insurer_reports import (
    _last_day_of_service,
    append_safe,
    as_date,
    autosize,
    bold_header,
    report_employees,
)
from app.services.leave_pricing_resolver import leave_sell_eligible
from app.services.roster_attributes import (
    DEPENDANT_ID_KEYS,
    EMPLOYEE_ID_KEYS,
    REL_KEYS,
    first_value,
)
from app.services.underwriting import (
    free_cover_limits,
    load_cases,
    report_uw_amounts,
)

# Which slice of the roster the sheet covers. Mirrors the incumbent's
# "Employee Status: All | Active Only" report parameter.
#
# `all` is deliberately EVERY row in the year, not `report_employees`' active +
# in-period leavers: that narrower set is the insurer's billable population, and
# a broker asking for "all" is asking who is on file — including the leaver who
# went out before the period opened, whose absence otherwise looks like a data
# loss rather than a filter.
STATUS_ALL = "all"
STATUS_ACTIVE = "active"
EMPLOYEE_STATUSES = frozenset({STATUS_ALL, STATUS_ACTIVE})

_ENTITY_KEYS = ("entity", "company", "subsidiary")

# Columns 1-28, in the incumbent's own names and order. Do not reorder, and do
# not "fix" the vocabulary: the staff id is `User ID` here and `Staff ID` on the
# dependant sheet because that is how the file being replaced prints them.
REFERENCE_EMPLOYEE_HEADER = [
    "Entity", "User ID", "Employee Name", "Identification No.",
    "Date of Birth", "Gender", "Marital Status",
    "Foreigner Employment Pass", "Nationality", "Monthly Salary",
    "Date of Hire", "Confirmation Date", "Effective Date",
    "Last Day of Service", "Category", "Division", "Department",
    "Cost Centre", "Email Address", "Mobile Phone", "Bank Code",
    "Branch Code", "Bank Account No.", "Company Description",
    "Location Description", "Current Job Grade", "Person Class", "Remarks",
]

# Ours, appended after the clone rather than interleaved. See the module
# docstring for why each one survived the cut.
EXTRA_EMPLOYEE_HEADER = [
    "Employee Status", "Designation",
    "Has Insurance Cover Last Year", "Eligible to Sell Leave",
]

REFERENCE_DEPENDANT_HEADER = [
    "Entity", "Staff ID", "Employee Name", "Employee's Identification No.",
    "Dependant Name", "Dependant's Identification No.", "Relationship",
    "Date of Marriage", "Gender", "Date of Birth", "Effective Date",
    "Remarks", "Deletion Date",
]

EXTRA_DEPENDANT_TAIL = ["Dependant Status"]

# 1-based indexes into REFERENCE_EMPLOYEE_HEADER / REFERENCE_DEPENDANT_HEADER
# whose cells carry a real date or a real number, so Excel is told to draw them
# as one. The incumbent's file prints dates `dd/mm/yyyy` and money `#,##0.00`;
# openpyxl otherwise defaults a date to ISO and — because `first_value` returns
# a string — would store the salary as TEXT, which no broker can sum.
_DATE_FMT = "dd/mm/yyyy"
_MONEY_FMT = "#,##0.00"
_EMPLOYEE_DATE_COLUMNS = (5, 11, 12, 13, 14)
_EMPLOYEE_MONEY_COLUMNS = (10,)
# Date of Marriage, Date of Birth, Effective Date and Deletion Date.
_DEPENDANT_DATE_COLUMNS = (8, 10, 11, 13)


def _number(raw: object) -> object:
    """A roster figure as a NUMBER where it is one, else untouched.

    Rosters carry salaries as "5,500", "5500.00" or a real float depending on
    how the sheet was typed. A value we cannot read as a number is printed as
    it stands rather than blanked — "TBC" in a salary cell is worth seeing.
    """
    if raw in (None, ""):
        return ""
    if isinstance(raw, bool):
        return str(raw)
    if isinstance(raw, (int, float)):
        return raw
    try:
        return float(str(raw).replace(",", "").strip())
    except ValueError:
        return str(raw)


def _format_columns(ws, dates: tuple[int, ...], money: tuple[int, ...]) -> None:
    for row in ws.iter_rows(min_row=2):
        for idx in dates:
            cell = row[idx - 1]
            if isinstance(cell.value, date):
                cell.number_format = _DATE_FMT
        for idx in money:
            cell = row[idx - 1]
            if isinstance(cell.value, (int, float)) and not isinstance(
                cell.value, bool
            ):
                cell.number_format = _MONEY_FMT


def normalize_employee_status(value: str | None) -> str:
    """Coerce the report parameter, defaulting to ``all``.

    Defaulting to ``all`` matches the incumbent — and matters: their file has
    658 rows where an active-only read of the same roster has 518, so a broker
    reconciling the two must get the same population by default.
    """
    wanted = (value or STATUS_ALL).strip().lower()
    return wanted if wanted in EMPLOYEE_STATUSES else STATUS_ALL


def listing_employees(
    db: Session, py: PolicyYear, employee_status: str = STATUS_ALL
) -> list[Employee]:
    if normalize_employee_status(employee_status) == STATUS_ACTIVE:
        return [
            e for e in report_employees(db, py)
            if e.status == EMPLOYEE_STATUS_ACTIVE
        ]
    return list(
        db.execute(
            select(Employee)
            .where(Employee.policy_year_id == py.id)
            .order_by(Employee.employee_name, Employee.staff_id)
        )
        .scalars()
        .all()
    )


@dataclass(frozen=True)
class _MemberIdColumn:
    insurer: str
    header: str


def _member_id_columns(db: Session, py: PolicyYear) -> list[_MemberIdColumn]:
    """One `{Insurer} Member ID` column per insurer the year is placed with.

    The insurer listing has exactly one because it IS one insurer's sheet. A
    built-in listing spans them, and a member carries a different id with each
    (`attribute_values["insurer_member_ids"]` is keyed by insurer name), so
    collapsing them into one column would print whichever happened to be found
    first and silently lose the rest.
    """
    return [
        _MemberIdColumn(insurer=name, header=f"{name} Member ID")
        for name in configured_insurers_for_year(db, py)
    ]


def build_built_in_employee_listing(
    db: Session,
    py: PolicyYear,
    masked: bool = True,
    employee_status: str = STATUS_ALL,
) -> Workbook:
    """Full-company employee listing with each product's default level."""
    wanted_status = normalize_employee_status(employee_status)
    blocks = product_blocks(db, py)
    employees = listing_employees(db, py, wanted_status)
    coverage, _ = _employee_coverage(db, py, employees, blocks)
    cases = load_cases(db, py.id)
    fcl_by_product = free_cover_limits(db, py.id)
    prior_people = _prior_year_people(db, py)
    id_columns = _member_id_columns(db, py)

    header = [
        *REFERENCE_EMPLOYEE_HEADER,
        *EXTRA_EMPLOYEE_HEADER,
        *(c.header for c in id_columns),
    ]
    money_columns = list(_EMPLOYEE_MONEY_COLUMNS)
    for b in blocks:
        if b.lump_sum:
            header += [
                f"{b.report_code} EE Default Plan ID",
                f"{b.report_code} EE Last Accepted Sum Assured",
            ]
            # The sum assured is the one product column carrying a real figure
            # (the dependant amounts render through `_money` as text).
            money_columns.append(len(header))
            # The household's dependant option level, where the product offers
            # one. Deliberately NO "Last Accepted" for these: acceptance is a
            # per-dependant underwriting decision and one employee row cannot
            # carry three children's outcomes. That figure lives on the
            # dependant listing, per life, where it is unambiguous.
            for role in ("spouse", "child"):
                if role in b.role_options:
                    header.append(
                        f"{b.report_code} {_ROLE_CODES[role]} Default Plan ID"
                    )
        else:
            header += [
                f"{b.report_code} Default Plan ID",
                f"{b.report_code} Default Group Option",
            ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    append_safe(ws, header)
    bold_header(ws)

    for emp in employees:
        attrs = emp.attribute_values or {}
        row: list[object] = [
            first_value(attrs, _ENTITY_KEYS) or "",
            emp.staff_id,
            emp.employee_name or "",
            _ident(attrs, EMPLOYEE_ID_KEYS, masked),
            as_date(first_value(attrs, ("date_of_birth", "dob"))),
            first_value(attrs, ("gender", "sex")) or "",
            first_value(attrs, ("marital_status",)) or "",
            first_value(attrs, ("pass",)) or "",
            first_value(attrs, ("nationality",)) or "",
            _number(first_value(attrs, ("salary",))),
            as_date(first_value(attrs, ("date_of_hire", "hire_date"))),
            as_date(first_value(attrs, ("confirmation_date",))),
            as_date(first_value(attrs, ("effective_date",))),
            _last_day_of_service(emp),
            first_value(attrs, ("category",)) or "",
            first_value(attrs, ("division",)) or "",
            first_value(attrs, ("department",)) or "",
            first_value(attrs, ("cost_centre",)) or "",
            first_value(attrs, ("email", "email_address")) or "",
            first_value(attrs, ("mobile", "mobile_phone")) or "",
            first_value(attrs, ("bank_code",)) or "",
            first_value(attrs, ("branch_code",)) or "",
            first_value(attrs, ("bank_account_no",)) or "",
            first_value(attrs, ("company_description",)) or "",
            first_value(attrs, ("location_description",)) or "",
            first_value(attrs, ("job_grade", "grade")) or "",
            # The incumbent prints "Employee" on every row of this sheet — it is
            # what tells an employee line apart from a dependant one in their
            # combined extracts. Defaulted rather than left blank so the column
            # means the same thing when a roster has never carried it.
            first_value(attrs, ("person_class",)) or "Employee",
            first_value(attrs, ("remarks",)) or "",
            # The status column exists because the sheet defaults to `all`: a
            # leaver and an active member are otherwise distinguishable only by
            # a Last Day of Service that many rosters leave blank.
            (emp.status or "").title(),
            first_value(attrs, ("designation",)) or "",
            _prior_cover_flag(emp, prior_people),
            # `_flag`, not a local Yes/No: the column beside it
            # (`_prior_cover_flag`) already renders through it, and two
            # spellings of a boolean on one sheet reads as two different facts.
            _flag(leave_sell_eligible(emp)),
        ]
        row += [member_id_for_insurer(attrs, c.insurer) for c in id_columns]

        per_product = coverage.get(emp.id, {})
        for b in blocks:
            cov = per_product.get(b.product.id)
            if b.lump_sum:
                if cov is None or cov.eligible is None:
                    row += [cov.basis_display if cov else "", ""]
                else:
                    _pending, accepted = report_uw_amounts(
                        cov.eligible,
                        fcl_by_product.get(b.product.id),
                        cases.get((emp.id, b.product.id)),
                    )
                    row += [cov.plan_code or cov.basis_display or "", accepted]
                for role in ("spouse", "child"):
                    if role not in b.role_options:
                        continue
                    amount = _dependant_amount(b, cov, role) if cov else None
                    row.append(_money(amount) if amount is not None else "")
            else:
                if cov is None:
                    row += ["No Coverage", ""]
                else:
                    row += [cov.plan_code or cov.plan_label or "", cov.grouping]
        append_safe(ws, row)

    _format_columns(ws, _EMPLOYEE_DATE_COLUMNS, tuple(money_columns))
    autosize(ws)
    return wb


def build_built_in_dependant_listing(
    db: Session,
    py: PolicyYear,
    masked: bool = True,
    employee_status: str = STATUS_ALL,
) -> Workbook:
    """Full-company dependant listing.

    Carries NO product columns, matching the incumbent's built-in sheet: what a
    broker checks here is who is on file and with what details. Which products
    cover them is the insurer listing's question, and it is answered there per
    insurer rather than smeared across one sheet.

    Unlike the insurer dependant listing, an UNCOVERED dependant still appears.
    A dependant nobody covers is usually a data problem — a pending self-add, a
    missing election — and filtering them out hides exactly the rows a broker
    opened this sheet to find.
    """
    wanted_status = normalize_employee_status(employee_status)
    employees = listing_employees(db, py, wanted_status)
    by_id = {e.id: e for e in employees}
    id_columns = _member_id_columns(db, py)

    deps = list(
        db.execute(
            select(Dependant)
            .where(Dependant.employee_id.in_(list(by_id) or [""]))
            .order_by(Dependant.employee_id, Dependant.id)
        )
        .scalars()
        .all()
    )

    header = [
        *REFERENCE_DEPENDANT_HEADER,
        *(c.header for c in id_columns),
        *EXTRA_DEPENDANT_TAIL,
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dependants"
    append_safe(ws, header)
    bold_header(ws)

    for dep in deps:
        emp = by_id.get(dep.employee_id)
        if emp is None:
            continue
        dattrs = dep.attribute_values or {}
        row: list[object] = [
            first_value(dattrs, _ENTITY_KEYS)
            or first_value(emp.attribute_values or {}, _ENTITY_KEYS)
            or "",
            emp.staff_id,
            emp.employee_name or "",
            _ident(emp.attribute_values or {}, EMPLOYEE_ID_KEYS, masked),
            first_value(dattrs, ("dependant_name", "name")) or "",
            _ident(dattrs, DEPENDANT_ID_KEYS, masked),
            first_value(dattrs, REL_KEYS) or "",
            as_date(first_value(dattrs, ("date_of_marriage",))),
            first_value(dattrs, ("gender", "sex")) or "",
            as_date(first_value(dattrs, ("date_of_birth", "dob"))),
            as_date(first_value(dattrs, ("effective_date",))),
            first_value(dattrs, ("remarks",)) or "",
            # ONE deletion date, under the incumbent's name for it. Prefer the
            # system's record of the removal; fall back to the date the roster
            # STATED, which is what ADC reads and is all there is while a
            # termination is uploaded but not yet applied. Printing both as
            # separate columns made the sheet un-round-trippable — two headers
            # competing for the same `termination_date` key on re-upload, where
            # only the leftmost wins and a pending date could be blanked.
            dep.terminated_effective
            or as_date(first_value(dattrs, ("termination_date",))),
        ]
        # A dependant's insurer member id usually sits on the EMPLOYEE row —
        # rosters rarely repeat it per life — so fall back the same way the
        # panel e-card does rather than printing a blank.
        row += [
            member_id_for_insurer(dattrs, c.insurer)
            or member_id_for_insurer(emp.attribute_values or {}, c.insurer)
            for c in id_columns
        ]
        row.append((dep.status or "").replace("_", " ").title())
        append_safe(ws, row)

    _format_columns(ws, _DEPENDANT_DATE_COLUMNS, ())
    autosize(ws)
    return wb


def built_in_filename(kind: str, today: date | None = None) -> str:
    return f"built-in-{kind}-listing-report-{(today or business_today()):%Y%m%d}.xlsx"

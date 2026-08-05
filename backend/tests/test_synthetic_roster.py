"""The synthetic roster generator must produce files the parsers can read."""
from __future__ import annotations

from pathlib import Path

import pytest

from app.services.roster_parser import (
    EMPLOYEE_COLUMN_MAP,
    _build_column_map,
    parse_dependant_workbook,
    parse_employee_workbook,
)
from tests.fixtures.generate_synthetic_roster import generate


def test_column_map_binds_each_attribute_once() -> None:
    # A secondary "Job Grade ..." header must not steal the job_grade binding
    # from the exact "Job Grade" column (else it overwrites the real code).
    header = ["Staff ID", "Category", "Job Grade", "Job Grade Band"]
    out = _build_column_map(header, EMPLOYEE_COLUMN_MAP)
    assert out[2] == "job_grade"  # exact match wins
    assert 3 not in out  # secondary column left unmapped, not a duplicate binding
    assert out[0] == "staff_id"
    assert out[1] == "category"


def test_column_map_still_allows_loose_prefix() -> None:
    # Loose prefix match still works when there's no exact header.
    out = _build_column_map(["Monthly Salary (SGD)"], EMPLOYEE_COLUMN_MAP)
    assert out[0] == "salary"


def test_column_map_binds_entity_column() -> None:
    # The CDL/STM "Entity" column carries the employing legal entity — it
    # drives the insured-entity gate in matching for multi-subsidiary schemes.
    out = _build_column_map(["Entity", "Staff ID", "Category"], EMPLOYEE_COLUMN_MAP)
    assert out[0] == "entity"
    # A "Company Email" header must NOT be mistaken for the entity (there is
    # deliberately no bare "company" alias).
    out = _build_column_map(["Company Email", "Staff ID"], EMPLOYEE_COLUMN_MAP)
    assert out.get(0) != "entity"


def test_excel_reader_coerces_dates_to_iso() -> None:
    # Excel date/datetime cells must store as "YYYY-MM-DD" (no "00:00:00" tail
    # that later blanks the fact-find age tables); non-dates pass through.
    from datetime import date, datetime

    from app.services.excel_reader import _coerce

    assert _coerce(datetime(1958, 2, 19, 0, 0)) == "1958-02-19"
    assert _coerce(date(1990, 5, 1)) == "1990-05-01"
    assert _coerce("Manager") == "Manager"
    assert _coerce(5000) == 5000
    assert _coerce(None) is None


def test_iso_date_normalizes_stored_values() -> None:
    from app.services.roster_attributes import iso_date

    assert iso_date("1962-01-27 00:00:00") == "1962-01-27"  # legacy stored form
    assert iso_date("1990-05-01") == "1990-05-01"
    assert iso_date(None) is None
    assert iso_date("not a date") == "not a date"  # surfaced, not dropped


@pytest.fixture(scope="module")
def synthetic_pair(tmp_path_factory: pytest.TempPathFactory) -> tuple[Path, Path]:
    tmp = tmp_path_factory.mktemp("synthetic")
    return generate(rows=25, out_dir=tmp)


def test_generated_employee_workbook_parses(synthetic_pair: tuple[Path, Path]) -> None:
    emp_path, _ = synthetic_pair
    records = parse_employee_workbook(emp_path)
    assert len(records) == 25
    assert all(r.staff_id.startswith("SYN") for r in records)
    assert all(r.employee_name for r in records)
    assert all("category" in r.attributes for r in records)


def test_generated_dependant_workbook_parses(synthetic_pair: tuple[Path, Path]) -> None:
    _, dep_path = synthetic_pair
    records = parse_dependant_workbook(dep_path)
    # Roughly half of 25 employees get 1-2 dependants → expect some records.
    assert len(records) > 0
    assert all(r.employee_staff_id and r.employee_staff_id.startswith("SYN") for r in records)


def test_generator_is_deterministic(tmp_path: Path) -> None:
    a_emp, _ = generate(rows=10, out_dir=tmp_path / "a")
    b_emp, _ = generate(rows=10, out_dir=tmp_path / "b")
    def _key(r):
        return (r.staff_id, r.employee_name, r.attributes.get("category"))
    a = [_key(r) for r in parse_employee_workbook(a_emp)]
    b = [_key(r) for r in parse_employee_workbook(b_emp)]
    assert a == b


# ── Sheet selection ─────────────────────────────────────────────────────────
#
# Both downloadable templates are two-sheet workbooks (Employees / Dependants)
# and both roster tabs offer both, so each parser must read ITS OWN sheet.
# Reading sheet 0 unconditionally made a dependant upload of either template
# parse the employee sheet through DEPENDANT_COLUMN_MAP — Staff ID and Employee
# Name resolve there, so the "no identifying column" guard passes and every
# employee imports as a nameless, DOB-less dependant.


def _two_sheet_workbook(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Staff ID", "Employee Name", "Date of Birth", "Category"])
    emp.append(["E001", "Alice Tan", "1990-01-02", "Executive"])
    emp.append(["E002", "Bob Lim", "1985-06-11", "Manager"])

    dep = wb.create_sheet("Dependants")
    dep.append(["Staff ID", "Employee Name", "Dependant Name", "Relationship"])
    dep.append(["E001", "Alice Tan", "Cara Tan", "Spouse"])
    wb.save(path)


def test_each_parser_reads_its_own_sheet(tmp_path: Path) -> None:
    book = tmp_path / "member-listing-template.xlsx"
    _two_sheet_workbook(book)

    employees = parse_employee_workbook(book)
    assert [e.staff_id for e in employees] == ["E001", "E002"]

    dependants = parse_dependant_workbook(book)
    # One real dependant — NOT one row per employee off the first sheet.
    assert len(dependants) == 1
    assert dependants[0].attributes["dependant_name"] == "Cara Tan"
    assert dependants[0].employee_staff_id == "E001"


def test_single_sheet_workbook_still_parses(tmp_path: Path) -> None:
    # Rosters in the wild carry one arbitrarily-named sheet; those must keep
    # falling back to the first sheet exactly as before.
    from openpyxl import Workbook

    emp_book = tmp_path / "roster.xlsx"
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.active.append(["Staff ID", "Employee Name"])
    wb.active.append(["E009", "Dana Ng"])
    wb.save(emp_book)
    assert [e.staff_id for e in parse_employee_workbook(emp_book)] == ["E009"]

    dep_book = tmp_path / "deps.xlsx"
    wb2 = Workbook()
    wb2.active.title = "STM Dependants Listing"
    wb2.active.append(["Staff ID", "Dependant Name", "Relationship"])
    wb2.active.append(["E009", "Evan Ng", "Child"])
    wb2.save(dep_book)
    deps = parse_dependant_workbook(dep_book)
    assert len(deps) == 1
    assert deps[0].attributes["dependant_name"] == "Evan Ng"


def test_employee_sheet_never_imports_as_dependants(tmp_path: Path) -> None:
    # The other half of the same failure: a SINGLE-sheet employee roster dropped
    # on the Dependants upload. `_read_sheet` can't help (there is no Dependants
    # sheet to prefer), so the row guard has to reject it — Staff ID and
    # Employee Name alone are the SPONSOR's identity, not a dependant's.
    from openpyxl import Workbook

    book = tmp_path / "employee-roster.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(["Staff ID", "Employee Name", "Date of Birth", "Gender"])
    ws.append(["E001", "Alice Tan", "1990-01-02", "Female"])
    ws.append(["E002", "Bob Lim", "1985-06-11", "Male"])
    wb.save(book)

    assert parse_dependant_workbook(book) == []
    # ...while the employee parser still reads it normally.
    assert len(parse_employee_workbook(book)) == 2


def test_dependant_row_keeps_parsing_without_a_name(tmp_path: Path) -> None:
    # The guard tests for a dependant-only COLUMN, not a name. A row carrying
    # just an NRIC, or just relationship + DOB, is a legitimate dependant —
    # test_roster_dedup regression-tests the latter round-tripping the upload.
    from openpyxl import Workbook

    book = tmp_path / "deps.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Dependants"
    ws.append(["Staff ID", "Dependant's Identification No.", "Relationship", "Date of Birth"])
    ws.append(["E001", "S1234567D", "Spouse", "1990-03-04"])
    ws.append(["E002", "", "Child", "2014-05-05"])
    wb.save(book)

    deps = parse_dependant_workbook(book)
    assert len(deps) == 2
    assert deps[0].attributes["dependant_id_no"] == "S1234567D"
    assert deps[1].attributes["relationship"] == "Child"

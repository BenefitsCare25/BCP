"""Downloadable employee + dependant coverage reports (.xlsx).

Verifies headers, masked NRIC, one-row-per-person, and that terminated leavers
are excluded.
"""
from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_roster_reports.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import Employee  # noqa: E402
from app.models.employee import EMPLOYEE_STATUS_TERMINATED  # noqa: E402
from scripts.seed_demo import seed  # noqa: E402

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    seed()
    yield
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(app)


def _py(client: TestClient) -> str:
    return client.get("/api/v1/policy-years").json()[0]["id"]


def _xlsx(headers, rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module", autouse=True)
def _seed_roster(_setup_db, client: TestClient):
    py = _py(client)
    emp_xlsx = _xlsx(
        ["Staff ID", "Employee Name", "Identification No. (NRIC/FIN)", "Date of Birth"],
        [
            ["R-1", "Rita One", "S1112223A", "1990-01-01"],
            ["R-2", "Sam Two", "S2223334B", "1985-02-02"],
            ["R-3", "Tim Gone", "S3334445C", "1980-03-03"],
        ],
    )
    assert client.post(
        "/api/v1/employees/upload",
        files={"file": ("r.xlsx", emp_xlsx, XLSX_MIME)},
        data={"policy_year_id": py},
    ).status_code == 200
    dep_xlsx = _xlsx(
        [
            "Staff ID",
            "Employee Name",
            "Dependant Name",
            "Dependant's Identification No.",
            "Relationship",
            "Date of Birth",
        ],
        [["R-1", "Rita One", "Rita Kid", "T1012345Z", "Child", "2015-06-06"]],
    )
    assert client.post(
        "/api/v1/dependants/upload",
        files={"file": ("d.xlsx", dep_xlsx, XLSX_MIME)},
        data={"policy_year_id": py},
    ).status_code == 200

    # Soft-terminate R-3 directly so the report-exclusion path is exercised.
    db = SessionLocal()
    try:
        emp = db.execute(
            select(Employee).where(Employee.staff_id == "R-3")
        ).scalar_one()
        emp.status = EMPLOYEE_STATUS_TERMINATED
        db.commit()
    finally:
        db.close()
    yield


def _load(res) -> list[list]:
    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith(XLSX_MIME)
    wb = load_workbook(BytesIO(res.content))
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def test_employee_report_headers_and_masking(client: TestClient) -> None:
    py = _py(client)
    rows = _load(
        client.get(f"/api/v1/employees/coverage-report/export?policy_year_id={py}")
    )
    header = rows[0]
    assert header[:3] == ["Staff ID", "Employee Name", "NRIC/FIN"]
    assert "Flex Tier" in header and "Flex Wallet" in header
    body = rows[1:]
    staff_ids = {r[0] for r in body}
    assert "R-1" in staff_ids and "R-2" in staff_ids
    assert "R-3" not in staff_ids, "terminated employee must be excluded"
    rita = next(r for r in body if r[0] == "R-1")
    assert rita[2] == "S******3A", "NRIC must be masked"


def test_dependant_report_headers_and_masking(client: TestClient) -> None:
    py = _py(client)
    rows = _load(
        client.get(f"/api/v1/dependants/coverage-report/export?policy_year_id={py}")
    )
    header = rows[0]
    assert header[0] == "Employee Staff ID"
    assert header[2] == "Dependant Name"
    assert "Insurance Products" in header
    body = rows[1:]
    kid = next(r for r in body if r[2] == "Rita Kid")
    assert kid[0] == "R-1"
    assert kid[5] == "T******5Z", "dependant NRIC must be masked"

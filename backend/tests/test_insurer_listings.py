"""Insurer employee/dependant listings, underwriting cases, member-listing
template, readiness, and the insurer/report-label config surfaces.

Fixture scheme: one insurer ("TestSure") underwriting a lump-sum product (TLIF,
numeric basis + Spouse/Child dependant-scope levels) and a schedule product
(TMD2, plan with an insurer report label + family rate tiers); one orphan
product (TORP) with no insurer for the readiness check.
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_insurer_listings.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import date  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Category,
    Client,
    Dependant,
    Employee,
    Plan,
    PolicyYear,
    Product,
    ProductTerm,
    UnderwritingCase,
    UnderwritingReview,
)
from app.models.policy_year import PolicyYearStatus  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-0000000i1000"
PY_ID = "00000000-0000-0000-0000-0000000i1001"
LIF_PROD = "00000000-0000-0000-0000-0000000i1010"
MED_PROD = "00000000-0000-0000-0000-0000000i1011"
ORP_PROD = "00000000-0000-0000-0000-0000000i1012"
LIF_CAT = "00000000-0000-0000-0000-0000000i1020"
LIF_SP_CAT = "00000000-0000-0000-0000-0000000i1021"
LIF_CH_CAT = "00000000-0000-0000-0000-0000000i1022"
MED_CAT = "00000000-0000-0000-0000-0000000i1023"
MED_PLAN = "00000000-0000-0000-0000-0000000i1030"
EMP_FAMILY = "00000000-0000-0000-0000-0000000i1101"
EMP_SOLO = "00000000-0000-0000-0000-0000000i1102"
DEP_SPOUSE = "00000000-0000-0000-0000-0000000i1201"
DEP_CHILD = "00000000-0000-0000-0000-0000000i1202"


def _user() -> CurrentUser:
    return CurrentUser(
        user_id="00000000-0000-0000-0000-0000000i10ff",
        broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role="broker_admin",
    )


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Listing Co", broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.flush()
        s.add(PolicyYear(
            id=PY_ID, client_id=CLIENT_ID, year=2034,
            start_date=date(2034, 1, 1), end_date=date(2034, 12, 31),
            status=PolicyYearStatus.active,
        ))
        s.add_all([
            Product(
                id=LIF_PROD, client_id=CLIENT_ID, code="TLIF",
                display_name="Test Life", insurer="TestSure",
                has_dependants=False,
            ),
            Product(
                id=MED_PROD, client_id=CLIENT_ID, code="TMD2",
                display_name="Test Medical", insurer="TestSure",
                has_dependants=True,
                product_metadata={"report_code": "TMED"},
            ),
            Product(
                id=ORP_PROD, client_id=CLIENT_ID, code="TORP",
                display_name="Test Orphan", insurer=None,
            ),
        ])
        s.flush()
        s.add_all([
            Category(
                id=LIF_CAT, policy_year_id=PY_ID, product_id=LIF_PROD,
                display_name="All staff", raw_description="All staff",
                plan_assignments={"plan_code": "1", "basis": "100000"},
                source="manual", status="confirmed",
            ),
            Category(
                id=LIF_SP_CAT, policy_year_id=PY_ID, product_id=LIF_PROD,
                display_name="Spouse", raw_description="Spouse",
                plan_assignments={
                    "plan_code": "1", "basis": "60000",
                    "member_scope": "dependant",
                },
                source="manual", status="confirmed",
            ),
            Category(
                id=LIF_CH_CAT, policy_year_id=PY_ID, product_id=LIF_PROD,
                display_name="Child", raw_description="Child",
                plan_assignments={
                    "plan_code": "1", "basis": "50000",
                    "member_scope": "dependant",
                },
                source="manual", status="confirmed",
            ),
            Category(
                id=MED_CAT, policy_year_id=PY_ID, product_id=MED_PROD,
                display_name="All staff medical", raw_description="All staff",
                plan_assignments={
                    "plan_code": "2",
                    "rate_tiers": {"EO": {"rate": 1.0}, "ES": {"rate": 2.0},
                                   "EC": {"rate": 2.0}, "EF": {"rate": 3.0}},
                },
                source="manual", status="confirmed",
            ),
        ])
        s.add(Plan(
            id=MED_PLAN, product_id=MED_PROD, policy_year_id=PY_ID,
            code="2", display_name="Plan 2",
            report_label="1 Bed Restr Hosp / S$80,000",
        ))
        s.add(ProductTerm(
            policy_year_id=PY_ID, product_id=LIF_PROD, free_cover_limit=50000.0,
        ))
        matched = [
            {"category_id": LIF_CAT, "product_code": "TLIF",
             "method": "rule", "confidence": 1.0},
            {"category_id": MED_CAT, "product_code": "TMD2",
             "method": "rule", "confidence": 1.0},
        ]
        s.add_all([
            Employee(
                id=EMP_FAMILY, client_id=CLIENT_ID, policy_year_id=PY_ID,
                staff_id="IL-1", employee_name="Fam Ily",
                attribute_values={
                    "entity": "Listing Co Pte Ltd",
                    "id_no": "S1234567D",
                    "insurer_member_ids": {"TestSure": "TS-001"},
                    "leave_sell_eligible": False,
                },
                derived_attribute_values={}, matched_categories=matched,
                source="csv_import", status="active",
            ),
            Employee(
                id=EMP_SOLO, client_id=CLIENT_ID, policy_year_id=PY_ID,
                staff_id="IL-2", employee_name="So Lo",
                attribute_values={}, derived_attribute_values={},
                matched_categories=matched, source="csv_import", status="active",
            ),
        ])
        s.flush()
        s.add_all([
            Dependant(
                id=DEP_SPOUSE, client_id=CLIENT_ID, policy_year_id=PY_ID,
                employee_id=EMP_FAMILY,
                attribute_values={
                    "dependant_name": "Spo Use", "relationship": "Spouse",
                    "gender": "Female", "date_of_birth": "1990-01-01",
                },
                status="active",
            ),
            Dependant(
                id=DEP_CHILD, client_id=CLIENT_ID, policy_year_id=PY_ID,
                employee_id=EMP_FAMILY,
                attribute_values={
                    "dependant_name": "Chi Ld", "relationship": "Child",
                    "gender": "Male", "date_of_birth": "2015-06-06",
                },
                status="active",
            ),
        ])
        s.commit()
    yield
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture
def client() -> TestClient:
    app.dependency_overrides[get_current_user] = _user
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture
def viewer_client() -> TestClient:
    app.dependency_overrides[get_current_user] = lambda: CurrentUser(
        user_id="00000000-0000-0000-0000-0000000i10fe",
        broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role="broker_viewer",
    )
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture(autouse=True)
def _reset_uw():
    yield
    with SessionLocal() as s:
        s.query(UnderwritingCase).delete()
        s.query(UnderwritingReview).delete()
        s.commit()


def _sheet_rows(content: bytes, sheet: int = 0) -> list[dict]:
    wb = load_workbook(BytesIO(content))
    ws = wb.worksheets[sheet]
    data = list(ws.iter_rows(values_only=True))
    header = list(data[0])
    return [dict(zip(header, r, strict=False)) for r in data[1:]]


# ── Employee listing ─────────────────────────────────────────────────────────


def test_employee_listing_columns_and_values(client: TestClient) -> None:
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
    )
    assert res.status_code == 200, res.text
    rows = {r["Staff ID"]: r for r in _sheet_rows(res.content)}
    fam = rows["IL-1"]
    # Lump-sum block (internal code TLIF, no report_code → as-is).
    assert fam["TLIF EE Basis of Cover"] == "S$ 100,000"
    assert fam["TLIF EE Eligible Sum Insured"] == 100000
    # TLIF has FCL 50k: eligible 100k exceeds it, so 50k is auto-accepted and
    # 50k is pending U/W — reflected from the LIVE FCL without any refresh run.
    assert fam["TLIF EE Sum Insured Pending U/W"] == 50000
    assert fam["TLIF EE Last Accepted Sum Insured"] == 50000
    # Schedule block uses the report_code (TMD2 → TMED) + the plan report label.
    assert fam["TMED Plan/Basis of Cover"] == "1 Bed Restr Hosp / S$80,000"
    assert fam["TMED Family Grouping"] == "EF"
    assert rows["IL-2"]["TMED Family Grouping"] == "EO"
    # Per-insurer member ID + the leave flag + masked NRIC.
    assert fam["TestSure Member ID"] == "TS-001"
    assert fam["Eligible to Sell Leave"] == "false"
    assert rows["IL-2"]["Eligible to Sell Leave"] == "true"
    assert fam["Identification No."] == "S******7D"
    assert fam["Policy Period"] == "1 Jan 2034 to 31 Dec 2034"


def test_unknown_insurer_404s(client: TestClient) -> None:
    # An unknown/mistyped insurer must 404, not silently return a real-looking
    # file with every coverage column missing (misleading-empty report).
    assert (
        client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=Nobody"
        ).status_code
        == 404
    )
    assert (
        client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=Nobody"
        ).status_code
        == 404
    )
    # Configured insurer matches case-insensitively.
    assert (
        client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=testsure"
        ).status_code
        == 200
    )


# ── Dependant listing ────────────────────────────────────────────────────────


def test_dependant_listing_role_blocks(client: TestClient) -> None:
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=TestSure"
    )
    assert res.status_code == 200, res.text
    rows = {r["Dependant Name"]: r for r in _sheet_rows(res.content)}
    assert set(rows) == {"Spo Use", "Chi Ld"}
    spouse = rows["Spo Use"]
    assert spouse["TLIF SP Basis of Cover"] == "S$ 60,000"
    assert spouse["TLIF SP Eligible Sum Insured"] == 60000
    # Child columns stay blank on a spouse row.
    assert spouse["TLIF CH Basis of Cover"] in (None, "")
    assert spouse["TMED Plan/Basis of Cover"] == "1 Bed Restr Hosp / S$80,000"
    child = rows["Chi Ld"]
    assert child["TLIF CH Basis of Cover"] == "S$ 50,000"
    assert child["TLIF CH Eligible Sum Insured"] == 50000


def test_covered_dependant_appears_when_option_unresolved(client: TestClient) -> None:
    # Two child option levels + no per-member election → the level can't be
    # resolved, but a covered child must still appear (blank CH sum insured),
    # never be silently dropped from the insurer listing.
    cat2 = "00000000-0000-0000-0000-0000000i1304"
    child2 = "00000000-0000-0000-0000-0000000i1305"
    with SessionLocal() as s:
        s.add(Category(
            id=cat2, policy_year_id=PY_ID, product_id=LIF_PROD,
            display_name="Child", raw_description="Child",
            plan_assignments={
                "plan_code": "1", "basis": "30000", "member_scope": "dependant",
            },
            source="manual", status="confirmed",
        ))
        s.add(Dependant(
            id=child2, client_id=CLIENT_ID, policy_year_id=PY_ID,
            employee_id=EMP_SOLO,
            attribute_values={
                "dependant_name": "Ambi Child", "relationship": "Child",
                "date_of_birth": "2016-01-01",
            },
            status="active",
        ))
        s.commit()
    try:
        res = client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=TestSure"
        )
        rows = {r["Dependant Name"]: r for r in _sheet_rows(res.content)}
        assert "Ambi Child" in rows  # covered → present, not dropped
        assert rows["Ambi Child"]["TLIF CH Eligible Sum Insured"] in (None, "")
    finally:
        with SessionLocal() as s:
            for oid, model in ((child2, Dependant), (cat2, Category)):
                obj = s.get(model, oid)
                if obj is not None:
                    s.delete(obj)
            s.commit()


def test_in_period_terminated_dependant_included(client: TestClient) -> None:
    # A dependant who terminated mid-year is still on the insurer listing (with
    # a deletion date); a pre-period leaver is excluded.
    in_period = "00000000-0000-0000-0000-0000000i1301"
    pre_period = "00000000-0000-0000-0000-0000000i1302"
    with SessionLocal() as s:
        s.add_all([
            Dependant(
                id=in_period, client_id=CLIENT_ID, policy_year_id=PY_ID,
                employee_id=EMP_SOLO,
                attribute_values={
                    "dependant_name": "Leaver Spouse", "relationship": "Spouse",
                    "date_of_birth": "1988-02-02",
                },
                status="terminated", terminated_effective=date(2034, 6, 30),
            ),
            Dependant(
                id=pre_period, client_id=CLIENT_ID, policy_year_id=PY_ID,
                employee_id=EMP_SOLO,
                attribute_values={
                    "dependant_name": "Ancient Child", "relationship": "Child",
                    "date_of_birth": "2010-01-01",
                },
                status="terminated", terminated_effective=date(2033, 6, 30),
            ),
        ])
        s.commit()
    try:
        res = client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=TestSure"
        )
        rows = {r["Dependant Name"]: r for r in _sheet_rows(res.content)}
        assert "Leaver Spouse" in rows
        assert "Ancient Child" not in rows
        leaver = rows["Leaver Spouse"]
        assert leaver["TLIF SP Eligible Sum Insured"] == 60000
        assert str(leaver["Deletion Date"]).startswith("2034-06-30")
        # The sponsoring employee's family grouping reflects the (leaving) spouse.
        emp_rows = {
            r["Staff ID"]: r
            for r in _sheet_rows(
                client.get(
                    f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
                ).content
            )
        }
        assert emp_rows["IL-2"]["TMED Family Grouping"] == "ES"
    finally:
        with SessionLocal() as s:
            for did in (in_period, pre_period):
                dep = s.get(Dependant, did)
                if dep is not None:
                    s.delete(dep)
            s.commit()


def test_unclassified_relationship_not_reported_employee_only(
    client: TestClient,
) -> None:
    # A covered dependant with a relationship that doesn't classify as
    # spouse/child must still lift the employee off "EO" (billed on the
    # with-dependants side → EC), never contradict the dependant listing.
    ward = "00000000-0000-0000-0000-0000000i1303"
    with SessionLocal() as s:
        s.add(Dependant(
            id=ward, client_id=CLIENT_ID, policy_year_id=PY_ID,
            employee_id=EMP_SOLO,
            attribute_values={
                "dependant_name": "Ward Person", "relationship": "Ward",
                "date_of_birth": "2012-03-03",
            },
            status="active",
        ))
        s.commit()
    try:
        emp_rows = {
            r["Staff ID"]: r
            for r in _sheet_rows(
                client.get(
                    f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
                ).content
            )
        }
        assert emp_rows["IL-2"]["TMED Family Grouping"] == "EC"
    finally:
        with SessionLocal() as s:
            dep = s.get(Dependant, ward)
            if dep is not None:
                s.delete(dep)
            s.commit()


# ── Underwriting ─────────────────────────────────────────────────────────────


def test_underwriting_refresh_and_decision_flow(client: TestClient) -> None:
    res = client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    assert res.status_code == 200, res.text
    body = res.json()
    # Both employees are at 100k vs FCL 50k; spouse 60k and child 50k → only
    # the spouse exceeds (50k is AT the limit, not above).
    assert body["opened"] == 3

    queue = client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()
    # Reviews group by (life, insurer): IL-1 the member, IL-1's spouse, IL-2.
    assert queue["total"] == 3 and queue["open"] == 3
    by_subject = {
        (i["staff_id"], i["subject_type"]): i for i in queue["items"]
    }
    fam = by_subject[("IL-1", "employee")]
    assert fam["insurer"] == "TestSure"
    assert fam["relationship"] == "Self"
    assert fam["status"] == "pending_requirements"
    fam_line = fam["cases"][0]
    assert fam_line["requested_si"] == 100000
    assert fam_line["guaranteed_si"] == 50000  # auto-covered at FCL
    assert fam_line["accepted_si"] == 50000
    assert fam_line["pending_si"] == 50000
    spouse = by_subject[("IL-1", "dependant")]
    assert spouse["relationship"] == "Spouse"
    assert spouse["cases"][0]["requested_si"] == 60000
    assert spouse["cases"][0]["pending_si"] == 10000

    # Listing reflects the pending amounts.
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
    )
    rows = {r["Staff ID"]: r for r in _sheet_rows(res.content)}
    assert rows["IL-1"]["TLIF EE Sum Insured Pending U/W"] == 50000
    assert rows["IL-1"]["TLIF EE Last Accepted Sum Insured"] == 50000

    # Broker records the insurer's acceptance at 80k.
    res = client.patch(
        f"/api/v1/underwriting/cases/{fam_line['id']}",
        json={
            "status": "approved_standard",
            "accepted_si": 80000,
            "remarks": "Loaded",
        },
    )
    assert res.status_code == 200, res.text
    assert res.json()["cases"][0]["pending_si"] == 0

    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
    )
    rows = {r["Staff ID"]: r for r in _sheet_rows(res.content)}
    assert rows["IL-1"]["TLIF EE Sum Insured Pending U/W"] == 0
    assert rows["IL-1"]["TLIF EE Last Accepted Sum Insured"] == 80000

    # Case workflow status + requirements live on the review.
    res = client.patch(
        f"/api/v1/underwriting/reviews/{fam['id']}",
        json={"status": "pending_insurer", "requirements": "Full medical report"},
    )
    assert res.status_code == 200, res.text
    assert res.json()["status"] == "pending_insurer"
    assert res.json()["requirements"] == "Full medical report"

    # Re-sync keeps the decided case (no duplicate, nothing reopened).
    res = client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    assert res.json()["opened"] == 0


def test_fcl_change_resyncs_pending_case(client: TestClient) -> None:
    # A pending case auto-accepts at the FCL; moving the FCL (without any change
    # to eligible SI) must re-sync the auto-accepted amount, not leave it stale.
    # The product-terms PUT now runs the sync itself — no manual refresh needed.
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    try:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 70000},
        )
        queue = client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()
        emp = next(
            i for i in queue["items"]
            if i["staff_id"] == "IL-1" and i["subject_type"] == "employee"
        )
        line = emp["cases"][0]
        assert line["guaranteed_si"] == 70000  # re-synced to the new FCL
        assert line["accepted_si"] == 70000
        assert line["pending_si"] == 30000
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 50000},
        )


def test_age_trigger_new_hire_vs_existing(client: TestClient) -> None:
    """NEL age gate: a new hire above the age gets guaranteed 0 (whole SI
    underwritten); an existing life keeps last year's covered SI and only the
    increase is pending."""
    from app.models import UnderwritingReview

    PREV_PY = "00000000-0000-0000-0000-0000000i1901"
    PREV_CAT = "00000000-0000-0000-0000-0000000i1902"
    PREV_EMP = "00000000-0000-0000-0000-0000000i1903"
    NEW_EMP = "00000000-0000-0000-0000-0000000i1904"
    with SessionLocal() as s:
        # Previous benefit year: IL-2 was covered at 80k on the same product.
        s.add(PolicyYear(
            id=PREV_PY, client_id=CLIENT_ID, year=2033,
            start_date=date(2033, 1, 1), end_date=date(2033, 12, 31),
            status=PolicyYearStatus.archived,
        ))
        s.flush()
        s.add(Category(
            id=PREV_CAT, policy_year_id=PREV_PY, product_id=LIF_PROD,
            display_name="All staff", raw_description="All staff",
            plan_assignments={"plan_code": "1", "basis": "80000"},
            source="manual", status="confirmed",
        ))
        s.flush()
        s.add(Employee(
            id=PREV_EMP, client_id=CLIENT_ID, policy_year_id=PREV_PY,
            staff_id="IL-2", employee_name="So Lo",
            attribute_values={"date_of_birth": "1960-06-01"},
            derived_attribute_values={},
            matched_categories=[{
                "category_id": PREV_CAT, "product_code": "TLIF",
                "method": "rule", "confidence": 1.0,
            }],
            source="csv_import", status="active",
        ))
        # Current year: IL-2 (existing, ANB 75 at 2034-01-01) + IL-9 (new hire,
        # same age) both at 100k eligible.
        emp2 = s.get(Employee, EMP_SOLO)
        emp2.attribute_values = {
            **(emp2.attribute_values or {}), "date_of_birth": "1960-06-01",
        }
        s.add(Employee(
            id=NEW_EMP, client_id=CLIENT_ID, policy_year_id=PY_ID,
            staff_id="IL-9", employee_name="New Hire",
            attribute_values={"date_of_birth": "1960-06-01"},
            derived_attribute_values={},
            matched_categories=[{
                "category_id": LIF_CAT, "product_code": "TLIF",
                "method": "rule", "confidence": 1.0,
            }],
            source="csv_import", status="active",
        ))
        s.commit()
    try:
        res = client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"nel_age_limit": 70},
        )
        assert res.status_code == 200, res.text
        assert res.json()["nel_age_limit"] == 70

        queue = client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()
        by_subject = {
            (i["staff_id"], i["subject_type"]): i for i in queue["items"]
        }
        # Existing life: guaranteed = last year's covered SI (80k), only the
        # 20k increase pending.
        existing = by_subject[("IL-2", "employee")]["cases"][0]
        assert existing["guaranteed_si"] == 80000
        assert existing["pending_si"] == 20000
        # New hire above the age gate: nothing guaranteed, whole SI pending.
        new_hire = by_subject[("IL-9", "employee")]["cases"][0]
        assert new_hire["guaranteed_si"] == 0
        assert new_hire["pending_si"] == 100000
        # IL-1 (no DOB → age unknown) stays on the plain SI trigger.
        fam = by_subject[("IL-1", "employee")]["cases"][0]
        assert fam["guaranteed_si"] == 50000
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"nel_age_limit": None},
        )
        with SessionLocal() as s:
            s.query(UnderwritingCase).delete()
            s.query(UnderwritingReview).delete()
            for model, oid in (
                (Employee, NEW_EMP), (Employee, PREV_EMP),
                (Category, PREV_CAT), (PolicyYear, PREV_PY),
            ):
                obj = s.get(model, oid)
                if obj is not None:
                    s.delete(obj)
            emp2 = s.get(Employee, EMP_SOLO)
            attrs = dict(emp2.attribute_values or {})
            attrs.pop("date_of_birth", None)
            emp2.attribute_values = attrs
            s.commit()


def test_salary_multiple_basis_feeds_underwriting(client: TestClient) -> None:
    """A '24 times basic monthly salary' basis resolves against the roster
    salary, so salary-based lump-sum products trigger underwriting too."""
    with SessionLocal() as s:
        cat = s.get(Category, LIF_CAT)
        original = dict(cat.plan_assignments or {})
        cat.plan_assignments = {
            **original, "basis": "24 times basic monthly salary",
        }
        emp = s.get(Employee, EMP_FAMILY)
        emp.attribute_values = {**(emp.attribute_values or {}), "salary": "5,000"}
        s.commit()
    try:
        client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
        queue = client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()
        by_subject = {
            (i["staff_id"], i["subject_type"]): i for i in queue["items"]
        }
        # 24 x 5000 = 120k vs FCL 50k → 70k pending.
        fam = by_subject[("IL-1", "employee")]["cases"][0]
        assert fam["requested_si"] == 120000
        assert fam["pending_si"] == 70000
        # IL-2 has no salary on file → SI unresolvable → no employee case.
        assert ("IL-2", "employee") not in by_subject
    finally:
        with SessionLocal() as s:
            cat = s.get(Category, LIF_CAT)
            cat.plan_assignments = original
            emp = s.get(Employee, EMP_FAMILY)
            attrs = dict(emp.attribute_values or {})
            attrs.pop("salary", None)
            emp.attribute_values = attrs
            s.commit()


def test_nel_autofill_fills_blanks_only() -> None:
    from app.services.product_terms import autofill_nel_terms

    with SessionLocal() as s:
        # LIF has FCL 50000 set manually and no age: amount must NOT be
        # overwritten, age fills in.
        assert autofill_nel_terms(s, PY_ID, LIF_PROD, 999999.0, 70) is True
        s.flush()
        term = s.execute(
            select(ProductTerm).where(
                ProductTerm.policy_year_id == PY_ID,
                ProductTerm.product_id == LIF_PROD,
            )
        ).scalar_one()
        assert term.free_cover_limit == 50000.0
        assert term.nel_age_limit == 70
        # A product with no term row gets one created with both values.
        assert autofill_nel_terms(s, PY_ID, ORP_PROD, 300000.0, 70) is True
        s.flush()
        orp = s.execute(
            select(ProductTerm).where(
                ProductTerm.policy_year_id == PY_ID,
                ProductTerm.product_id == ORP_PROD,
            )
        ).scalar_one()
        assert orp.free_cover_limit == 300000.0 and orp.nel_age_limit == 70
        s.rollback()


def test_empty_prior_year_is_not_treated_as_a_cohort_of_new_hires(
    client: TestClient,
) -> None:
    """A prior policy year that carries no roster (config-only, or not yet
    uploaded) proves nothing about who is new. Trusting it would mark EVERY
    life a new hire, and the age gate would then underwrite each senior's whole
    sum insured instead of just the increase."""
    EMPTY_PY = "00000000-0000-0000-0000-0000000i1801"
    with SessionLocal() as s:
        s.add(PolicyYear(
            id=EMPTY_PY, client_id=CLIENT_ID, year=2033,
            start_date=date(2033, 1, 1), end_date=date(2033, 12, 31),
            status=PolicyYearStatus.archived,
        ))
        emp = s.get(Employee, EMP_FAMILY)
        emp.attribute_values = {
            **(emp.attribute_values or {}), "date_of_birth": "1955-01-01",
        }
        s.commit()
    try:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"nel_age_limit": 70},
        )
        review = next(
            i for i in client.get(
                f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
            ).json()["items"]
            if i["staff_id"] == "IL-1" and i["subject_type"] == "employee"
        )
        line = review["cases"][0]
        # Unknown history → the FCL floor, NOT a guaranteed-0 new hire.
        assert line["guaranteed_si"] == 50000
        assert line["pending_si"] == 50000
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"nel_age_limit": None},
        )
        with SessionLocal() as s:
            emp = s.get(Employee, EMP_FAMILY)
            attrs = dict(emp.attribute_values or {})
            attrs.pop("date_of_birth", None)
            emp.attribute_values = attrs
            year = s.get(PolicyYear, EMPTY_PY)
            if year is not None:
                s.delete(year)
            s.commit()


def test_approval_without_amount_keeps_the_approved_excess(
    client: TestClient,
) -> None:
    """Approving with no explicit figure means the insurer took the request as
    asked — defaulting to the guaranteed floor would mark the case decided
    (pending → 0) while silently dropping the approved excess from cover."""
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    review = next(
        i for i in client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()["items"]
        if i["staff_id"] == "IL-1" and i["subject_type"] == "employee"
    )
    line = review["cases"][0]
    assert line["guaranteed_si"] == 50000 and line["requested_si"] == 100000

    res = client.patch(
        f"/api/v1/underwriting/cases/{line['id']}",
        json={"status": "approved_standard"},
    )
    assert res.status_code == 200, res.text
    out = res.json()["cases"][0]
    assert out["accepted_si"] == 100000  # the full request, not the floor
    assert out["pending_si"] == 0
    rows = {
        r["Staff ID"]: r for r in _sheet_rows(
            client.get(
                f"/api/v1/policy-years/{PY_ID}/reports/employee-listing"
                "?insurer=TestSure"
            ).content
        )
    }
    assert rows["IL-1"]["TLIF EE Last Accepted Sum Insured"] == 100000

    # A rejection is the opposite default: the excess is refused, so only the
    # guaranteed amount stays in force.
    res = client.patch(
        f"/api/v1/underwriting/cases/{line['id']}", json={"status": "rejected"}
    )
    assert res.json()["cases"][0]["accepted_si"] == 50000


def test_sync_keeps_reviews_carrying_broker_notes(client: TestClient) -> None:
    """Retiring a life's pending lines must not bin the requirements text and
    workflow position the broker typed — those are cancelled, not deleted."""
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    review = client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()["items"][0]
    client.patch(
        f"/api/v1/underwriting/reviews/{review['id']}",
        json={"status": "pending_insurer", "requirements": "Full medical report"},
    )
    try:
        # Raise the FCL past everyone → every pending line is moot.
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 1_000_000},
        )
        queue = client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()
        kept = next((i for i in queue["items"] if i["id"] == review["id"]), None)
        assert kept is not None, "a review with broker notes was deleted"
        assert kept["status"] == "cancelled"
        assert kept["requirements"] == "Full medical report"
        assert kept["cases"] == []
        # Untouched reviews are still cleaned up.
        assert len(queue["items"]) == 1
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 50000},
        )


def test_reset_product_term_clears_underwriting(client: TestClient) -> None:
    """DELETE drops the row's NEL values, so the cases are moot. Undecided
    lines hold a guaranteed-SI snapshot, so without a re-sync the listing keeps
    reporting a pending excess for a product with no limit."""
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    assert client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()["total"] >= 1
    try:
        assert client.delete(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}"
        ).status_code == 204
        assert client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()["total"] == 0
        rows = {
            r["Staff ID"]: r for r in _sheet_rows(
                client.get(
                    f"/api/v1/policy-years/{PY_ID}/reports/employee-listing"
                    "?insurer=TestSure"
                ).content
            )
        }
        assert rows["IL-1"]["TLIF EE Sum Insured Pending U/W"] == 0
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 50000},
        )


def test_scoped_refresh_leaves_other_households_alone() -> None:
    """A per-member trigger (one enrollment confirm) must not retire cases for
    members whose coverage it never recomputed."""
    from app.models import PolicyYear
    from app.services.underwriting import refresh_underwriting_cases

    with SessionLocal() as s:
        py = s.get(PolicyYear, PY_ID)
        refresh_underwriting_cases(s, py)
        s.commit()
        before = s.query(UnderwritingCase).count()
        assert before >= 2

        # Scope to IL-1 only. IL-2's case must survive untouched even though
        # this run never looked at them.
        refresh_underwriting_cases(s, py, {EMP_FAMILY})
        s.commit()
        assert s.query(UnderwritingCase).count() == before
        solo = s.query(UnderwritingCase).filter_by(employee_id=EMP_SOLO).one()
        assert solo.guaranteed_si == 50000
        s.query(UnderwritingCase).delete()
        s.query(UnderwritingReview).delete()
        s.commit()


def test_orphan_cases_are_adopted_on_read(client: TestClient) -> None:
    """Rows written before the review model carry review_id NULL and no data
    migration backfills them (it can't reach per-firm schemas) — the queue must
    surface them rather than reporting an empty page over live data."""
    with SessionLocal() as s:
        s.add(UnderwritingCase(
            client_id=CLIENT_ID, policy_year_id=PY_ID, product_id=LIF_PROD,
            employee_id=EMP_FAMILY, eligible_si=100000.0, accepted_si=50000.0,
            status="pending",  # legacy row: no review_id, no guaranteed_si
        ))
        s.commit()
    queue = client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()
    assert queue["total"] == 1
    review = queue["items"][0]
    assert review["insurer"] == "TestSure"
    line = review["cases"][0]
    # The legacy FCL in accepted_si is read as the guarantee, so the excess is
    # still reported pending rather than silently vanishing.
    assert line["guaranteed_si"] == 50000 and line["pending_si"] == 50000
    # And it is decidable (it used to 409 for having no review).
    res = client.patch(
        f"/api/v1/underwriting/cases/{line['id']}",
        json={"status": "approved_standard", "accepted_si": 90000},
    )
    assert res.status_code == 200, res.text
    assert res.json()["cases"][0]["accepted_si"] == 90000


# ── Readiness ────────────────────────────────────────────────────────────────


def test_readiness_reports_config_gaps(client: TestClient) -> None:
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/readiness")
    assert res.status_code == 200
    body = res.json()
    assert body["insurers"] == ["TestSure"]
    # TORP has categories? No — products without categories don't block.
    assert "TORP" not in body["products_without_insurer"]
    # The labeled plan is not flagged.
    assert body["plans_missing_report_label"] == []
    assert body["employees_missing_member_id"]["TestSure"] == 1  # IL-2
    assert body["employees_missing_nric"] == 1  # IL-2


# ── Member-listing template ──────────────────────────────────────────────────


def test_member_listing_template_round_trips(client: TestClient) -> None:
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/member-listing-template"
    )
    assert res.status_code == 200, res.text
    wb = load_workbook(BytesIO(res.content))
    assert wb.sheetnames == ["Employees", "Dependants"]
    emp_rows = _sheet_rows(res.content, sheet=0)
    fam = next(r for r in emp_rows if r["Staff ID"] == "IL-1")
    # Pre-filled with the current roster, UNMASKED (it re-imports).
    assert fam["Identification No. (NRIC/FIN)"] == "S1234567D"
    assert fam["TestSure Member ID"] == "TS-001"
    assert "Eligible to Sell Leave" in fam
    dep_rows = _sheet_rows(res.content, sheet=1)
    assert {r["Dependant Name"] for r in dep_rows} == {"Spo Use", "Chi Ld"}

    # Every employee header round-trips through the upload parser: all 31
    # attribute columns map (the "<Insurer> Member ID" column is handled by
    # the dynamic member-ID pass, not the alias map).
    from app.services.roster_parser import (
        EMPLOYEE_COLUMN_MAP,
        _build_column_map,
        _member_id_columns,
    )
    header = list(fam.keys())
    mapped = _build_column_map(header, EMPLOYEE_COLUMN_MAP)
    member_cols = _member_id_columns(header)
    assert len(mapped) + len(member_cols) == len(header)
    assert list(member_cols.values()) == ["TestSure"]


# ── Config surfaces ──────────────────────────────────────────────────────────


def test_product_report_code_patch(client: TestClient) -> None:
    res = client.patch(
        f"/api/v1/schemas/products/{LIF_PROD}", json={"report_code": "GTL"}
    )
    assert res.status_code == 200, res.text
    assert res.json()["report_code"] == "GTL"
    # The listing column heading follows.
    listing = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
    )
    header = next(
        load_workbook(BytesIO(listing.content)).active.iter_rows(values_only=True)
    )
    assert "GTL EE Basis of Cover" in header
    # Clear it again (null removes the metadata override).
    res = client.patch(
        f"/api/v1/schemas/products/{LIF_PROD}", json={"report_code": None}
    )
    assert res.json()["report_code"] is None


def test_plan_report_label_editable_on_active_year(client: TestClient) -> None:
    # Configuration is editable on every year (the activation lock was removed).
    res = client.patch(
        f"/api/v1/plans/{MED_PLAN}", json={"report_label": "4 Bed / S$60,000"}
    )
    assert res.status_code == 200, res.text
    assert res.json()["report_label"] == "4 Bed / S$60,000"
    # A config field (display_name) is also editable on an active year now.
    res = client.patch(
        f"/api/v1/plans/{MED_PLAN}",
        json={"report_label": "x", "display_name": "New name"},
    )
    assert res.status_code == 200, res.text
    assert res.json()["display_name"] == "New name"
    # Restore.
    client.patch(
        f"/api/v1/plans/{MED_PLAN}",
        json={"report_label": "1 Bed Restr Hosp / S$80,000"},
    )


def test_fcl_only_term_update_allowed_on_active_year(client: TestClient) -> None:
    res = client.put(
        f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
        json={"free_cover_limit": 55000},
    )
    assert res.status_code == 200, res.text
    assert res.json()["free_cover_limit"] == 55000
    # Coverage dates are also editable on an active year now (no lock).
    res = client.put(
        f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
        json={"coverage_start": "2034-01-01", "coverage_end": "2034-12-31"},
    )
    assert res.status_code == 200, res.text
    # Restore.
    client.put(
        f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
        json={"free_cover_limit": 50000, "coverage_start": None, "coverage_end": None},
    )


# ── Leave sell eligibility ───────────────────────────────────────────────────


def test_leave_sell_eligible_resolver() -> None:
    from app.services.leave_pricing_resolver import leave_sell_eligible

    assert leave_sell_eligible(Employee(attribute_values={})) is True
    assert leave_sell_eligible(
        Employee(attribute_values={"leave_sell_eligible": False})
    ) is False
    assert leave_sell_eligible(
        Employee(attribute_values={"leave_sell_eligible": "No"})
    ) is False
    assert leave_sell_eligible(
        Employee(attribute_values={"leave_sell_eligible": "true"})
    ) is True


# ── Audit hardening (2026-07-18) ─────────────────────────────────────────────


def test_formula_injection_neutralized(client: TestClient) -> None:
    # A roster value starting with a formula leader (= + - @) must be stored as
    # literal text (apostrophe-prefixed), never a live formula, in the workbook
    # the broker emails to the insurer.
    with SessionLocal() as s:
        emp = s.get(Employee, EMP_SOLO)
        emp.employee_name = '=HYPERLINK("http://evil")'
        emp.attribute_values = {**(emp.attribute_values or {}), "remarks": "@SUM(A1)"}
        s.commit()
    try:
        res = client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
        )
        row = next(r for r in _sheet_rows(res.content) if r["Staff ID"] == "IL-2")
        assert row["Employee Name"].startswith("'=")
        assert row["Remarks"].startswith("'@")
    finally:
        with SessionLocal() as s:
            emp = s.get(Employee, EMP_SOLO)
            emp.employee_name = "So Lo"
            attrs = dict(emp.attribute_values or {})
            attrs.pop("remarks", None)
            emp.attribute_values = attrs
            s.commit()


def test_lump_sum_classification_by_salary_multiple_basis() -> None:
    # A life product whose basis is a salary multiple (non-numeric) is still a
    # lump-sum product and must get the SI columns, not schedule-plan columns.
    from app.services.insurer_listings import product_blocks

    salary_prod = "00000000-0000-0000-0000-0000000i1013"
    salary_cat = "00000000-0000-0000-0000-0000000i1024"
    with SessionLocal() as s:
        s.add(Product(
            id=salary_prod, client_id=CLIENT_ID, code="TSAL",
            display_name="Salary Life", insurer="TestSure", has_dependants=False,
        ))
        s.flush()
        s.add(Category(
            id=salary_cat, policy_year_id=PY_ID, product_id=salary_prod,
            display_name="All", raw_description="All",
            plan_assignments={"plan_code": "1",
                              "basis": "36 times basic monthly salary"},
            source="manual", status="confirmed",
        ))
        s.commit()
    try:
        with SessionLocal() as s:
            py = s.get(PolicyYear, PY_ID)
            block = next(
                b for b in product_blocks(s, py) if b.product.code == "TSAL"
            )
            assert block.lump_sum is True
    finally:
        with SessionLocal() as s:
            for oid, model in ((salary_cat, Category), (salary_prod, Product)):
                obj = s.get(model, oid)
                if obj is not None:
                    s.delete(obj)
            s.commit()


def test_report_uw_amounts_refresh_independent() -> None:
    from app.models.underwriting_case import UnderwritingCase, UnderwritingStatus
    from app.services.underwriting import report_uw_amounts

    # No decision + no FCL → fully accepted.
    assert report_uw_amounts(100000, None, None) == (0.0, 100000)
    # No decision, over FCL → excess pending, FCL accepted (no refresh needed).
    assert report_uw_amounts(100000, 50000, None) == (50000, 50000)
    # No decision, within FCL → fully accepted.
    assert report_uw_amounts(40000, 50000, None) == (0.0, 40000)
    # A decision wins: accepted figure stands, capped at eligible, nothing pending.
    accepted = UnderwritingCase(
        accepted_si=80000, eligible_si=100000,
        status=UnderwritingStatus.approved_standard,
    )
    assert report_uw_amounts(100000, 50000, accepted) == (0.0, 80000)
    declined = UnderwritingCase(
        accepted_si=50000, eligible_si=100000, status=UnderwritingStatus.rejected,
    )
    assert report_uw_amounts(100000, 50000, declined) == (0.0, 50000)
    # Legacy vocabulary (pre review model) still reads as decided.
    legacy = UnderwritingCase(
        accepted_si=80000, eligible_si=100000, status=UnderwritingStatus.accepted,
    )
    assert report_uw_amounts(100000, 50000, legacy) == (0.0, 80000)
    # An undecided line with a computed guarantee (age trigger / last covered)
    # is in force at that guarantee, not the FCL.
    pending = UnderwritingCase(
        accepted_si=0.0, guaranteed_si=0.0, eligible_si=100000,
        status=UnderwritingStatus.pending,
    )
    assert report_uw_amounts(100000, 50000, pending) == (100000, 0.0)
    # Postponed = still awaiting the insurer — excess stays pending.
    postponed = UnderwritingCase(
        accepted_si=50000, guaranteed_si=50000, eligible_si=100000,
        status=UnderwritingStatus.postponed,
    )
    assert report_uw_amounts(100000, 50000, postponed) == (50000, 50000)


def test_decide_case_validation_and_capping(client: TestClient) -> None:
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    review = client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()["items"][0]
    case_id = review["cases"][0]["id"]

    # Bad status → 422 (legacy vocabulary is not accepted on writes).
    for bad in ("maybe", "accepted", "declined"):
        assert client.patch(
            f"/api/v1/underwriting/cases/{case_id}", json={"status": bad}
        ).status_code == 422, bad
    # Non-existent case → 404.
    assert client.patch(
        "/api/v1/underwriting/cases/does-not-exist",
        json={"status": "approved_standard"},
    ).status_code == 404
    # accepted_si is capped at eligible_si.
    res = client.patch(
        f"/api/v1/underwriting/cases/{case_id}",
        json={"status": "approved_standard", "accepted_si": 9_999_999},
    )
    line = res.json()["cases"][0]
    assert line["accepted_si"] == line["requested_si"]
    assert line["decided_on"] is not None  # auto-stamped on a decision
    # Rejected → excess refused, the guaranteed amount stays in force.
    res = client.patch(
        f"/api/v1/underwriting/cases/{case_id}",
        json={"status": "rejected", "remarks": "excess refused"},
    )
    line = res.json()["cases"][0]
    assert line["pending_si"] == 0
    assert line["status"] == "rejected"
    assert line["accepted_si"] == line["guaranteed_si"]
    # Postponed = undecided: excess back to pending, decided_on cleared.
    res = client.patch(
        f"/api/v1/underwriting/cases/{case_id}", json={"status": "postponed"}
    )
    line = res.json()["cases"][0]
    assert line["pending_si"] == line["requested_si"] - line["guaranteed_si"]
    assert line["decided_on"] is None

    # Review status validation.
    assert client.patch(
        f"/api/v1/underwriting/reviews/{review['id']}", json={"status": "nope"}
    ).status_code == 422
    assert client.patch(
        "/api/v1/underwriting/reviews/does-not-exist",
        json={"status": "completed"},
    ).status_code == 404


def test_refresh_removes_stale_pending_case(client: TestClient) -> None:
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    before = client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()["total"]
    assert before >= 1
    try:
        # Raise FCL above every eligible SI → all pending cases become moot.
        # The product-terms PUT re-syncs in the same transaction, so the queue
        # (reviews included) empties without a manual refresh.
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 1_000_000},
        )
        assert client.get(
            f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
        ).json()["total"] == 0
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{LIF_PROD}",
            json={"free_cover_limit": 50000},
        )


def test_viewer_unmasked_403_on_all_pii_downloads(viewer_client: TestClient) -> None:
    for path in (
        f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure&masked=false",
        f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=TestSure&masked=false",
        f"/api/v1/policy-years/{PY_ID}/reports/member-listing-template",
    ):
        assert viewer_client.get(path).status_code == 403, path
    # Masked listing is fine for a viewer.
    assert viewer_client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/employee-listing?insurer=TestSure"
    ).status_code == 200


def test_dependant_listing_masked_vs_unmasked(client: TestClient) -> None:
    masked = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=TestSure"
    )
    # The fixture dependants carry no NRIC, so assert on the employee's NRIC
    # column which the dependant listing echoes.
    row = next(r for r in _sheet_rows(masked.content) if r["Dependant Name"] == "Spo Use")
    assert row["Employee's Identification No."] == "S******7D"
    unmasked = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/dependant-listing?insurer=TestSure&masked=false"
    )
    row = next(
        r for r in _sheet_rows(unmasked.content) if r["Dependant Name"] == "Spo Use"
    )
    assert row["Employee's Identification No."] == "S1234567D"


def test_roster_parser_coercion_and_member_ids() -> None:
    from app.services.roster_parser import (
        INSURER_MEMBER_ID_KEY,
        _coerce_attr,
        _collect_member_ids,
        _member_id_columns,
    )

    # Flags → real bools; codes → clean strings (float 81.0 not "81.0").
    assert _coerce_attr("leave_sell_eligible", "No") is False
    assert _coerce_attr("prior_year_cover", "Yes") is True
    assert _coerce_attr("leave_sell_eligible", "Unknown") == "Unknown"  # kept raw
    assert _coerce_attr("branch_code", 81.0) == "81"
    assert _coerce_attr("bank_account_no", 123456789.0) == "123456789"

    header = ["Staff ID", "AIA Member ID", "Zurich Member ID", "Name"]
    cols = _member_id_columns(header)
    assert cols == {1: "AIA", 2: "Zurich"}
    attrs: dict = {}
    _collect_member_ids(["E1", 2427617201, "ZU-9", "Bob"], cols, attrs)
    assert attrs[INSURER_MEMBER_ID_KEY] == {"AIA": "2427617201", "Zurich": "ZU-9"}


def test_decide_case_cross_tenant_404(client: TestClient) -> None:
    # A case created for this client must not be decidable by a user from
    # another client (user_owns → _deny_cross_tenant → 404).
    client.post(f"/api/v1/policy-years/{PY_ID}/underwriting/refresh")
    review = client.get(
        f"/api/v1/policy-years/{PY_ID}/underwriting/cases"
    ).json()["items"][0]
    review_id = review["id"]
    case_id = review["cases"][0]["id"]

    def other() -> CurrentUser:
        return CurrentUser(
            user_id="00000000-0000-0000-0000-0000000i10fd",
            broker_firm_id=DEMO_BROKER_FIRM_ID,
            client_id="00000000-0000-0000-0000-0000000i9999", role="broker_admin",
        )
    app.dependency_overrides[get_current_user] = other
    try:
        res = TestClient(app).patch(
            f"/api/v1/underwriting/cases/{case_id}",
            json={"status": "approved_standard"},
        )
        assert res.status_code == 404
        res = TestClient(app).patch(
            f"/api/v1/underwriting/reviews/{review_id}",
            json={"status": "completed"},
        )
        assert res.status_code == 404
    finally:
        app.dependency_overrides[get_current_user] = _user

"""Activity + portal-access reports.

The load-bearing case here is CROSS-CLIENT LEAKAGE: `auth_events`,
`member_accounts` and `auth_mfa` are CONTROL tables living in `public`, so the
Postgres `search_path` does not scope them the way it scopes every other
tenant table. Their `client_id` filter is the only boundary, and forgetting it
is invisible on SQLite in dev — every one of these reports is exercised against
a second client's rows for exactly that reason.
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_activity_reports.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import UTC, date, datetime, timedelta  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    AuditLog,
    AuthEvent,
    AuthMfa,
    Client,
    Employee,
    MemberAccount,
    PolicyYear,
    User,
)
from app.models.auth import SUBJECT_MEMBER, SUBJECT_USER  # noqa: E402
from app.models.policy_year import PolicyYearStatus  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-0000000ac000"
OTHER_CLIENT_ID = "00000000-0000-0000-0000-0000000ac0a0"
PY_ID = "00000000-0000-0000-0000-0000000ac001"
OTHER_PY_ID = "00000000-0000-0000-0000-0000000ac0a1"
USER_ID = "00000000-0000-0000-0000-0000000ac0ff"

EMP_ACTIVE = "00000000-0000-0000-0000-0000000ac101"
EMP_INVITED = "00000000-0000-0000-0000-0000000ac102"
EMP_UNSENT = "00000000-0000-0000-0000-0000000ac103"
EMP_NONE = "00000000-0000-0000-0000-0000000ac104"
EMP_LEAVER = "00000000-0000-0000-0000-0000000ac105"

ACC_ACTIVE = "00000000-0000-0000-0000-0000000ac201"
ACC_INVITED = "00000000-0000-0000-0000-0000000ac202"
ACC_UNSENT = "00000000-0000-0000-0000-0000000ac203"
ACC_LEAVER = "00000000-0000-0000-0000-0000000ac204"
ACC_OTHER = "00000000-0000-0000-0000-0000000ac2a0"

# Anchor every timestamp to one instant so the date-range assertions are stable
# regardless of when the suite runs.
NOW = datetime.now(UTC).replace(hour=12, minute=0, second=0, microsecond=0)
TODAY = NOW.date()
# The widest range the endpoint accepts (its cap is 366 days), used by the
# leakage tests so they sweep every seeded event rather than the 30-day default.
WIDE_START = (TODAY - timedelta(days=364)).isoformat()


def _user(role: str = "broker_admin") -> CurrentUser:
    return CurrentUser(
        user_id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role=role,
    )


def _emp(eid: str, staff: str, name: str, **kw) -> Employee:
    defaults = dict(
        client_id=CLIENT_ID, policy_year_id=PY_ID, staff_id=staff,
        employee_name=name, attribute_values={}, derived_attribute_values={},
        matched_categories=[], source="csv_import", status="active",
    )
    defaults.update(kw)
    return Employee(id=eid, **defaults)


def _account(aid: str, staff: str, client_id: str = CLIENT_ID, **kw) -> MemberAccount:
    defaults = dict(client_id=client_id, staff_id=staff, status="active")
    defaults.update(kw)
    return MemberAccount(id=aid, **defaults)


def _event(subject_id: str | None, when: datetime, **kw) -> AuthEvent:
    defaults = dict(
        event_type="login_success", outcome="success", surface="portal",
        subject_type=SUBJECT_MEMBER, subject_id=subject_id,
        client_id=CLIENT_ID, occurred_at=when,
    )
    defaults.update(kw)
    return AuthEvent(**defaults)


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add_all([
            Client(id=CLIENT_ID, name="Activity Co", slug="activity-co",
                   broker_firm_id=DEMO_BROKER_FIRM_ID),
            Client(id=OTHER_CLIENT_ID, name="Other Co", slug="other-co",
                   broker_firm_id=DEMO_BROKER_FIRM_ID),
            User(id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
                 email="ops@activity.co", display_name="Ola Ops",
                 role="broker_admin", status="active"),
        ])
        s.flush()
        s.add_all([
            PolicyYear(id=PY_ID, client_id=CLIENT_ID, year=2036,
                       start_date=date(2036, 1, 1), end_date=date(2036, 12, 31),
                       status=PolicyYearStatus.active),
            PolicyYear(id=OTHER_PY_ID, client_id=OTHER_CLIENT_ID, year=2036,
                       start_date=date(2036, 1, 1), end_date=date(2036, 12, 31),
                       status=PolicyYearStatus.active),
        ])
        s.flush()
        s.add_all([
            _emp(EMP_ACTIVE, "AC-1", "Amy Active",
                 member_account_id=ACC_ACTIVE,
                 attribute_values={
                     "entity": "Activity Co Pte Ltd",
                     "email": "amy@activity.co",
                     "mobile": "+6591000001",
                     "category": "Manager",
                     "date_of_hire": "2020-02-01",
                 }),
            _emp(EMP_INVITED, "AC-2", "Ben Invited",
                 attribute_values={"entity": "Activity Co Pte Ltd"}),
            _emp(EMP_UNSENT, "AC-3", "Cal Unsent",
                 attribute_values={"entity": "Activity Co Pte Ltd"}),
            _emp(EMP_NONE, "AC-4", "Dee Noaccount",
                 attribute_values={"entity": "Activity Co Pte Ltd"}),
            # A leaver whose ACCOUNT carries no display name: the roster is the
            # only place their name exists.
            _emp(EMP_LEAVER, "AC-5", "Eli Leaver",
                 member_account_id=ACC_LEAVER,
                 status="terminated", terminated_effective=date(2036, 6, 30),
                 attribute_values={"entity": "Activity Co Pte Ltd"}),
        ])
        s.add_all([
            _account(ACC_ACTIVE, "AC-1", display_name="Amy Active",
                     email="amy@activity.co", system_login_id="M-0001",
                     status="active", invite_sent_at=NOW - timedelta(days=40),
                     last_sign_in_at=NOW - timedelta(days=1)),
            _account(ACC_INVITED, "AC-2", display_name="Ben Invited",
                     status="invited", invite_sent_at=NOW - timedelta(days=5)),
            # Provisioned but the mailer never accepted the message.
            _account(ACC_UNSENT, "AC-3", display_name="Cal Unsent",
                     status="invited", invite_sent_at=None),
            _account(ACC_LEAVER, "AC-5", display_name=None, status="active"),
            _account(ACC_OTHER, "OT-1", client_id=OTHER_CLIENT_ID,
                     display_name="Zoe Other", status="active"),
        ])
        s.flush()
        s.add(AuthMfa(subject_type=SUBJECT_MEMBER, subject_id=ACC_ACTIVE,
                      totp_secret_enc="x", confirmed_at=NOW))
        s.add_all([
            _event(ACC_ACTIVE, NOW - timedelta(days=2)),
            # Last instant of the range's final day — a `<= end` comparison
            # against a timestamp column drops this one.
            _event(ACC_ACTIVE, NOW.replace(hour=23, minute=59)),
            # Name lives only on the roster for this one.
            _event(ACC_LEAVER, NOW - timedelta(days=3)),
            # Failed login with no resolved subject: we store only a hash of
            # what was typed, so the row is anonymous but must still appear.
            _event(None, NOW - timedelta(days=1), event_type="login_fail",
                   outcome="fail", subject_type=None, ip="203.0.113.9"),
            # HR surface.
            _event(USER_ID, NOW - timedelta(days=4), surface="hr",
                   subject_type=SUBJECT_USER),
            # Outside the default 30-day window.
            _event(ACC_ACTIVE, NOW - timedelta(days=200)),
            # ANOTHER CLIENT — must never appear.
            _event(ACC_OTHER, NOW - timedelta(days=2), client_id=OTHER_CLIENT_ID),
        ])
        s.add_all([
            AuditLog(client_id=CLIENT_ID, user_id=USER_ID, actor_type="user",
                     action="run_matching", entity_type="policy_year",
                     entity_id=PY_ID, created_at=NOW - timedelta(days=2)),
            AuditLog(client_id=CLIENT_ID, member_account_id=ACC_ACTIVE,
                     actor_type="member", action="create", entity_type="claim",
                     entity_id="c1", employee_id=EMP_ACTIVE,
                     created_at=NOW - timedelta(days=1)),
            AuditLog(client_id=OTHER_CLIENT_ID, user_id=USER_ID,
                     actor_type="user", action="delete", entity_type="employee",
                     entity_id="nope", created_at=NOW - timedelta(days=1)),
        ])
        s.commit()
    app.dependency_overrides[get_current_user] = lambda: _user()
    yield
    app.dependency_overrides.clear()


@pytest.fixture()
def client():
    with TestClient(app) as c:
        yield c


def _sheet(resp):
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


def _get(client, path: str, **params):
    return client.get(f"/api/v1/policy-years/{PY_ID}/reports/{path}", params=params)


# ── Portal activity ──────────────────────────────────────────────────────────

def test_portal_activity_lists_every_surface_and_outcome(client):
    header, rows = _sheet(_get(client, "portal-activity"))
    assert header[0] == "Entity"
    assert "Outcome" in header
    activities = [r[header.index("Activity")] for r in rows]
    assert "Portal Login" in activities
    # A failed login is the reason a security report is opened; it must not be
    # filtered out just because it resolves to no subject.
    assert "Failed Login" in activities
    user_types = {r[header.index("User Type")] for r in rows}
    assert {"employee", "hr"} <= user_types


def test_portal_activity_excludes_other_clients(client):
    """`auth_events` is a control table — client_id is the ONLY tenant scope."""
    _, rows = _sheet(_get(client, "portal-activity", start=WIDE_START, end=TODAY.isoformat()))
    names = {r[2] for r in rows}
    assert "Zoe Other" not in names


def test_portal_activity_resolves_name_from_account_then_roster(client):
    header, rows = _sheet(_get(client, "portal-activity"))
    name_at = header.index("Employee Name")
    staff_at = header.index("Staff ID")
    by_staff = {r[staff_at]: r for r in rows if r[staff_at]}
    # Account display name.
    assert by_staff["AC-1"][name_at] == "Amy Active"
    # Account has no display name — the roster supplies it. Resolving from the
    # roster ALONE would blank every leaver's sign-ins.
    assert by_staff["AC-5"][name_at] == "Eli Leaver"
    assert by_staff["AC-1"][0] == "Activity Co Pte Ltd"


def test_portal_activity_range_includes_the_last_days_final_event(client):
    """An event at 23:59 on the end date is inside an inclusive range."""
    header, rows = _sheet(
        _get(client, "portal-activity",
             start=TODAY.isoformat(), end=TODAY.isoformat())
    )
    stamps = [r[header.index("Timestamp")] for r in rows]
    assert any(s and s.hour == 23 for s in stamps)


def test_portal_activity_default_window_excludes_ancient_events(client):
    _, rows = _sheet(_get(client, "portal-activity"))
    default_count = len(rows)
    _, all_rows = _sheet(
        _get(client, "portal-activity", start=WIDE_START, end=TODAY.isoformat())
    )
    assert len(all_rows) > default_count


def test_portal_activity_rejects_inverted_and_oversized_ranges(client):
    bad = _get(client, "portal-activity", start="2036-06-01", end="2036-01-01")
    assert bad.status_code == 400
    huge = _get(client, "portal-activity", start="2000-01-01", end=TODAY.isoformat())
    assert huge.status_code == 400


def test_portal_activity_is_audited(client):
    assert _get(client, "portal-activity").status_code == 200
    with SessionLocal() as s:
        rows = s.query(AuditLog).filter(
            AuditLog.entity_type == "activity_report"
        ).all()
        assert any(
            (r.after or {}).get("report") == "portal-activity" for r in rows
        )


# ── Company activity ─────────────────────────────────────────────────────────

def test_company_activity_lists_audit_rows_and_resolves_actors(client):
    header, rows = _sheet(_get(client, "company-activity"))
    actions = {r[header.index("Action")] for r in rows}
    assert "Run Matching" in actions
    actor_types = {r[header.index("Actor Type")] for r in rows}
    assert {"Platform user", "Member"} <= actor_types
    member_row = next(
        r for r in rows if r[header.index("Actor Type")] == "Member"
    )
    assert member_row[header.index("Actor")] == "Amy Active"
    assert member_row[header.index("Employee Staff ID")] == "AC-1"


def test_company_activity_excludes_other_clients(client):
    _, rows = _sheet(
        _get(client, "company-activity", start=WIDE_START, end=TODAY.isoformat())
    )
    # The other client's row is a `delete` on `employee`; nothing here is.
    assert not any(r[4] == "Employee" for r in rows)


# ── Portal access ────────────────────────────────────────────────────────────

def test_portal_access_is_roster_first(client):
    """An employee with no account is the row that matters most."""
    header, rows = _sheet(_get(client, "portal-access"))
    by_staff = {r[header.index("Staff ID")]: r for r in rows}
    assert set(by_staff) == {"AC-1", "AC-2", "AC-3", "AC-4", "AC-5"}
    assert by_staff["AC-4"][header.index("Status")] == "Not provisioned"


def test_portal_access_separates_unsent_invites_from_sent_ones(client):
    """Both are `invited` on the account; only one needs the button pressed."""
    header, rows = _sheet(_get(client, "portal-access"))
    status_at = header.index("Status")
    by_staff = {r[header.index("Staff ID")]: r for r in rows}
    assert by_staff["AC-2"][status_at] == "Invited"
    assert by_staff["AC-3"][status_at] == "Invite not sent"
    assert by_staff["AC-1"][status_at] == "Active"


def test_portal_access_profile_link_carries_the_company(client):
    header, rows = _sheet(_get(client, "portal-access"))
    links = {r[header.index("ProfileLink")] for r in rows}
    assert len(links) == 1
    # Host-based or path-based depending on `tenant_mode`; either way the
    # company must be IN the link (see `member_invite.portal_sign_in_url`).
    assert "activity-co" in links.pop()


def test_portal_access_reports_login_id_and_two_factor(client):
    header, rows = _sheet(_get(client, "portal-access"))
    by_staff = {r[header.index("Staff ID")]: r for r in rows}
    assert by_staff["AC-1"][header.index("Login ID")] == "M-0001"
    assert by_staff["AC-1"][header.index("Two-Factor")] == "Yes"
    assert by_staff["AC-2"][header.index("Two-Factor")] in (None, "")


def test_portal_access_excludes_other_clients_accounts(client):
    header, rows = _sheet(_get(client, "portal-access"))
    assert "OT-1" not in {r[header.index("Staff ID")] for r in rows}

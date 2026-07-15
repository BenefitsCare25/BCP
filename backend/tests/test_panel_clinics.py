"""Panel clinic locator — workbook parsing, broker CRUD/upload/tagging, the
member locator endpoint, broker preview parity, and tenant/member isolation.
"""
from __future__ import annotations

import os
from datetime import date
from io import BytesIO
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_panel_clinics.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from app.core.auth import (  # noqa: E402
    DEMO_BROKER_FIRM_ID,
    DEMO_CLIENT_ID,
    CurrentUser,
    get_current_user,
)
from app.core.portal_auth import issue_member_token  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Client,
    Employee,
    MemberAccount,
    PanelClinic,
    PanelListing,
    PolicyYear,
)
from app.models.member_account import MEMBER_STATUS_ACTIVE  # noqa: E402
from app.models.policy_year import PolicyYearStatus  # noqa: E402
from app.services.panel_clinics import (  # noqa: E402
    PanelParseError,
    haversine_km,
    parse_panel_workbook,
)
from scripts.seed_demo import seed  # noqa: E402

PY_A = "00000000-0000-0000-0000-00000000pc01"
EMP_ALICE = "00000000-0000-0000-0000-00000000pc02"
ACC_ALICE = "00000000-0000-0000-0000-00000000pc03"

CLIENT_B_ID = "00000000-0000-0000-0000-00000000pcb0"
LISTING_B = "00000000-0000-0000-0000-00000000pcb1"
PY_B = "00000000-0000-0000-0000-00000000pcb2"

HEADERS = [
    "Code", "Name", "Zone", "Area", "Specialty", "Doctor",
    "Address1", "Address2", "Address3", "PostalCode", "Country",
    "PhoneNumber", "MonToFri", "Saturday", "Sunday", "PublicHoliday",
    "Latitude", "Longitude", "GoogleMapURL",
]

# Yishun (north) and Raffles Place (downtown) — distance sorting anchors.
ROW_YISHUN = [
    "C1", "YISHUN CLINIC", "NORTH REGION", "YISHUN", None, "DR LEE",
    "618 YISHUN RING ROAD", None, None, "760618", "SINGAPORE",
    "62353490 - FIRST DAY OF OPERATION: 01 MAY 2021",
    "MON-FRI: 8AM - 1PM", "SAT: 8AM - 1PM", "SUN: 9AM - 12PM", "PH: CLOSED",
    1.4187, 103.8357, None,
]
ROW_RAFFLES = [
    "C2", "RAFFLES PLACE CLINIC", "CENTRAL REGION", "RAFFLES PLACE", None, None,
    "1 RAFFLES PLACE", "#04-01", None, "048616", "SINGAPORE", "61234567",
    "MON-FRI: 9AM - 6PM", "SAT: CLOSED", "SUN: CLOSED", "PH: CLOSED",
    1.2839, 103.8515, "https://maps.google.com/?q=1.2839,103.8515",
]
ROW_JB = [
    "M1", "JB POLIKLINIK", "Malaysia", "Malaysia", None, None,
    "Blk 114 Jalan Perwira 1", None, None, "81300 Johor", "MALAYSIA",
    "02-07-5583303", "24 hours", "24 hours", "24 hours", "24 hours",
    1.5226, 103.6633, None,
]


def _workbook_bytes(rows: list[list], headers: list | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers or HEADERS)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _user_a() -> CurrentUser:
    return CurrentUser(
        user_id="00000000-0000-0000-0000-000000000001",
        broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=DEMO_CLIENT_ID,
        role="broker_admin",
    )


def _user_b() -> CurrentUser:
    return CurrentUser(
        user_id="00000000-0000-0000-0000-0000000000b1",
        broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_B_ID,
        role="broker_admin",
    )


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    seed()
    with SessionLocal() as session:
        session.add(
            PolicyYear(
                id=PY_A,
                client_id=DEMO_CLIENT_ID,
                year=2027,
                start_date=date(2027, 3, 1),
                end_date=date(2028, 2, 28),
                status=PolicyYearStatus.active,
            )
        )
        session.add(
            Client(id=CLIENT_B_ID, name="Client B", broker_firm_id=DEMO_BROKER_FIRM_ID)
        )
        session.flush()
        session.add(
            PolicyYear(
                id=PY_B,
                client_id=CLIENT_B_ID,
                year=2027,
                start_date=date(2027, 1, 1),
                end_date=date(2027, 12, 31),
                status=PolicyYearStatus.draft,
            )
        )
        session.add(
            MemberAccount(
                id=ACC_ALICE,
                client_id=DEMO_CLIENT_ID,
                email="alice@a.test",
                staff_id="S-1",
                status=MEMBER_STATUS_ACTIVE,
            )
        )
        session.add(
            Employee(
                id=EMP_ALICE,
                client_id=DEMO_CLIENT_ID,
                policy_year_id=PY_A,
                staff_id="S-1",
                employee_name="Alice",
                member_account_id=ACC_ALICE,
                attribute_values={},
                derived_attribute_values={},
                source="csv_import",
                status="active",
            )
        )
        # A listing PINNED to client B (legacy scoping) — the cross-tenant
        # probe target. Listings created via the API are library entries
        # (client_id NULL) visible to every company.
        session.add(
            PanelListing(
                id=LISTING_B,
                client_id=CLIENT_B_ID,
                insurer="GE-SG",
                panel_provider="Adept",
                country="SG",
                clinic_type="gp",
            )
        )
        session.commit()
    yield
    with SessionLocal() as session:
        session.query(MemberAccount).filter(
            MemberAccount.id == ACC_ALICE
        ).delete()
        session.query(PanelListing).filter(
            PanelListing.client_id.in_([DEMO_CLIENT_ID, CLIENT_B_ID])
        ).delete()
        for year_id in (PY_A, PY_B):
            py = session.get(PolicyYear, year_id)
            if py is not None:
                session.delete(py)
        client_b = session.get(Client, CLIENT_B_ID)
        if client_b is not None:
            session.delete(client_b)
        session.commit()
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture
def broker_a() -> TestClient:
    app.dependency_overrides[get_current_user] = _user_a
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture
def broker_b() -> TestClient:
    app.dependency_overrides[get_current_user] = _user_b
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture
def anon() -> TestClient:
    return TestClient(app)


def _member_auth() -> dict[str, str]:
    token, _ = issue_member_token(ACC_ALICE, DEMO_CLIENT_ID)
    return {"Authorization": f"Bearer {token}"}


def _upload(client: TestClient, listing_id: str, rows: list[list]) -> dict:
    res = client.post(
        f"/api/v1/panel-listings/{listing_id}/upload",
        files={
            "file": (
                "panel.xlsx",
                _workbook_bytes(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert res.status_code == 200, res.text
    return res.json()


# ── Workbook parsing ─────────────────────────────────────────────────────────


def test_parse_workbook_maps_headers_and_cleans(tmp_path: Path) -> None:
    path = tmp_path / "panel.xlsx"
    rows = [
        ROW_YISHUN,
        # Missing name → skipped but counted.
        ["C9", None, "NORTH REGION", "YISHUN", None, None, "X", None, None,
         "760618", "SINGAPORE", "6123", "9-5", None, None, None, 1.4, 103.8, None],
        # Out-of-range coords → row kept, coords dropped.
        ["C3", "BAD COORDS CLINIC", "EAST  REGION ", "BEDOK", None, None,
         "1 BEDOK ROAD", None, None, 469675.0, "SINGAPORE", "6123",
         "9-5", None, None, None, 91.0, 103.8, None],
    ]
    path.write_bytes(_workbook_bytes(rows))
    result = parse_panel_workbook(path)

    assert result.rows_total == 3
    assert result.skipped_no_name == 1
    assert result.missing_coordinates == 1
    assert len(result.clinics) == 2

    yishun = result.clinics[0]
    assert yishun.name == "YISHUN CLINIC"
    assert yishun.hours == {
        "mon_fri": "MON-FRI: 8AM - 1PM",
        "sat": "SAT: 8AM - 1PM",
        "sun": "SUN: 9AM - 12PM",
        "public_holiday": "PH: CLOSED",
    }
    # Map URL synthesized from coordinates when the cell is blank.
    assert yishun.google_map_url == "https://maps.google.com/?q=1.4187,103.8357"

    bad = result.clinics[1]
    assert bad.latitude is None and bad.longitude is None
    assert bad.google_map_url is None
    assert bad.zone == "EAST REGION"  # whitespace collapsed
    assert bad.postal_code == "469675"  # float tail stripped


def test_parse_workbook_without_name_column_fails(tmp_path: Path) -> None:
    path = tmp_path / "bad.xlsx"
    path.write_bytes(_workbook_bytes([["x", "y"]], headers=["Foo", "Bar"]))
    with pytest.raises(PanelParseError):
        parse_panel_workbook(path)


def test_haversine_known_distance() -> None:
    # Raffles Place → Yishun is ~15 km.
    dist = haversine_km(1.2839, 103.8515, 1.4187, 103.8357)
    assert 14.0 < dist < 16.5


# ── Broker CRUD + upload + tagging ───────────────────────────────────────────


def test_full_broker_flow_and_member_locator(broker_a: TestClient, anon: TestClient) -> None:
    # Create SG GP + MY GP listings.
    res = broker_a.post(
        "/api/v1/panel-listings",
        json={
            "insurer": "AIA-SG",
            "panel_provider": "Alliance",
            "country": "sg",  # normalized upward
            "clinic_type": "GP",  # normalized downward
        },
    )
    assert res.status_code == 201, res.text
    sg = res.json()
    assert sg["country"] == "SG" and sg["clinic_type"] == "gp"
    assert sg["type_label"] == "SG GP"
    assert sg["clinic_count"] == 0

    res = broker_a.post(
        "/api/v1/panel-listings",
        json={
            "insurer": "AIA-SG",
            "panel_provider": "Alliance",
            "country": "MY",
            "clinic_type": "gp",
        },
    )
    assert res.status_code == 201
    my = res.json()
    assert my["type_label"] == "JB GP"

    # Duplicate combo → 409.
    res = broker_a.post(
        "/api/v1/panel-listings",
        json={
            "insurer": "AIA-SG",
            "panel_provider": "Alliance",
            "country": "SG",
            "clinic_type": "gp",
        },
    )
    assert res.status_code == 409
    assert res.json()["detail"]["code"] == "duplicate_panel_listing"

    # Upload clinics.
    up = _upload(broker_a, sg["id"], [ROW_YISHUN, ROW_RAFFLES])
    assert up["imported"] == 2 and up["rows_total"] == 2
    assert up["listing"]["clinic_count"] == 2
    _upload(broker_a, my["id"], [ROW_JB])

    # Broker clinic preview + search.
    res = broker_a.get(f"/api/v1/panel-listings/{sg['id']}/clinics", params={"q": "raffles"})
    assert res.status_code == 200
    assert [c["name"] for c in res.json()] == ["RAFFLES PLACE CLINIC"]

    # Download round-trips through the parser.
    res = broker_a.get(f"/api/v1/panel-listings/{sg['id']}/download")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert wb.active.max_row == 3  # header + 2 clinics

    # Member sees nothing until the listings are tagged to their year.
    res = anon.get("/api/v1/portal/clinics", headers=_member_auth())
    assert res.status_code == 200
    assert res.json()["total"] == 0
    assert res.json()["filters"]["clinic_types"] == []

    # Tag both listings to the active year.
    res = broker_a.put(
        f"/api/v1/policy-years/{PY_A}/panels",
        json={"panel_listing_ids": [sg["id"], my["id"]]},
    )
    assert res.status_code == 200
    assert set(res.json()["panel_listing_ids"]) == {sg["id"], my["id"]}
    res = broker_a.get(f"/api/v1/policy-years/{PY_A}/panels")
    assert set(res.json()["panel_listing_ids"]) == {sg["id"], my["id"]}

    # Member locator: all clinics, facets present.
    res = anon.get("/api/v1/portal/clinics", headers=_member_auth())
    body = res.json()
    assert body["total"] == 3
    assert body["located"] is False
    facets = {(f["country"], f["clinic_type"]): f for f in body["filters"]["clinic_types"]}
    assert facets[("SG", "gp")]["count"] == 2
    assert facets[("SG", "gp")]["label"] == "SG GP"
    assert facets[("MY", "gp")]["label"] == "JB GP"
    assert "YISHUN" in body["filters"]["areas"]

    # Distance sort from Raffles Place: Raffles → Yishun → JB.
    res = anon.get(
        "/api/v1/portal/clinics",
        params={"lat": 1.2839, "lng": 103.8515},
        headers=_member_auth(),
    )
    body = res.json()
    assert body["located"] is True
    names = [c["name"] for c in body["items"]]
    assert names == ["RAFFLES PLACE CLINIC", "YISHUN CLINIC", "JB POLIKLINIK"]
    assert body["items"][0]["distance_km"] < 1.0
    assert body["items"][1]["distance_km"] == pytest.approx(15.1, abs=1.0)

    # Country/type filter → JB only.
    res = anon.get(
        "/api/v1/portal/clinics",
        params={"clinic_type": "gp", "country": "MY"},
        headers=_member_auth(),
    )
    assert [c["name"] for c in res.json()["items"]] == ["JB POLIKLINIK"]

    # Text + area filters.
    res = anon.get("/api/v1/portal/clinics", params={"q": "760618"}, headers=_member_auth())
    assert [c["name"] for c in res.json()["items"]] == ["YISHUN CLINIC"]
    res = anon.get("/api/v1/portal/clinics", params={"area": "yishun"}, headers=_member_auth())
    assert [c["name"] for c in res.json()["items"]] == ["YISHUN CLINIC"]

    # Broker preview mirrors the member response exactly.
    portal = anon.get(
        "/api/v1/portal/clinics",
        params={"lat": 1.2839, "lng": 103.8515},
        headers=_member_auth(),
    ).json()
    preview = broker_a.get(
        f"/api/v1/employees/{EMP_ALICE}/portal-preview/clinics",
        params={"lat": 1.2839, "lng": 103.8515},
    ).json()
    assert preview == portal

    # Re-upload replaces (not appends) the clinic list.
    up = _upload(broker_a, sg["id"], [ROW_RAFFLES])
    assert up["imported"] == 1
    res = anon.get("/api/v1/portal/clinics", headers=_member_auth())
    assert res.json()["total"] == 2  # 1 SG + 1 JB

    # A rejected workbook leaves the previous list untouched.
    res = broker_a.post(
        f"/api/v1/panel-listings/{sg['id']}/upload",
        files={
            "file": (
                "bad.xlsx",
                _workbook_bytes([["a"]], headers=["Foo"]),
                "application/octet-stream",
            )
        },
    )
    assert res.status_code == 422
    res = broker_a.get(f"/api/v1/panel-listings/{sg['id']}/clinics")
    assert len(res.json()) == 1

    # PATCH relabel.
    res = broker_a.patch(
        f"/api/v1/panel-listings/{sg['id']}", json={"label": "AIA Alliance SG GP"}
    )
    assert res.status_code == 200
    assert res.json()["display_label"] == "AIA Alliance SG GP"

    # Delete the MY listing — clinics + tag go with it.
    res = broker_a.delete(f"/api/v1/panel-listings/{my['id']}")
    assert res.status_code == 204
    res = broker_a.get(f"/api/v1/policy-years/{PY_A}/panels")
    assert res.json()["panel_listing_ids"] == [sg["id"]]
    res = anon.get("/api/v1/portal/clinics", headers=_member_auth())
    assert res.json()["total"] == 1
    with SessionLocal() as session:
        assert (
            session.query(PanelClinic)
            .filter(PanelClinic.panel_listing_id == my["id"])
            .count()
            == 0
        )


def test_validation_rejects_unknown_type_and_country(broker_a: TestClient) -> None:
    res = broker_a.post(
        "/api/v1/panel-listings",
        json={
            "insurer": "X",
            "panel_provider": "Y",
            "country": "US",
            "clinic_type": "gp",
        },
    )
    assert res.status_code == 422
    res = broker_a.post(
        "/api/v1/panel-listings",
        json={
            "insurer": "X",
            "panel_provider": "Y",
            "country": "SG",
            "clinic_type": "hospital",
        },
    )
    assert res.status_code == 422


# ── Isolation ────────────────────────────────────────────────────────────────


def test_cross_tenant_listing_access_404(broker_a: TestClient) -> None:
    for method, path in [
        ("get", f"/api/v1/panel-listings/{LISTING_B}/clinics"),
        ("get", f"/api/v1/panel-listings/{LISTING_B}/download"),
        ("delete", f"/api/v1/panel-listings/{LISTING_B}"),
    ]:
        res = getattr(broker_a, method)(path)
        assert res.status_code == 404, f"{method} {path} → {res.status_code}"
    res = broker_a.patch(
        f"/api/v1/panel-listings/{LISTING_B}", json={"label": "hijack"}
    )
    assert res.status_code == 404


def test_cross_tenant_listing_cannot_be_tagged(broker_a: TestClient) -> None:
    res = broker_a.put(
        f"/api/v1/policy-years/{PY_A}/panels",
        json={"panel_listing_ids": [LISTING_B]},
    )
    assert res.status_code == 404
    # And nothing was written.
    res = broker_a.get(f"/api/v1/policy-years/{PY_A}/panels")
    assert LISTING_B not in res.json()["panel_listing_ids"]


def test_library_listing_visible_to_every_client(broker_b: TestClient) -> None:
    """Library entries (client_id NULL) appear for every company; another
    tenant's PINNED row never does."""
    res = broker_b.get("/api/v1/panel-listings")
    assert res.status_code == 200
    listings = res.json()
    ids = {listing["id"] for listing in listings}
    assert LISTING_B in ids  # B's own pinned row
    # The library listing created through client A's flow is visible to B too.
    assert any(
        listing["insurer"] == "AIA-SG" and listing["panel_provider"] == "Alliance"
        for listing in listings
    )


def test_pinned_listing_hidden_from_other_tenant(broker_a: TestClient) -> None:
    res = broker_a.get("/api/v1/panel-listings")
    assert res.status_code == 200
    assert LISTING_B not in {listing["id"] for listing in res.json()}


def test_portal_clinics_requires_member_token(anon: TestClient) -> None:
    assert anon.get("/api/v1/portal/clinics").status_code == 401


# ── Review-fix regressions ───────────────────────────────────────────────────


def _sg_listing_id(client: TestClient) -> str:
    listings = client.get("/api/v1/panel-listings").json()
    return next(
        listing["id"]
        for listing in listings
        if listing["insurer"] == "AIA-SG" and listing["country"] == "SG"
    )


def test_patch_whitespace_only_field_is_422(broker_a: TestClient) -> None:
    """A blank-after-strip field must fail request validation (422), never
    reach the in-handler PanelListingIn construction (which would 500)."""
    listing_id = _sg_listing_id(broker_a)
    res = broker_a.patch(
        f"/api/v1/panel-listings/{listing_id}", json={"insurer": "   "}
    )
    assert res.status_code == 422
    res = broker_a.patch(
        f"/api/v1/panel-listings/{listing_id}", json={"panel_provider": " "}
    )
    assert res.status_code == 422


def test_broker_preview_search_escapes_like_and_matches_doctor(
    broker_a: TestClient,
) -> None:
    listing_id = _sg_listing_id(broker_a)
    _upload(broker_a, listing_id, [ROW_YISHUN, ROW_RAFFLES])

    # LIKE metacharacters match literally: no clinic contains "_" or "%".
    for q in ("_", "%", "a_c"):
        res = broker_a.get(
            f"/api/v1/panel-listings/{listing_id}/clinics", params={"q": q}
        )
        assert res.status_code == 200
        assert res.json() == [], f"q={q!r} over-matched"

    # Doctor names are searchable, same as the member locator.
    res = broker_a.get(
        f"/api/v1/panel-listings/{listing_id}/clinics", params={"q": "dr lee"}
    )
    assert [c["name"] for c in res.json()] == ["YISHUN CLINIC"]


def test_areas_facet_respects_type_filter(
    broker_a: TestClient, anon: TestClient
) -> None:
    """The area dropdown never offers areas that don't exist under the
    selected clinic-type chip (no dead-end combinations)."""
    res = broker_a.post(
        "/api/v1/panel-listings",
        json={
            "insurer": "AIA-SG",
            "panel_provider": "Alliance",
            "country": "MY",
            "clinic_type": "gp",
        },
    )
    assert res.status_code == 201
    my_id = res.json()["id"]
    _upload(broker_a, my_id, [ROW_JB])
    sg_id = _sg_listing_id(broker_a)
    res = broker_a.put(
        f"/api/v1/policy-years/{PY_A}/panels",
        json={"panel_listing_ids": [sg_id, my_id]},
    )
    assert res.status_code == 200

    body = anon.get("/api/v1/portal/clinics", headers=_member_auth()).json()
    assert "Malaysia" in body["filters"]["areas"]  # unfiltered → all areas

    body = anon.get(
        "/api/v1/portal/clinics",
        params={"clinic_type": "gp", "country": "SG"},
        headers=_member_auth(),
    ).json()
    assert "Malaysia" not in body["filters"]["areas"]
    assert "YISHUN" in body["filters"]["areas"]

    body = anon.get(
        "/api/v1/portal/clinics",
        params={"clinic_type": "gp", "country": "MY"},
        headers=_member_auth(),
    ).json()
    assert body["filters"]["areas"] == ["Malaysia"]
    # The type chips stay unfiltered by design.
    assert len(body["filters"]["clinic_types"]) == 2


# ── Multi-company enablement + policy-year carry-over ───────────────────────


def test_enable_listing_across_companies(broker_a: TestClient) -> None:
    """One dialog enables a library listing for several companies at once —
    each on its own target policy year — without switching the active client."""
    sg_id = _sg_listing_id(broker_a)

    res = broker_a.get(f"/api/v1/panel-listings/{sg_id}/companies")
    assert res.status_code == 200
    by_client = {c["client_id"]: c for c in res.json()}
    assert DEMO_CLIENT_ID in by_client and CLIENT_B_ID in by_client
    assert by_client[DEMO_CLIENT_ID]["policy_year_id"] == PY_A  # active year
    assert by_client[CLIENT_B_ID]["policy_year_id"] == PY_B  # latest draft
    assert by_client[DEMO_CLIENT_ID]["enabled"] is True  # tagged earlier

    # Enable for BOTH companies in one call.
    res = broker_a.put(
        f"/api/v1/panel-listings/{sg_id}/companies",
        json={"client_ids": [DEMO_CLIENT_ID, CLIENT_B_ID]},
    )
    assert res.status_code == 200
    by_client = {c["client_id"]: c for c in res.json()}
    assert by_client[DEMO_CLIENT_ID]["enabled"] is True
    assert by_client[CLIENT_B_ID]["enabled"] is True
    res = broker_a.get(f"/api/v1/policy-years/{PY_A}/panels")
    assert sg_id in res.json()["panel_listing_ids"]

    # Disable for client B only — demo keeps its tag.
    res = broker_a.put(
        f"/api/v1/panel-listings/{sg_id}/companies",
        json={"client_ids": [DEMO_CLIENT_ID]},
    )
    by_client = {c["client_id"]: c for c in res.json()}
    assert by_client[CLIENT_B_ID]["enabled"] is False
    assert by_client[DEMO_CLIENT_ID]["enabled"] is True

    # Unknown / inaccessible company → 404, nothing written.
    res = broker_a.put(
        f"/api/v1/panel-listings/{sg_id}/companies",
        json={"client_ids": ["00000000-0000-0000-0000-00000000nope"]},
    )
    assert res.status_code == 404


def test_enable_for_company_without_policy_year_422() -> None:
    """A company with no usable policy year can't be enabled — explicit 422,
    not a silent skip."""
    from app.core.auth import get_current_user

    with SessionLocal() as session:
        yearless = Client(
            id="00000000-0000-0000-0000-00000000pcc0",
            name="Yearless Co",
            broker_firm_id=DEMO_BROKER_FIRM_ID,
        )
        session.add(yearless)
        session.commit()
    app.dependency_overrides[get_current_user] = _user_a
    try:
        client = TestClient(app)
        sg_id = _sg_listing_id(client)
        res = client.put(
            f"/api/v1/panel-listings/{sg_id}/companies",
            json={"client_ids": ["00000000-0000-0000-0000-00000000pcc0"]},
        )
        assert res.status_code == 422
        assert res.json()["detail"]["code"] == "client_has_no_policy_year"
    finally:
        app.dependency_overrides.pop(get_current_user, None)
        with SessionLocal() as session:
            row = session.get(Client, "00000000-0000-0000-0000-00000000pcc0")
            if row is not None:
                session.delete(row)
                session.commit()


def test_new_policy_year_inherits_panel_tags(broker_a: TestClient) -> None:
    """Creating a new policy year copies the previous year's panel selections,
    so the company's clinic locator survives renewals without re-ticking."""
    before = set(
        broker_a.get(f"/api/v1/policy-years/{PY_A}/panels").json()[
            "panel_listing_ids"
        ]
    )
    assert before, "precondition: the active year has panel tags"

    res = broker_a.post(
        "/api/v1/policy-years",
        json={"start_date": "2028-03-01", "end_date": "2029-02-28"},
    )
    assert res.status_code == 201, res.text
    new_year_id = res.json()["id"]
    inherited = set(
        broker_a.get(f"/api/v1/policy-years/{new_year_id}/panels").json()[
            "panel_listing_ids"
        ]
    )
    assert inherited == before
    # Clean up so the extra year can't skew other modules' expectations.
    with SessionLocal() as session:
        row = session.get(PolicyYear, new_year_id)
        if row is not None:
            session.delete(row)
            session.commit()

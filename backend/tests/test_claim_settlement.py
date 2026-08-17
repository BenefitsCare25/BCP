"""The insurer settlement leg: references, dispatch, payment, SLA, reports.

The regression that matters most here is the UTILIZATION one. `approved` used
to be terminal, so every money test compared to it directly. Now a claim moves
on to `sent_to_insurer` and `paid`, and any surviving `== approved` comparison
silently hands the member back a limit they have already spent.
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_claim_settlement.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import UTC, date, datetime, timedelta  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.core.clock import today as business_today  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Claim,
    ClaimMessage,
    Client,
    Employee,
    MemberAccount,
    PolicyYear,
    User,
)
from app.models.claim import (  # noqa: E402
    CLAIM_STATUS_APPROVED,
    CLAIM_STATUS_PAID,
    CLAIM_STATUS_SENT_TO_INSURER,
    CLAIM_STATUS_SUBMITTED,
    LIVE_STATUSES,
    SETTLED_STATUSES,
)
from app.models.policy_year import PolicyYearStatus  # noqa: E402
from app.services.claim_settlement import (  # noqa: E402
    days_over_deadline,
    insurer_days,
    mint_reference_no,
    reference_prefix,
)
from app.services.utilization import PENDING_STATUSES  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-0000000c5000"
PY_ID = "00000000-0000-0000-0000-0000000c5001"
USER_ID = "00000000-0000-0000-0000-0000000c50ff"
EMP = "00000000-0000-0000-0000-0000000c5101"
ACC = "00000000-0000-0000-0000-0000000c5201"

NOW = datetime.now(UTC)


def _user(role: str = "broker_admin") -> CurrentUser:
    return CurrentUser(
        user_id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role=role,
    )


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Settle Co", slug="settle-co",
                     broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.add(User(id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
                   email="ops@settle.co", display_name="Sam Ops",
                   role="broker_admin", status="active"))
        s.flush()
        s.add(PolicyYear(id=PY_ID, client_id=CLIENT_ID, year=2038,
                         start_date=date(2038, 1, 1), end_date=date(2038, 12, 31),
                         status=PolicyYearStatus.active))
        s.add(MemberAccount(id=ACC, client_id=CLIENT_ID, staff_id="ST-1",
                            email="mem@settle.co", status="active"))
        s.flush()
        s.add(Employee(id=EMP, client_id=CLIENT_ID, policy_year_id=PY_ID,
                       staff_id="ST-1", employee_name="Mel Member",
                       member_account_id=ACC, attribute_values={},
                       derived_attribute_values={}, matched_categories=[],
                       source="csv_import", status="active"))
        s.commit()
    app.dependency_overrides[get_current_user] = lambda: _user()
    yield
    app.dependency_overrides.clear()


@pytest.fixture()
def client():
    with TestClient(app) as c:
        yield c


def _claim(session, **kw) -> Claim:
    defaults = dict(
        client_id=CLIENT_ID, policy_year_id=PY_ID, employee_id=EMP,
        claim_kind="insured", product_code="GHS", claim_type="Hospital",
        incurred_date=date(2038, 3, 1), amount_claimed=500.0, currency="SGD",
        status=CLAIM_STATUS_APPROVED, amount_approved=500.0,
        decided_at=NOW - timedelta(days=5), origin="portal",
    )
    defaults.update(kw)
    c = Claim(**defaults)
    session.add(c)
    session.flush()
    return c


# ── The status-set invariants ────────────────────────────────────────────────

def test_settled_statuses_are_not_pending():
    """The whole point of SETTLED_STATUSES: PENDING is derived by subtraction,
    so a new settled status lands in it by default and silently un-spends the
    member's limit."""
    assert CLAIM_STATUS_SENT_TO_INSURER not in PENDING_STATUSES
    assert CLAIM_STATUS_PAID not in PENDING_STATUSES
    assert SETTLED_STATUSES.isdisjoint(PENDING_STATUSES)


def test_settled_statuses_are_live():
    """A settled claim still consumes the limit and still blocks a duplicate
    invoice."""
    assert SETTLED_STATUSES <= LIVE_STATUSES


def test_settlement_does_not_give_the_limit_back():
    """The behavioural half of the invariant above.

    Approve 500, send it, then have the insurer pay it. At no point does the
    approved total drop — before the fix, moving off `approved` reclassified
    the money as `pending`, which utilization reports beside the limit and
    never subtracts.
    """
    from app.services.utilization import _bucket_sums

    def sums_for(status: str) -> dict[str, float | int]:
        claim = Claim(
            client_id=CLIENT_ID, policy_year_id=PY_ID, employee_id=EMP,
            claim_kind="insured", product_code="GHS", claim_type="Hospital",
            incurred_date=date(2038, 3, 1), amount_claimed=500.0,
            amount_approved=500.0, status=status,
        )
        return _bucket_sums([claim])[("GHS", None)]

    for status in (
        CLAIM_STATUS_APPROVED,
        CLAIM_STATUS_SENT_TO_INSURER,
        CLAIM_STATUS_PAID,
    ):
        row = sums_for(status)
        assert row["approved"] == 500.0, status
        assert row["pending"] == 0.0, status


# ── Reference numbers ────────────────────────────────────────────────────────

def test_reference_is_minted_once_and_never_changes(client):
    with SessionLocal() as s:
        c = _claim(s, status="draft", amount_approved=None, decided_at=None)
        first = mint_reference_no(s, c)
        s.commit()
        again = mint_reference_no(s, c)
    assert first.startswith("SETTLECO-")
    # Idempotent: a needs_info resubmission runs back through submit_claim, and
    # a reference that changed is one the member can no longer quote.
    assert again == first


def test_references_increment(client):
    with SessionLocal() as s:
        a = mint_reference_no(s, _claim(s, status="draft"))
        s.commit()
        b = mint_reference_no(s, _claim(s, status="draft"))
        s.commit()
    assert int(b.rsplit("-", 1)[1]) == int(a.rsplit("-", 1)[1]) + 1


def test_a_shared_prefix_does_not_wedge_the_second_company(client):
    """Two companies in one firm can reduce to the SAME reference prefix.

    A prefix is the first 8 alphanumerics of the slug, so "Settle Co" and
    "Settle Co Group" both give `SETTLECO` — and `ix_claims_reference_no` is
    unique across the whole schema, not per client. When the sequence was read
    per CLIENT, the second company's very first claim proposed a number the
    first company already held, and the retry recomputed the identical
    candidate; every claim that company ever filed 503'd. The sequence has to be
    read from the same scope the constraint enforces.
    """
    other_id = "00000000-0000-0000-0000-0000000c5900"
    with SessionLocal() as s:
        s.add(Client(id=other_id, name="Settle Co Group",
                     slug="settle-co-group", broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.flush()
        mine = mint_reference_no(s, _claim(s, status="draft"))
        s.commit()

        theirs = mint_reference_no(
            s, _claim(s, status="draft", client_id=other_id)
        )
        s.commit()

    assert reference_prefix(Client(name="Settle Co Group",
                                   slug="settle-co-group")) == "SETTLECO"
    assert theirs != mine
    # One ascending series across the shared prefix — which is what keeps it
    # unique. A per-client counter cannot, because the constraint is not.
    assert int(theirs.rsplit("-", 1)[1]) == int(mine.rsplit("-", 1)[1]) + 1


def test_reference_prefix_falls_back_rather_than_being_blank():
    """A bare number is not recognisable as a claim reference on a phone call."""
    assert reference_prefix(None) == "CLM"
    assert reference_prefix(Client(name="!!!", slug=None)) == "CLM"


# ── Dispatch and payment ─────────────────────────────────────────────────────

def test_send_to_insurer_sets_a_deadline_even_when_none_is_given(client):
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    res = client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["status"] == CLAIM_STATUS_SENT_TO_INSURER
    # An unbounded claim is one nobody chases.
    assert body["insurer_deadline_on"] is not None


def test_only_an_approved_claim_can_be_sent(client):
    with SessionLocal() as s:
        c = _claim(s, status=CLAIM_STATUS_SUBMITTED, amount_approved=None)
        s.commit()
        cid = c.id
    res = client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    assert res.status_code == 409
    assert res.json()["detail"]["code"] == "invalid_transition"


def test_payment_records_the_insurers_own_figure(client):
    """A shortfall between approved and paid is the reason to reconcile."""
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    res = client.post(
        f"/api/v1/claims/{cid}/payment",
        json={"paid_on": "2038-04-10", "amount": 420.0},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["status"] == CLAIM_STATUS_PAID
    assert body["payment_amount"] == 420.0
    # What we approved is NOT overwritten.
    assert body["amount_approved"] == 500.0


def test_overpayment_requires_then_accepts_explicit_acknowledgement(client):
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    body = {"paid_on": "2038-04-10", "amount": 520.0}
    warning = client.post(f"/api/v1/claims/{cid}/payment", json=body)
    assert warning.status_code == 409, warning.text
    assert warning.json()["detail"] == {
        "code": "payment_exceeds_approval",
        "message": (
            "The insurer payment exceeds the approved amount. Resend with "
            "acknowledge_overpayment=true to record the exception."
        ),
        "approved": 500.0,
        "payment": 520.0,
    }
    accepted = client.post(
        f"/api/v1/claims/{cid}/payment",
        json={**body, "acknowledge_overpayment": True},
    )
    assert accepted.status_code == 200, accepted.text
    assert accepted.json()["payment_amount"] == 520.0


def test_a_zero_settlement_is_accepted(client):
    """Fully offset against an excess is a real advice; refusing it would
    strand the claim in `sent_to_insurer` forever."""
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    res = client.post(
        f"/api/v1/claims/{cid}/payment",
        json={"paid_on": "2038-04-10", "amount": 0},
    )
    assert res.status_code == 200, res.text
    assert res.json()["payment_amount"] == 0


def test_a_zero_payment_notice_says_zero(client):
    """`or` instead of `is not None` told the member they were paid the full
    approved amount when the insurer paid nothing."""
    from app.models.claim_message import EVENT_PAID
    from app.services.claim_messages import _system_copy

    with SessionLocal() as s:
        c = _claim(s, status=CLAIM_STATUS_PAID, amount_approved=500.0,
                   payment_amount=0.0, paid_on=date(2038, 4, 10))
        s.flush()
        _subject, body = _system_copy(c, EVENT_PAID, None)
    assert "500" not in body


def test_references_are_unique_per_client(client):
    """`mint_reference_no` reads the max then writes one past it, which races.
    The unique index is what actually stops a duplicate reaching the ledger."""
    from sqlalchemy.exc import IntegrityError

    with SessionLocal() as s:
        first = _claim(s, status="draft")
        ref = mint_reference_no(s, first)
        s.commit()
        clash = _claim(s, status="draft")
        clash.reference_no = ref
        with pytest.raises(IntegrityError):
            s.commit()
        s.rollback()


def test_payment_notifies_the_member_separately_from_approval(client):
    """"Approved" and "the money is in my account" are weeks apart and only the
    second ends the member's wait."""
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    client.post(f"/api/v1/claims/{cid}/payment", json={"paid_on": "2038-04-10"})
    with SessionLocal() as s:
        events = [
            m.event for m in s.query(ClaimMessage).filter(
                ClaimMessage.claim_id == cid
            ).all()
        ]
    assert "paid" in events


def test_dispatch_posts_no_member_notice(client):
    """Forwarding to the insurer is our workflow, not a new decision."""
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    with SessionLocal() as s:
        assert s.query(ClaimMessage).filter(
            ClaimMessage.claim_id == cid
        ).count() == 0


# ── SLA counters ─────────────────────────────────────────────────────────────

def test_insurer_days_keeps_counting_on_an_unpaid_claim():
    """Blank here would make an overdue claim look like one never sent."""
    c = Claim(
        client_id=CLIENT_ID, policy_year_id=PY_ID, employee_id=EMP,
        claim_kind="insured", claim_type="Hospital",
        incurred_date=date(2038, 3, 1), amount_claimed=1.0,
        sent_to_insurer_at=datetime.now(UTC) - timedelta(days=9),
    )
    assert insurer_days(c) == 9


def test_the_insurer_clock_stops_when_they_decline():
    """A declined claim never gets a `paid_on`. Falling back to today made its
    overdue count climb every night forever, parking a settled claim at the top
    of the broker's overdue list."""
    from datetime import timedelta as _td

    decided = datetime.now(UTC) - _td(days=40)
    c = Claim(
        client_id=CLIENT_ID, policy_year_id=PY_ID, employee_id=EMP,
        claim_kind="insured", claim_type="Hospital",
        incurred_date=date(2038, 3, 1), amount_claimed=1.0,
        status="rejected",
        sent_to_insurer_at=datetime.now(UTC) - _td(days=60),
        # The deadline is a `date` COLUMN, so it has to be stated in the same
        # calendar the counters read the timestamps in — business, not UTC
        # (`core/clock.py`). Built from `(now(UTC) - 30d).date()` this fixture
        # agreed with itself only while the server ran in UTC, and disagreed by
        # a day for the eight hours a Singapore day runs ahead of one.
        insurer_deadline_on=business_today() - _td(days=30),
        decided_at=decided,
    )
    # 60 days sent → 40 days decided = 20, and it stays 20 tomorrow.
    assert insurer_days(c) == 20
    assert days_over_deadline(c) == -10


def test_days_over_deadline_is_signed():
    c = Claim(
        client_id=CLIENT_ID, policy_year_id=PY_ID, employee_id=EMP,
        claim_kind="insured", claim_type="Hospital",
        incurred_date=date(2038, 3, 1), amount_claimed=1.0,
        insurer_deadline_on=date(2038, 5, 1), paid_on=date(2038, 4, 20),
    )
    assert days_over_deadline(c) == -11  # still in time


# ── Assessment detail ────────────────────────────────────────────────────────

def test_assessment_is_a_partial_update(client):
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"admission_date": "2038-03-01", "discharge_date": "2038-03-04"},
    )
    # A second form editing only the sector must not blank the dates.
    res = client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"hospital_type": "government"}
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["hospital_type"] == "government"
    assert body["admission_date"] == "2038-03-01"
    assert body["discharge_date"] == "2038-03-04"


def test_assessment_rejects_a_bad_sector_and_inverted_dates(client):
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    bad = client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"hospital_type": "clinic"}
    )
    assert bad.status_code == 422
    inverted = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"admission_date": "2038-03-05", "discharge_date": "2038-03-01"},
    )
    assert inverted.status_code == 422


def test_assessment_checks_dates_against_the_stored_pair(client):
    """The endpoint is a PARTIAL update, so an inverted pair is most easily
    created one field at a time — the body alone always looks fine."""
    with SessionLocal() as s:
        c = _claim(s)
        s.commit()
        cid = c.id
    ok = client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"admission_date": "2038-03-05"}
    )
    assert ok.status_code == 200, ok.text
    # Only the discharge date is in this body, and it precedes the admission
    # date already on the row.
    bad = client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"discharge_date": "2038-03-01"}
    )
    assert bad.status_code == 422


# ── Reports ──────────────────────────────────────────────────────────────────

CLAIMS_REGISTER = f"/api/v1/policy-years/{PY_ID}/reports/workbooks/claims-register"

# The three insurance-claim scopes and the per-employee view are SHEETS of one
# workbook now, not four downloads. The scope is still one builder and a filter
# (`claims_reports._claim_rows`); only the packaging changed.
_CLAIM_SHEET = {
    "all": "All Claims",
    "adjudication": "Adjudication",
    "inpatient": "Inpatient",
    "outpatient": "Outpatient",
    "employee": "By Employee",
}


def _sheet(resp, title: str):
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    assert title in wb.sheetnames, wb.sheetnames
    rows = list(wb[title].iter_rows(values_only=True))
    return rows[0], rows[1:]


def _claims_sheet(client, scope="all", **params):
    return _sheet(
        client.get(CLAIMS_REGISTER, params=params), _CLAIM_SHEET[scope]
    )


def test_insurance_claims_report_carries_the_servicing_columns(client):
    with SessionLocal() as s:
        c = _claim(s, reference_no="SETTLECO-000900")
        s.commit()
        cid = c.id
    client.post(f"/api/v1/claims/{cid}/send-to-insurer", json={})
    client.post(f"/api/v1/claims/{cid}/payment", json={"paid_on": "2038-04-10"})
    header, rows = _claims_sheet(client)
    for col in (
        "Reference No.", "Date Sent to Insurer", "Deadline Date for Insurer",
        "Payment Date", "No. of days for Tracking Insurer", "Days Over Deadline",
        "First Document Receive Date", "Verified Date",
    ):
        assert col in header
    row = next(r for r in rows if r[header.index("Reference No.")] == "SETTLECO-000900")
    assert row[header.index("Status")] == "Paid"


def test_outpatient_sheet_drops_the_inpatient_only_columns(client):
    """Three columns blank on every row read as missing data, not as N/A."""
    header, _ = _claims_sheet(client, "outpatient")
    assert "Admission Date" not in header
    assert "Hospital Type" not in header
    assert "Referral Letter" in header


def test_inpatient_scope_filters_by_product_not_sub_type(client):
    with SessionLocal() as s:
        _claim(s, product_code="GOGP", reference_no="SETTLECO-000901")
        s.commit()
    header, rows = _claims_sheet(client, "inpatient")
    refs = {r[header.index("Reference No.")] for r in rows}
    assert "SETTLECO-000901" not in refs


def test_employee_claims_sheet_covers_flex_and_insured(client):
    with SessionLocal() as s:
        _claim(s, claim_kind="flex", product_code=None,
               flex_category_name="Wellness", claim_type="Wellness",
               reference_no="SETTLECO-000902")
        s.commit()
    header, rows = _claims_sheet(client, "employee")
    categories = {r[header.index("Claim Category")] for r in rows}
    assert {"Flexible Benefits", "Insurance"} <= categories


def test_flex_claims_are_absent_from_the_insurance_sheet(client):
    """A flex claim is funded by the member's wallet — there is no insurer."""
    header, rows = _claims_sheet(client)
    refs = {r[header.index("Reference No.")] for r in rows}
    assert "SETTLECO-000902" not in refs


def test_needs_info_reports_as_pending_documents(client):
    with SessionLocal() as s:
        _claim(s, status="needs_info", amount_approved=None,
               reference_no="SETTLECO-000903")
        s.commit()
    header, rows = _claims_sheet(client)
    row = next(r for r in rows if r[header.index("Reference No.")] == "SETTLECO-000903")
    assert row[header.index("Status")] == "Pending Documents"


def test_tax_and_cpf_default_to_no(client):
    """No is the ordinary payroll treatment and the assessment form's default,
    so an untouched claim reports No rather than blank.

    The point is that the sheet and the form AGREE: the form offers only
    Yes/No, so a blank column under a control reading "No" would be two answers
    to one question, and payroll acts on whichever it is holding.
    """
    with SessionLocal() as s:
        _claim(s, reference_no="SETTLECO-000904")
        s.commit()
    header, rows = _claims_sheet(client)
    row = next(r for r in rows if r[header.index("Reference No.")] == "SETTLECO-000904")
    assert row[header.index("TAX")] == "No"
    assert row[header.index("CPF")] == "No"


def test_hospital_sector_is_derived_from_the_provider(client):
    """The sector was already computed from the provider by the intake autofill
    and the review's document check — but `Claim.hospital_type`, the column the
    report prints, is written by NOTHING, so the sheet was blank on every row
    and a manual dropdown was the only thing that could ever fill it."""
    with SessionLocal() as s:
        _claim(s, reference_no="SETTLECO-000910",
               provider_name="Raffles Hospital")
        _claim(s, reference_no="SETTLECO-000911",
               provider_name="Singapore General Hospital")
        # An assessor's stored value OVERRIDES the derivation — an overseas
        # admission, or a hospital the registry does not list.
        _claim(s, reference_no="SETTLECO-000912",
               provider_name="Bumrungrad International",
               hospital_type="private")
        s.commit()
    header, rows = _claims_sheet(client)
    by_ref = {r[header.index("Reference No.")]: r for r in rows}
    col = header.index("Hospital Type")
    assert by_ref["SETTLECO-000910"][col] == "Private/Overseas"
    assert by_ref["SETTLECO-000911"][col] == "Government"
    assert by_ref["SETTLECO-000912"][col] == "Private/Overseas"


def test_the_form_is_served_the_same_derivation_the_report_prints(client):
    """Served, never re-derived in TypeScript: the dropdown's default names the
    sector the sheet will print, so the two cannot disagree."""
    with SessionLocal() as s:
        c = _claim(s, provider_name="Mount Alvernia Hospital")
        s.commit()
        cid = c.id
    body = client.get(f"/api/v1/claims/{cid}").json()
    assert body["hospital_type_derived"] == "private"
    # The stored override stays empty — null means "derive", not "unassessed".
    assert body["hospital_type"] is None


def test_settlement_dates_can_be_corrected_after_the_fact(client):
    """`send-to-insurer` and `payment` are TRANSITIONS, each legal from one
    status only — so a claim already at `paid` had lost the controls that set
    these and could never have a wrong date fixed, or a missing one filled in.

    The amendment writes the record and leaves the STATUS alone: re-running the
    transition would repost the member's "your claim has been paid" notice for
    a typo correction.
    """
    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_PAID,
            paid_on=date(2038, 4, 10),
            payment_amount=500.0,
        )
        s.commit()
        cid = c.id

    res = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={
            "sent_to_insurer_on": "2038-03-20",
            "insurer_deadline_on": "2038-04-19",
            "payment_amount": 450.0,
        },
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["status"] == CLAIM_STATUS_PAID
    assert body["sent_to_insurer_at"].startswith("2038-03-20")
    assert body["insurer_deadline_on"] == "2038-04-19"
    assert body["payment_amount"] == 450.0
    # The SLA counters are derived, so correcting the dates recomputes them
    # rather than leaving a stored copy behind: 20 Mar → 10 Apr.
    assert body["insurer_days"] == 21
    assert body["days_over_deadline"] == -9


def test_assessment_overpayment_warning_is_json_and_acknowledgeable(client):
    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_PAID,
            sent_to_insurer_at=NOW - timedelta(days=20),
            paid_on=date(2038, 4, 10),
            payment_amount=500.0,
        )
        s.commit()
        cid = c.id
    warning = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"payment_amount": 520.0},
    )
    assert warning.status_code == 409, warning.text
    assert warning.json()["detail"]["approved"] == 500.0
    assert warning.json()["detail"]["payment"] == 520.0
    accepted = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"payment_amount": 520.0, "acknowledge_overpayment": True},
    )
    assert accepted.status_code == 200, accepted.text
    assert accepted.json()["payment_amount"] == 520.0


def test_an_amendment_audits_the_figure_it_replaced(client):
    """The `before` snapshot must be taken BEFORE the amendment writes.

    Read afterwards it was read off an already-mutated claim, so a broker
    correcting a payment from 1,200.00 to 120.00 logged before=120, after=120 —
    losing the only figure the audit row existed to preserve. Money and dates
    are exactly what this trail is for.
    """
    from app.models import AuditLog

    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_PAID,
            sent_to_insurer_at=NOW - timedelta(days=20),
            paid_on=date(2038, 4, 10),
            payment_amount=1200.0,
        )
        s.commit()
        cid = c.id

    assert client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"payment_amount": 120.0}
    ).status_code == 200

    with SessionLocal() as s:
        row = (
            s.query(AuditLog)
            .filter(AuditLog.action == "claim.assessment",
                    AuditLog.entity_id == cid)
            .order_by(AuditLog.created_at.desc())
            .first()
        )
    assert row.before["payment_amount"] == "1200.00"
    assert row.after["payment_amount"] == "120.00"


def test_a_dispatch_only_correction_is_still_audited(client):
    """`sent_to_insurer_on` is a request field whose COLUMN is
    `sent_to_insurer_at`. Excluded from the snapshot by name, a request that
    corrected only the dispatch date wrote before={} / after={} — no record."""
    from app.models import AuditLog

    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_SENT_TO_INSURER,
            sent_to_insurer_at=datetime(2038, 3, 1, tzinfo=UTC),
        )
        s.commit()
        cid = c.id

    assert client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"sent_to_insurer_on": "2038-03-05"},
    ).status_code == 200

    with SessionLocal() as s:
        row = (
            s.query(AuditLog)
            .filter(AuditLog.action == "claim.assessment",
                    AuditLog.entity_id == cid)
            .order_by(AuditLog.created_at.desc())
            .first()
        )
    assert row.before["sent_to_insurer_at"].startswith("2038-03-01")
    assert row.after["sent_to_insurer_at"].startswith("2038-03-05")


def test_an_insurer_declined_claim_can_still_be_corrected(client):
    """`sent_to_insurer → rejected` is a real transition, and such a claim WAS
    dispatched. Gating on `status in {sent_to_insurer, paid}` refused it with
    "…once the claim has been sent to the insurer" — which is false, and it is
    precisely the "wrong date, no way back" case the amendment exists for."""
    with SessionLocal() as s:
        c = _claim(
            s,
            status="rejected",
            amount_approved=None,
            sent_to_insurer_at=NOW - timedelta(days=10),
        )
        s.commit()
        cid = c.id
    res = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"insurer_deadline_on": "2038-05-01"},
    )
    assert res.status_code == 200, res.text
    assert res.json()["insurer_deadline_on"] == "2038-05-01"


def test_a_payment_cannot_be_recorded_by_amendment_on_an_unpaid_claim(client):
    """Writing `paid_on` onto a claim still with the insurer does not move the
    status, but `_insurer_clock_stop` reads it FIRST — so the SLA counters
    freeze and an UNPAID claim silently drops off the overdue list a broker
    works, while it is still pending against the member's limit."""
    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_SENT_TO_INSURER,
            sent_to_insurer_at=NOW - timedelta(days=40),
            insurer_deadline_on=(NOW - timedelta(days=10)).date(),
        )
        s.commit()
        cid = c.id
    assert client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"paid_on": "2038-04-10"}
    ).status_code == 409
    # Still overdue, still being chased.
    with SessionLocal() as s:
        assert s.get(Claim, cid).paid_on is None


def test_a_dispatch_date_cannot_be_cleared_out_from_under_the_status(client):
    """A cleared date input sends null, so this is one keystroke away. Cleared,
    `insurer_days` goes blank while `days_over_deadline` keeps counting against
    a deadline nothing now explains."""
    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_SENT_TO_INSURER,
            sent_to_insurer_at=NOW - timedelta(days=5),
        )
        s.commit()
        cid = c.id
    assert client.patch(
        f"/api/v1/claims/{cid}/assessment", json={"sent_to_insurer_on": None}
    ).status_code == 422
    with SessionLocal() as s:
        assert s.get(Claim, cid).sent_to_insurer_at is not None


def test_settlement_dates_are_refused_before_the_insurer_leg(client):
    """Backfilling a dispatch date onto a claim that was never sent would
    invent a history the SLA counters then report on."""
    with SessionLocal() as s:
        c = _claim(s, status=CLAIM_STATUS_APPROVED)
        s.commit()
        cid = c.id
    res = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"sent_to_insurer_on": "2038-03-20"},
    )
    assert res.status_code == 409


def test_an_amended_deadline_cannot_precede_the_dispatch_on_the_row(client):
    """Checked against the EFFECTIVE pair, not just what the request carried —
    a partial update that moves only the deadline is exactly how it comes to
    precede a dispatch date already stored."""
    with SessionLocal() as s:
        c = _claim(
            s,
            status=CLAIM_STATUS_SENT_TO_INSURER,
            sent_to_insurer_at=NOW,
            insurer_deadline_on=(NOW + timedelta(days=30)).date(),
        )
        s.commit()
        cid = c.id
    res = client.patch(
        f"/api/v1/claims/{cid}/assessment",
        json={"insurer_deadline_on": (NOW - timedelta(days=5)).date().isoformat()},
    )
    assert res.status_code == 422


def test_the_adjudication_register_is_a_sheet_of_the_workbook(client):
    """Its own Reports row was deleted in the consolidation, which left the
    endpoint behind it reachable from nothing. It is NOT redundant with All
    Claims — it is the only sheet carrying the claim id and the invoice number,
    which is the key a broker reconciles a disputed line against."""
    header, _ = _claims_sheet(client, "adjudication")
    assert "Invoice No." in header
    assert "Claim ID" in header


def test_viewer_cannot_pull_unmasked_claims(client):
    app.dependency_overrides[get_current_user] = lambda: _user("broker_viewer")
    try:
        res = client.get(CLAIMS_REGISTER, params={"masked": "false"})
        assert res.status_code == 403
    finally:
        app.dependency_overrides[get_current_user] = lambda: _user()

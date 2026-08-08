"""Flex wallet ledger, utilisation summary and the leaver reports.

The invariant under test throughout: the ledger, its summary and the leaver
sheet are three views of ONE resolution (`flex_pricing_resolver.summarize_
employee`). If they ever disagree about a member's balance, the derived-not-
stored design has failed at the only thing it was chosen for.
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_flex_ledger.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import UTC, date, datetime  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Claim,
    Client,
    Employee,
    PolicyYear,
    User,
)
from app.models.claim import CLAIM_STATUS_APPROVED, CLAIM_STATUS_PAID  # noqa: E402
from app.models.policy_year import PolicyYearStatus  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-0000000f1000"
PY_ID = "00000000-0000-0000-0000-0000000f1001"
USER_ID = "00000000-0000-0000-0000-0000000f10ff"

EMP_WALLET = "00000000-0000-0000-0000-0000000f1101"
EMP_LEAVER = "00000000-0000-0000-0000-0000000f1102"
EMP_NOFLEX = "00000000-0000-0000-0000-0000000f1103"
EMP_PRE_PERIOD = "00000000-0000-0000-0000-0000000f1104"
EMP_LATE_LEAVER = "00000000-0000-0000-0000-0000000f1105"
EMP_NOFLEX_LEAVER = "00000000-0000-0000-0000-0000000f1106"

NOW = datetime.now(UTC)


def _user(role: str = "broker_admin") -> CurrentUser:
    return CurrentUser(
        user_id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role=role,
    )


def _emp(eid: str, staff: str, name: str, **kw) -> Employee:
    defaults = dict(
        client_id=CLIENT_ID, policy_year_id=PY_ID, staff_id=staff,
        employee_name=name, attribute_values={}, derived_attribute_values={},
        matched_categories=[], source="csv_import", status="active",
    )
    defaults.update(kw)
    return Employee(id=eid, **defaults)


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Flexi Co", slug="flexi-co",
                     broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.add(User(id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
                   email="ops@flexi.co", display_name="Fay Ops",
                   role="broker_admin", status="active"))
        s.flush()
        s.add(PolicyYear(id=PY_ID, client_id=CLIENT_ID, year=2039,
                         start_date=date(2039, 1, 1), end_date=date(2039, 12, 31),
                         status=PolicyYearStatus.active))
        s.flush()
        s.add_all([
            _emp(EMP_WALLET, "FX-1", "Wanda Wallet",
                 flex_wallet_amount=2000.0, flex_currency="SGD",
                 flex_tier_name="Manager", flex_family_status="EO",
                 flex_assigned_at=NOW,
                 attribute_values={
                     "entity": "Flexi Co Pte Ltd",
                     "id_no": "S1111111A",
                     "category": "Manager",
                     "date_of_hire": "2020-01-06",
                 }),
            _emp(EMP_LEAVER, "FX-2", "Leo Leaver",
                 status="terminated", terminated_effective=date(2039, 6, 30),
                 flex_wallet_amount=1000.0, flex_currency="SGD",
                 flex_assigned_at=NOW,
                 attribute_values={
                     "entity": "Flexi Co Pte Ltd",
                     "id_no": "S2222222B",
                     "category": "Officer",
                 }),
            # No wallet, no leave, no pricing → flex does not apply.
            _emp(EMP_NOFLEX, "FX-3", "Nora Noflex",
                 attribute_values={"entity": "Flexi Co Pte Ltd"}),
            # Left BEFORE the period, and the date lives only in the roster —
            # `terminated_effective` is written by the listing sync, so a roster
            # imported any other way carries it here instead.
            _emp(EMP_PRE_PERIOD, "FX-4", "Percy Priorleaver",
                 status="terminated",
                 attribute_values={"entity": "Flexi Co Pte Ltd",
                                   "last_day_of_service": "2038-03-01"}),
            # Left AFTER the period closed: covered for all of 2039, so a leaver
            # of the NEXT year's sheet — but still on the insurer listing.
            _emp(EMP_LATE_LEAVER, "FX-5", "Lena Lateleaver",
                 status="terminated", terminated_effective=date(2040, 1, 15),
                 attribute_values={"entity": "Flexi Co Pte Ltd"}),
            # An in-period leaver flex does not apply to. Sorts FIRST by name.
            _emp(EMP_NOFLEX_LEAVER, "FX-6", "Amos Ahead",
                 status="terminated", terminated_effective=date(2039, 9, 30),
                 attribute_values={"entity": "Flexi Co Pte Ltd",
                                   "category": "Officer"}),
        ])
        s.flush()
        s.add_all([
            # Settled — spends the wallet.
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_WALLET, claim_kind="flex",
                  flex_category_name="Wellness", claim_type="Wellness",
                  incurred_date=date(2039, 2, 1), amount_claimed=300.0,
                  amount_approved=300.0, currency="SGD",
                  status=CLAIM_STATUS_APPROVED, decided_at=NOW,
                  reference_no="FLEXICO-000001"),
            # PAID — still spent. Before SETTLED_STATUSES this fell into
            # "pending" and handed the wallet back.
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_WALLET, claim_kind="flex",
                  flex_category_name="Dental", claim_type="Dental",
                  incurred_date=date(2039, 3, 1), amount_claimed=200.0,
                  amount_approved=200.0, payment_amount=200.0,
                  paid_on=date(2039, 4, 1), currency="SGD",
                  status=CLAIM_STATUS_PAID, decided_at=NOW,
                  reference_no="FLEXICO-000002"),
            # In flight — reported, never subtracted.
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_WALLET, claim_kind="flex",
                  flex_category_name="Optical", claim_type="Optical",
                  incurred_date=date(2039, 5, 1), amount_claimed=150.0,
                  currency="SGD", status="submitted",
                  reference_no="FLEXICO-000003"),
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_LEAVER, claim_kind="flex",
                  flex_category_name="Wellness", claim_type="Wellness",
                  incurred_date=date(2039, 5, 20), amount_claimed=120.0,
                  amount_approved=120.0, currency="SGD",
                  status=CLAIM_STATUS_APPROVED, decided_at=NOW,
                  reference_no="FLEXICO-000004"),
            # Reclassified as a LOG case AFTER the member submitted it, so it
            # still carries the member's own descriptive label.
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_LEAVER, claim_kind="insured",
                  product_code="GHS",
                  claim_type="Emergency Accidental Outpatient Treatment",
                  case_type="log",
                  incurred_date=date(2039, 6, 1), amount_claimed=90.0,
                  currency="SGD", status="submitted",
                  reference_no="FLEXICO-000005"),
            # Incurred AFTER their last day (30 Jun) — outside the cover they held.
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_LEAVER, claim_kind="insured",
                  product_code="GHS", claim_type="GP Consultation",
                  incurred_date=date(2039, 7, 15), amount_claimed=45.0,
                  currency="SGD", status="submitted",
                  reference_no="FLEXICO-000006"),
            Claim(client_id=CLIENT_ID, policy_year_id=PY_ID,
                  employee_id=EMP_NOFLEX_LEAVER, claim_kind="insured",
                  product_code="GHS", claim_type="GP Consultation",
                  incurred_date=date(2039, 8, 1), amount_claimed=60.0,
                  currency="SGD", status="submitted",
                  reference_no="FLEXICO-000007"),
        ])
        s.commit()
    app.dependency_overrides[get_current_user] = lambda: _user()
    yield
    app.dependency_overrides.clear()


@pytest.fixture()
def client():
    with TestClient(app) as c:
        yield c


def _sheet(resp, title: str):
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    assert title in wb.sheetnames, wb.sheetnames
    rows = list(wb[title].iter_rows(values_only=True))
    return rows[0], rows[1:]


# Each of these was a download of its own; they are sheets of two workbooks
# now. Same builders — see `test_a_grafted_sheet_matches_its_standalone_builder`.
_IN_WORKBOOK = {
    "wallet-utilisation": ("flex-wallet", "Ledger"),
    "wallet-utilisation-summary": ("flex-wallet", "Summary"),
    "leaver-summary": ("leavers", "Summary"),
    "leaver-details": ("leavers", "Claims"),
}


def _get(client, report, **params):
    key, title = _IN_WORKBOOK[report]
    return _sheet(
        client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/workbooks/{key}", params=params
        ),
        title,
    )


# ── Ledger ───────────────────────────────────────────────────────────────────

def test_ledger_emits_a_dated_row_per_movement(client):
    header, rows = _get(client, "wallet-utilisation")
    mine = [r for r in rows if r[header.index("Staff ID")] == "FX-1"]
    descriptions = [r[header.index("Description")] for r in mine]
    assert "Wallet Allocation" in descriptions
    assert descriptions.count("Claim Payment") == 2  # approved + paid
    assert all(r[header.index("Date of Transaction")] is not None for r in mine)


def test_ledger_keeps_the_incumbents_unbuilt_columns_blank(client):
    """They are blank in the incumbent's own live file too — kept for a
    column-for-column diff, never invented."""
    header, rows = _get(client, "wallet-utilisation")
    for col in ("B/F Allocation Amt", "Deals Amt", "Salary Deduction"):
        assert col in header
        assert all(r[header.index(col)] in (None, "") for r in rows)


def test_ledger_skips_members_flex_does_not_apply_to(client):
    _, rows = _get(client, "wallet-utilisation")
    assert "FX-3" not in {r[1] for r in rows}


# ── Summary ──────────────────────────────────────────────────────────────────

def test_summary_subtracts_settled_claims_and_not_pending_ones(client):
    """300 approved + 200 paid = 500 spent; the 150 in flight is reported
    separately and never subtracted."""
    header, rows = _get(client, "wallet-utilisation-summary")
    row = next(r for r in rows if r[header.index("Staff ID")] == "FX-1")
    assert row[header.index("Claims Payment Amt")] == 500.0
    assert row[header.index("Pending Claims Payment Amt")] == 150.0
    assert row[header.index("Total Allocation Amt")] == 2000.0
    assert row[header.index("Balance Available Allocation Amt")] == 1500.0


def test_summary_terms_add_up_to_the_total(client):
    """The printed terms must reconcile with the printed total — the defect the
    broker coverage pane had when the leave trade was left out."""
    header, rows = _get(client, "wallet-utilisation-summary")
    row = next(r for r in rows if r[header.index("Staff ID")] == "FX-1")

    def cell(name):
        return row[header.index(name)] or 0

    terms = (
        cell("Selection Amt")
        + cell("Claims Payment Amt")
        + cell("Buy Leave Amt")
        - cell("Sell Leave Amt")
    )
    assert round(terms, 2) == cell("Total Utilized Amt")
    assert round(
        cell("Total Allocation Amt") - cell("Total Utilized Amt"), 2
    ) == cell("Balance Available Allocation Amt")


def test_summary_and_ledger_agree(client):
    """Two views of one resolution — the whole reason the ledger is derived."""
    lheader, lrows = _get(client, "wallet-utilisation")
    sheader, srows = _get(client, "wallet-utilisation-summary")
    ledger_claims = sum(
        r[lheader.index("Claims Payment Amt")] or 0
        for r in lrows
        if r[lheader.index("Staff ID")] == "FX-1"
    )
    summary_row = next(r for r in srows if r[sheader.index("Staff ID")] == "FX-1")
    assert ledger_claims == summary_row[sheader.index("Claims Payment Amt")]


def test_summary_masks_nric_by_default(client):
    header, rows = _get(client, "wallet-utilisation-summary")
    row = next(r for r in rows if r[header.index("Staff ID")] == "FX-1")
    assert row[header.index("Identification No.")] != "S1111111A"


# ── Leavers ──────────────────────────────────────────────────────────────────

def _row(header, rows, staff: str):
    return next(r for r in rows if r[header.index("Staff ID")] == staff)


def test_leaver_summary_lists_only_in_period_leavers(client):
    header, rows = _get(client, "leaver-summary")
    assert {r[header.index("Staff ID")] for r in rows} == {"FX-2", "FX-6"}


def test_leaver_benefit_end_is_their_last_day_not_the_year_end(client):
    """The whole point of the sheet: cover stops when they leave."""
    header, rows = _get(client, "leaver-summary")
    row = _row(header, rows, "FX-2")
    assert row[header.index("Benefit End Date")] == datetime(2039, 6, 30)
    assert row[header.index("Last Day of Service")] == datetime(2039, 6, 30)


def test_leaver_summary_carries_the_final_wallet_position(client):
    header, rows = _get(client, "leaver-summary")
    row = _row(header, rows, "FX-2")
    assert row[header.index("Total Allocation Amt")] == 1000.0
    assert row[header.index("Claims Payment Amt")] == 120.0
    assert row[header.index("Balance Available Allocation Amt")] == 880.0


def test_a_pre_period_leaver_is_excluded_on_the_rosters_own_date(client):
    """`terminated_effective` is written by the listing sync alone; a roster
    imported any other way carries the leaving date in `attribute_values`. The
    period filter used to read the column while every sheet PRINTED the resolved
    value, so a 2038 leaver landed on the 2039 sheet with 2038 beside them."""
    header, rows = _get(client, "leaver-summary")
    assert "FX-4" not in {r[header.index("Staff ID")] for r in rows}


def test_a_leaver_who_left_after_the_period_is_not_this_years_leaver(client):
    """Covered for the whole of 2039, so they belong to 2040's sheet — but the
    insurer listing still needs them, which is why only this sheet drops them."""
    header, rows = _get(client, "leaver-summary")
    assert "FX-5" not in {r[header.index("Staff ID")] for r in rows}

    from app.services.insurer_reports import report_employees
    with SessionLocal() as s:
        py = s.get(PolicyYear, PY_ID)
        reportable = {e.staff_id for e in report_employees(s, py)}
    assert "FX-5" in reportable, "the insurer listing still has to off-bill them"
    assert "FX-4" not in reportable, "a pre-period leaver is out of both"


def test_a_leaver_flex_does_not_apply_to_still_gets_a_row(client):
    """The utilisation summary skips them; this sheet must not. Their cover
    window is the reason the row exists — the wallet columns are simply blank."""
    header, rows = _get(client, "leaver-summary")
    row = _row(header, rows, "FX-6")
    assert row[header.index("Benefit End Date")] == datetime(2039, 9, 30)
    assert row[header.index("Total Allocation Amt")] is None


def test_leaver_details_lists_their_claims_only(client):
    header, rows = _get(client, "leaver-details")
    refs = {r[header.index("Reference No.")] for r in rows}
    assert refs == {
        "FLEXICO-000004", "FLEXICO-000005", "FLEXICO-000006", "FLEXICO-000007",
    }
    flex_row = next(
        r for r in rows if r[header.index("Reference No.")] == "FLEXICO-000004"
    )
    assert flex_row[header.index("Claim Category")] == "Flexible Benefits"


def test_a_log_case_keeps_its_label_and_is_marked_in_its_own_column(client):
    """`log_cases.set_case_type` deliberately does not rewrite `claim_type` — it
    is what the MEMBER sees as the title of their claim. This sheet used to
    overwrite it with "LOG", which was the one place that label was destroyed."""
    header, rows = _get(client, "leaver-details")
    row = next(
        r for r in rows if r[header.index("Reference No.")] == "FLEXICO-000005"
    )
    assert row[header.index("Claim Type")] == (
        "Emergency Accidental Outpatient Treatment"
    )
    assert row[header.index("LOG")] == "Yes"
    assert next(
        r for r in rows if r[header.index("Reference No.")] == "FLEXICO-000006"
    )[header.index("LOG")] == "No"


def test_a_claim_incurred_after_cover_end_is_flagged(client):
    """The sheet's premise, made checkable — and the end date sits beside the
    flag so a reader can verify it without opening the Summary sheet."""
    header, rows = _get(client, "leaver-details")
    after = next(
        r for r in rows if r[header.index("Reference No.")] == "FLEXICO-000006"
    )
    assert after[header.index("Incurred After Cover End")] == "Yes"
    assert after[header.index("Benefit End Date")] == datetime(2039, 6, 30)
    within = next(
        r for r in rows if r[header.index("Reference No.")] == "FLEXICO-000005"
    )
    assert within[header.index("Incurred After Cover End")] == "No"


def test_both_leaver_sheets_list_people_in_the_same_order(client):
    """Two sheets of one workbook. The claims were ordered by `employee_id` — an
    opaque identifier — so neither sheet read as sorted and they disagreed."""
    sheader, srows = _get(client, "leaver-summary")
    dheader, drows = _get(client, "leaver-details")
    summary_order = [r[sheader.index("Staff ID")] for r in srows]
    seen: list[str] = []
    for r in drows:
        staff = r[dheader.index("Staff ID")]
        if staff not in seen:
            seen.append(staff)
    assert seen == [s for s in summary_order if s in seen]


def test_leaver_reports_are_audited(client):
    """Every download emits PII, so every download is logged — the composite
    included. It records the WORKBOOK, since that is what left the building."""
    from app.models import AuditLog
    _get(client, "leaver-summary")
    with SessionLocal() as s:
        pulled = {
            (r.after or {}).get("workbook")
            for r in s.query(AuditLog).filter(
                AuditLog.entity_type == "report_workbook"
            ).all()
        }
    assert "leavers" in pulled


# ── Composite workbooks ──────────────────────────────────────────────────────
#
# These replaced the zip "report sets". A zip of two workbooks and a workbook of
# two sheets carry the same bytes; only the second can be cross-referenced in
# place and keeps its sheet names once it is forwarded on.

def test_workbook_listing_names_every_sheet(client):
    """The sheet list is SERVED. The Reports page prints what is inside a
    workbook before a broker downloads it, and a broker files against what the
    page said was inside — so it must not be a frontend constant that can drift
    from the composer."""
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/workbooks")
    assert res.status_code == 200, res.text
    by_key = {b["key"]: b for b in res.json()}
    wallet = by_key["flex-wallet"]
    assert [s["title"] for s in wallet["sheets"]] == ["Summary", "Ledger"]
    assert all(s["description"] for s in wallet["sheets"])
    assert wallet["requires_insurer"] is False
    assert by_key["insurer-submission"]["requires_insurer"] is True
    # A workbook with no identification number on any sheet says so, rather
    # than offering a masking toggle that would govern nothing.
    assert by_key["activity-access"]["supports_masking"] is False
    assert by_key["activity-access"]["supports_date_range"] is True


def test_workbook_download_is_one_file_with_named_sheets(client):
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/workbooks/flex-wallet"
    )
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers["content-type"]
    wb = load_workbook(BytesIO(res.content))
    # Named, and in declaration order — and NOT "Sheet1", which every one of
    # these workbooks used to be called.
    assert wb.sheetnames == ["Summary", "Ledger"]


def test_a_grafted_sheet_matches_its_standalone_builder(client):
    """The composite is a COPY of the single-sheet builder's output, never a
    reimplementation — which is what stops the workbook and the file it came
    from disagreeing about a member's wallet."""
    from app.services.flex_ledger import build_utilisation_summary_workbook

    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/workbooks/flex-wallet"
    )
    grafted = list(
        load_workbook(BytesIO(res.content))["Summary"].iter_rows(values_only=True)
    )
    with SessionLocal() as s:
        py = s.get(PolicyYear, PY_ID)
        wb = build_utilisation_summary_workbook(s, py, masked=True)
    # Round-trip the standalone workbook through a save/load too, so the two
    # sides are compared as Excel stores them. An in-memory sheet still holds
    # `date` and `float`; the file format has neither — every date reads back as
    # a datetime and a whole float as an int. Comparing an unsaved workbook
    # against a downloaded one fails on that alone and says nothing about the
    # graft.
    buf = BytesIO()
    wb.save(buf)
    direct = list(
        load_workbook(BytesIO(buf.getvalue())).active.iter_rows(values_only=True)
    )
    assert grafted == direct


def test_insurer_workbook_refuses_without_an_insurer(client):
    """An all-blank coverage column reads as "there is nothing to submit"."""
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/workbooks/insurer-submission"
    )
    assert res.status_code == 400


def test_insurer_workbook_refuses_an_unconfigured_insurer(client):
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/workbooks/insurer-submission",
        params={"insurer": "Not An Insurer"},
    )
    assert res.status_code == 404


def test_unknown_workbook_404s(client):
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/workbooks/nonsense")
    assert res.status_code == 404


def test_workbook_viewer_cannot_pull_unmasked(client):
    app.dependency_overrides[get_current_user] = lambda: _user("broker_viewer")
    try:
        res = client.get(
            f"/api/v1/policy-years/{PY_ID}/reports/workbooks/flex-wallet",
            params={"masked": "false"},
        )
        assert res.status_code == 403
    finally:
        app.dependency_overrides[get_current_user] = lambda: _user()


# ── Pro-ration on the sheets ─────────────────────────────────────────────────

def test_the_proration_columns_are_blank_when_nothing_was_prorated(client):
    """The columns only carry a figure where it differs from the one beside it.
    Every member here holds a full annual allowance, so both stay empty rather
    than restating `Total Allocation Amt` and a `12/12` nobody needs."""
    for kind in ("leaver-summary", "wallet-utilisation-summary"):
        header, rows = _get(client, kind)
        assert "Annual Allocation Amt" in header
        for r in rows:
            assert r[header.index("Annual Allocation Amt")] is None
            assert not r[header.index("Pro-ration")]


def test_a_prorated_wallet_prints_its_annual_figure_and_the_fraction(client):
    """A reduced number with nothing explaining it is unauditable, and this is
    the sheet people argue over."""
    with SessionLocal() as s:
        emp = s.get(Employee, EMP_LEAVER)
        emp.flex_wallet_amount = 500.0
        emp.flex_proration = {
            "basis": "months_served", "factor": 0.5, "served": 6, "total": 12,
            "full_amount": 1000.0,
            "period_start": "2039-01-01", "period_end": "2039-12-31",
        }
        s.commit()
    try:
        header, rows = _get(client, "leaver-summary")
        row = _row(header, rows, "FX-2")
        assert row[header.index("Total Allocation Amt")] == 500.0
        assert row[header.index("Annual Allocation Amt")] == 1000.0
        assert row[header.index("Pro-ration")] == "6/12 months"
    finally:
        with SessionLocal() as s:
            emp = s.get(Employee, EMP_LEAVER)
            emp.flex_wallet_amount, emp.flex_proration = 1000.0, None
            s.commit()


def test_the_balance_never_prints_negative(client):
    """A flex wallet pays UP TO the limit — a member with S$500 left who presents
    a S$700 bill utilises S$500 and pays the rest themselves. So "overspent" is
    not a state the product can be in, and a negative balance would be an
    indication of something that cannot happen. Reachable on paper only because
    pro-ration binds FORWARD: it can shrink an allowance below what was already
    reimbursed, and it never reaches back for that money."""
    with SessionLocal() as s:
        emp = s.get(Employee, EMP_LEAVER)
        emp.flex_wallet_amount = 50.0  # below the 120.00 already claimed
        s.commit()
    try:
        header, rows = _get(client, "leaver-summary")
        row = _row(header, rows, "FX-2")
        assert row[header.index("Total Utilized Amt")] == 120.0
        assert row[header.index("Balance Available Allocation Amt")] == 0.0
    finally:
        with SessionLocal() as s:
            s.get(Employee, EMP_LEAVER).flex_wallet_amount = 1000.0
            s.commit()


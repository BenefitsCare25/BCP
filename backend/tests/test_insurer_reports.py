"""Insurer reports — benefit-selection + buy/sell-leave listing.

Covers the status-vocabulary mapping, in-period leaver inclusion (pre-period
leavers excluded), the leave columns, masked/unmasked NRIC gating (viewer
403), and the export audit trail.
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_insurer_reports.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import UTC, date, datetime, timedelta  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    AuditLog,
    Client,
    Employee,
    Enrollment,
    EnrollmentWindow,
    LeaveElection,
    PolicyYear,
    User,
)
from app.models.enrollment import EnrollmentStatus  # noqa: E402
from app.models.enrollment_window import WindowStatus  # noqa: E402
from app.models.policy_year import PolicyYearStatus  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-00000012e000"
PY_ID = "00000000-0000-0000-0000-00000012e001"
WINDOW_OLD = "00000000-0000-0000-0000-00000012e010"
WINDOW_NEW = "00000000-0000-0000-0000-00000012e011"
USER_ID = "00000000-0000-0000-0000-00000012e0ff"

EMP_CONFIRMED = "00000000-0000-0000-0000-00000012e101"
EMP_PROGRESS = "00000000-0000-0000-0000-00000012e102"
EMP_UNTOUCHED = "00000000-0000-0000-0000-00000012e103"
EMP_LEAVER = "00000000-0000-0000-0000-00000012e104"
EMP_OLD_LEAVER = "00000000-0000-0000-0000-00000012e105"
EMP_DEEMED = "00000000-0000-0000-0000-00000012e106"


def _user(role: str = "broker_admin") -> CurrentUser:
    return CurrentUser(
        user_id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role=role,
    )


def _emp(eid: str, staff: str, name: str, **kw) -> Employee:
    defaults = dict(
        client_id=CLIENT_ID, policy_year_id=PY_ID, staff_id=staff,
        employee_name=name, attribute_values={}, derived_attribute_values={},
        matched_categories=[], source="csv_import", status="active",
    )
    defaults.update(kw)
    return Employee(id=eid, **defaults)


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    now = datetime.now(UTC)
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Report Co", broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.add(User(
            id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
            email="processor@report.co", display_name="Pat Processor",
            role="broker_admin", status="active",
        ))
        s.flush()
        s.add(PolicyYear(
            id=PY_ID, client_id=CLIENT_ID, year=2033,
            start_date=date(2033, 1, 1), end_date=date(2033, 12, 31),
            status=PolicyYearStatus.active,
        ))
        s.flush()
        s.add_all([
            _emp(
                EMP_CONFIRMED, "IR-1", "Alice Confirmed",
                attribute_values={
                    "entity": "Report Co Pte Ltd",
                    "id_no": "S1234567D",
                    "date_of_hire": "2020-03-01",
                },
                flex_currency="SGD",
            ),
            _emp(EMP_PROGRESS, "IR-2", "Bob Progress"),
            _emp(EMP_UNTOUCHED, "IR-3", "Cara Untouched"),
            _emp(
                EMP_LEAVER, "IR-4", "Dan Leaver",
                status="terminated", terminated_effective=date(2033, 5, 31),
            ),
            _emp(
                EMP_OLD_LEAVER, "IR-5", "Eve Ancient",
                status="terminated", terminated_effective=date(2032, 6, 30),
            ),
            _emp(EMP_DEEMED, "IR-6", "Fay Deemed"),
        ])
        # An older closed window that must NOT drive the report…
        s.add(EnrollmentWindow(
            id=WINDOW_OLD, policy_year_id=PY_ID, client_id=CLIENT_ID,
            name="OE 2032", opens_at=now - timedelta(days=200),
            closes_at=now - timedelta(days=170), status=WindowStatus.closed,
        ))
        # …and the latest window the report should read.
        s.add(EnrollmentWindow(
            id=WINDOW_NEW, policy_year_id=PY_ID, client_id=CLIENT_ID,
            name="OE 2033", opens_at=now - timedelta(days=10),
            closes_at=now + timedelta(days=10), status=WindowStatus.open,
            allow_leave=True,
        ))
        s.flush()
        confirmed = Enrollment(
            id="00000000-0000-0000-0000-00000012e201",
            window_id=WINDOW_NEW, policy_year_id=PY_ID, client_id=CLIENT_ID,
            employee_id=EMP_CONFIRMED, status=EnrollmentStatus.confirmed,
            submitted_at=now - timedelta(days=3),
            confirmed_at=now - timedelta(days=1), confirmed_by=USER_ID,
        )
        s.add(confirmed)
        s.add(Enrollment(
            id="00000000-0000-0000-0000-00000012e202",
            window_id=WINDOW_NEW, policy_year_id=PY_ID, client_id=CLIENT_ID,
            employee_id=EMP_PROGRESS, status=EnrollmentStatus.in_progress,
        ))
        s.add(Enrollment(
            id="00000000-0000-0000-0000-00000012e203",
            window_id=WINDOW_NEW, policy_year_id=PY_ID, client_id=CLIENT_ID,
            employee_id=EMP_DEEMED, status=EnrollmentStatus.deemed,
        ))
        # A row in the OLD window that would flip Cara to confirmed if the
        # report picked the wrong window.
        s.add(Enrollment(
            id="00000000-0000-0000-0000-00000012e204",
            window_id=WINDOW_OLD, policy_year_id=PY_ID, client_id=CLIENT_ID,
            employee_id=EMP_UNTOUCHED, status=EnrollmentStatus.confirmed,
            submitted_at=now - timedelta(days=180),
        ))
        s.flush()
        s.add(LeaveElection(
            enrollment_id=confirmed.id, policy_year_id=PY_ID,
            client_id=CLIENT_ID, employee_id=EMP_CONFIRMED,
            action="sell", days=2.0, flex_amount=500.0, status="confirmed",
        ))
        s.commit()
    yield
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture
def client() -> TestClient:
    app.dependency_overrides[get_current_user] = _user
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture
def viewer_client() -> TestClient:
    app.dependency_overrides[get_current_user] = lambda: _user("broker_viewer")
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


def _rows(res) -> list[dict]:
    ws = load_workbook(BytesIO(res.content)).active
    data = list(ws.iter_rows(values_only=True))
    header = list(data[0])
    return [dict(zip(header, r, strict=False)) for r in data[1:]]


def _by_staff(res) -> dict[str, dict]:
    return {r["Staff ID"]: r for r in _rows(res)}


def test_status_vocabulary_and_selection_flags(client: TestClient) -> None:
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection")
    assert res.status_code == 200, res.text
    rows = _by_staff(res)

    alice = rows["IR-1"]
    assert alice["Status"] == "Processed"
    assert (alice["Employee Selected"], alice["Employee Submitted"]) == ("Yes", "Yes")
    assert alice["ProcessedByName"] == "Pat Processor"
    assert alice["ProcessedOn"] is not None

    bob = rows["IR-2"]
    assert bob["Status"] == "In Progress"
    assert (bob["Employee Selected"], bob["Employee Submitted"]) == ("Yes", "No")

    # Cara has NO row in the latest window (only the stale old one) → Not Started.
    cara = rows["IR-3"]
    assert cara["Status"] == "Not Started"
    assert (cara["Employee Selected"], cara["Employee Submitted"]) == ("No", "No")

    fay = rows["IR-6"]
    assert fay["Status"] == "Processed"
    assert (fay["Employee Selected"], fay["Employee Submitted"]) == ("No", "No")


def test_leaver_inclusion_window(client: TestClient) -> None:
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection")
    rows = _by_staff(res)
    # In-period leaver included, with the last day; pre-period leaver dropped.
    assert "IR-4" in rows
    assert str(rows["IR-4"]["Last Day of Service"]).startswith("2033-05-31")
    assert "IR-5" not in rows


def test_leave_columns(client: TestClient) -> None:
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection")
    rows = _by_staff(res)
    alice = rows["IR-1"]
    assert alice["Buy or Sell Leave"] == "sell"
    assert alice["Days to Sell"] == 2
    assert alice["Days to Buy"] in (None, "")
    assert alice["Currency"] == "SGD"
    assert alice["PriceTag"] == 500
    # No leave election → all leave cells blank.
    bob = rows["IR-2"]
    assert bob["Buy or Sell Leave"] in (None, "")
    assert bob["PriceTag"] in (None, "")


def test_demographics_and_masked_nric(client: TestClient) -> None:
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection")
    alice = _by_staff(res)["IR-1"]
    assert alice["Entity"] == "Report Co Pte Ltd"
    assert str(alice["Date of Hire"]).startswith("2020-03-01")
    nric = alice["Identification No."]
    assert nric == "S******7D"


def test_unmasked_nric_for_write_role(client: TestClient) -> None:
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection?masked=false"
    )
    assert res.status_code == 200
    assert _by_staff(res)["IR-1"]["Identification No."] == "S1234567D"


def test_viewer_unmasked_403_masked_ok(viewer_client: TestClient) -> None:
    res = viewer_client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection?masked=false"
    )
    assert res.status_code == 403
    res = viewer_client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection"
    )
    assert res.status_code == 200
    assert _by_staff(res)["IR-1"]["Identification No."] == "S******7D"


def test_download_is_audited(client: TestClient) -> None:
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection?masked=false"
    )
    assert res.status_code == 200
    with SessionLocal() as s:
        rows = (
            s.query(AuditLog)
            .filter(
                AuditLog.entity_type == "insurer_report",
                AuditLog.entity_id == PY_ID,
            )
            .all()
        )
        assert rows and all(
            r.after["report"] == "benefit-selection" for r in rows
        )
        # This download was unmasked — the choice must be on the record.
        assert any(r.after["masked"] is False for r in rows)


def test_pinned_window_id_overrides_latest(client: TestClient) -> None:
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/benefit-selection"
        f"?window_id={WINDOW_OLD}"
    )
    rows = _by_staff(res)
    # In the OLD window Cara confirmed and Alice never enrolled.
    assert rows["IR-3"]["Status"] == "Processed"
    assert rows["IR-1"]["Status"] == "Not Started"

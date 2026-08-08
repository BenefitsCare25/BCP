"""Report version retention: versioned series + movement, latest-mode supersede.

Fixture mirrors test_insurer_listings: one insurer ("TestSure") over a lump-sum
product (TLIF) + a schedule product (TMD2), a family employee + a solo employee.
"""
from __future__ import annotations

import os
from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_report_versions.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.core.storage import get_storage  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Category,
    Client,
    Dependant,
    Employee,
    Plan,
    PolicyYear,
    Product,
    ProductTerm,
    ReportVersion,
    User,
)
from app.models.employee import EMPLOYEE_STATUS_TERMINATED  # noqa: E402
from app.models.policy_year import PolicyYearStatus  # noqa: E402
from app.services.report_registry import REGISTRY, build_report_bytes  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-0000000r2000"
PY_ID = "00000000-0000-0000-0000-0000000r2001"
LIF_PROD = "00000000-0000-0000-0000-0000000r2010"
MED_PROD = "00000000-0000-0000-0000-0000000r2011"
LIF_CAT = "00000000-0000-0000-0000-0000000r2020"
MED_CAT = "00000000-0000-0000-0000-0000000r2023"
MED_PLAN = "00000000-0000-0000-0000-0000000r2030"
EMP_FAMILY = "00000000-0000-0000-0000-0000000r2101"
EMP_SOLO = "00000000-0000-0000-0000-0000000r2102"
DEP_SPOUSE = "00000000-0000-0000-0000-0000000r2201"

_MATCHED = [
    {"category_id": MED_CAT, "product_code": "TMD2", "method": "rule", "confidence": 1.0},
    {"category_id": LIF_CAT, "product_code": "TLIF", "method": "rule", "confidence": 1.0},
]


def _user() -> CurrentUser:
    return CurrentUser(
        user_id="00000000-0000-0000-0000-0000000r20ff",
        broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role="broker_admin",
    )


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Versions Co", broker_firm_id=DEMO_BROKER_FIRM_ID))
        # A real user row so the history can print WHO generated a version —
        # `generated_by_user_id` was stored and served and rendered nowhere.
        s.add(User(id="00000000-0000-0000-0000-0000000r20ff",
                   broker_firm_id=DEMO_BROKER_FIRM_ID, email="rita@versions.co",
                   display_name="Rita Reports", role="broker_admin",
                   status="active"))
        s.flush()
        s.add(PolicyYear(
            id=PY_ID, client_id=CLIENT_ID, year=2034,
            start_date=date(2034, 1, 1), end_date=date(2034, 12, 31),
            status=PolicyYearStatus.active,
        ))
        s.add_all([
            Product(id=LIF_PROD, client_id=CLIENT_ID, code="TLIF",
                    display_name="Test Life", insurer="TestSure", has_dependants=False),
            Product(id=MED_PROD, client_id=CLIENT_ID, code="TMD2",
                    display_name="Test Medical", insurer="TestSure", has_dependants=True,
                    product_metadata={"report_code": "TMED"}),
        ])
        s.flush()
        s.add_all([
            Category(id=LIF_CAT, policy_year_id=PY_ID, product_id=LIF_PROD,
                     display_name="All staff", raw_description="All staff",
                     plan_assignments={"plan_code": "1", "basis": "100000"},
                     source="manual", status="confirmed"),
            Category(id=MED_CAT, policy_year_id=PY_ID, product_id=MED_PROD,
                     display_name="All staff medical", raw_description="All staff",
                     plan_assignments={"plan_code": "2",
                                       "rate_tiers": {"EO": {"rate": 1.0}, "ES": {"rate": 2.0},
                                                      "EC": {"rate": 2.0}, "EF": {"rate": 3.0}}},
                     source="manual", status="confirmed"),
        ])
        s.add(Plan(id=MED_PLAN, product_id=MED_PROD, policy_year_id=PY_ID,
                   code="2", display_name="Plan 2", report_label="1 Bed / S$80,000"))
        s.add(ProductTerm(policy_year_id=PY_ID, product_id=LIF_PROD, free_cover_limit=50000.0))
        s.add_all([
            Employee(id=EMP_FAMILY, client_id=CLIENT_ID, policy_year_id=PY_ID,
                     staff_id="RV-1", employee_name="Fam Ily",
                     attribute_values={"entity": "Versions Co", "id_no": "S1234567D",
                                       "insurer_member_ids": {"TestSure": "TS-001"}},
                     derived_attribute_values={}, matched_categories=_MATCHED,
                     source="csv_import", status="active"),
            Employee(id=EMP_SOLO, client_id=CLIENT_ID, policy_year_id=PY_ID,
                     staff_id="RV-2", employee_name="So Lo",
                     attribute_values={"id_no": "S7654321D",
                                       "insurer_member_ids": {"TestSure": "TS-002"}},
                     derived_attribute_values={}, matched_categories=_MATCHED,
                     source="csv_import", status="active"),
        ])
        s.flush()
        s.add(Dependant(id=DEP_SPOUSE, client_id=CLIENT_ID, policy_year_id=PY_ID,
                        employee_id=EMP_FAMILY,
                        attribute_values={"dependant_name": "Spo Use", "relationship": "Spouse",
                                          "date_of_birth": "1990-01-01"},
                        status="active"))
        s.commit()
    yield
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture
def client() -> TestClient:
    app.dependency_overrides[get_current_user] = _user
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


def _sheet(content: bytes):
    wb = load_workbook(BytesIO(content))
    ws = wb.worksheets[0]
    return [[c.value for c in row] for row in ws.iter_rows()]


def _create(client, report_type, **params):
    return client.post(
        f"/api/v1/policy-years/{PY_ID}/report-versions",
        json={"report_type": report_type, **params},
    )


# ── Registry dispatch coverage ───────────────────────────────────────────────


def test_every_report_type_builds_bytes() -> None:
    with SessionLocal() as s:
        py = s.get(PolicyYear, PY_ID)
        for rt in REGISTRY:
            params = {"masked": True}
            if REGISTRY[rt].scope == "insurer":
                params["insurer"] = "TestSure"
            data = build_report_bytes(s, py, rt, params)
            assert isinstance(data, bytes) and len(data) > 0, rt


# ── Versioned series + movement ──────────────────────────────────────────────


def test_versioned_flow_and_movement(client: TestClient) -> None:
    # v1 of the employee listing for TestSure.
    res = _create(client, "employee_listing", insurer="TestSure")
    assert res.status_code == 200, res.text
    v1 = res.json()
    assert v1["version_no"] == 1
    assert v1["summary"]["employee_count"] == 2  # RV-1 + RV-2 covered by TestSure

    # It downloads as a real xlsx (PK zip magic).
    dl = client.get(f"/api/v1/report-versions/{v1['id']}/download")
    assert dl.status_code == 200
    assert dl.content[:2] == b"PK"

    # Freshly saved → not stale.
    st = client.get(
        f"/api/v1/policy-years/{PY_ID}/report-versions/status",
        params={"report_type": "employee_listing", "scope_key": "testsure"},
    ).json()
    assert st["latest"]["id"] == v1["id"]
    assert st["is_stale"] is False

    # Saving again with no change is a no-op — no duplicate v2.
    again = _create(client, "employee_listing", insurer="TestSure")
    assert again.json()["unchanged"] is True
    assert again.json()["version_no"] == 1

    # Roster movement: add a covered hire, terminate the solo employee. Stamp a
    # clearly-later updated_at so the staleness gate is deterministic on SQLite's
    # second-granularity clock.
    later = datetime.now(UTC) + timedelta(hours=1)
    with SessionLocal() as s:
        s.add(Employee(
            id="00000000-0000-0000-0000-0000000r2103",
            client_id=CLIENT_ID, policy_year_id=PY_ID, staff_id="RV-3",
            employee_name="New Hire",
            attribute_values={"id_no": "S9999999Z",
                              "insurer_member_ids": {"TestSure": "TS-003"}},
            derived_attribute_values={}, matched_categories=_MATCHED,
            source="adc", status="active", updated_at=later,
        ))
        solo = s.get(Employee, EMP_SOLO)
        solo.status = EMPLOYEE_STATUS_TERMINATED
        solo.terminated_effective = date(2034, 6, 30)
        solo.updated_at = later
        s.commit()

    # Now stale.
    st = client.get(
        f"/api/v1/policy-years/{PY_ID}/report-versions/status",
        params={"report_type": "employee_listing", "scope_key": "testsure"},
    ).json()
    assert st["is_stale"] is True

    # v2 captures the new membership.
    res2 = _create(client, "employee_listing", insurer="TestSure")
    assert res2.status_code == 200, res2.text
    v2 = res2.json()
    assert v2["version_no"] == 2

    # Movement v1 → v2: the hire is an addition, the leaver a deletion.
    mv = client.get(
        f"/api/v1/report-versions/{v2['id']}/movement", params={"since": v1["id"]}
    )
    assert mv.status_code == 200
    text = " ".join(
        str(c) for row in _sheet(mv.content) for c in row if c is not None
    )
    assert "ADDITIONS (1)" in text
    assert "New Hire" in text
    assert "DELETIONS (1)" in text
    assert "So Lo" in text


def test_masked_and_unmasked_are_distinct_versions(client: TestClient) -> None:
    # A masked save and an unmasked save of the SAME membership must NOT collapse
    # into one version, so the broker can retain both an internal (masked) and an
    # insurer (unmasked) copy. The dedup fingerprint is of the rendered BYTES, and
    # the masked NRIC differs in the cells, so they separate naturally.
    r_masked = _create(client, "dependant_listing", insurer="TestSure", masked=True)
    assert r_masked.status_code == 200, r_masked.text
    v_masked = r_masked.json()
    assert v_masked["unchanged"] is False

    r_unmasked = _create(client, "dependant_listing", insurer="TestSure", masked=False)
    assert r_unmasked.status_code == 200, r_unmasked.text
    v_unmasked = r_unmasked.json()
    assert v_unmasked["unchanged"] is False  # NOT collapsed onto the masked one
    assert v_unmasked["version_no"] == v_masked["version_no"] + 1
    assert v_unmasked["summary"]["masked"] is False

    # Re-saving the unmasked one with no change IS a no-op.
    again = _create(client, "dependant_listing", insurer="TestSure", masked=False)
    assert again.json()["unchanged"] is True


def test_viewer_cannot_download_a_retained_unmasked_listing(client: TestClient) -> None:
    """A retained version holds the same PII the live endpoint gates.

    `reports.py` refuses unmasked NRIC/FIN to a `broker_viewer`, but the
    retained blob was reachable with no such check — so a viewer could pull an
    unmasked listing a colleague had saved. The POST guard doesn't help: the
    router sits behind `require_write_access`, so the GET is the only path a
    viewer can take.
    """
    masked_v = _create(client, "employee_listing", insurer="TestSure", masked=True).json()
    unmasked_v = _create(
        client, "employee_listing", insurer="TestSure", masked=False
    ).json()

    def as_viewer() -> CurrentUser:
        return CurrentUser(
            user_id="00000000-0000-0000-0000-0000000r20fe",
            broker_firm_id=DEMO_BROKER_FIRM_ID,
            client_id=CLIENT_ID, role="broker_viewer",
        )

    app.dependency_overrides[get_current_user] = as_viewer
    try:
        blocked = client.get(f"/api/v1/report-versions/{unmasked_v['id']}/download")
        assert blocked.status_code == 403, blocked.text
        # The movement workbook is built from the same listings — gated too.
        moved = client.get(f"/api/v1/report-versions/{unmasked_v['id']}/movement")
        assert moved.status_code == 403, moved.text
        # A masked version stays available to viewers.
        allowed = client.get(f"/api/v1/report-versions/{masked_v['id']}/download")
        assert allowed.status_code == 200, allowed.text
    finally:
        app.dependency_overrides[get_current_user] = _user


# ── Latest-mode supersede ────────────────────────────────────────────────────


def _placement_rows():
    with SessionLocal() as s:
        return (
            s.execute(
                select(ReportVersion).where(
                    ReportVersion.policy_year_id == PY_ID,
                    ReportVersion.report_type == "placement_slip",
                )
            )
            .scalars()
            .all()
        )


def test_latest_mode_noop_then_supersede(client: TestClient) -> None:
    r1 = _create(client, "placement_slip")
    assert r1.status_code == 200, r1.text
    assert r1.json()["version_no"] == 1
    assert r1.json()["unchanged"] is False
    old_path = _placement_rows()[0].storage_path

    # Re-saving with no config change is a no-op — no duplicate row.
    r2 = _create(client, "placement_slip")
    assert r2.json()["unchanged"] is True
    rows = _placement_rows()
    assert len(rows) == 1 and rows[0].version_no == 1

    # A real config change supersedes: new version, prior blob deleted.
    with SessionLocal() as s:
        s.add(Category(
            id="00000000-0000-0000-0000-0000000r2024", policy_year_id=PY_ID,
            product_id=LIF_PROD, display_name="Extra cohort",
            raw_description="Extra cohort",
            plan_assignments={"plan_code": "1", "basis": "25000"},
            source="manual", status="confirmed",
        ))
        s.commit()

    r3 = _create(client, "placement_slip")
    assert r3.json()["unchanged"] is False
    assert r3.json()["version_no"] == 2
    rows = _placement_rows()
    assert len(rows) == 1 and rows[0].version_no == 2

    with pytest.raises(FileNotFoundError):
        get_storage().read(old_path)


def test_content_signature_ignores_write_timestamp():
    """The xlsx signature must be stable across re-saves of identical data (the
    volatile docProps/core.xml timestamp differs) but change with the data — and
    must not re-parse the workbook (openpyxl load) to do it."""
    import time as _time

    from openpyxl import Workbook

    from app.services.report_versions import _content_signature

    class _Spec:
        fmt = "xlsx"

    def _build(rows):
        wb = Workbook()
        ws = wb.active
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    a = _build([["Name", "ID"], ["Alice", 1]])
    _time.sleep(1.1)  # force a different core.xml modified timestamp
    a_again = _build([["Name", "ID"], ["Alice", 1]])
    b = _build([["Name", "ID"], ["Bob", 2]])

    sig_a = _content_signature(_Spec, a)
    assert sig_a is not None
    assert sig_a == _content_signature(_Spec, a_again)  # timestamp ignored
    assert sig_a != _content_signature(_Spec, b)  # data change caught


# ── Retention on download (the "Save version" button is gone) ────────────────


def _download(client, key="insurer-submission", **params):
    """A SUBMISSION pull — unmasked, which is what retention keys on."""
    return client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/workbooks/{key}",
        params={"insurer": "TestSure", "masked": "false", **params},
    )


def _versions(client, report_type):
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/report-versions",
        params={"report_type": report_type, "scope_key": "testsure"},
    )
    assert res.status_code == 200, res.text
    return res.json()


def test_downloading_a_submission_files_a_version(client: TestClient) -> None:
    """The download IS the record. Retention used to need a separate click, so
    the archive held whatever someone remembered to press Save on."""
    before = len(_versions(client, "insurer_submission"))
    res = _download(client)
    assert res.status_code == 200
    assert res.headers["X-Inspro-Report-Filed"].startswith("v")
    after = _versions(client, "insurer_submission")
    assert len(after) == before + 1
    assert after[0]["summary"]["member_count"] > 0
    # Who pulled it, resolved — the id was stored and served and rendered
    # nowhere, so "who sent this" needed a UUID lookup by hand.
    assert after[0]["generated_by"] == "Rita Reports"


def test_a_masked_pull_is_a_preview_and_files_nothing(client: TestClient) -> None:
    """The masked copy is an internal preview — an insurer matches members on
    the identification number, which is why unmasked output is gated as "for
    insurer submission only". Retaining previews put them in the same numbered
    series, so "Last sent v5" could name a file nobody sent, and one pulled
    after a roster change CLEARED the changed-since badge."""
    before = _versions(client, "insurer_submission")
    res = _download(client, masked="true")
    assert res.status_code == 200
    assert "X-Inspro-Report-Filed" not in res.headers
    assert len(_versions(client, "insurer_submission")) == len(before)


def test_a_viewer_download_cannot_write_or_prune(client: TestClient) -> None:
    """`require_write_access` only blocks non-read methods, so retention on a
    GET let the role documented as read-only write rows and blobs — and
    `prune_series` DELETES past the cap, through the same GET."""
    assert _download(client).status_code == 200  # something on file to evict
    before = _versions(client, "insurer_submission")

    def _viewer() -> CurrentUser:
        return CurrentUser(
            user_id="00000000-0000-0000-0000-0000000r20fe",
            broker_firm_id=DEMO_BROKER_FIRM_ID,
            client_id=CLIENT_ID, role="broker_viewer",
        )

    app.dependency_overrides[get_current_user] = _viewer
    try:
        # Unmasked is refused outright; masked is allowed and must file nothing.
        assert _download(client).status_code == 403
        masked = _download(client, masked="true")
        assert masked.status_code == 200
        assert "X-Inspro-Report-Filed" not in masked.headers
    finally:
        app.dependency_overrides[get_current_user] = _user
    assert _versions(client, "insurer_submission") == before


def test_an_unchanged_redownload_files_nothing_new(client: TestClient) -> None:
    """Deduped on content, so the archive grows with data changes and not with
    clicks — which is what makes retaining on every download affordable. The
    header says so, because the record line looks identical either way."""
    assert _download(client).status_code == 200
    first = _versions(client, "insurer_submission")
    res = _download(client)
    assert res.headers["X-Inspro-Report-Filed"].startswith("unchanged:v")
    assert len(_versions(client, "insurer_submission")) == len(first)


def test_an_internal_register_logs_but_retains_nothing(client: TestClient) -> None:
    """Retention is for what a THIRD PARTY acts on. Keeping a blob per pull of a
    working document banks NRIC-bearing files for nothing."""
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/reports/workbooks/member-register",
        params={"masked": "false"},
    )
    assert res.status_code == 200, res.text
    with SessionLocal() as s:
        from app.models import AuditLog
        rows = [
            r for r in s.query(AuditLog).all()
            if (r.after or {}).get("workbook") == "member-register"
        ]
    assert rows, "the pull must still be logged"
    assert all("report_version_id" not in (r.after or {}) for r in rows)


def test_history_merges_the_superseded_series(client: TestClient) -> None:
    """Retiring a report type must not orphan the record of what was submitted
    under it — the bytes are the point and nothing else reaches them."""
    assert _create(client, "employee_listing", insurer="TestSure").status_code == 200
    assert _download(client).status_code == 200
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/report-versions",
        params={
            "report_type": "insurer_submission,employee_listing",
            "scope_key": "testsure",
        },
    )
    assert res.status_code == 200, res.text
    types = {r["report_type"] for r in res.json()}
    assert {"insurer_submission", "employee_listing"} <= types
    # Merged by DATE: version numbers restart per series, so ordering on them
    # would interleave nonsensically.
    stamps = [r["created_at"] for r in res.json()]
    assert stamps == sorted(stamps, reverse=True)


def test_a_repeated_report_type_is_deduped_and_bounded(client: TestClient) -> None:
    """`_spec_or_404` rejects an unknown type but not a repeated one, and each
    entry costs a query returning up to MAX_LIMIT rows."""
    res = client.get(
        f"/api/v1/policy-years/{PY_ID}/report-versions",
        params={
            "report_type": "insurer_submission,insurer_submission",
            "scope_key": "testsure",
        },
    )
    assert res.status_code == 200, res.text
    ids = [r["id"] for r in res.json()]
    assert len(ids) == len(set(ids)), "deduplicated, not listed twice"

    flood = client.get(
        f"/api/v1/policy-years/{PY_ID}/report-versions",
        params={"report_type": ",".join(f"t{i}" for i in range(50))},
    )
    assert flood.status_code == 400


def test_a_pruned_baseline_refuses_instead_of_inventing_a_diff(
    client: TestClient,
) -> None:
    """`compute_movement` reads a missing predecessor as "initial submission —
    everything is an addition", which is right for v1 and catastrophic for a
    version whose predecessor was pruned: every member would be listed under
    ADDITIONS and none under DELETIONS, i.e. wrong in the direction a broker
    acts on."""
    from app.services.report_versions import list_versions, prune_series

    assert _download(client).status_code == 200
    # A second submission off changed data, so there is something to prune under.
    with SessionLocal() as s:
        emp = s.get(Employee, EMP_SOLO)
        attrs = dict(emp.attribute_values or {})
        attrs["salary"] = float(attrs.get("salary") or 0) + 1000
        emp.attribute_values = attrs
        s.commit()
    assert _download(client).status_code == 200

    with SessionLocal() as s:
        py = s.get(PolicyYear, PY_ID)
        assert len(list_versions(s, py, "insurer_submission", "testsure")) >= 2
        prune_series(s, py, "insurer_submission", "testsure", 1)
        s.commit()
        survivor = list_versions(s, py, "insurer_submission", "testsure")[0]
        assert survivor.version_no > 1, "the pruned one was its baseline"
        survivor_id = survivor.id

    res = client.get(f"/api/v1/report-versions/{survivor_id}/movement")
    assert res.status_code == 409, res.text
    assert res.json()["detail"]["code"] == "baseline_pruned"


def test_retention_keeps_only_the_newest_versions() -> None:
    """Nothing pruned before the download became the retention event, when a
    version cost a deliberate click."""
    from app.services.report_registry import RETENTION_KEEP
    from app.services.report_versions import list_versions, prune_series

    with SessionLocal() as s:
        py = s.get(PolicyYear, PY_ID)
        prune_series(s, py, "insurer_submission", "testsure", 1)
        s.flush()
        remaining = list_versions(s, py, "insurer_submission", "testsure")
        s.rollback()
    assert len(remaining) <= 1
    assert RETENTION_KEEP == 24

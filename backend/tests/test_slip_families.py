"""Synthetic per-layout tests for the generalized slip parsing.

Each test builds the minimal sheet-row shape a real workbook uses (VDL tier
headers below "Rate :", dependant-only tier schemes, Hartree Spouse/Child
columns, GPA dependant option categories, location-scoped participation, GBT
annotated premiums, phantom-column workbooks) so the quirks stay covered in CI
without committing any client file.
"""
from __future__ import annotations

from pathlib import Path

from app.services import placement_slip_parser as shim
from app.services.excel_reader import MAX_SCAN_COLS, Cell, open_workbook
from app.services.slip_parsing.models import ExtractedCategory
from app.services.slip_parsing.participation import parse_participation
from app.services.slip_parsing.rates import (
    DEPENDANT_TIER_KEYS,
    _enrich_with_rates,
    extract_rate_section,
)
from app.services.slip_parsing.walk import _Columns, _walk_data_rows

# ── Tier header on the row BELOW "Rate :" (VDL GHS shape) ────────────────────


def test_tier_header_below_rate_anchor_parses_tiered() -> None:
    rows: list[list[Cell]] = [
        [None, "Rate :"],
        [None, "VDL Enabling Technologies", None, None, "EO", None, "ES", None, "EC", None, "EF"],
        [None, None, None, "Plan", "Rate", "Premium", "Rate", "Premium", "Rate",
         "Premium", "Rate", "Premium"],
        [None, None, None, "B3", 265.0, 5035.0, 530.0, 0.0, 530.0, 0.0, 848.0, 0.0],
        [None, None, None, "A", 210.0, 40320.0, 420.0, 0.0, 420.0, 0.0, 672.0, 0.0],
        ["Annual Premium (sbj to GST) :", None, 182348.0],
    ]
    rate_rows, labels = extract_rate_section(rows)
    assert labels is None  # EO/ES/EC/EF are the canonical labels already
    assert [r.key for r in rate_rows] == ["B3", "A"]
    assert all(r.rate_basis == "tiered" for r in rate_rows)
    assert rate_rows[0].rate_tiers == {
        "EO": {"rate": 265.0, "premium": 5035.0},
        "ES": {"rate": 530.0, "premium": 0.0},
        "EC": {"rate": 530.0, "premium": 0.0},
        "EF": {"rate": 848.0, "premium": 0.0},
    }


# ── Dependant-only tier scheme (VDL "GHS - Dependants": EO | SO | CO | SC) ───


def test_dependant_only_tiers_keep_dependant_keys() -> None:
    rows: list[list[Cell]] = [
        [None, "Rate :"],
        [None, "1. VDL Enabling", None, None, "EO", None, "SO", None, "CO", None, "SC"],
        [None, None, None, "Plan", "Rate", "Premium", "Rate", "Premium", "Rate",
         "Premium", "Rate", "Premium"],
        # EO cell is a cross-sheet reference — tolerated, yields no EO tier.
        [None, None, None, "B3", "refer to local tab", None, 407.0, 0.0, 407.0, 0.0, 678.0, 0.0],
        [None, None, None, "A", None, None, 324.0, 0.0, 324.0, 0.0, 541.0, 0.0],
        ["Annual Premium (sbj to GST) :", None, 0.0],
    ]
    rate_rows, _labels = extract_rate_section(rows)
    assert [r.key for r in rate_rows] == ["B3", "A"]
    tiers = rate_rows[0].rate_tiers
    assert tiers is not None
    assert set(tiers) == {"SO", "CO", "SC"}  # EO reference cell yields nothing
    assert tiers["SO"] == {"rate": 407.0, "premium": 0.0}
    # dependant-only keys must never be folded onto composite ES/EC/EF keys
    assert set(tiers) <= DEPENDANT_TIER_KEYS


# ── Hartree "EO | Spouse | Child" columns → SO/CO + preserved labels ─────────


def test_spouse_child_columns_canonicalize_with_labels() -> None:
    rows: list[list[Cell]] = [
        [None, "Rate :"],
        [None, "Hartree Partners", None, "Plan", "EO", None, "Spouse", None, "Child"],
        [None, None, None, None, "Rate", "Premium", "Rate", "Premium", "Rate", "Premium"],
        [None, None, None, "1", 500.0, 10500.0, 380.0, 0.0, 260.0, 0.0],
    ]
    rate_rows, labels = extract_rate_section(rows)
    assert labels == {"SO": "Spouse", "CO": "Child"}
    tiers = rate_rows[0].rate_tiers
    assert tiers is not None
    assert set(tiers) == {"EO", "SO", "CO"}
    assert tiers["SO"]["rate"] == 380.0
    assert tiers["CO"]["rate"] == 260.0


# ── GBT: annotated text premium keeps amount + note ──────────────────────────


def test_gbt_text_premium_amount_and_note() -> None:
    rows: list[list[Cell]] = [
        ["Rate :", "Insured", None, "Category / Name", None, "Annual Premium"],
        [
            None,
            "City Developments Limited",
            None,
            "Senior Management on authorised Journey",
            None,
            "$3,169.80 (Subject to Minimum Policy Premium of S$500)",
        ],
        ["Annual Premium (GST exempt) :", None, 3169.8],
    ]
    rate_rows, _labels = extract_rate_section(rows)
    assert len(rate_rows) == 1
    assert rate_rows[0].rate_basis == "annual_flat"
    assert rate_rows[0].annual_premium == 3169.8
    assert rate_rows[0].premium_note is not None
    assert "Minimum Policy Premium" in rate_rows[0].premium_note

    cat = ExtractedCategory(
        insured="City Developments Limited",
        category="Senior Management on authorised Journey",
        participation="",
        plan_code="",
        source_row=2,
    )
    # The premium is a POLICY-level figure printed once — sibling categories
    # (e.g. "All Other Employees on authorised Journey") must inherit it
    # instead of showing blank financials.
    sibling = ExtractedCategory(
        insured="",
        category="All Other Employees on authorised Journey",
        participation="",
        plan_code="",
        source_row=3,
    )
    enriched = _enrich_with_rates((cat, sibling), rate_rows)
    for c in enriched:
        assert c.annual_premium == 3169.8
        assert c.rate_basis == "annual_flat"
        assert c.premium_note is not None


# ── Location-scoped participation ────────────────────────────────────────────


def test_participation_scope_extraction() -> None:
    sg = parse_participation("Compulsory - SG Office")
    assert sg.employee == "compulsory"
    assert sg.scope == "SG Office"

    flex = parse_participation("Voluntary Flex - SG Office")
    assert flex.employee == "voluntary"
    assert flex.scope == "SG Office"

    thai = parse_participation("Compulsory - Thai Office")
    assert thai.scope == "Thai Office"

    # direction and audience tails are NOT scopes
    assert parse_participation("Voluntary - Downgrade / Upgrade").scope is None
    assert parse_participation(
        "Compulsory - Employees / Voluntary - Dependents"
    ).scope is None
    assert parse_participation("Compulsory").scope is None


def test_walk_captures_location_scope() -> None:
    cols = _Columns(insured=1, category=3, participation=5, plan=-1, num_employees=6)
    rows: list[list[Cell]] = [
        ["", "Insured", "", "Category", "", "Participation", "No. of employees"],
        ["", "CDL", "", "CEO, Deputy CEO, CSO", "", "Compulsory - SG Office", 2],
        ["", "", "", "All Employees", "", "Compulsory - Thai Office", 2],
    ]
    cats = _walk_data_rows(rows, 0, cols, product_code="GPA")
    assert [c.location_scope for c in cats] == ["SG Office", "Thai Office"]


# ── GPA dependant option categories ──────────────────────────────────────────


def test_gpa_spouse_child_option_rows_are_dependant_scope() -> None:
    cols = _Columns(insured=1, category=3, participation=5, plan=-1, num_employees=-1)
    rows: list[list[Cell]] = [
        ["", "Insured", "", "Category", "", "Participation"],
        ["", "CDL", "", "CEO, Deputy CEO, CSO (Option 1)", "", "Voluntary Flex - SG Office"],
        ["", "", "", "Spouse (Option 1)", "", ""],
        ["", "", "", "Spouse (Option 2)", "", ""],
        ["", "", "", "Child (Option 1)", "", ""],
    ]
    cats = _walk_data_rows(rows, 0, cols, product_code="GPA")
    assert [c.category for c in cats] == [
        "CEO, Deputy CEO, CSO (Option 1)",
        "Spouse (Option 1)",
        "Spouse (Option 2)",
        "Child (Option 1)",
    ]
    assert [c.member_scope for c in cats] == [None, "dependant", "dependant", "dependant"]


def test_dependants_sheet_categories_are_dependant_scope() -> None:
    cols = _Columns(insured=1, category=3, participation=5, plan=7)
    rows: list[list[Cell]] = [
        ["", "Insured", "", "Category", "", "Participation", "", "Plan"],
        ["", "VDL", "", "Grade 40 & above (S Pass) eligible dependants", "", "Voluntary", "", "B3"],
        ["", "", "", "Board of Directors eligible dependants", "", "", "", "B1"],
    ]
    cats = _walk_data_rows(rows, 0, cols, product_code="GHS-DEPENDANTS")
    assert all(c.member_scope == "dependant" for c in cats)


# ── Merged-cell column shift (VDL WICA): header over the insured column ─────


def test_walk_realigns_category_column_under_merged_header() -> None:
    # "Basis of Cover :" row doubles as the column header, with "Category"
    # printed over the insured column; the real category text sits at col 3.
    cols = _Columns(insured=-1, category=1, participation=4, plan=-1, num_employees=5)
    rows: list[list[Cell]] = [
        ["Basis of Cover : ", "Category", "", "", "Participation", "*No. of employees"],
        ["", "VDL Enabling Technologies", "", "Management / Admin", "Compulsory", 112],
        ["", "", "", "Manufacturing Assistants", "", 133],
        ["", "", "", "Forklift Drivers", "", 14],
    ]
    cats = _walk_data_rows(rows, 0, cols, product_code="WICA")
    assert [c.category for c in cats] == [
        "Management / Admin", "Manufacturing Assistants", "Forklift Drivers",
    ]
    assert all(c.insured == "VDL Enabling Technologies" for c in cats)


def test_flat_rates_realign_key_column_and_premium_rate_header() -> None:
    # VDL GBT: "Premium Rate (Per Trip)" is a RATE column (text data), the
    # actual premium lives under "Annual Flat Premium".
    rows: list[list[Cell]] = [
        ["Rate :", "Insured", None, "Category / Name", None, "Plan / Region",
         "Premium Rate (Per Trip)", "Annual Flat Premium"],
        [None, "1. VDL Enabling", None, "Expats & Employees who travel", None,
         "Plan A - International / Asia", "International Regional", 2051.5],
        ["Annual Premium (GST exempt) :", None, None, 2051.5],
    ]
    rate_rows, _ = extract_rate_section(rows)
    assert len(rate_rows) == 1
    assert rate_rows[0].key == "Expats & Employees who travel"
    assert rate_rows[0].annual_premium == 2051.5


def test_flat_rates_per_1000_detected_on_deferred_label_row() -> None:
    # Hartree GTL: "Rate :" alone, the per-$1,000 label one row below.
    rows: list[list[Cell]] = [
        [None, "Rate :"],
        [None, "Insured", None, "Category", "* Sum Insured ( SI )",
         "Rate per $1,000 sum insured", "Annual Premium"],
        [None, "Hartree Partners", None, "All Employees", 54343320, 1.1, 59777.652],
    ]
    rate_rows, _ = extract_rate_section(rows)
    assert len(rate_rows) == 1
    assert rate_rows[0].rate_basis == "per_1000_si"
    assert rate_rows[0].rate == 1.1


def test_tiered_composite_plan_keys_expand_to_categories() -> None:
    # CBRE GHS: rate rows keyed "1A/1B" must reach both plan-1A and plan-1B
    # basis-of-cover categories.
    rows: list[list[Cell]] = [
        ["Rate :", None, None, None, "EO", None, "ES"],
        [None, "Insured", None, "Plan", "Rate", "Premium", "Rate", "Premium"],
        [None, "CBRE PTE LTD", None, "1A/1B", 438.0, 6570.0, 985.5, 3942.0],
    ]
    rate_rows, _ = extract_rate_section(rows)
    cats = tuple(
        ExtractedCategory(
            insured="CBRE PTE LTD", category=name, participation="Compulsory",
            plan_code=code, source_row=i,
        )
        for i, (name, code) in enumerate(
            [("CEO and Executive Directors", "1A"), ("Grandfathered levels", "1B")]
        )
    )
    enriched = _enrich_with_rates(cats, rate_rows)
    assert all(e.rate_basis == "tiered" for e in enriched)
    assert all(e.rate_tiers and e.rate_tiers["EO"]["rate"] == 438.0 for e in enriched)


# ── Phantom-column workbooks stay bounded ────────────────────────────────────


def test_phantom_columns_are_clamped(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "WICA"
    ws.cell(row=1, column=1, value="Basis of Cover :")
    ws.cell(row=2, column=2, value="Insured")
    ws.cell(row=2, column=4, value="Category")
    ws.cell(row=2, column=6, value="Participation")
    ws.cell(row=3, column=2, value="Hartree Partners")
    ws.cell(row=3, column=4, value="Non-Manual Staffs")
    ws.cell(row=3, column=6, value="Compulsory")
    # Stray formatting far to the right — the phantom-column shape openpyxl
    # reports as max_column ≈ 16k on the real Hartree workbook.
    ws.cell(row=1, column=16000, value=" ")
    path = tmp_path / "phantom.xlsx"
    wb.save(path)

    with open_workbook(path) as book:
        sheet = book.sheet("WICA")
    assert all(len(row) <= MAX_SCAN_COLS for row in sheet.rows)
    # The bounded read still yields the real content.
    assert any("Non-Manual Staffs" in str(c) for row in sheet.rows for c in row if c)


# ── Unknown product codes: needs_classification + broker override ───────────


def test_unknown_code_flags_needs_classification_and_resolver_clears_it(
    tmp_path: Path,
) -> None:
    from openpyxl import Workbook

    from app.services.slip_parsing.dispatch import parse_placement_slip
    from app.services.slip_reconcile import reconcile_slip

    wb = Workbook()
    ws = wb.active
    ws.title = "GNEW"
    ws.cell(row=1, column=1, value="Group New Product")
    ws.cell(row=3, column=1, value="Basis of Cover :")
    ws.cell(row=4, column=2, value="Insured")
    ws.cell(row=4, column=4, value="Category")
    ws.cell(row=4, column=6, value="Participation")
    ws.cell(row=5, column=2, value="Acme Pte Ltd")
    ws.cell(row=5, column=4, value="All Employees Grade A")
    ws.cell(row=5, column=6, value="Compulsory")
    path = tmp_path / "unknown_product.xlsx"
    wb.save(path)

    parsed = parse_placement_slip(path, client_label="test")
    assert len(parsed.products) == 1
    assert parsed.products[0].registry_known is False
    diag = reconcile_slip(parsed).diagnostics[0]
    assert diag.needs_classification is True
    assert diag.needs_attention is True
    assert any("unclassified" in i for i in diag.issues)

    # A stored broker classification (via the resolver) clears the flag and
    # applies the chosen layout family.
    def resolver(code: str) -> dict | None:
        if code == "GNEW":
            return {"form_profile": "sum_assured", "layout_family": "si_based"}
        return None

    parsed2 = parse_placement_slip(
        path, client_label="test", classification_resolver=resolver
    )
    assert parsed2.products[0].registry_known is True
    assert parsed2.products[0].layout_family == "si_based"
    diag2 = reconcile_slip(parsed2).diagnostics[0]
    assert diag2.needs_classification is False


# ── Import-surface parity for the shim ───────────────────────────────────────


def test_parser_shim_exports_full_surface() -> None:
    for name in shim.__all__:
        assert hasattr(shim, name), f"placement_slip_parser is missing {name!r}"

"""Placement + quotation slip exports — configured products rendered to .xlsx.

Covers the slip-shaped sheet layout (header label block, Basis of Cover,
tiered/flat rate tables, voluntary age-banded rates, annual-premium total,
SOB plan fold), the ProductTerm coverage override, quotation-mode blanking
(insurer + every rate/premium cell), and the export audit trail.
"""
from __future__ import annotations

import os
import re
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_placement_slip_export.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import date  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    AuditLog,
    Category,
    Client,
    Dependant,
    Employee,
    Plan,
    PolicyYear,
    Product,
    ProductSetup,
    ProductTerm,
    User,
)
from app.models.policy_year import PolicyYearStatus  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-00000015e000"
PY_ID = "00000000-0000-0000-0000-00000015e001"
GHS_ID = "00000000-0000-0000-0000-00000015e010"
GTL_ID = "00000000-0000-0000-0000-00000015e011"
GDN_ID = "00000000-0000-0000-0000-00000015e012"
GDN_CAT_ID = "00000000-0000-0000-0000-00000015e020"
USER_ID = "00000000-0000-0000-0000-00000015e0ff"


def _user() -> CurrentUser:
    return CurrentUser(
        user_id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role="broker_admin",
    )


def _sob(item_value: str) -> dict:
    return {
        "items": [
            {
                "number": "1", "name": "Daily Room & Board",
                "value": item_value, "note": None,
                "limits": [{"label": "Maximum no. of days", "value": "120 days"}],
                "sub_items": [], "properties": {},
            },
        ]
    }


def _seed_roster_product(s) -> None:
    """A product whose figures must come from the roster, not the slip.

    Its category STATES 99 lives at a 50,000 basis (so the stored group sum
    insured is 4,950,000); four members actually match it, one household per
    composite tier. Everything the sheet prints for it must follow the roster.
    A ProductSetup supplies the header/eligibility wording, and a ProductTerm
    the non-evidence limits.
    """
    s.add(Category(
        id=GDN_CAT_ID, policy_year_id=PY_ID, product_id=GDN_ID, priority=4,
        display_name="All Employees", raw_description="All Employees",
        participation_detail={"employee": "compulsory", "raw": "Compulsory"},
        plan_assignments={
            "plan_code": "1", "insured": "Slip Co Pte Ltd",
            "num_employees": 99, "basis": "50000", "sum_insured": 4_950_000.0,
            "premium_rate": 0.1, "rate_basis": "per_1000_si",
        },
    ))
    s.add(ProductTerm(
        policy_year_id=PY_ID, product_id=GDN_ID,
        free_cover_limit=250000.0, nel_age_limit=70,
    ))
    s.add(ProductSetup(
        policy_year_id=PY_ID, product_code="GDN", template_version=1,
        answers={
            "header": {
                "policyholder": "Slip Co Pte Ltd",
                "address": "1 Raffles Place, Singapore",
                "business": "Widget Manufacturing",
                "admin_basis": "Headcount basis",
            },
            "eligibility": {
                "eligibility": "All full time & permanent employees",
                "eligibility_date": "Upon employment",
                "last_entry_age": "68",
                # Not on the slip's own ladder — must still be exported, under
                # the label the product's template declares for it.
                "employee_age_limit": "74",
            },
        },
    ))
    matched = [{"category_id": GDN_CAT_ID, "product_code": "GDN"}]
    # One household per composite tier, so the count block exercises the whole
    # ladder: alone (EO), spouse (ES), child (EC), both (EF).
    household = [
        ("E1", ()), ("E2", ("Spouse",)), ("E3", ("Child",)),
        ("E4", ("Spouse", "Child")),
    ]
    for i, (staff, _rels) in enumerate(household):
        s.add(Employee(
            id=f"00000000-0000-0000-0000-0000001{i}e030",
            client_id=CLIENT_ID, policy_year_id=PY_ID,
            staff_id=staff, employee_name=f"Member {staff}",
            attribute_values={}, matched_categories=matched, status="active",
        ))
    s.flush()  # employees must exist before their dependants reference them
    for i, (_staff, rels) in enumerate(household):
        for rel in rels:
            s.add(Dependant(
                client_id=CLIENT_ID, policy_year_id=PY_ID,
                employee_id=f"00000000-0000-0000-0000-0000001{i}e030",
                attribute_values={"relationship": rel}, status="active",
            ))


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Slip Co", broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.add(User(
            id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
            email="broker@slip.co", display_name="Broker One",
            role="broker_admin", status="active",
        ))
        s.flush()
        s.add(PolicyYear(
            id=PY_ID, client_id=CLIENT_ID, year=2034,
            start_date=date(2034, 1, 1), end_date=date(2034, 12, 31),
            status=PolicyYearStatus.draft,
        ))
        s.add_all([
            Product(
                id=GHS_ID, client_id=CLIENT_ID, code="GHS",
                display_name="Group Hospital & Surgical", insurer="AIA",
            ),
            Product(
                id=GTL_ID, client_id=CLIENT_ID, code="GTL",
                display_name="Group Term Life", insurer="Singlife",
            ),
            # A product with a live roster behind it — the counts, tier split,
            # dependants and sum insured on its sheet all come from the roster.
            Product(
                id=GDN_ID, client_id=CLIENT_ID, code="GDN",
                display_name="Group Dental", insurer="AIA", has_dependants=True,
            ),
        ])
        s.flush()
        _seed_roster_product(s)
        # GTL renews off-cycle — its sheet must show the term window, not the PY's.
        s.add(ProductTerm(
            policy_year_id=PY_ID, product_id=GTL_ID,
            coverage_start=date(2034, 4, 1), coverage_end=date(2035, 3, 31),
        ))
        s.add_all([
            Category(
                policy_year_id=PY_ID, product_id=GHS_ID, priority=1,
                display_name="All Executives", raw_description="All Executives",
                participation_detail={"employee": "compulsory", "raw": "Compulsory"},
                plan_assignments={
                    "plan_code": "1", "insured": "Slip Co Pte Ltd",
                    "num_employees": 40, "sum_insured": 100000.0,
                    "annual_premium": 12000.0,
                    "rate_tiers": {
                        "EO": {"rate": 250.0, "premium": 10000.0},
                        "SO": {"rate": 300.0, "premium": 2000.0},
                    },
                    "tier_labels": {"SO": "Spouse"},
                },
            ),
            Category(
                policy_year_id=PY_ID, product_id=GTL_ID, priority=2,
                display_name="All Staff", raw_description="All Staff",
                participation_model="voluntary",
                plan_assignments={
                    "plan_code": "1", "rate_basis": "age_banded",
                    "voluntary_rates": [
                        {"label": "16 - 30", "min": 16, "max": 30, "rate": 0.55},
                        {"label": "31 - 35", "min": 31, "max": 35, "rate": 0.72},
                    ],
                },
            ),
            Category(
                policy_year_id=PY_ID, product_id=None, priority=3,
                display_name="Mystery cohort", raw_description="Mystery cohort",
            ),
        ])
        # Two GHS plans with differing R&B values → two SOB columns; the GTL
        # plan carries cover data but no schedule → Plan Details table only.
        s.add_all([
            Plan(
                product_id=GHS_ID, policy_year_id=PY_ID, code="1",
                display_name="Plan 1", benefit_schedule=_sob("S$650 per day"),
            ),
            Plan(
                product_id=GHS_ID, policy_year_id=PY_ID, code="2",
                display_name="Plan 2", benefit_schedule=_sob("S$450 per day"),
            ),
            Plan(
                product_id=GTL_ID, policy_year_id=PY_ID, code="1",
                display_name="Plan 1",
                cover_description="24x basic monthly salary",
                annual_policy_limit="S$500,000",
            ),
        ])
        s.commit()
    yield
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture
def client() -> TestClient:
    app.dependency_overrides[get_current_user] = _user
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_current_user, None)


def _download(client: TestClient, kind: str = "placement"):
    res = client.get(f"/api/v1/policy-years/{PY_ID}/reports/{kind}-slip")
    assert res.status_code == 200, res.text
    assert res.content[:2] == b"PK"
    return load_workbook(BytesIO(res.content))


def _cells(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _row_index(rows: list[list], label: str, col: int = 0) -> int:
    return next(i for i, r in enumerate(rows) if r[col] == label)


def _blankish(v) -> bool:
    return v in (None, "")


def test_workbook_shape_and_overview(client: TestClient) -> None:
    wb = _download(client)
    assert wb.sheetnames == ["Overview", "GDN", "GHS", "GTL", "Unassigned"]

    rows = _cells(wb["Overview"])
    assert rows[0][0] == "Placement Slip — Configured Products"
    by_code = {r[0]: r for r in rows if r and r[0] in ("GDN", "GHS", "GTL")}
    # …code, product, insurer, period, categories, plans, members, premium.
    # No roster matches GHS, so its members fall back to the slip's stated 40;
    # the premium is the sum of the tier premiums the sheet prints (10k + 2k).
    assert by_code["GHS"][1:] == [
        "Group Hospital & Surgical", "AIA",
        "01 Jan 2034 to 31 Dec 2034", 1, 2, 40, 12000.0,
    ]
    # GDN's three matched members win over its stated 99.
    assert by_code["GDN"][6] == 4
    # Coverage window comes from the ProductTerm override, not the PY span.
    assert by_code["GTL"][3] == "01 Apr 2034 to 31 Mar 2035"
    # The premium total reconciles with the per-product sheets.
    total = next(r for r in rows if r and r[0] == "Total")
    assert total[7] == pytest.approx(
        sum(r[7] for r in by_code.values() if isinstance(r[7], (int, float)))
    )


def test_header_label_block(client: TestClient) -> None:
    rows = _cells(_download(client)["GHS"])
    assert rows[0][0] == "Group Hospital & Surgical"

    def value_of(label: str):
        return rows[_row_index(rows, label)][2]

    assert value_of("Policyholder :") == "Slip Co"
    assert value_of("Insured :") == "Slip Co Pte Ltd"
    assert value_of("Period of Insurance :") == "01 Jan 2034 to 31 Dec 2034"
    assert value_of("Insurer :") == "AIA"
    # Unknown slip fields are emitted as labelled blanks for the broker.
    for label in ("Group :", "Pool :", "Eligibility :",
                  "Type of Administration :"):
        assert _blankish(value_of(label))
    # The unstored terms appear as labelled blank rows too.
    assert any(r[0] == "Non Evidence Limit :" for r in rows)


def test_basis_of_cover_and_tiered_rates(client: TestClient) -> None:
    rows = _cells(_download(client)["GHS"])
    basis_i = _row_index(rows, "Basis of Cover :")
    assert rows[basis_i + 1][1:8] == [
        "Insured", "Category", "Participation",
        "Plan", "* No. of employees", "* Sum Insured (S$)", None,
    ]
    assert rows[basis_i + 2][1:7] == [
        "Slip Co Pte Ltd", "All Executives", "Compulsory", "1", 40, 100000,
    ]

    rate_i = _row_index(rows, "Rate :")
    # Tier codes sit above their Rate/Premium column pairs.
    assert (rows[rate_i][3], rows[rate_i][5]) == ("EO", "SO")
    assert rows[rate_i + 1][1:8] == [
        "Insured", "Category", "Plan", "Rate", "Premium", "Rate", "Premium",
    ]
    assert rows[rate_i + 2][1:8] == [
        "Slip Co Pte Ltd", "All Executives", "1", 250, 10000, 300, 2000,
    ]

    total_i = _row_index(rows, "Annual Premium (GST-exclusive) :")
    assert rows[total_i][2] == 12000


def test_roster_drives_counts_tiers_and_sum_insured(client: TestClient) -> None:
    """The roster's answer replaces the slip's stated figures, and the sum
    insured is recomputed from it so cover and headcount can't disagree."""
    rows = _cells(_download(client)["GDN"])
    basis_i = _row_index(rows, "Basis of Cover :")
    # A dependant-covering product splits its count block by tier, so the header
    # spans two rows: the block label, then the tier codes beneath it.
    assert rows[basis_i + 1][5:11] == [
        "* No. of members", None, None, None, None, "* No. of dependants",
    ]
    assert rows[basis_i + 2][5:11] == ["EO", "ES", "EC", "EF", "Total", None]
    # One household per tier — alone / spouse / child / both — so 4 members and
    # 4 dependants. The stated 99 lives and 4,950,000 sum insured are BOTH
    # superseded: 4 members x the 50,000 basis = 200,000.
    # …EO, ES, EC, EF, total, dependants, basis, sum insured.
    assert rows[basis_i + 3][5:13] == [1, 1, 1, 1, 4, 4, 50000, 200000]
    # …and the sheet says where its figures came from. No generated-on date:
    # report_versions fingerprints the workbook to skip re-saving an unchanged
    # report, and a date would break that guard every calendar day.
    note = rows[basis_i + 4][1]
    assert note.startswith("* Member counts are from the current roster")
    assert not re.search(r"\d{4}", note)

    # The premium follows the recomputed cover, not the stale one.
    rate_i = _row_index(rows, "Rate :")
    assert rows[rate_i + 2][3:6] == [200000, 0.1, 20.0]


def test_stated_figures_kept_when_nothing_matches(client: TestClient) -> None:
    """A category the roster never matched keeps the slip's stated headcount —
    publishing "0 lives" against a priced cover would be worse — and the
    footnote says so instead of implying the whole table is live."""
    rows = _cells(_download(client)["GHS"])
    basis_i = _row_index(rows, "Basis of Cover :")
    assert rows[basis_i + 2][5] == 40
    note = rows[basis_i + 3][1]
    assert "stated on the placement slip" in note


def test_unrebuildable_cover_is_disclosed_not_silently_mixed(
    client: TestClient,
) -> None:
    """A salary-relative basis has no per-member amount, so when a matched
    member has no salary on file the group cover can't be rebuilt. The live
    headcount then sits beside the slip's sum insured — two different
    populations in one row — which the footnote must say out loud."""
    with SessionLocal() as s:
        cat = s.get(Category, GDN_CAT_ID)
        original = dict(cat.plan_assignments)
        cat.plan_assignments = {
            **original, "basis": "12 times basic monthly salary",
        }
        s.commit()
    try:
        rows = _cells(_download(client)["GDN"])
        basis_i = _row_index(rows, "Basis of Cover :")
        # The count is live…
        assert rows[basis_i + 3][9] == 4
        # …the cover is not, and the sheet says so.
        assert rows[basis_i + 3][12] == 4_950_000
        assert "could not be recomputed from the roster" in rows[basis_i + 4][1]
    finally:
        with SessionLocal() as s:
            s.get(Category, GDN_CAT_ID).plan_assignments = original
            s.commit()


def test_header_and_terms_filled_from_configuration(client: TestClient) -> None:
    """Everything the platform actually stores reaches the sheet: the captured
    header/eligibility wording, fields beyond the slip's own ladder, and the
    non-evidence limits."""
    rows = _cells(_download(client)["GDN"])

    def value_of(label: str):
        return rows[_row_index(rows, label)][2]

    assert value_of("Policyholder :") == "Slip Co Pte Ltd"
    assert value_of("Address :") == "1 Raffles Place, Singapore"
    assert value_of("Business :") == "Widget Manufacturing"
    assert value_of("Eligibility :") == "All full time & permanent employees"
    assert value_of("Eligibility Date :") == "Upon employment"
    assert value_of("Last entry age :") == "68"
    assert value_of("Type of Administration :") == "Headcount basis"
    # A captured field the slip's ladder doesn't name still exports, labelled
    # by the product form rather than dropped.
    assert value_of("Employee Age Limit :") == "74"
    # Both NEL gates are configured, so the terms row states them.
    assert value_of("Non Evidence Limit :") == (
        "S$250,000; underwriting from age 70 (ANB)"
    )


def test_quotation_rate_table_skips_unpriceable_rows(client: TestClient) -> None:
    """A voluntary option states the cover a member COULD elect — until someone
    does there is no sum insured to rate, so it never becomes an empty rate row.
    The aggregate the insurer quotes against is stated instead."""
    with SessionLocal() as s:
        s.add(Category(
            policy_year_id=PY_ID, product_id=GDN_ID, priority=5,
            display_name="All Employees (Option 1)",
            raw_description="All Employees (Option 1)",
            participation_detail={"employee": "voluntary", "raw": "Voluntary"},
            plan_assignments={
                "plan_code": "2", "insured": "Slip Co Pte Ltd",
                "basis": "80000", "premium_rate": 0.1,
                "rate_basis": "per_1000_si",
            },
        ))
        s.commit()
    try:
        rows = _cells(_download(client, kind="quotation")["GDN"])
        rate_i = _row_index(rows, "Rate :")
        categories = [
            r[2] for r in rows[rate_i + 2:] if r[2] not in (None, "")
        ]
        assert "All Employees" in categories
        assert "All Employees (Option 1)" not in categories
        # Every rate/premium cell is blank for the quoting insurer, but the
        # aggregate cover stays — it is what they price against.
        assert _blankish(rows[rate_i + 2][4])
        assert rows[rate_i + 2][3] == 200000
    finally:
        with SessionLocal() as s:
            s.query(Category).filter(
                Category.display_name == "All Employees (Option 1)"
            ).delete()
            s.commit()


def test_sob_fold_and_plan_details(client: TestClient) -> None:
    wb = _download(client)
    ghs = _cells(wb["GHS"])
    sob_i = _row_index(ghs, "SCHEDULE OF BENEFITS / PLAN")
    # Two plans with differing values stay two columns.
    assert ghs[sob_i + 1][:4] == [
        "No.", "Benefit / Definition", "Plan 1", "Plan 2",
    ]
    assert ghs[sob_i + 2][:4] == [
        "1", "Daily Room & Board", "S$650 per day", "S$450 per day",
    ]
    assert ghs[sob_i + 3][1:3] == [
        "    • Maximum no. of days", "120 days",
    ]
    qualifier_row = sob_i + 4  # openpyxl row number; ``ghs`` is zero-based
    assert f"C{qualifier_row}:D{qualifier_row}" in {
        str(rng) for rng in wb["GHS"].merged_cells.ranges
    }

    gtl = _cells(wb["GTL"])
    details_i = _row_index(gtl, "Plan Details")
    assert gtl[details_i + 2][:3] == ["1", "24x basic monthly salary", "S$500,000"]


def test_sob_uses_platform_labels_and_renders_every_qualifier() -> None:
    """The export keeps the full materialized Plan schedules, but uses the
    Product Setup column mapping for insurer-facing headings. Non-limit
    properties (especially outpatient copays) must not disappear."""
    from openpyxl import Workbook

    from app.services.slip_export.sob import write_sob
    from app.services.slip_export.styles import (
        finalize_sheet,
        set_compact_product_widths,
    )

    def schedule(value: str, private: str, copay: str) -> dict:
        return {
            "items": [
                {
                    "number": "A",
                    "name": "Consultation",
                    "value": value,
                    "note": "Includes medication",
                    "limits": [],
                    "sub_items": [
                        {
                            "key": "(a)",
                            "name": "Panel clinic",
                            "value": "As charged",
                            "limits": [],
                            "properties": {},
                        }
                    ],
                    "properties": {
                        "per_visit_private": private,
                        "co_payment_private": copay,
                    },
                },
                {
                    "number": "B",
                    "name": "Curated extra row",
                    "value": "YES",
                    "limits": [],
                    "sub_items": [],
                    "properties": {},
                },
            ]
        }

    plans = [
        Plan(
            code="U01",
            display_name="Plan U01",
            benefit_schedule=schedule("YES", "120", "As charged"),
        ),
        Plan(code="1", display_name="Plan 1", benefit_schedule=schedule("YES", "100", "20%")),
    ]
    # Deliberately contains only one draft item: the renderer must use this for
    # labels/mapping only, never replace the fuller Plan schedules with it.
    answers = {
        "sob": {
            "columns": [
                {"id": "a", "label": "PLAN 1", "plan_codes": ["1"]},
                {"id": "b", "label": "PLAN U01", "plan_codes": ["U01"]},
            ],
            "items": [{"number": "A", "name": "Consultation"}],
        }
    }
    ws = Workbook().active
    write_sob(ws, plans, None, answers=answers, quotation=True)
    rows = _cells(ws)
    sob_i = _row_index(
        rows, "SCHEDULE OF BENEFITS / DEFINITIONS / INSURER RESPONSE"
    )
    assert rows[sob_i + 1][:5] == [
        "No.", "Benefit / Definition", "PLAN 1", "PLAN U01",
        "Insurer Response",
    ]
    assert any(row[1] == "Curated extra row" for row in rows)
    per_visit = next(row for row in rows if row[1] == "    • Per visit — Private Hospital")
    assert per_visit[2:4] == ["100", "120"]
    copay = next(row for row in rows if row[1] == "    • Co-payment — Private Hospital")
    assert copay[2:4] == ["20%", "As charged"]
    nested = next(row for row in rows if row[1] == "    Panel clinic")
    assert nested[:4] == [
        "(a)", "    Panel clinic", "As charged", "As charged",
    ]

    # The source workbook stores outpatient group indices as negative values
    # with an accounting format that displays parentheses. The export must
    # reproduce the visible index, never expose the raw negative sentinel.
    group_plan = Plan(
        code="1",
        display_name="Plan 1",
        benefit_schedule={
            "items": [
                {
                    "number": "-1", "name": "Panel", "value": "YES",
                    "limits": [], "sub_items": [], "properties": {},
                }
            ]
        },
    )
    group_ws = Workbook().active
    write_sob(group_ws, [group_plan], None, quotation=True)
    assert next(row[0] for row in _cells(group_ws) if row[1] == "Panel") == "( 1 )"

    long_plan = Plan(
        code="1",
        display_name="Plan 1",
        benefit_schedule={
            "items": [
                {
                    "number": "1", "name": "Long definition",
                    "value": "word " * 180,
                    "limits": [], "sub_items": [], "properties": {},
                }
            ]
        },
    )
    long_ws = Workbook().active
    write_sob(long_ws, [long_plan], None, quotation=True)
    set_compact_product_widths(long_ws)
    finalize_sheet(long_ws, "Quotation Slip")
    long_row = next(
        row[0].row for row in long_ws.iter_rows() if row[1].value == "Long definition"
    )
    assert long_ws.row_dimensions[long_row].height > 90

    # A compact one-plan sheet must not orphan the SOB header at a page foot.
    compact_ws = Workbook().active
    compact_ws.cell(row=25, column=1, value="Setup content")
    write_sob(compact_ws, [plans[0]], None, quotation=True)
    assert len(compact_ws.row_breaks.brk) == 1


def test_export_has_print_ready_sheet_formatting(client: TestClient) -> None:
    wb = _download(client, kind="quotation")
    for ws in wb.worksheets:
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        assert ws.print_area
        assert ws.sheet_view.showGridLines is False
        assert any(str(merged).startswith("A1:") for merged in ws.merged_cells.ranges)
    assert wb["Overview"].freeze_panes == "A8"
    assert len(wb["GHS"].row_breaks.brk) == 0
    # Product sheets reserve A for compact SOB indexing. Long header values use
    # merged cells instead of making every row inherit an oversized A/C column.
    assert wb["GHS"].column_dimensions["A"].width <= 9
    assert wb["GHS"].column_dimensions["B"].width <= 30
    assert any(
        rng.min_col == 3 and rng.max_col > 3
        for rng in wb["GHS"].merged_cells.ranges
    )


def test_voluntary_rates_block(client: TestClient) -> None:
    gtl = _cells(_download(client)["GTL"])
    # Off-cycle term drives the sheet's period.
    assert gtl[_row_index(gtl, "Period of Insurance :")][2] == "01 Apr 2034 to 31 Mar 2035"
    vol_i = _row_index(gtl, "Voluntary Rates", col=1)
    assert gtl[vol_i + 1][1:3] == [
        "Based on Age Last Birthday", "Rate per 1,000 Sum assured (S$)",
    ]
    bands = {r[1]: r[2] for r in gtl[vol_i + 2 : vol_i + 4]}
    assert bands == {"16 - 30": 0.55, "31 - 35": 0.72}


def test_quotation_mode_blanks_insurer_and_rates(client: TestClient) -> None:
    wb = _download(client, kind="quotation")
    overview = _cells(wb["Overview"])
    assert overview[0][0] == "Quotation Slip — Configured Products"
    ghs_row = next(r for r in overview if r[0] == "GHS")
    assert _blankish(ghs_row[2])  # insurer left for the quoting insurer

    ghs = _cells(wb["GHS"])
    assert _blankish(ghs[_row_index(ghs, "Insurer :")][2])
    rate_i = _row_index(ghs, "Rate :")
    data = ghs[rate_i + 2]
    # Structure kept (insured/category/plan), every rate/premium cell blank.
    assert data[1:4] == ["Slip Co Pte Ltd", "All Executives", "1"]
    assert all(_blankish(v) for v in data[4:8])
    assert _blankish(ghs[_row_index(ghs, "Annual Premium (GST-exclusive) :")][2])

    gtl = _cells(wb["GTL"])
    vol_i = _row_index(gtl, "Voluntary Rates", col=1)
    assert all(_blankish(r[2]) for r in gtl[vol_i + 2 : vol_i + 4])
    # Basis of Cover figures stay — the insurer quotes off them.
    basis_i = _row_index(ghs, "Basis of Cover :")
    assert ghs[basis_i + 2][6] == 100000


def test_flat_premium_rows_reconcile_with_total() -> None:
    """The 'Annual Premium' total is the sum of the premiums actually PRINTED —
    block copies collapse to one shown+counted row, distinct premiums sum, and
    derived per-row premiums (equal SI) each show AND count (never blanked)."""
    from openpyxl import Workbook

    from app.services.slip_export.context import SlipContext
    from app.services.slip_export.rates import _derived_premium, _write_flat_rates

    with SessionLocal() as s:
        ctx = SlipContext(policy_year=s.get(PolicyYear, PY_ID), mode="placement")

    def cat(pa: dict) -> Category:
        return Category(
            policy_year_id=PY_ID, display_name="x", raw_description="x",
            plan_assignments=pa,
        )

    def total_and_shown(cats: list[Category]) -> tuple[float | None, list[float]]:
        ws = Workbook().active
        total = _write_flat_rates(
            ws, cats, ctx, with_label=False, blank=False, insured_default="",
        )
        shown = [
            row[5]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[2] != "Total" and isinstance(row[5], (int, float))
        ]
        return total, shown

    # GCGP-style: 4 cohorts all carry the plan's 186,732 → printed + counted once.
    total, shown = total_and_shown(
        [cat({"insured": "A", "annual_premium": 186732.0}) for _ in range(4)]
    )
    assert total == 186732.0 and shown == [186732.0]

    # Distinct per-category premiums all show and sum.
    total, shown = total_and_shown([
        cat({"insured": "A", "annual_premium": 100.0}),
        cat({"insured": "A", "annual_premium": 200.0}),
    ])
    assert total == 300.0 and sorted(shown) == [100.0, 200.0]

    # The derived premium is rated on the RESOLVED sum insured, so it always
    # prices off the same headcount the Basis-of-Cover table published.
    assert _derived_premium(
        {"rate_basis": "per_1000_si", "premium_rate": 0.072}, 4_000_000.0
    ) == 288.0
    # Derived premiums are genuine per-row figures: two equal-SI cohorts each
    # print 72 and the total is 144 — rows reconcile with the total.
    total, shown = total_and_shown([
        cat({"insured": "A", "rate_basis": "per_1000_si",
             "premium_rate": 0.072, "sum_insured": 1_000_000.0})
        for _ in range(2)
    ])
    assert total == 144.0 and shown == [72.0, 72.0]

    # An annotated note (string) prints once but still contributes its numeric
    # annual_premium to the total; the repeated block copy is blanked.
    total, shown = total_and_shown([
        cat({"insured": "A", "annual_premium": 3169.8,
             "premium_note": "$3,169.80 (min S$500)"}),
        cat({"insured": "A", "annual_premium": 3169.8,
             "premium_note": "$3,169.80 (min S$500)"}),
    ])
    assert total == 3169.8 and shown == []  # both cells hold the string note


def test_policy_number_operational_and_exported(client: TestClient) -> None:
    """Policy numbers are insurer-issued AFTER placement: settable on the
    current (active) year — as is all configuration now (the lock was removed);
    the placement slip shows it, the quotation leaves it blank (goes to
    prospective insurers)."""
    with SessionLocal() as s:
        s.get(PolicyYear, PY_ID).status = PolicyYearStatus.active
        s.commit()
    try:
        res = client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{GHS_ID}",
            json={"policy_number": "POL-2034-001"},
        )
        assert res.status_code == 200, res.text
        assert res.json()["policy_number"] == "POL-2034-001"
        # Coverage dates are editable on an active year too (no lock).
        assert client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{GHS_ID}",
            json={"coverage_start": "2034-02-01", "coverage_end": "2034-12-31"},
        ).status_code == 200

        ghs = _cells(_download(client)["GHS"])
        assert ghs[_row_index(ghs, "Policy No. :")][2] == "POL-2034-001"
        ghs_q = _cells(_download(client, kind="quotation")["GHS"])
        assert _blankish(ghs_q[_row_index(ghs_q, "Policy No. :")][2])
    finally:
        client.put(
            f"/api/v1/policy-years/{PY_ID}/product-terms/{GHS_ID}",
            json={
                "policy_number": None,
                "coverage_start": None,
                "coverage_end": None,
            },
        )
        with SessionLocal() as s:
            s.get(PolicyYear, PY_ID).status = PolicyYearStatus.draft
            s.commit()


def test_unassigned_sheet_and_audit(client: TestClient) -> None:
    wb = _download(client)
    _download(client, kind="quotation")
    unassigned = _cells(wb["Unassigned"])
    assert unassigned[0][0] == "Unassigned categories"
    assert any(r[2] == "Mystery cohort" for r in unassigned if len(r) > 2)

    with SessionLocal() as s:
        reports = {
            (e.after or {}).get("report")
            for e in s.query(AuditLog)
            .filter(
                AuditLog.entity_type == "placement_slip",
                AuditLog.action == "export",
                AuditLog.entity_id == PY_ID,
            )
            .all()
        }
        assert {"placement-slip", "quotation-slip"} <= reports

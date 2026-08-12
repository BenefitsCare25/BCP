"""Built-in (non-insurer) member listings.

The point of these sheets is that they span every insurer and every person on
file, so the tests that matter are the two ways that can silently fail: a
product block dropped because it belongs to the "wrong" insurer, and a row
dropped because the person or the dependant is not billable.
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_built_in_listings.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from datetime import date  # noqa: E402
from io import BytesIO  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.auth import DEMO_BROKER_FIRM_ID, CurrentUser, get_current_user  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Category,
    Client,
    Dependant,
    Employee,
    Plan,
    PolicyYear,
    Product,
    ProductSetup,
    User,
)
from app.models.policy_year import PolicyYearStatus  # noqa: E402

CLIENT_ID = "00000000-0000-0000-0000-0000000b1000"
PY_ID = "00000000-0000-0000-0000-0000000b1001"
USER_ID = "00000000-0000-0000-0000-0000000b10ff"

PROD_GHS = "00000000-0000-0000-0000-0000000b1301"
PROD_GTL = "00000000-0000-0000-0000-0000000b1302"
CAT_GHS = "00000000-0000-0000-0000-0000000b1401"
CAT_GTL = "00000000-0000-0000-0000-0000000b1402"

EMP_ACTIVE = "00000000-0000-0000-0000-0000000b1101"
EMP_LEAVER_IN = "00000000-0000-0000-0000-0000000b1102"
EMP_LEAVER_OLD = "00000000-0000-0000-0000-0000000b1103"
DEP_COVERED = "00000000-0000-0000-0000-0000000b1201"
DEP_PENDING = "00000000-0000-0000-0000-0000000b1202"


def _user(role: str = "broker_admin") -> CurrentUser:
    return CurrentUser(
        user_id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
        client_id=CLIENT_ID, role=role,
    )


def _emp(eid: str, staff: str, name: str, **kw) -> Employee:
    defaults = dict(
        client_id=CLIENT_ID, policy_year_id=PY_ID, staff_id=staff,
        employee_name=name, attribute_values={}, derived_attribute_values={},
        matched_categories=[], source="csv_import", status="active",
    )
    defaults.update(kw)
    return Employee(id=eid, **defaults)


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    from scripts.seed_demo import seed
    seed()
    with SessionLocal() as s:
        s.add(Client(id=CLIENT_ID, name="Builtin Co", slug="builtin-co",
                     broker_firm_id=DEMO_BROKER_FIRM_ID))
        s.add(User(id=USER_ID, broker_firm_id=DEMO_BROKER_FIRM_ID,
                   email="ops@builtin.co", display_name="Bo Ops",
                   role="broker_admin", status="active"))
        s.flush()
        s.add(PolicyYear(id=PY_ID, client_id=CLIENT_ID, year=2037,
                         start_date=date(2037, 1, 1), end_date=date(2037, 12, 31),
                         status=PolicyYearStatus.active))
        s.flush()
        # Two products placed with DIFFERENT insurers — the whole point of a
        # built-in listing is that both appear on one sheet.
        s.add_all([
            Product(id=PROD_GHS, client_id=CLIENT_ID, code="GHS",
                    display_name="Group Hospital & Surgical",
                    product_metadata={"report_code": "GHS"}),
            Product(id=PROD_GTL, client_id=CLIENT_ID, code="GTL",
                    display_name="Group Term Life",
                    product_metadata={"report_code": "GTL"}),
        ])
        s.flush()
        for code, insurer in (("GHS", "AIA"), ("GTL", "Zurich")):
            s.add(ProductSetup(
                policy_year_id=PY_ID, product_code=code,
                answers={"header": {"insurer": insurer}},
            ))
        s.add_all([
            Category(id=CAT_GHS, policy_year_id=PY_ID,
                     product_id=PROD_GHS, display_name="All Employees",
                     raw_description="All Employees",
                     plan_assignments={"plan_code": "Plan1"}),
            Category(id=CAT_GTL, policy_year_id=PY_ID,
                     product_id=PROD_GTL, display_name="All Employees",
                     raw_description="All Employees",
                     plan_assignments={"plan_code": "100K", "basis": "S$ 100,000"}),
        ])
        s.add_all([
            Plan(policy_year_id=PY_ID, product_id=PROD_GHS,
                 code="Plan1", display_name="Plan 1",
                 report_label="1 Bed Pte Hosp"),
        ])
        s.add_all([
            _emp(EMP_ACTIVE, "BI-1", "Ann Active",
                 attribute_values={
                     "entity": "Builtin Co Pte Ltd",
                     "id_no": "S1234567D",
                     "category": "All Employees",
                     # A roster types salary however the sheet was typed —
                     # "5,500" must still print as a number.
                     "salary": "5,500",
                     "company_description": "Builtin Co BHQ",
                     "insurer_member_ids": {"AIA": "AIA-99", "Zurich": "ZUR-77"},
                 },
                 matched_categories=[
                     {"category_id": CAT_GHS, "product_id": PROD_GHS,
                      "product_code": "GHS", "method": "rule"},
                     {"category_id": CAT_GTL, "product_id": PROD_GTL,
                      "product_code": "GTL", "method": "rule"},
                 ]),
            _emp(EMP_LEAVER_IN, "BI-2", "Ben Leaver",
                 status="terminated", terminated_effective=date(2037, 6, 30)),
            # Left BEFORE the period: excluded from the insurer listing, but
            # "who is on file" must still include them.
            _emp(EMP_LEAVER_OLD, "BI-3", "Cy Ancient",
                 status="terminated", terminated_effective=date(2036, 3, 31)),
        ])
        s.flush()
        s.add_all([
            Dependant(id=DEP_COVERED, client_id=CLIENT_ID, policy_year_id=PY_ID,
                      employee_id=EMP_ACTIVE, status="active",
                      attribute_values={"dependant_name": "Dara Dep",
                                        "relationship": "Spouse",
                                        "dependant_id_no": "S7654321Z"}),
            # A portal self-add awaiting approval — covered by nobody yet.
            Dependant(id=DEP_PENDING, client_id=CLIENT_ID, policy_year_id=PY_ID,
                      employee_id=EMP_ACTIVE, status="pending_approval",
                      attribute_values={"dependant_name": "Eli Pending",
                                        "relationship": "Child"}),
        ])
        s.commit()
    app.dependency_overrides[get_current_user] = lambda: _user()
    yield
    app.dependency_overrides.clear()


@pytest.fixture()
def client():
    with TestClient(app) as c:
        yield c


# Both listings are now SHEETS of the Member Register workbook rather than
# files of their own — same builders, one download. The tests reach them by
# sheet name, which is also the assertion that the sheet is named at all: the
# whole point of the composite is that a broker opening it can tell the tabs
# apart, and every one of these workbooks used to be called "Sheet1".
MEMBER_REGISTER = f"/api/v1/policy-years/{PY_ID}/reports/workbooks/member-register"


def _sheet(resp, title: str):
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    assert title in wb.sheetnames, wb.sheetnames
    rows = list(wb[title].iter_rows(values_only=True))
    return rows[0], rows[1:]


def _emp_sheet(client, **params):
    return _sheet(client.get(MEMBER_REGISTER, params=params), "Employees")


def _dep_sheet(client, **params):
    return _sheet(client.get(MEMBER_REGISTER, params=params), "Dependants")


def test_employee_sheet_opens_with_the_incumbent_column_block(client):
    """Columns 1-28 are a column-for-column clone of the file being replaced.

    Asserted literally, not against the module constant: the two files are
    diffed side by side during the migration, so a rename here has to be a
    deliberate edit in two places rather than a constant quietly following the
    code that reads it.
    """
    header, _ = _emp_sheet(client)
    assert list(header[:28]) == [
        "Entity", "User ID", "Employee Name", "Identification No.",
        "Date of Birth", "Gender", "Marital Status",
        "Foreigner Employment Pass", "Nationality", "Monthly Salary",
        "Date of Hire", "Confirmation Date", "Effective Date",
        "Last Day of Service", "Category", "Division", "Department",
        "Cost Centre", "Email Address", "Mobile Phone", "Bank Code",
        "Branch Code", "Bank Account No.", "Company Description",
        "Location Description", "Current Job Grade", "Person Class", "Remarks",
    ]
    # Ours follow the clone rather than being interleaved into it.
    assert header[28] == "Employee Status"


def test_dependant_sheet_opens_with_the_incumbent_column_block(client):
    header, _ = _dep_sheet(client)
    assert list(header[:13]) == [
        "Entity", "Staff ID", "Employee Name", "Employee's Identification No.",
        "Dependant Name", "Dependant's Identification No.", "Relationship",
        "Date of Marriage", "Gender", "Date of Birth", "Effective Date",
        "Remarks", "Deletion Date",
    ]
    assert header[13] == "Termination Date"


def test_salary_is_a_number_and_person_class_defaults(client):
    """Salary must be summable, and Person Class must mean the same thing on a
    roster that has never carried the column."""
    header, rows = _emp_sheet(client)
    row = next(r for r in rows if r[1] == "BI-1")
    assert row[header.index("Monthly Salary")] == 5500.0
    assert row[header.index("Person Class")] == "Employee"
    assert row[header.index("Company Description")] == "Builtin Co BHQ"


def test_employee_listing_spans_every_insurer(client):
    """One sheet, both insurers' products — the reason this report exists."""
    header, _ = _emp_sheet(client)
    assert "GHS Default Plan ID" in header
    assert "GHS Default Group Option" in header
    assert "GTL EE Default Plan ID" in header
    assert "GTL EE Last Accepted Sum Assured" in header


def test_employee_listing_has_a_member_id_column_per_insurer(client):
    """A member holds a different id with each insurer; one column loses all
    but the first."""
    header, rows = _emp_sheet(client)
    assert "AIA Member ID" in header
    assert "Zurich Member ID" in header
    row = next(r for r in rows if r[1] == "BI-1")
    assert row[header.index("AIA Member ID")] == "AIA-99"
    assert row[header.index("Zurich Member ID")] == "ZUR-77"


def test_employee_status_all_includes_pre_period_leavers(client):
    """`all` is everyone on file, which is wider than the insurer listing's
    active + in-period-leaver population."""
    header, rows = _emp_sheet(client, employee_status="all")
    staff = {r[1] for r in rows}
    assert staff == {"BI-1", "BI-2", "BI-3"}
    leaver = next(r for r in rows if r[1] == "BI-3")
    assert leaver[header.index("Employee Status")] == "Terminated"


def test_employee_status_active_drops_every_leaver(client):
    _, rows = _emp_sheet(client, employee_status="active")
    assert {r[1] for r in rows} == {"BI-1"}


def test_employee_status_defaults_to_all(client):
    _, defaulted = _emp_sheet(client)
    _, explicit = _emp_sheet(client, employee_status="all")
    assert len(defaulted) == len(explicit) == 3


def test_unknown_employee_status_falls_back_to_all(client):
    """A typo'd parameter must not silently narrow the population."""
    _, rows = _emp_sheet(client, employee_status="nonsense")
    assert len(rows) == 3


def test_employee_listing_masks_nric_by_default(client):
    header, rows = _emp_sheet(client)
    row = next(r for r in rows if r[1] == "BI-1")
    assert row[header.index("Identification No.")] != "S1234567D"
    _, unmasked = _emp_sheet(client, masked="false")
    row = next(r for r in unmasked if r[1] == "BI-1")
    assert row[header.index("Identification No.")] == "S1234567D"


def test_viewer_cannot_pull_unmasked(client):
    app.dependency_overrides[get_current_user] = lambda: _user("broker_viewer")
    try:
        res = client.get(
            MEMBER_REGISTER,
            params={"masked": "false"},
        )
        assert res.status_code == 403
    finally:
        app.dependency_overrides[get_current_user] = lambda: _user()


def test_dependant_listing_includes_uncovered_dependants(client):
    """A dependant nobody covers is usually the data problem being hunted."""
    header, rows = _dep_sheet(client)
    names = {r[header.index("Dependant Name")] for r in rows}
    assert names == {"Dara Dep", "Eli Pending"}
    pending = next(r for r in rows if r[header.index("Dependant Name")] == "Eli Pending")
    assert pending[header.index("Dependant Status")] == "Pending Approval"


def test_dependant_listing_carries_no_product_columns(client):
    """Coverage is the insurer listing's question, answered there per insurer."""
    header, _ = _dep_sheet(client)
    assert not any("Default Plan ID" in h for h in header)
    assert not any("Family Grouping" in h for h in header)


def test_dependant_member_id_falls_back_to_the_employee_row(client):
    """Rosters put the insurer member id on the employee row, not per life."""
    header, rows = _dep_sheet(client)
    row = next(r for r in rows if r[header.index("Dependant Name")] == "Dara Dep")
    assert row[header.index("AIA Member ID")] == "AIA-99"


def test_dependant_listing_follows_the_employee_status_filter(client):
    _, rows = _dep_sheet(client, employee_status="active")
    assert len(rows) == 2

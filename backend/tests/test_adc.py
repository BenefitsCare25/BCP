"""Roster movement DERIVED from an uploaded member listing: preview + apply.

There is no Action column and no movement template — the broker uploads the
listing and the diff classifies every row. The load-bearing case is
`test_unchanged_listing_is_a_no_op`: the file the system hands out must, sent
straight back, report nothing to do.
"""
from __future__ import annotations

import os
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

import pytest

TEST_DB = Path(__file__).parent / "_test_adc.db"
os.environ["INSPRO_DATABASE_URL"] = f"sqlite:///{TEST_DB}"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.clock import today as business_today  # noqa: E402
from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import Dependant, Employee  # noqa: E402
from app.models.dependant import DEPENDANT_STATUS_TERMINATED  # noqa: E402
from app.models.employee import EMPLOYEE_STATUS_TERMINATED  # noqa: E402
from scripts.seed_demo import seed  # noqa: E402

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EMP_COLS = [
    "Staff ID", "Employee Name", "Identification No. (NRIC/FIN)",
    "Date of Birth", "Category", "Last Day of Service",
]
DEP_COLS = [
    "Staff ID", "Employee Name", "Dependant Name",
    "Dependant's Identification No.", "Relationship", "Date of Birth",
    "Termination Date",
]

PAST = (business_today() - timedelta(days=30)).isoformat()
FUTURE = (business_today() + timedelta(days=30)).isoformat()

# The three seeded employees, as listing rows. Tests derive their files from
# this so "everyone still here" is expressed once.
ROSTER = [
    ["A-1", "Anna Lim", "S1111111A", "1990-01-01", "Executive", ""],
    ["A-2", "Ben Ong", "S2222222B", "1985-02-02", "Manager", ""],
    ["A-3", "Cara Tan", "S3333333C", "1988-03-03", "Executive", ""],
]


@pytest.fixture(scope="module", autouse=True)
def _setup_db():
    if TEST_DB.exists():
        TEST_DB.unlink()
    Base.metadata.create_all(bind=engine)
    seed()
    yield
    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(app)


def _py(client: TestClient) -> str:
    return client.get("/api/v1/policy-years").json()[0]["id"]


def _listing(emp_rows: list[list], dep_rows: list[list] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(EMP_COLS)
    for r in emp_rows:
        ws.append(r)
    if dep_rows is not None:
        dws = wb.create_sheet("Dependants")
        dws.append(DEP_COLS)
        for r in dep_rows:
            dws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _preview(client, py, content):
    return client.post(
        f"/api/v1/policy-years/{py}/adc/preview",
        files={"file": ("listing.xlsx", content, XLSX_MIME)},
    )


def _apply(client, py, content, *, terminate_missing: bool = False):
    return client.post(
        f"/api/v1/policy-years/{py}/adc/apply",
        files={"file": ("listing.xlsx", content, XLSX_MIME)},
        data={"terminate_missing": str(terminate_missing).lower()},
    )


@pytest.fixture(scope="module", autouse=True)
def _seed_roster(_setup_db, client: TestClient):
    py = _py(client)
    res = _apply(client, py, _listing(ROSTER))
    assert res.status_code == 200, res.text
    assert res.json()["added"] == 3
    yield


def _active(staff_id: str) -> Employee:
    db = SessionLocal()
    try:
        return db.execute(
            select(Employee).where(Employee.staff_id == staff_id)
        ).scalar_one()
    finally:
        db.close()


# ── The core invariant ──────────────────────────────────────────────────────


def test_unchanged_listing_is_a_no_op(client: TestClient) -> None:
    """Sending the roster back unchanged must propose nothing.

    Everything else here depends on this: if an untouched round-trip produced
    changes, every real upload would bury its genuine movements in noise.
    """
    py = _py(client)
    body = _preview(client, py, _listing(ROSTER)).json()
    assert body["counts"]["additions"] == 0
    assert body["counts"]["changes"] == 0
    assert body["counts"]["deletions"] == 0
    assert body["counts"]["missing"] == 0
    assert body["counts"]["unchanged"] == 3


# ── Classification ──────────────────────────────────────────────────────────


def test_preview_derives_add_change_and_missing(client: TestClient) -> None:
    py = _py(client)
    content = _listing([
        ["A-9", "New Hire", "S9999999Z", "1995-05-05", "Executive", ""],
        ["A-1", "Anna Lim-Wong", "S1111111A", "1990-01-01", "Manager", ""],
        # A-2 and A-3 are simply absent — inferred, never applied by default.
    ])
    body = _preview(client, py, content).json()
    assert body["counts"]["additions"] == 1
    assert body["counts"]["changes"] == 1
    assert body["counts"]["deletions"] == 0
    assert body["counts"]["missing"] == 2
    assert body["counts"]["roster_total"] == 3
    fields = {d["field"] for d in body["changes"][0]["field_diffs"]}
    assert {"category", "employee_name"} & fields
    assert {op["staff_id"] for op in body["missing"]} == {"A-2", "A-3"}


def test_past_leaving_date_is_a_stated_termination(client: TestClient) -> None:
    py = _py(client)
    rows = [list(r) for r in ROSTER]
    rows[1][5] = PAST  # A-2 left
    body = _preview(client, py, _listing(rows)).json()
    assert body["counts"]["deletions"] == 1
    assert body["counts"]["missing"] == 0
    assert body["deletions"][0]["staff_id"] == "A-2"
    assert body["deletions"][0]["effective"] == PAST


def test_future_leaving_date_does_not_terminate(client: TestClient) -> None:
    """Someone on notice is still covered — terminating early would drop them
    from the insurer listing while the policy still owes them cover."""
    py = _py(client)
    rows = [list(r) for r in ROSTER]
    rows[1][5] = FUTURE
    body = _preview(client, py, _listing(rows)).json()
    assert body["counts"]["deletions"] == 0
    assert body["counts"]["changes"] == 1  # recorded as a field change instead


# ── Missing is opt-in ───────────────────────────────────────────────────────


def test_preview_does_not_mutate(client: TestClient) -> None:
    py = _py(client)
    _preview(client, py, _listing([ROSTER[0]]))
    assert _active("A-2").status == "active"
    assert _active("A-3").status == "active"


def test_apply_leaves_missing_people_alone_by_default(client: TestClient) -> None:
    """The safety property: a partial file must never end anyone's cover."""
    py = _py(client)
    res = _apply(client, py, _listing([ROSTER[0]]))
    assert res.status_code == 200, res.text
    assert res.json()["missing_terminated"] == 0
    assert _active("A-2").status == "active"
    assert _active("A-3").status == "active"


def test_apply_terminates_missing_only_on_opt_in(client: TestClient) -> None:
    py = _py(client)
    res = _apply(
        client, py, _listing([ROSTER[0], ROSTER[1]]), terminate_missing=True
    )
    assert res.status_code == 200, res.text
    assert res.json()["missing_terminated"] == 1
    a3 = _active("A-3")
    assert a3.status == EMPLOYEE_STATUS_TERMINATED
    assert a3.terminated_effective == business_today()
    assert _active("A-1").status == "active"
    assert _active("A-2").status == "active"


def test_terminated_person_is_not_missing_and_is_not_resurrected(
    client: TestClient,
) -> None:
    """A-3 is terminated. Absent, they must not re-appear as missing; present,
    an upload must neither reinstate them nor create a second row."""
    py = _py(client)
    body = _preview(client, py, _listing([ROSTER[0], ROSTER[1]])).json()
    assert body["counts"]["missing"] == 0

    body = _preview(client, py, _listing(ROSTER)).json()
    assert body["counts"]["additions"] == 0, "must not duplicate a leaver"
    assert body["counts"]["already_terminated"] == 1
    assert _active("A-3").status == EMPLOYEE_STATUS_TERMINATED


def test_employees_only_file_never_touches_dependants(client: TestClient) -> None:
    """A workbook with no Dependants sheet must not conclude every dependant
    left — missing-detection is scoped to the kinds the file actually covers."""
    py = _py(client)
    db = SessionLocal()
    try:
        db.add(
            Dependant(
                client_id=_active("A-1").client_id,
                policy_year_id=py,
                employee_id=_active("A-1").id,
                attribute_values={"dependant_name": "Solo Kid", "relationship": "Child"},
                link_method="staff_id",
            )
        )
        db.commit()
    finally:
        db.close()

    body = _preview(client, py, _listing([ROSTER[0], ROSTER[1]])).json()
    assert body["counts"]["missing"] == 0
    assert not any(op["record_type"] == "dependant" for op in body["missing"])


# ── Apply ───────────────────────────────────────────────────────────────────


def test_apply_adds_changes_and_terminates(client: TestClient) -> None:
    py = _py(client)
    rows = [
        ["A-1", "Anna Lim-Wong", "S1111111A", "1990-01-01", "Manager", ""],
        ["A-2", "Ben Ong", "S2222222B", "1985-02-02", "Manager", PAST],
        ["NEW-1", "New Hire", "S9999999Z", "1995-05-05", "Executive", ""],
    ]
    res = _apply(client, py, _listing(rows, [
        ["A-1", "Anna Lim-Wong", "Anna Kid", "T1512345Z", "Child", "2015-07-07", ""],
    ]))
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["added"] == 2  # 1 employee + 1 dependant
    assert body["changed"] == 1
    assert body["deleted"] == 1
    assert body["missing_terminated"] == 0

    new_hire = _active("NEW-1")
    assert new_hire.national_id_normalized == "S9999999Z"
    assert new_hire.source == "listing_sync"

    changed = _active("A-1")
    assert changed.employee_name == "Anna Lim-Wong"
    assert (changed.attribute_values or {}).get("category") == "Manager"

    left = _active("A-2")
    assert left.status == EMPLOYEE_STATUS_TERMINATED
    assert left.terminated_effective == date.fromisoformat(PAST)

    db = SessionLocal()
    try:
        kid = db.execute(
            select(Dependant).where(Dependant.national_id_normalized == "T1512345Z")
        ).scalar_one()
        assert kid.employee_id == changed.id
    finally:
        db.close()


def test_reapplying_the_same_file_changes_nothing(client: TestClient) -> None:
    py = _py(client)
    rows = [["NEW-1", "New Hire", "S9999999Z", "1995-05-05", "Executive", ""]]
    res = _apply(client, py, _listing(rows))
    assert res.status_code == 200
    body = res.json()
    assert body["added"] == 0 and body["changed"] == 0
    assert body["unchanged"] == 1
    db = SessionLocal()
    try:
        found = db.execute(
            select(Employee).where(Employee.national_id_normalized == "S9999999Z")
        ).scalars().all()
        assert len(found) == 1, "an unchanged row must not create a second person"
    finally:
        db.close()


def test_same_file_new_hire_dependant_links(client: TestClient) -> None:
    """A new employee + their dependant in one file: the dependant must link to
    the employee added in the same run, not land unlinked."""
    py = _py(client)
    res = _apply(client, py, _listing(
        [["NH-1", "New Parent", "S7777777G", "1992-09-09", "Executive", ""]],
        [["NH-1", "New Parent", "NH Kid", "T1712345Z", "Child", "2017-08-08", ""]],
    ))
    assert res.status_code == 200, res.text
    db = SessionLocal()
    try:
        kid = db.execute(
            select(Dependant).where(Dependant.national_id_normalized == "T1712345Z")
        ).scalar_one()
        assert kid.employee_id == _active("NH-1").id
    finally:
        db.close()


def test_dependant_leaving_date_terminates_the_dependant(client: TestClient) -> None:
    py = _py(client)
    res = _apply(client, py, _listing(
        [["NH-1", "New Parent", "S7777777G", "1992-09-09", "Executive", ""]],
        [["NH-1", "New Parent", "NH Kid", "T1712345Z", "Child", "2017-08-08", PAST]],
    ))
    assert res.status_code == 200, res.text
    assert res.json()["deleted"] == 1
    db = SessionLocal()
    try:
        kid = db.execute(
            select(Dependant).where(Dependant.national_id_normalized == "T1712345Z")
        ).scalar_one()
        assert kid.status == DEPENDANT_STATUS_TERMINATED
    finally:
        db.close()


# ── Identity guards ─────────────────────────────────────────────────────────


def test_row_without_identifier_is_flagged(client: TestClient) -> None:
    """A whitespace-only Staff ID survives the parser but identifies nobody."""
    py = _py(client)
    body = _preview(client, py, _listing(
        [["   ", "Nameless Row", "", "", "Executive", ""]]
    )).json()
    assert body["counts"]["additions"] == 0
    assert any("No Staff ID or NRIC" in i["message"] for i in body["issues"])


def test_rows_the_parser_drops_are_counted(client: TestClient) -> None:
    """A row with no Staff ID at all never reaches the diff — the parser drops
    it. Importing 1 of 2 rows and reporting success is how a roster goes quietly
    wrong, so the drop is counted and shown."""
    py = _py(client)
    body = _preview(client, py, _listing([
        ["", "No Staff Id", "", "1990-01-01", "Executive", ""],
        ["OK-1", "Has Staff Id", "", "1990-01-01", "Executive", ""],
    ])).json()
    assert body["counts"]["additions"] == 1
    assert body["counts"]["dropped_rows"] == 1


def test_row_naming_two_different_people_is_rejected(client: TestClient) -> None:
    """A-1's staff id with A-2's NRIC. Resolving on whichever key is checked
    first would rewrite the wrong person — which is what a mistyped NRIC looks
    like — so the row is refused instead."""
    py = _py(client)
    body = _preview(client, py, _listing(
        [["A-1", "Anna", "S2222222B", "1990-01-01", "Manager", ""]]
    )).json()
    assert body["counts"]["changes"] == 0
    assert any("two different employees" in i["message"] for i in body["issues"])


def test_one_fresh_nric_on_two_rows_is_rejected(client: TestClient) -> None:
    """Two rows assigning the SAME new NRIC to different people. There is no DB
    unique constraint to catch it, so the second row is skipped."""
    py = _py(client)
    body = _preview(client, py, _listing([
        ["A-1", "Anna", "S9090909Z", "1990-01-01", "Manager", ""],
        ["NEW-1", "New Hire", "S9090909Z", "1995-05-05", "Manager", ""],
    ])).json()
    assert any("two different employees in this file" in i["message"]
               for i in body["issues"])


def test_repeated_addition_in_one_file_is_flagged(client: TestClient) -> None:
    py = _py(client)
    body = _preview(client, py, _listing([
        ["DUP-1", "Twice Over", "S8888888H", "1991-04-04", "Executive", ""],
        ["DUP-1", "Twice Over", "S8888888H", "1991-04-04", "Executive", ""],
    ])).json()
    assert body["counts"]["additions"] == 1
    assert any("Repeated in this file" in i["message"] for i in body["issues"])


def test_terminated_excluded_from_default_list(client: TestClient) -> None:
    py = _py(client)
    listing = client.get(f"/api/v1/employees?policy_year_id={py}&limit=200").json()
    assert "A-2" not in {e["staff_id"] for e in listing["items"]}
    term = client.get(
        f"/api/v1/employees?policy_year_id={py}&limit=200&status=terminated"
    ).json()
    assert "A-2" in {e["staff_id"] for e in term["items"]}


# ── Round-trip fidelity ─────────────────────────────────────────────────────
#
# Everything below is a way the EXPORT and the PARSER disagreed about a value
# nobody edited. Each one put phantom changes in front of a broker who had
# changed one field — and the real edit is what gets lost in a list of 18.


def test_formula_guard_round_trips(client: TestClient) -> None:
    """`insurer_reports.safe_cell` prefixes a leading `+` with an apostrophe so
    Excel can't run it as a formula. The listing is re-uploaded now, so the
    parser must undo it — otherwise a Malaysian mobile reads as a change every
    single time, and applying one writes the apostrophe into the roster."""
    from app.services.insurer_reports import safe_cell
    from app.services.roster_parser import unescape_formula_guard

    guarded = safe_cell("+60186448967")
    assert guarded == "'+60186448967"
    assert unescape_formula_guard(guarded) == "+60186448967"
    # A value that legitimately starts with an apostrophe is untouched.
    assert unescape_formula_guard("'Reilly") == "'Reilly"


def test_comma_in_a_stored_name_is_not_a_change(client: TestClient) -> None:
    """`_normalize_name` collapses commas on the way in, so a stored
    "Lee Puay Tin, Ammie" re-reads without one. That is the parser's doing, not
    the broker's."""
    py = _py(client)
    db = SessionLocal()
    try:
        emp = db.execute(
            select(Employee).where(Employee.staff_id == "A-1")
        ).scalar_one()
        emp.employee_name = "Anna Lim-Wong, Junior"
        db.commit()
    finally:
        db.close()

    rows = [["A-1", "Anna Lim-Wong Junior", "S1111111A", "1990-01-01", "Manager", ""]]
    body = _preview(client, py, _listing(rows)).json()
    assert body["counts"]["changes"] == 0
    assert body["counts"]["unchanged"] == 1
    # The stored value keeps its comma — a round-trip must not be a write.
    assert _active("A-1").employee_name == "Anna Lim-Wong, Junior"


def test_float_tail_on_a_stored_code_is_not_a_change(client: TestClient) -> None:
    """Excel yields numeric cells as floats, so a mobile stored as 91234567.0
    re-reads as "91234567"."""
    py = _py(client)
    db = SessionLocal()
    try:
        emp = db.execute(
            select(Employee).where(Employee.staff_id == "A-1")
        ).scalar_one()
        emp.attribute_values = {**(emp.attribute_values or {}), "mobile": "91234567.0"}
        db.commit()
    finally:
        db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append([*EMP_COLS, "Mobile"])
    ws.append(["A-1", "Anna Lim-Wong Junior", "S1111111A", "1990-01-01",
               "Manager", "", "91234567"])
    buf = BytesIO()
    wb.save(buf)
    body = _preview(client, py, buf.getvalue()).json()
    assert body["counts"]["changes"] == 0


def test_two_identical_records_both_match(client: TestClient) -> None:
    """A roster really can hold two records with an IDENTICAL identity (CDL has
    two dependants sharing employee, name, DOB and relationship, with no NRIC).
    Indexing one per key made the second unmatchable: both file rows resolved to
    the first, and the second was reported MISSING FROM A FILE IT WAS IN — one
    tick away from terminating a live dependant."""
    py = _py(client)
    emp = _active("A-1")
    db = SessionLocal()
    try:
        for _ in range(2):
            db.add(
                Dependant(
                    client_id=emp.client_id,
                    policy_year_id=py,
                    employee_id=emp.id,
                    attribute_values={
                        "dependant_name": "Twin Kid",
                        "relationship": "Child",
                        "date_of_birth": "2020-05-05",
                    },
                    link_method="staff_id",
                )
            )
        db.commit()
    finally:
        db.close()

    twin = ["A-1", "Anna Lim-Wong Junior", "Twin Kid", "", "Child", "2020-05-05", ""]
    body = _preview(client, py, _listing([ROSTER[0]], [twin, twin])).json()
    assert not any(
        op["name"] == "Twin Kid" for op in body["missing"]
    ), "both identical dependants must match, not one be reported missing"
    assert body["counts"]["additions"] == 0, "must not add a third copy"


def test_file_repeating_a_person_more_often_than_the_roster_is_flagged(
    client: TestClient,
) -> None:
    """Three copies of a dependant the roster holds twice: the third is a
    repeated row, not a new person."""
    py = _py(client)
    twin = ["A-1", "Anna Lim-Wong Junior", "Twin Kid", "", "Child", "2020-05-05", ""]
    body = _preview(client, py, _listing([ROSTER[0]], [twin, twin, twin])).json()
    assert body["counts"]["additions"] == 0
    assert any("Repeated in this file" in i["message"] for i in body["issues"])


# ── Skipped rows must not become terminations ───────────────────────────────


def test_a_skipped_row_does_not_make_its_people_missing(client: TestClient) -> None:
    """The sharpest edge in this module.

    A row skipped as an ISSUE still NAMES people. Left out of the seen set they
    fell into "Not in this file" — which tells the broker they are "not named
    anywhere in this upload" — so one mistyped NRIC put both the intended
    employee AND the NRIC's real owner one tick away from termination.
    """
    py = _py(client)
    # A-1's staff id carrying A-3's NRIC: a conflict, skipped.
    rows = [
        ["A-1", "Anna", "S3333333C", "1990-01-01", "Manager", ""],
        ["A-3", "Cara Tan", "S3333333C", "1988-03-03", "Executive", ""],
    ]
    body = _preview(client, py, _listing(rows)).json()
    assert body["counts"]["issues"] >= 1
    named = {"A-1", "A-3"}
    assert not (named & {op["staff_id"] for op in body["missing"]}), (
        "a person named by a skipped row is not absent from the file"
    )


def test_dropped_rows_block_the_termination_opt_in(client: TestClient) -> None:
    """Absence is only evidence when the whole file was read. The preview
    reports the unreadable rows; the sheet withholds the tick on them."""
    py = _py(client)
    body = _preview(client, py, _listing([
        ["", "No Staff Id", "", "1990-01-01", "Executive", ""],
        ROSTER[0],
    ])).json()
    assert body["counts"]["dropped_rows"] == 1
    assert body["counts"]["missing"] > 0


# ── Round-trip fidelity, continued ──────────────────────────────────────────


def test_a_stale_leaving_date_does_not_re_terminate(client: TestClient) -> None:
    """`Last Day of Service` is a column of the listing template, so an ACTIVE
    employee carrying an old past date would be proposed for termination every
    time anyone downloaded the file and changed one salary. Only a NEWLY stated
    date terminates."""
    py = _py(client)
    db = SessionLocal()
    try:
        emp = db.execute(
            select(Employee).where(Employee.staff_id == "A-1")
        ).scalar_one()
        emp.attribute_values = {
            **(emp.attribute_values or {}),
            "last_day_of_service": PAST,
        }
        db.commit()
    finally:
        db.close()

    rows = [["A-1", "Anna Lim-Wong Junior", "S1111111A", "1990-01-01",
             "Manager", PAST]]
    body = _preview(client, py, _listing(rows)).json()
    assert body["counts"]["deletions"] == 0, "the date was already on file"
    assert _active("A-1").status == "active"

    # ...but changing it to a different past date IS a statement.
    newer = (business_today() - timedelta(days=5)).isoformat()
    rows[0][5] = newer
    body = _preview(client, py, _listing(rows)).json()
    assert body["counts"]["deletions"] == 1
    assert body["deletions"][0]["effective"] == newer


def test_insurer_member_ids_merge_instead_of_replacing(client: TestClient) -> None:
    """The template writes one column per insurer configured on the CURRENT
    year, so overwriting the dict wholesale permanently drops an id held under
    an insurer that has since been reconfigured."""
    py = _py(client)
    db = SessionLocal()
    try:
        emp = db.execute(
            select(Employee).where(Employee.staff_id == "A-1")
        ).scalar_one()
        emp.attribute_values = {
            **(emp.attribute_values or {}),
            "insurer_member_ids": {"AIA": "AAA-1", "Retired Insurer": "OLD-9"},
        }
        db.commit()
    finally:
        db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append([*EMP_COLS, "AIA Member ID"])
    ws.append(["A-1", "Anna Lim-Wong Junior", "S1111111A", "1990-01-01",
               "Manager", "", "AAA-2"])
    buf = BytesIO()
    wb.save(buf)
    res = _apply(client, py, buf.getvalue())
    assert res.status_code == 200, res.text

    ids = (_active("A-1").attribute_values or {}).get("insurer_member_ids") or {}
    assert ids.get("AIA") == "AAA-2", "the stated id updates"
    assert ids.get("Retired Insurer") == "OLD-9", "an unexported id survives"


def test_dependant_roster_never_imports_as_employees(tmp_path: Path) -> None:
    """Mirror of the employee-sheet guard: a single-sheet DEPENDANT roster
    dropped on the Employees upload falls back to sheet 0, where Staff ID and
    Employee Name resolve — so every row would import as an employee carrying
    the DEPENDANT's date of birth."""
    from app.services.roster_parser import parse_employee_workbook

    book = tmp_path / "dependants-only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Listing"
    ws.append(["Staff ID", "Employee Name", "Dependant Name", "Relationship",
               "Date of Birth"])
    ws.append(["A-1", "Anna Lim", "Kid One", "Child", "2015-01-01"])
    wb.save(book)

    assert parse_employee_workbook(book) == []


def test_apply_refuses_a_stale_termination_set(client: TestClient) -> None:
    """`terminate_missing` is confirmed against a list the broker READ, but
    apply re-runs the diff. Someone joining in between (a portal dependant
    approval, another broker's upload) would silently enlarge the set that
    actually gets terminated — so apply compares a digest and refuses."""
    py = _py(client)
    content = _listing([ROSTER[0]])
    preview = _preview(client, py, content).json()
    digest = preview["missing_digest"]
    assert digest, "the preview must fingerprint the set it showed"

    res = client.post(
        f"/api/v1/policy-years/{py}/adc/apply",
        files={"file": ("listing.xlsx", content, XLSX_MIME)},
        data={"terminate_missing": "true", "missing_digest": "not-the-same"},
    )
    assert res.status_code == 409, res.text
    assert res.json()["detail"]["code"] == "stale_listing_preview"

    # The matching digest goes through.
    res = client.post(
        f"/api/v1/policy-years/{py}/adc/apply",
        files={"file": ("listing.xlsx", content, XLSX_MIME)},
        data={"terminate_missing": "true", "missing_digest": digest},
    )
    assert res.status_code == 200, res.text


# ── Advisory checks ─────────────────────────────────────────────────────────


def test_a_mistyped_nric_warns_without_blocking_the_row(client: TestClient) -> None:
    """The checksum advisory has to reach the path brokers actually USE.

    It was first wired only into `POST /employees/upload`, which no UI calls —
    the roster is changed through this listing sync — so it was unreachable in
    practice. A warning is also not an issue: the row still applies.
    """
    py = _py(client)
    content = _listing([
        # S2222222H is the valid form, so H -> Q is a single-character typo.
        ["A-8", "Typo Tan", "S2222222Q", "1990-02-02", "Executive", ""],
        # Correct checksum — must NOT warn.
        ["A-9", "Clean Chua", "S1111111D", "1991-03-03", "Executive", ""],
    ])
    body = _preview(client, py, content).json()

    warnings = [w for w in body["warnings"] if w["record_type"] == "employee"]
    assert len(warnings) == 1, warnings
    assert "checksum" in warnings[0]["message"]
    # Masked in transit — a wrong ID is still personal data.
    assert "S2222222Q" not in warnings[0]["message"]
    # Advisory, not a refusal: the row is still proposed, and it is NOT an issue.
    assert any(op["staff_id"] == "A-8" for op in body["additions"])
    assert all("checksum" not in i["message"] for i in body["issues"])


def test_a_non_nric_identifier_never_warns(client: TestClient) -> None:
    """A foreign passport or work-pass number is not an NRIC at all, so it fails
    the checksum for a reason that is not an error. Warning on those would fire
    for every foreign hire and train brokers to ignore the advisory."""
    py = _py(client)
    content = _listing([
        ["A-7", "Overseas Ong", "E1234567X", "1992-04-04", "Executive", ""],
    ])
    body = _preview(client, py, content).json()
    assert body["warnings"] == []



def test_a_login_user_id_never_steals_the_staff_id_column() -> None:
    """`User ID` is the incumbent's name for the staff id — but plenty of HR
    extracts carry a login "User ID" NEXT TO the real "Staff ID".

    The staff id is the roster's primary key: bind the wrong column and every
    employee arrives keyed by their login, so ADC matches nobody, proposes the
    whole roster as joiners and the people already on file as missing. The
    fallback spelling must lose to the primary one whatever order they sit in.
    """
    from app.services.roster_parser import EMPLOYEE_COLUMN_MAP, _build_column_map

    both = ["Entity", "User ID", "Employee Name", "Staff ID", "Date of Birth"]
    mapped = _build_column_map(both, EMPLOYEE_COLUMN_MAP)
    assert mapped[3] == "staff_id", mapped
    assert mapped.get(1) != "staff_id", mapped
    # Every other column still binds — the fallback loses its attribute, not
    # the whole pass.
    assert mapped[0] == "entity"
    assert mapped[2] == "employee_name"
    assert mapped[4] == "date_of_birth"

    # Alone, it IS the staff id — the incumbent's own export has no other.
    only = ["Entity", "User ID", "Employee Name", "Date of Birth"]
    assert _build_column_map(only, EMPLOYEE_COLUMN_MAP)[1] == "staff_id"
